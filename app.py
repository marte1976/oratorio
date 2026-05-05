from __future__ import annotations

import html
import hashlib
import hmac
import json
import mimetypes
import os
import re
import secrets
import shutil
import sqlite3
import subprocess
import uuid
import zipfile
import unicodedata
from calendar import monthrange
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from functools import lru_cache
from http.cookies import SimpleCookie
from io import BytesIO
from pathlib import Path
from socketserver import ThreadingMixIn
from urllib.parse import parse_qs, quote, unquote, urlencode, urlsplit, urlunsplit
from wsgiref.simple_server import WSGIServer, make_server


BASE_DIR = Path(__file__).resolve().parent
VERSION_FILE = BASE_DIR / "VERSION.txt"
DEFAULT_DB_PATH = BASE_DIR / "database" / "gestione_associazione.sqlite"
DB_PATH = Path(os.environ.get("ASSOCIAZIONE_DB_PATH", str(DEFAULT_DB_PATH)))
SCHEMA_PATH = BASE_DIR / "database" / "schema_associazione.sql"
STATIC_DIR = BASE_DIR / "static"
COMUNI_JSON_PATH = BASE_DIR / "data" / "comuni.json"
GUIDE_MEDIA_JSON_PATH = BASE_DIR / "data" / "guida_multimedia.json"
LOGO_URL = "/static/logo-ca.jpg"
APP_NAME = "Oratorio Carlo Acutis"
APP_SUBTITLE = ""
ESTATE_LABEL = "Campo estivo"
ORATORIO_LABEL = "Oratorio"
OUTPUT_DIR = BASE_DIR / "outputs"
UPDATE_DIR = OUTPUT_DIR / "aggiornamenti"
ASSOCIATI_IMPORT_DIR = OUTPUT_DIR / "import-associati"
TUTORIAL_VIDEO_DIR = OUTPUT_DIR / "tutorial"
ACTIVITY_LOG_XLS_PATH = OUTPUT_DIR / "registro_attivita.xls"
ACTIVITY_LOG_RETENTION_YEARS = 2
BACKUP_DIR = OUTPUT_DIR / "backups"
BACKUP_ROOT_NAME = "OratorioCarloAcutis"
BACKUP_EXCLUDED_TOP_LEVEL = {"dist", "__pycache__", ".git", ".pytest_cache", ".mypy_cache"}
EXPORT_SCRIPT = BASE_DIR / "scripts" / "export_report.mjs"
TUTORIAL_VIDEO_SCRIPT = BASE_DIR / "scripts" / "genera_tutorial_video.ps1"
LOCAL_RUNTIME_ROOT = BASE_DIR / "runtime"
LOCAL_NODE_BIN = LOCAL_RUNTIME_ROOT / "node" / "bin" / "node.exe"
LOCAL_ARTIFACT_TOOL_MODULE = (
    LOCAL_RUNTIME_ROOT
    / "node"
    / "node_modules"
    / "@oai"
    / "artifact-tool"
    / "dist"
    / "artifact_tool.mjs"
)
DEFAULT_NODE_BIN = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "node"
    / "bin"
    / "node.exe"
)
DEFAULT_ARTIFACT_TOOL_MODULE = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "node"
    / "node_modules"
    / "@oai"
    / "artifact-tool"
    / "dist"
    / "artifact_tool.mjs"
)
NODE_BIN = Path(
    os.environ.get(
        "ASSOCIAZIONE_NODE_BIN",
        str(LOCAL_NODE_BIN if LOCAL_NODE_BIN.exists() else DEFAULT_NODE_BIN),
    )
)
ARTIFACT_TOOL_MODULE = Path(
    os.environ.get(
        "ASSOCIAZIONE_ARTIFACT_TOOL_MODULE",
        str(LOCAL_ARTIFACT_TOOL_MODULE if LOCAL_ARTIFACT_TOOL_MODULE.exists() else DEFAULT_ARTIFACT_TOOL_MODULE),
    )
)


def load_app_version() -> str:
    try:
        value = VERSION_FILE.read_text(encoding="utf-8").strip()
        return value or "2026.05.02.1"
    except OSError:
        return "2026.05.02.1"


APP_VERSION = load_app_version()
DEFAULT_DOCUMENT_TYPE = "Carta d'identit\u00e0"
DEFAULT_DOCUMENT_TYPE_SQL = DEFAULT_DOCUMENT_TYPE.replace("'", "''")
DOCUMENT_TYPE_VALUES = (DEFAULT_DOCUMENT_TYPE, "Patente di guida")

PROGRESSIVE_ENTITIES = {
    "associati": {
        "table": "associati",
        "column": "numero_progressivo",
        "code_column": "codice_associato",
        "prefix": "ASS",
        "label": "Associato",
    },
    "corsi": {
        "table": "corsi",
        "column": "numero_progressivo",
        "code_column": "codice_corso",
        "prefix": "COR",
        "label": "Corso",
    },
    "campi_estivi": {
        "table": "campi_estivi",
        "column": "numero_progressivo",
        "code_column": "codice_campo",
        "prefix": "CES",
        "label": ESTATE_LABEL,
    },
    "eventi": {
        "table": "eventi",
        "column": "numero_progressivo",
        "code_column": "codice_evento",
        "prefix": "EVT",
        "label": "Evento",
    },
    "oratorio": {
        "table": "oratorio",
        "column": "numero_progressivo",
        "code_column": "codice_oratorio",
        "prefix": "ORA",
        "label": ORATORIO_LABEL,
    },
}

LOCKED_CRUD_ENTITIES = {
    "tipologie_corsi": "Le tipologie corsi sono fisse e non possono essere modificate.",
    "campi_estivi": "Il Campo estivo annuale non puo essere modificato o eliminato da questa area.",
}

MONTH_NAMES = {
    1: "Gennaio",
    2: "Febbraio",
    3: "Marzo",
    4: "Aprile",
    5: "Maggio",
    6: "Giugno",
    7: "Luglio",
    8: "Agosto",
    9: "Settembre",
    10: "Ottobre",
    11: "Novembre",
    12: "Dicembre",
}

CARICA_VALUES = (
    "Associato",
    "Presidente",
    "Vice Presidente",
    "Segretario",
    "Tesoriere",
    "Consigliere",
    "Consigliere spirituale",
)
CONSIGLIO_DIRETTIVO_ORDER = (
    "Presidente",
    "Vice Presidente",
    "Tesoriere",
    "Segretario",
    "Consigliere",
    "Consigliere spirituale",
)
CONSIGLIO_DIRETTIVO_LABELS = {
    "Presidente": "Presidente",
    "Vice Presidente": "Vice Presidente",
    "Tesoriere": "Tesoriere",
    "Segretario": "Segretario",
    "Consigliere": "Consigliere",
    "Consigliere spirituale": "Consigliere Spirituale",
}

REPORT_RECIPIENT_CARICHE = CARICA_VALUES[1:]
SESSION_COOKIE_NAME = "oratorio_session"
SESSION_DURATION_DAYS = 30
PASSWORD_HASH_ITERATIONS = 210000
ADMIN_ONLY_REPORT_KEYS = {"registro-attivita"}

ASSOCIATI_IMPORT_FIELDS: list[tuple[str, str]] = [
    ("nome", "Nome"),
    ("cognome", "Cognome"),
    ("codice_fiscale", "Codice fiscale"),
    ("data_nascita", "Data nascita"),
    ("sesso", "Sesso"),
    ("comune_nascita", "Comune di nascita"),
    ("provincia_nascita", "Provincia di nascita"),
    ("email", "Email"),
    ("telefono", "Cellulare"),
    ("indirizzo", "Indirizzo"),
    ("cap", "CAP"),
    ("citta", "CittÃ "),
    ("provincia", "Provincia"),
    ("impiego", "Impiego"),
    ("data_prima_iscrizione", "Data prima iscrizione"),
    ("stato_associato", "Stato associato"),
    ("carica", "Carica"),
    ("liberatoria_video", "Liberatoria Video"),
    ("patologie", "Patologie, allergie, intolleranze alimentari ed eventuali terapie in corso"),
    ("genitore_tutore_nome", "Genitore/Tutore nome"),
    ("genitore_tutore_cognome", "Genitore/Tutore cognome"),
    ("genitore_tutore_cellulare", "Genitore/Tutore cellulare"),
    ("genitore_tutore_email", "Genitore/Tutore email"),
    ("genitore_tutore_impiego", "Genitore/Tutore impiego"),
    ("genitore_tutore_tipo_documento", "Genitore/Tutore tipo documento"),
    ("genitore_tutore_numero_documento", "Genitore/Tutore numero documento"),
    ("prelievo_altro_genitore_nome", "Altro genitore nome"),
    ("prelievo_altro_genitore_cognome", "Altro genitore cognome"),
    ("prelievo_altro_genitore_cellulare", "Altro genitore cellulare"),
    ("prelievo_altro_genitore_impiego", "Altro genitore impiego"),
    ("prelievo_altro_genitore_tipo_documento", "Altro genitore tipo documento"),
    ("prelievo_altro_genitore_numero_documento", "Altro genitore numero documento"),
    ("prelievo_altra_persona_nome", "Altra persona nome"),
    ("prelievo_altra_persona_cognome", "Altra persona cognome"),
    ("prelievo_altra_persona_cellulare", "Altra persona cellulare"),
    ("prelievo_altra_persona_tipo_documento", "Altra persona tipo documento"),
    ("prelievo_altra_persona_numero_documento", "Altra persona numero documento"),
    ("note", "Note"),
]

CODICE_FISCALE_MONTHS = {
    "A": 1,
    "B": 2,
    "C": 3,
    "D": 4,
    "E": 5,
    "H": 6,
    "L": 7,
    "M": 8,
    "P": 9,
    "R": 10,
    "S": 11,
    "T": 12,
}

CODICE_FISCALE_MONTH_CODES = {value: key for key, value in CODICE_FISCALE_MONTHS.items()}

CODICE_FISCALE_ODD_MAP = {
    "0": 1,
    "1": 0,
    "2": 5,
    "3": 7,
    "4": 9,
    "5": 13,
    "6": 15,
    "7": 17,
    "8": 19,
    "9": 21,
    "A": 1,
    "B": 0,
    "C": 5,
    "D": 7,
    "E": 9,
    "F": 13,
    "G": 15,
    "H": 17,
    "I": 19,
    "J": 21,
    "K": 2,
    "L": 4,
    "M": 18,
    "N": 20,
    "O": 11,
    "P": 3,
    "Q": 6,
    "R": 8,
    "S": 12,
    "T": 14,
    "U": 16,
    "V": 10,
    "W": 22,
    "X": 25,
    "Y": 24,
    "Z": 23,
}

CODICE_FISCALE_EVEN_MAP = {
    **{str(number): number for number in range(10)},
    **{chr(ord("A") + index): index for index in range(26)},
}

CODICE_FISCALE_CONTROL_CHARS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def associato_base_name_sql(alias: str = "a") -> str:
    prefix = f"{alias}." if alias else ""
    return f"trim(COALESCE({prefix}cognome, '') || ' ' || COALESCE({prefix}nome, ''))"


def associato_display_sql(alias: str = "a") -> str:
    prefix = f"{alias}." if alias else ""
    base_name = associato_base_name_sql(alias)
    return (
        f"{base_name} || CASE "
        f"WHEN COALESCE({prefix}data_nascita, '') <> '' AND julianday({prefix}data_nascita) IS NOT NULL "
        f"THEN ' (' || CAST((julianday('now') - julianday({prefix}data_nascita)) / 365.2425 AS INTEGER) || ' anni)' "
        f"ELSE '' END"
    )


REPORTING_VIEWS_SQL = f"""
DROP VIEW IF EXISTS v_incassi_totali;
DROP VIEW IF EXISTS v_riepilogo_associati;
DROP VIEW IF EXISTS v_scadenze_da_incassare;
DROP VIEW IF EXISTS v_eventi_saldo;
DROP VIEW IF EXISTS v_campi_estivi_saldo;
DROP VIEW IF EXISTS v_oratorio_saldo;
DROP VIEW IF EXISTS v_rate_corsi_saldo;
DROP VIEW IF EXISTS v_iscrizioni_corsi_saldo;
DROP VIEW IF EXISTS v_tesseramenti_saldo;

CREATE VIEW v_tesseramenti_saldo AS
SELECT
    t.id,
    t.anno_sociale,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    t.data_tesseramento,
    t.data_scadenza,
    t.importo_dovuto,
    COALESCE(SUM(pt.importo), 0) AS importo_pagato,
    t.importo_dovuto - COALESCE(SUM(pt.importo), 0) AS saldo_residuo,
    CASE
        WHEN COALESCE(SUM(pt.importo), 0) >= t.importo_dovuto THEN 'Pagato'
        WHEN COALESCE(SUM(pt.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento
FROM tesseramenti_annuali t
JOIN associati a ON a.id = t.associato_id
LEFT JOIN pagamenti_tesseramenti pt ON pt.tesseramento_id = t.id
GROUP BY t.id;

CREATE VIEW v_iscrizioni_corsi_saldo AS
SELECT
    ic.id,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    tc.codice_tipologia,
    COALESCE(tc.nome, 'Senza tipologia') AS tipologia_corso,
    c.codice_corso,
    c.nome AS corso,
    ic.data_iscrizione,
    ic.data_inizio,
    ic.data_fine,
    ic.stato_iscrizione,
    ic.quota_iscrizione,
    COALESCE(SUM(pic.importo), 0) AS importo_pagato_iscrizione,
    ic.quota_iscrizione - COALESCE(SUM(pic.importo), 0) AS saldo_iscrizione,
    CASE
        WHEN COALESCE(SUM(pic.importo), 0) >= ic.quota_iscrizione THEN 'Pagato'
        WHEN COALESCE(SUM(pic.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento_iscrizione,
    ic.quota_mensile
FROM iscrizioni_corsi ic
JOIN associati a ON a.id = ic.associato_id
JOIN corsi c ON c.id = ic.corso_id
LEFT JOIN tesseramenti_annuali t
    ON t.associato_id = ic.associato_id
   AND t.anno_sociale = CAST(substr(COALESCE(ic.data_iscrizione, ic.data_inizio, ''), 1, 4) AS INTEGER)
LEFT JOIN tipologie_corsi tc ON tc.id = c.tipologia_corso_id
LEFT JOIN pagamenti_iscrizioni_corsi pic ON pic.iscrizione_corso_id = ic.id
GROUP BY ic.id;

CREATE VIEW v_rate_corsi_saldo AS
SELECT
    r.id,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    tc.codice_tipologia,
    COALESCE(tc.nome, 'Senza tipologia') AS tipologia_corso,
    c.codice_corso,
    c.nome AS corso,
    r.anno,
    r.mese,
    printf('%04d-%02d', r.anno, r.mese) AS competenza,
    r.data_scadenza,
    r.importo_dovuto,
    COALESCE(SUM(prc.importo), 0) AS importo_pagato,
    r.importo_dovuto - COALESCE(SUM(prc.importo), 0) AS saldo_residuo,
    CASE
        WHEN COALESCE(SUM(prc.importo), 0) >= r.importo_dovuto THEN 'Pagato'
        WHEN COALESCE(SUM(prc.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento
FROM rate_corsi_mensili r
JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
JOIN associati a ON a.id = ic.associato_id
JOIN corsi c ON c.id = ic.corso_id
LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
LEFT JOIN tipologie_corsi tc ON tc.id = c.tipologia_corso_id
LEFT JOIN pagamenti_rate_corsi prc ON prc.rata_corso_id = r.id
GROUP BY r.id;

CREATE VIEW v_campi_estivi_saldo AS
SELECT
    ice.id,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    ce.codice_campo,
    ce.nome AS campo_estivo,
    ce.anno,
    ce.data_inizio,
    ce.data_fine,
    ice.data_iscrizione,
    ice.stato_iscrizione,
    ice.quota_partecipazione AS importo_dovuto,
    COALESCE(SUM(pce.importo), 0) AS importo_pagato,
    ice.quota_partecipazione - COALESCE(SUM(pce.importo), 0) AS saldo_residuo,
    CASE
        WHEN COALESCE(SUM(pce.importo), 0) >= ice.quota_partecipazione THEN 'Pagato'
        WHEN COALESCE(SUM(pce.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento
FROM iscrizioni_campi_estivi ice
JOIN associati a ON a.id = ice.associato_id
JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
LEFT JOIN pagamenti_campi_estivi pce ON pce.iscrizione_campo_id = ice.id
GROUP BY ice.id;

CREATE VIEW v_oratorio_saldo AS
SELECT
    io.id,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    o.codice_oratorio,
    o.nome AS oratorio,
    o.anno,
    o.data_inizio,
    o.data_fine,
    io.data_iscrizione,
    io.stato_iscrizione,
    io.quota_partecipazione AS importo_dovuto,
    COALESCE(SUM(po.importo), 0) AS importo_pagato,
    io.quota_partecipazione - COALESCE(SUM(po.importo), 0) AS saldo_residuo,
    CASE
        WHEN COALESCE(SUM(po.importo), 0) >= io.quota_partecipazione THEN 'Pagato'
        WHEN COALESCE(SUM(po.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento
FROM iscrizioni_oratorio io
JOIN associati a ON a.id = io.associato_id
JOIN oratorio o ON o.id = io.oratorio_id
LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
LEFT JOIN pagamenti_oratorio po ON po.iscrizione_oratorio_id = io.id
GROUP BY io.id;

CREATE VIEW v_eventi_saldo AS
SELECT
    ie.id,
    t.codice_tesseramento,
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    e.codice_evento,
    e.nome AS evento,
    e.tipologia,
    e.data_evento,
    ie.data_iscrizione,
    ie.stato_iscrizione,
    ie.quota_partecipazione AS importo_dovuto,
    COALESCE(SUM(pe.importo), 0) AS importo_pagato,
    ie.quota_partecipazione - COALESCE(SUM(pe.importo), 0) AS saldo_residuo,
    CASE
        WHEN COALESCE(SUM(pe.importo), 0) >= ie.quota_partecipazione THEN 'Pagato'
        WHEN COALESCE(SUM(pe.importo), 0) > 0 THEN 'Parziale'
        ELSE 'Da pagare'
    END AS stato_pagamento
FROM iscrizioni_eventi ie
JOIN associati a ON a.id = ie.associato_id
JOIN eventi e ON e.id = ie.evento_id
LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
LEFT JOIN pagamenti_eventi pe ON pe.iscrizione_evento_id = ie.id
GROUP BY ie.id;

CREATE VIEW v_scadenze_da_incassare AS
SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Tesseramento annuale' AS area,
    'Anno ' || anno_sociale AS riferimento,
    data_scadenza AS scadenza,
    importo_dovuto,
    importo_pagato,
    saldo_residuo,
    stato_pagamento
FROM v_tesseramenti_saldo
WHERE saldo_residuo > 0

UNION ALL

SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Corso - iscrizione' AS area,
    corso AS riferimento,
    data_iscrizione AS scadenza,
    quota_iscrizione AS importo_dovuto,
    importo_pagato_iscrizione AS importo_pagato,
    saldo_iscrizione AS saldo_residuo,
    stato_pagamento_iscrizione AS stato_pagamento
FROM v_iscrizioni_corsi_saldo
WHERE saldo_iscrizione > 0

UNION ALL

SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Corso - quota mensile' AS area,
    corso || ' ' || competenza AS riferimento,
    data_scadenza AS scadenza,
    importo_dovuto,
    importo_pagato,
    saldo_residuo,
    stato_pagamento
FROM v_rate_corsi_saldo
WHERE saldo_residuo > 0

UNION ALL

SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Campo estivo' AS area,
    campo_estivo AS riferimento,
    data_inizio AS scadenza,
    importo_dovuto,
    importo_pagato,
    saldo_residuo,
    stato_pagamento
FROM v_campi_estivi_saldo
WHERE saldo_residuo > 0

UNION ALL

SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Oratorio' AS area,
    oratorio AS riferimento,
    data_inizio AS scadenza,
    importo_dovuto,
    importo_pagato,
    saldo_residuo,
    stato_pagamento
FROM v_oratorio_saldo
WHERE saldo_residuo > 0

UNION ALL

SELECT
    associato_id,
    codice_tesseramento,
    associato,
    data_nascita,
    'Evento' AS area,
    evento AS riferimento,
    data_evento AS scadenza,
    importo_dovuto,
    importo_pagato,
    saldo_residuo,
    stato_pagamento
FROM v_eventi_saldo
WHERE saldo_residuo > 0;

CREATE VIEW v_riepilogo_associati AS
WITH movimenti AS (
    SELECT associato_id, importo_dovuto, importo_pagato
    FROM v_tesseramenti_saldo
    UNION ALL
    SELECT associato_id, quota_iscrizione AS importo_dovuto, importo_pagato_iscrizione AS importo_pagato
    FROM v_iscrizioni_corsi_saldo
    UNION ALL
    SELECT associato_id, importo_dovuto, importo_pagato
    FROM v_rate_corsi_saldo
    UNION ALL
    SELECT associato_id, importo_dovuto, importo_pagato
    FROM v_campi_estivi_saldo
    UNION ALL
    SELECT associato_id, importo_dovuto, importo_pagato
    FROM v_oratorio_saldo
    UNION ALL
    SELECT associato_id, importo_dovuto, importo_pagato
    FROM v_eventi_saldo
)
SELECT
    a.id AS associato_id,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    a.stato_associato,
    COALESCE(SUM(m.importo_dovuto), 0) AS totale_dovuto,
    COALESCE(SUM(m.importo_pagato), 0) AS totale_pagato,
    COALESCE(SUM(m.importo_dovuto), 0) - COALESCE(SUM(m.importo_pagato), 0) AS saldo_residuo
FROM associati a
LEFT JOIN movimenti m ON m.associato_id = a.id
GROUP BY a.id;

CREATE VIEW v_incassi_totali AS
SELECT
    'Tesseramento annuale' AS area,
    pt.data_pagamento,
    pt.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    'Anno ' || t.anno_sociale AS riferimento
FROM pagamenti_tesseramenti pt
JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
JOIN associati a ON a.id = t.associato_id
LEFT JOIN metodi_pagamento mp ON mp.id = pt.metodo_pagamento_id

UNION ALL

SELECT
    'Corso - iscrizione' AS area,
    pic.data_pagamento,
    pic.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    c.nome AS riferimento
FROM pagamenti_iscrizioni_corsi pic
JOIN iscrizioni_corsi ic ON ic.id = pic.iscrizione_corso_id
JOIN associati a ON a.id = ic.associato_id
JOIN corsi c ON c.id = ic.corso_id
LEFT JOIN metodi_pagamento mp ON mp.id = pic.metodo_pagamento_id

UNION ALL

SELECT
    'Corso - quota mensile' AS area,
    prc.data_pagamento,
    prc.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    c.nome || ' ' || printf('%04d-%02d', r.anno, r.mese) AS riferimento
FROM pagamenti_rate_corsi prc
JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
JOIN associati a ON a.id = ic.associato_id
JOIN corsi c ON c.id = ic.corso_id
LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id

UNION ALL

SELECT
    'Campo estivo' AS area,
    pce.data_pagamento,
    pce.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    ce.nome AS riferimento
FROM pagamenti_campi_estivi pce
JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
JOIN associati a ON a.id = ice.associato_id
JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
LEFT JOIN metodi_pagamento mp ON mp.id = pce.metodo_pagamento_id

UNION ALL

SELECT
    'Oratorio' AS area,
    po.data_pagamento,
    po.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    o.nome AS riferimento
FROM pagamenti_oratorio po
JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
JOIN associati a ON a.id = io.associato_id
JOIN oratorio o ON o.id = io.oratorio_id
LEFT JOIN metodi_pagamento mp ON mp.id = po.metodo_pagamento_id

UNION ALL

SELECT
    'Evento' AS area,
    pe.data_pagamento,
    pe.importo,
    mp.nome AS metodo_pagamento,
    a.codice_associato,
    {associato_display_sql('a')} AS associato,
    a.data_nascita,
    e.nome AS riferimento
FROM pagamenti_eventi pe
JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
JOIN associati a ON a.id = ie.associato_id
JOIN eventi e ON e.id = ie.evento_id
LEFT JOIN metodi_pagamento mp ON mp.id = pe.metodo_pagamento_id;
"""


def rebuild_reporting_views(connection: sqlite3.Connection) -> None:
    connection.executescript(REPORTING_VIEWS_SQL)


def initialize_database_if_missing() -> None:
    if DB_PATH.exists() or not SCHEMA_PATH.exists():
        return

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    try:
        connection.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
        connection.commit()
    finally:
        connection.close()


def table_columns(connection: sqlite3.Connection, table_name: str) -> set[str]:
    rows = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return {row[1] for row in rows}


def ensure_column(connection: sqlite3.Connection, table_name: str, column_sql: str, column_name: str) -> None:
    if column_name in table_columns(connection, table_name):
        return
    connection.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_sql}")


def ensure_quote_predefinite_supports_oratorio(connection: sqlite3.Connection) -> None:
    row = connection.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'quote_predefinite'"
    ).fetchone()
    if row is None:
        return
    table_sql = str(row["sql"] or "")
    if "'oratorio'" in table_sql:
        return

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS quote_predefinite_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            area TEXT NOT NULL CHECK (area IN ('tesseramenti', 'campi-estivi', 'oratorio')),
            descrizione TEXT NOT NULL,
            importo NUMERIC NOT NULL CHECK (importo >= 0),
            attiva INTEGER NOT NULL DEFAULT 1 CHECK (attiva IN (0, 1)),
            note TEXT,
            creato_il TEXT NOT NULL DEFAULT (datetime('now')),
            aggiornato_il TEXT NOT NULL DEFAULT (datetime('now'))
        )
        """
    )
    connection.execute(
        """
        INSERT INTO quote_predefinite_new (id, area, descrizione, importo, attiva, note, creato_il, aggiornato_il)
        SELECT id, area, descrizione, importo, attiva, note, creato_il, aggiornato_il
        FROM quote_predefinite
        """
    )
    connection.execute("DROP TABLE quote_predefinite")
    connection.execute("ALTER TABLE quote_predefinite_new RENAME TO quote_predefinite")


def cleanup_expired_sessions(connection: sqlite3.Connection) -> None:
    connection.execute("DELETE FROM sessioni_accesso WHERE scade_il <= ?", (current_timestamp(),))


def access_users_count() -> int:
    return int(fetch_scalar("SELECT COUNT(*) FROM utenti_accesso") or 0)


def bootstrap_admin_required() -> bool:
    return access_users_count() == 0


def create_access_user(
    connection: sqlite3.Connection,
    *,
    username: str,
    password: str,
    is_admin: bool,
    email_recupero: str | None = None,
) -> int:
    password_salt, password_hash, iterations = hash_password(password)
    cursor = connection.execute(
        """
        INSERT INTO utenti_accesso (
            username, password_hash, password_salt, password_iterations, is_admin, email_recupero
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            validate_username(username),
            password_hash,
            password_salt,
            iterations,
            1 if is_admin else 0,
            (email_recupero or "").strip() or None,
        ),
    )
    return int(cursor.lastrowid)


def set_access_user_password(connection: sqlite3.Connection, user_id: int, password: str) -> None:
    password_salt, password_hash, iterations = hash_password(password)
    connection.execute(
        """
        UPDATE utenti_accesso
        SET password_hash = ?, password_salt = ?, password_iterations = ?, aggiornato_il = ?
        WHERE id = ?
        """,
        (password_hash, password_salt, iterations, current_timestamp(), user_id),
    )


def access_user_row(user_id: int) -> sqlite3.Row | None:
    return fetch_one(
        """
        SELECT
            id,
            username,
            email_recupero,
            is_admin,
            attivo,
            creato_il,
            aggiornato_il,
            COALESCE(ultimo_accesso, '') AS ultimo_accesso
        FROM utenti_accesso
        WHERE id = ?
        """,
        (user_id,),
    )


def access_user_by_username(username: str) -> sqlite3.Row | None:
    return fetch_one(
        """
        SELECT *
        FROM utenti_accesso
        WHERE username = ?
        """,
        (validate_username(username),),
    )


def validate_recovery_email(email_value: str) -> str:
    email = (email_value or "").strip()
    if not email or "@" not in email:
        raise ValueError("Indica un'email di recupero valida.")
    return email


def create_user_session(connection: sqlite3.Connection, user_id: int) -> str:
    cleanup_expired_sessions(connection)
    token = secrets.token_urlsafe(32)
    expires_at = (datetime.now() + timedelta(days=SESSION_DURATION_DAYS)).replace(microsecond=0).isoformat(sep=" ")
    now_value = current_timestamp()
    connection.execute(
        """
        INSERT INTO sessioni_accesso (session_token, utente_id, scade_il)
        VALUES (?, ?, ?)
        """,
        (token, user_id, expires_at),
    )
    connection.execute(
        """
        UPDATE utenti_accesso
        SET ultimo_accesso = ?, aggiornato_il = ?
        WHERE id = ?
        """,
        (now_value, now_value, user_id),
    )
    return token


def current_user_from_cookies(cookies: dict[str, str]) -> dict[str, object] | None:
    token = cookies.get(SESSION_COOKIE_NAME, "")
    if not token:
        return None

    with get_connection() as connection:
        cleanup_expired_sessions(connection)
        row = connection.execute(
            """
            SELECT
                u.id,
                u.username,
                u.is_admin,
                u.attivo
            FROM sessioni_accesso s
            JOIN utenti_accesso u ON u.id = s.utente_id
            WHERE s.session_token = ?
              AND s.scade_il > ?
            """,
            (token, current_timestamp()),
        ).fetchone()
        if row is None:
            connection.commit()
            return None
        if not row["attivo"]:
            connection.execute("DELETE FROM sessioni_accesso WHERE session_token = ?", (token,))
            connection.commit()
            return None
        connection.commit()
        return {
            "id": int(row["id"]),
            "username": row["username"],
            "is_admin": bool(row["is_admin"]),
        }


def login_destination(target: str | None) -> str:
    candidate = (target or "").strip()
    if not candidate.startswith("/") or candidate.startswith("//"):
        return "/"
    return candidate


def format_progressive_code(entity_key: str, number: int) -> str:
    prefix = PROGRESSIVE_ENTITIES[entity_key]["prefix"]
    return f"{prefix}-{number:04d}"


def format_tesseramento_code(anno_sociale: int, number: int) -> str:
    return f"{anno_sociale % 100:02d}/{number}"


def resolve_tesseramento_progressive_number(
    connection: sqlite3.Connection,
    anno_sociale: int,
    *,
    preferred_number: int | None = None,
    exclude_id: int | None = None,
) -> int:
    params: list[object] = [anno_sociale]
    query = """
        SELECT numero_progressivo_anno
        FROM tesseramenti_annuali
        WHERE anno_sociale = ?
          AND numero_progressivo_anno IS NOT NULL
    """
    if exclude_id is not None:
        query += " AND id <> ?"
        params.append(exclude_id)
    used_numbers = {
        int(row[0])
        for row in connection.execute(query, tuple(params)).fetchall()
        if row[0] is not None and int(row[0]) > 0
    }
    if preferred_number is not None and preferred_number > 0 and preferred_number not in used_numbers:
        return preferred_number
    number = 1
    while number in used_numbers:
        number += 1
    return number


def assign_tesseramento_identifier(
    connection: sqlite3.Connection,
    anno_sociale: int,
    *,
    preferred_number: int | None = None,
    exclude_id: int | None = None,
) -> tuple[int, str]:
    number = resolve_tesseramento_progressive_number(
        connection,
        anno_sociale,
        preferred_number=preferred_number,
        exclude_id=exclude_id,
    )
    return number, format_tesseramento_code(anno_sociale, number)


def peek_next_tesseramento_code(anno_sociale: int) -> str:
    with get_connection() as connection:
        _, code = assign_tesseramento_identifier(connection, anno_sociale)
        return code


def backfill_tesseramento_codes(connection: sqlite3.Connection) -> None:
    rows = connection.execute(
        """
        SELECT id, anno_sociale, numero_progressivo_anno, codice_tesseramento
        FROM tesseramenti_annuali
        ORDER BY anno_sociale, COALESCE(data_tesseramento, ''), id
        """
    ).fetchall()

    used_by_year: dict[int, set[int]] = {}
    updates: list[tuple[int, str, int]] = []
    for row in rows:
        record_id = int(row["id"])
        anno_sociale = int(row["anno_sociale"])
        used_numbers = used_by_year.setdefault(anno_sociale, set())
        current_number = int(row["numero_progressivo_anno"]) if row["numero_progressivo_anno"] else None
        if current_number is not None and current_number > 0 and current_number not in used_numbers:
            number = current_number
        else:
            number = 1
            while number in used_numbers:
                number += 1
        used_numbers.add(number)
        code = format_tesseramento_code(anno_sociale, number)
        if row["numero_progressivo_anno"] != number or row["codice_tesseramento"] != code:
            updates.append((number, code, record_id))

    if updates:
        connection.executemany(
            """
            UPDATE tesseramenti_annuali
            SET numero_progressivo_anno = ?, codice_tesseramento = ?
            WHERE id = ?
            """,
            updates,
        )


def backfill_progressive_numbers(connection: sqlite3.Connection, entity_key: str) -> None:
    meta = PROGRESSIVE_ENTITIES[entity_key]
    rows = connection.execute(
        f"""
        SELECT id, {meta["column"]}, {meta["code_column"]}
        FROM {meta["table"]}
        ORDER BY id
        """
    ).fetchall()

    used_numbers: set[int] = set()
    next_number = 1
    updates: list[tuple[int, str, int]] = []

    for row in rows:
        current_number = row[1]
        current_code = row[2]
        if isinstance(current_number, int) and current_number > 0 and current_number not in used_numbers:
            number = current_number
        else:
            while next_number in used_numbers:
                next_number += 1
            number = next_number
            next_number += 1

        used_numbers.add(number)
        if current_number != number or not current_code:
            code = current_code or format_progressive_code(entity_key, number)
            updates.append((number, code, row[0]))

    if updates:
        connection.executemany(
            f"""
            UPDATE {meta["table"]}
            SET {meta["column"]} = ?, {meta["code_column"]} = COALESCE(NULLIF({meta["code_column"]}, ''), ?)
            WHERE id = ?
            """,
            updates,
        )

    ultimo_valore = max(used_numbers, default=0)
    connection.execute(
        """
        INSERT INTO sequenze_progressive (chiave, ultimo_valore)
        VALUES (?, ?)
        ON CONFLICT(chiave) DO UPDATE SET ultimo_valore = excluded.ultimo_valore
        """,
        (entity_key, ultimo_valore),
    )


def ensure_schema() -> None:
    initialize_database_if_missing()
    with get_connection() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sequenze_progressive (
                chiave TEXT PRIMARY KEY,
                ultimo_valore INTEGER NOT NULL DEFAULT 0
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS utenti_accesso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                password_salt TEXT NOT NULL,
                password_iterations INTEGER NOT NULL DEFAULT 210000,
                email_recupero TEXT,
                is_admin INTEGER NOT NULL DEFAULT 0 CHECK (is_admin IN (0, 1)),
                attivo INTEGER NOT NULL DEFAULT 1 CHECK (attivo IN (0, 1)),
                creato_il TEXT NOT NULL DEFAULT (datetime('now')),
                aggiornato_il TEXT NOT NULL DEFAULT (datetime('now')),
                ultimo_accesso TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS sessioni_accesso (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_token TEXT NOT NULL UNIQUE,
                utente_id INTEGER NOT NULL,
                creata_il TEXT NOT NULL DEFAULT (datetime('now')),
                scade_il TEXT NOT NULL,
                FOREIGN KEY (utente_id) REFERENCES utenti_accesso (id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS impostazioni_app (
                chiave TEXT PRIMARY KEY,
                valore TEXT,
                aggiornato_il TEXT NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS registro_attivita (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                data_ora TEXT NOT NULL DEFAULT (datetime('now')),
                username TEXT,
                associato_id INTEGER,
                associato_codice TEXT,
                associato_nominativo TEXT,
                nome_pc TEXT,
                ip_client TEXT,
                metodo_http TEXT,
                percorso TEXT,
                categoria TEXT,
                descrizione_attivita TEXT NOT NULL,
                dettaglio TEXT,
                esito TEXT,
                anno_lavoro INTEGER,
                user_agent TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS oratorio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_progressivo INTEGER UNIQUE,
                codice_oratorio TEXT NOT NULL UNIQUE,
                nome TEXT NOT NULL,
                anno INTEGER NOT NULL,
                data_inizio TEXT NOT NULL,
                data_fine TEXT NOT NULL,
                sede TEXT,
                quota_partecipazione_standard NUMERIC NOT NULL DEFAULT 0 CHECK (quota_partecipazione_standard >= 0),
                posti_massimi INTEGER,
                descrizione TEXT,
                attivo INTEGER NOT NULL DEFAULT 1 CHECK (attivo IN (0, 1)),
                note TEXT,
                creato_il TEXT NOT NULL DEFAULT (datetime('now')),
                aggiornato_il TEXT NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS iscrizioni_oratorio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                associato_id INTEGER NOT NULL,
                oratorio_id INTEGER NOT NULL,
                data_iscrizione TEXT NOT NULL DEFAULT (date('now')),
                quota_partecipazione NUMERIC NOT NULL CHECK (quota_partecipazione >= 0),
                stato_iscrizione TEXT NOT NULL DEFAULT 'Iscritto' CHECK (stato_iscrizione IN ('Iscritto', 'Lista attesa', 'Annullato')),
                note TEXT,
                creato_il TEXT NOT NULL DEFAULT (datetime('now')),
                aggiornato_il TEXT NOT NULL DEFAULT (datetime('now')),
                UNIQUE (associato_id, oratorio_id),
                FOREIGN KEY (associato_id) REFERENCES associati (id) ON DELETE CASCADE,
                FOREIGN KEY (oratorio_id) REFERENCES oratorio (id) ON DELETE CASCADE
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS pagamenti_oratorio (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                iscrizione_oratorio_id INTEGER NOT NULL,
                data_pagamento TEXT NOT NULL,
                importo NUMERIC NOT NULL CHECK (importo > 0),
                metodo_pagamento_id INTEGER,
                riferimento TEXT,
                gruppo_ricevuta TEXT,
                note TEXT,
                creato_il TEXT NOT NULL DEFAULT (datetime('now')),
                UNIQUE (iscrizione_oratorio_id),
                FOREIGN KEY (iscrizione_oratorio_id) REFERENCES iscrizioni_oratorio (id) ON DELETE CASCADE,
                FOREIGN KEY (metodo_pagamento_id) REFERENCES metodi_pagamento (id)
            )
            """
        )
        for meta in PROGRESSIVE_ENTITIES.values():
            ensure_column(
                connection,
                meta["table"],
                f'{meta["column"]} INTEGER',
                meta["column"],
            )

        ensure_column(
            connection,
            "pagamenti_rate_corsi",
            "gruppo_ricevuta TEXT",
            "gruppo_ricevuta",
        )
        ensure_column(
            connection,
            "pagamenti_tesseramenti",
            "gruppo_ricevuta TEXT",
            "gruppo_ricevuta",
        )
        ensure_column(
            connection,
            "pagamenti_campi_estivi",
            "gruppo_ricevuta TEXT",
            "gruppo_ricevuta",
        )
        ensure_column(
            connection,
            "pagamenti_eventi",
            "gruppo_ricevuta TEXT",
            "gruppo_ricevuta",
        )
        ensure_column(
            connection,
            "pagamenti_oratorio",
            "gruppo_ricevuta TEXT",
            "gruppo_ricevuta",
        )
        ensure_column(
            connection,
            "tesseramenti_annuali",
            "numero_progressivo_anno INTEGER",
            "numero_progressivo_anno",
        )
        ensure_column(
            connection,
            "tesseramenti_annuali",
            "codice_tesseramento TEXT",
            "codice_tesseramento",
        )
        ensure_column(
            connection,
            "corsi",
            "data_inizio TEXT",
            "data_inizio",
        )
        ensure_column(
            connection,
            "corsi",
            "data_fine TEXT",
            "data_fine",
        )
        ensure_column(
            connection,
            "associati",
            "comune_nascita TEXT",
            "comune_nascita",
        )
        ensure_column(
            connection,
            "associati",
            "provincia_nascita TEXT",
            "provincia_nascita",
        )
        ensure_column(
            connection,
            "associati",
            "sesso TEXT NOT NULL DEFAULT 'M'",
            "sesso",
        )
        ensure_column(
            connection,
            "associati",
            "carica TEXT NOT NULL DEFAULT 'Associato'",
            "carica",
        )
        ensure_column(
            connection,
            "associati",
            "impiego TEXT",
            "impiego",
        )
        ensure_column(
            connection,
            "associati",
            "liberatoria_video TEXT NOT NULL DEFAULT 'Si'",
            "liberatoria_video",
        )
        ensure_column(
            connection,
            "associati",
            "patologie TEXT",
            "patologie",
        )
        ensure_column(connection, "associati", "genitore_tutore_cognome TEXT", "genitore_tutore_cognome")
        ensure_column(connection, "associati", "genitore_tutore_nome TEXT", "genitore_tutore_nome")
        ensure_column(connection, "associati", "genitore_tutore_cellulare TEXT", "genitore_tutore_cellulare")
        ensure_column(connection, "associati", "genitore_tutore_email TEXT", "genitore_tutore_email")
        ensure_column(connection, "associati", "genitore_tutore_impiego TEXT", "genitore_tutore_impiego")
        ensure_column(connection, "associati", f"genitore_tutore_tipo_documento TEXT NOT NULL DEFAULT '{DEFAULT_DOCUMENT_TYPE_SQL}'", "genitore_tutore_tipo_documento")
        ensure_column(connection, "associati", "genitore_tutore_numero_documento TEXT", "genitore_tutore_numero_documento")
        ensure_column(connection, "associati", "prelievo_altro_genitore_nome TEXT", "prelievo_altro_genitore_nome")
        ensure_column(connection, "associati", "prelievo_altro_genitore_cognome TEXT", "prelievo_altro_genitore_cognome")
        ensure_column(connection, "associati", "prelievo_altro_genitore_cellulare TEXT", "prelievo_altro_genitore_cellulare")
        ensure_column(connection, "associati", "prelievo_altro_genitore_impiego TEXT", "prelievo_altro_genitore_impiego")
        ensure_column(connection, "associati", f"prelievo_altro_genitore_tipo_documento TEXT NOT NULL DEFAULT '{DEFAULT_DOCUMENT_TYPE_SQL}'", "prelievo_altro_genitore_tipo_documento")
        ensure_column(connection, "associati", "prelievo_altro_genitore_numero_documento TEXT", "prelievo_altro_genitore_numero_documento")
        ensure_column(connection, "associati", "prelievo_altra_persona_nome TEXT", "prelievo_altra_persona_nome")
        ensure_column(connection, "associati", "prelievo_altra_persona_cognome TEXT", "prelievo_altra_persona_cognome")
        ensure_column(connection, "associati", "prelievo_altra_persona_cellulare TEXT", "prelievo_altra_persona_cellulare")
        ensure_column(connection, "associati", f"prelievo_altra_persona_tipo_documento TEXT NOT NULL DEFAULT '{DEFAULT_DOCUMENT_TYPE_SQL}'", "prelievo_altra_persona_tipo_documento")
        ensure_column(connection, "associati", "prelievo_altra_persona_numero_documento TEXT", "prelievo_altra_persona_numero_documento")
        ensure_column(
            connection,
            "utenti_accesso",
            "email_recupero TEXT",
            "email_recupero",
        )
        ensure_column(
            connection,
            "registro_attivita",
            "associato_id INTEGER",
            "associato_id",
        )
        ensure_column(
            connection,
            "registro_attivita",
            "associato_codice TEXT",
            "associato_codice",
        )
        ensure_column(
            connection,
            "registro_attivita",
            "associato_nominativo TEXT",
            "associato_nominativo",
        )

        for entity_key, meta in PROGRESSIVE_ENTITIES.items():
            backfill_progressive_numbers(connection, entity_key)
            connection.execute(
                f"""
                CREATE UNIQUE INDEX IF NOT EXISTS idx_{meta["table"]}_{meta["column"]}
                ON {meta["table"]} ({meta["column"]})
                """
            )

        backfill_tesseramento_codes(connection)
        connection.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_tesseramenti_annuali_codice_tesseramento
            ON tesseramenti_annuali (codice_tesseramento)
            """
        )
        connection.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_tesseramenti_annuali_anno_progressivo
            ON tesseramenti_annuali (anno_sociale, numero_progressivo_anno)
            """
        )

        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_pagamenti_rate_corsi_gruppo
            ON pagamenti_rate_corsi (gruppo_ricevuta)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_pagamenti_tesseramenti_gruppo
            ON pagamenti_tesseramenti (gruppo_ricevuta)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_pagamenti_campi_estivi_gruppo
            ON pagamenti_campi_estivi (gruppo_ricevuta)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_pagamenti_eventi_gruppo
            ON pagamenti_eventi (gruppo_ricevuta)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_pagamenti_oratorio_gruppo
            ON pagamenti_oratorio (gruppo_ricevuta)
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS quote_predefinite (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                area TEXT NOT NULL CHECK (area IN ('tesseramenti', 'campi-estivi', 'oratorio')),
                descrizione TEXT NOT NULL,
                importo NUMERIC NOT NULL CHECK (importo >= 0),
                attiva INTEGER NOT NULL DEFAULT 1 CHECK (attiva IN (0, 1)),
                note TEXT,
                creato_il TEXT NOT NULL DEFAULT (datetime('now')),
                aggiornato_il TEXT NOT NULL DEFAULT (datetime('now'))
            )
            """
        )
        ensure_quote_predefinite_supports_oratorio(connection)
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_quote_predefinite_area
            ON quote_predefinite (area, attiva, descrizione)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_utenti_accesso_admin_attivo
            ON utenti_accesso (is_admin, attivo, username)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_sessioni_accesso_utente
            ON sessioni_accesso (utente_id, scade_il)
            """
        )
        connection.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_registro_attivita_data
            ON registro_attivita (data_ora DESC, id DESC)
            """
        )
        prune_old_activity_log_rows(connection)
        cleanup_expired_sessions(connection)
        rebuild_reporting_views(connection)
        connection.commit()
    export_activity_log_xls()


def reserve_progressive_number(connection: sqlite3.Connection, entity_key: str) -> int:
    row = connection.execute(
        "SELECT ultimo_valore FROM sequenze_progressive WHERE chiave = ?",
        (entity_key,),
    ).fetchone()
    current_value = int(row[0]) if row else 0
    next_value = current_value + 1
    connection.execute(
        """
        INSERT INTO sequenze_progressive (chiave, ultimo_valore)
        VALUES (?, ?)
        ON CONFLICT(chiave) DO UPDATE SET ultimo_valore = excluded.ultimo_valore
        """,
        (entity_key, next_value),
    )
    return next_value


def peek_next_progressive_code(entity_key: str) -> str:
    row = fetch_one(
        "SELECT ultimo_valore FROM sequenze_progressive WHERE chiave = ?",
        (entity_key,),
    )
    next_value = (int(row[0]) if row else 0) + 1
    return format_progressive_code(entity_key, next_value)


def generate_receipt_group_code() -> str:
    return f"GRP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"


def generate_multi_area_group_code() -> str:
    return f"MGR-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"


def current_work_year(query_params: dict[str, str]) -> int:
    raw_value = (query_params.get("anno_lavoro") or "").strip()
    if raw_value.isdigit():
        return int(raw_value)
    return date.today().year


def work_year_query(query_params: dict[str, str]) -> dict[str, str]:
    return {"anno_lavoro": str(current_work_year(query_params))}


def work_year_query_from_form(form_data: dict[str, str]) -> dict[str, str]:
    params: dict[str, str] = {}
    raw_value = normalized(form_data, "anno_lavoro", "")
    if raw_value.isdigit():
        params["anno_lavoro"] = raw_value
    if normalized(form_data, "vista", "") == "dati":
        params["vista"] = "dati"
    return params


def year_start_end(year: int) -> tuple[str, str]:
    return f"{year}-01-01", f"{year}-12-31"


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def ensure_backup_dir() -> None:
    ensure_output_dir()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def human_file_size(size_bytes: int) -> str:
    units = ["B", "KB", "MB", "GB", "TB"]
    amount = float(max(size_bytes, 0))
    unit_index = 0
    while amount >= 1024 and unit_index < len(units) - 1:
        amount /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(amount)} {units[unit_index]}"
    return f"{amount:.1f} {units[unit_index]}"


def backup_restore_batch_contents() -> str:
    return """@echo off
setlocal
powershell -NoProfile -ExecutionPolicy Bypass -File "%~dp0payload\\OratorioCarloAcutis\\scripts\\installa_pacchetto_locale.ps1" -TargetRoot "C:\\OratorioCarloAcutis" -ReplaceDatabase
if errorlevel 1 (
  echo.
  echo Ripristino non completato.
  pause
  exit /b 1
)
echo.
echo Ripristino completato.
pause
endlocal
"""


def backup_restore_readme_contents(backup_name: str) -> str:
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    return (
        f"{APP_NAME} - Backup completo\n"
        f"Backup: {backup_name}\n"
        f"Generato il: {generated_at}\n\n"
        "Contenuto del backup:\n"
        "- applicazione completa\n"
        "- database con tutti i dati\n"
        "- runtime locale Python e Node\n"
        "- script, static, data e file di avvio\n\n"
        "Ripristino sullo stesso PC:\n"
        "1. Chiudi il gestionale.\n"
        "2. Estrai tutto il contenuto dello ZIP in una cartella a scelta.\n"
        "3. Esegui 'Ripristina backup su questo PC.bat'.\n"
        "4. Accetta la richiesta di Windows se compare.\n"
        "5. Al termine riapri il gestionale dal collegamento Desktop.\n\n"
        "Ripristino su un altro PC:\n"
        "1. Copia questo ZIP sull'altro PC.\n"
        "2. Estrai tutto il contenuto dello ZIP.\n"
        "3. Esegui 'Ripristina backup su un altro PC.bat'.\n"
        "4. Accetta la richiesta di Windows se compare.\n"
        "5. Al termine apri il gestionale dal collegamento Desktop creato automaticamente.\n\n"
        "Il ripristino sovrascrive il database locale con quello presente nel backup, mantenendo intatte funzionalita e configurazione."
    )


def iter_backup_source_files() -> list[Path]:
    files: list[Path] = []
    for file_path in BASE_DIR.rglob("*"):
        if not file_path.is_file():
            continue
        relative_path = file_path.relative_to(BASE_DIR)
        if relative_path.parts and relative_path.parts[0] in BACKUP_EXCLUDED_TOP_LEVEL:
            continue
        if relative_path.parts[:2] in {
            ("outputs", "backups"),
            ("outputs", "aggiornamenti"),
            ("outputs", "import-associati"),
        }:
            continue
        if "__pycache__" in relative_path.parts:
            continue
        files.append(file_path)
    files.sort(key=lambda item: item.relative_to(BASE_DIR).as_posix())
    return files


def create_backup_archive() -> Path:
    ensure_backup_dir()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_name = f"{BACKUP_ROOT_NAME}-backup-{timestamp}.zip"
    backup_path = BACKUP_DIR / backup_name
    with zipfile.ZipFile(backup_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        archive.writestr("Ripristina backup su questo PC.bat", backup_restore_batch_contents())
        archive.writestr("Ripristina backup su un altro PC.bat", backup_restore_batch_contents())
        archive.writestr("LEGGIMI-RIPRISTINO.txt", backup_restore_readme_contents(backup_name))
        for file_path in iter_backup_source_files():
            archive.write(
                file_path,
                arcname=(Path("payload") / BACKUP_ROOT_NAME / file_path.relative_to(BASE_DIR)).as_posix(),
            )
    return backup_path


def delete_backup_archive(file_name: str) -> bool:
    ensure_backup_dir()
    clean_name = (file_name or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9._-]+\.zip", clean_name):
        return False
    backup_path = (BACKUP_DIR / clean_name).resolve()
    if BACKUP_DIR.resolve() not in backup_path.parents or not backup_path.is_file():
        return False
    backup_path.unlink()
    return True


def backup_archive_rows(query_params: dict[str, str]) -> list[dict[str, str]]:
    ensure_backup_dir()
    rows: list[dict[str, str]] = []
    for backup_path in sorted(BACKUP_DIR.glob("*.zip"), key=lambda item: item.stat().st_mtime, reverse=True):
        stats = backup_path.stat()
        rows.append(
            {
                "filename": backup_path.name,
                "created_at": datetime.fromtimestamp(stats.st_mtime).strftime("%d/%m/%Y %H:%M"),
                "size": human_file_size(int(stats.st_size)),
                "path": str(backup_path),
                "download_url": with_query(f"/download/backup/{quote(backup_path.name)}", work_year_query(query_params)),
            }
        )
    return rows


def ensure_update_dir() -> None:
    ensure_output_dir()
    UPDATE_DIR.mkdir(parents=True, exist_ok=True)


def ensure_associati_import_dir() -> None:
    ensure_output_dir()
    ASSOCIATI_IMPORT_DIR.mkdir(parents=True, exist_ok=True)


def ensure_tutorial_video_dir() -> None:
    ensure_output_dir()
    TUTORIAL_VIDEO_DIR.mkdir(parents=True, exist_ok=True)


def app_version_safe() -> str:
    return re.sub(r"[^0-9A-Za-z._-]+", "-", APP_VERSION).strip("-") or "0.0.0"


def create_update_package_archive() -> Path:
    ensure_update_dir()
    script_path = BASE_DIR / "scripts" / "crea_pacchetto_distribuzione.ps1"
    if not script_path.exists():
        raise ValueError("Script di creazione aggiornamento non trovato.")
    command = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(script_path),
        "-OutputRoot",
        str(UPDATE_DIR),
        "-UpdateOnly",
        "-CreateZip",
    ]
    try:
        subprocess.run(
            command,
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            check=True,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
    except subprocess.CalledProcessError as error:
        detail = (error.stderr or error.stdout or "").strip()
        raise ValueError(f"Creazione aggiornamento non riuscita. {detail}".strip()) from error
    packages = sorted(UPDATE_DIR.glob("*.zip"), key=lambda item: item.stat().st_mtime, reverse=True)
    if not packages:
        raise ValueError("Il pacchetto di aggiornamento non e stato generato.")
    return packages[0]


def update_package_rows(query_params: dict[str, str]) -> list[dict[str, str]]:
    ensure_update_dir()
    rows: list[dict[str, str]] = []
    for package_path in sorted(UPDATE_DIR.glob("*.zip"), key=lambda item: item.stat().st_mtime, reverse=True):
        stats = package_path.stat()
        version_match = re.search(r"-v([0-9A-Za-z._-]+)-\d{8}-\d{6}\.zip$", package_path.name)
        rows.append(
            {
                "filename": package_path.name,
                "version": version_match.group(1) if version_match else APP_VERSION,
                "created_at": datetime.fromtimestamp(stats.st_mtime).strftime("%d/%m/%Y %H:%M"),
                "size": human_file_size(int(stats.st_size)),
                "path": str(package_path),
                "download_url": with_query(f"/download/aggiornamento/{quote(package_path.name)}", work_year_query(query_params)),
            }
        )
    return rows


def version_sort_key(version: str) -> tuple[tuple[int, object], ...]:
    parts = re.split(r"[._-]+", str(version or "").strip())
    key: list[tuple[int, object]] = []
    for part in parts:
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part.lower()))
    return tuple(key)


def compare_versions(left: str, right: str) -> int:
    left_key = version_sort_key(left)
    right_key = version_sort_key(right)
    if left_key > right_key:
        return 1
    if left_key < right_key:
        return -1
    return 0


def extract_update_package_version(package_bytes: bytes, filename: str = "") -> str | None:
    try:
        with zipfile.ZipFile(BytesIO(package_bytes)) as archive:
            for member_name in archive.namelist():
                if Path(member_name).name.lower() == "version.txt":
                    version = archive.read(member_name).decode("utf-8").strip()
                    return version or None
    except zipfile.BadZipFile as exc:
        raise ValueError("Il file selezionato non e un pacchetto ZIP valido.") from exc
    version_match = re.search(r"-v([0-9A-Za-z._-]+)-\d{8}-\d{6}\.zip$", filename or "")
    if version_match:
        return version_match.group(1)
    return None


def store_uploaded_update_package(filename: str, package_bytes: bytes) -> Path:
    ensure_update_dir()
    safe_name = slugify(Path(filename or "aggiornamento").stem)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    package_path = UPDATE_DIR / f"upload-{timestamp}-{safe_name}.zip"
    package_path.write_bytes(package_bytes)
    return package_path


def schedule_update_package_install(package_path: Path) -> None:
    ensure_update_dir()
    staging_root = UPDATE_DIR / f"staging-{uuid.uuid4().hex}"
    helper_script_path = UPDATE_DIR / f"installa-aggiornamento-{uuid.uuid4().hex}.ps1"
    log_path = UPDATE_DIR / f"installa-aggiornamento-{uuid.uuid4().hex}.log"
    install_root = BASE_DIR
    pythonw_path = BASE_DIR / "runtime" / "python" / "pythonw.exe"
    app_path = BASE_DIR / "app.py"

    def ps_literal(value: str) -> str:
        return value.replace("'", "''")

    helper_script = f"""
$ErrorActionPreference = 'Stop'
$zipPath = '{ps_literal(str(package_path))}'
$stagingRoot = '{ps_literal(str(staging_root))}'
$logPath = '{ps_literal(str(log_path))}'
$installRoot = '{ps_literal(str(install_root))}'
$pythonwPath = '{ps_literal(str(pythonw_path))}'
$appPath = '{ps_literal(str(app_path))}'

function Write-Log([string]$Message) {{
    $timestamp = Get-Date -Format 'yyyy-MM-dd HH:mm:ss'
    Add-Content -LiteralPath $logPath -Value ("[$timestamp] " + $Message)
}}

function Test-IsAdministrator {{
    $identity = [Security.Principal.WindowsIdentity]::GetCurrent()
    $principal = [Security.Principal.WindowsPrincipal]::new($identity)
    return $principal.IsInRole([Security.Principal.WindowsBuiltinRole]::Administrator)
}}

try {{
    Write-Log "Avvio installazione aggiornamento da $zipPath"
    if (-not (Test-IsAdministrator)) {{
        Write-Log "Richiesta elevazione amministratore."
        $arguments = @(
            '-NoProfile',
            '-ExecutionPolicy', 'Bypass',
            '-File', ('"' + $PSCommandPath + '"')
        )
        Start-Process -FilePath 'powershell.exe' -ArgumentList $arguments -Verb RunAs | Out-Null
        exit
    }}

    Start-Sleep -Seconds 4
    if (Test-Path -LiteralPath $stagingRoot) {{
        Remove-Item -LiteralPath $stagingRoot -Recurse -Force
    }}
    New-Item -ItemType Directory -Path $stagingRoot -Force | Out-Null
    Expand-Archive -LiteralPath $zipPath -DestinationPath $stagingRoot -Force
    Write-Log "Pacchetto estratto in $stagingRoot"

    $payloadRoot = Get-ChildItem -Path $stagingRoot -Recurse -Directory | Where-Object {{
        $_.FullName -match '[\\\\/]payload[\\\\/]OratorioCarloAcutis$'
    }} | Select-Object -First 1
    if (-not $payloadRoot) {{
        throw 'Contenuto del pacchetto aggiornamento non valido.'
    }}

    Get-CimInstance Win32_Process | Where-Object {{
        $_.Name -eq 'pythonw.exe' -and $_.CommandLine -like ('*' + $appPath + '*')
    }} | ForEach-Object {{
        Stop-Process -Id $_.ProcessId -Force
    }}
    Write-Log "Processi del gestionale arrestati."
    Start-Sleep -Seconds 1

    $arguments = @(
        $payloadRoot.FullName,
        $installRoot,
        '/E',
        '/R:2',
        '/W:1',
        '/NFL',
        '/NDL',
        '/NJH',
        '/NJS',
        '/NP',
        '/XD',
        'outputs',
        '/XF',
        'gestione_associazione.sqlite'
    )
    Write-Log "Avvio copia file su $installRoot"
    & robocopy @arguments | Out-Null
    if ($LASTEXITCODE -ge 8) {{
        throw 'Installazione aggiornamento non riuscita durante la copia dei file.'
    }}
    Write-Log "Copia file completata."

    Start-Sleep -Seconds 1
    if (-not (Test-Path -LiteralPath $pythonwPath)) {{
        throw 'Python runtime non trovato dopo l''aggiornamento.'
    }}
    if (-not (Test-Path -LiteralPath $appPath)) {{
        throw 'File app.py non trovato dopo l''aggiornamento.'
    }}
    Start-Process -FilePath $pythonwPath -ArgumentList @($appPath) -WindowStyle Hidden
    Write-Log "Gestionale riavviato correttamente."
}} catch {{
    Write-Log ("ERRORE: " + $_.Exception.Message)
}}
"""
    helper_script_path.write_text(helper_script, encoding="utf-8")
    subprocess.Popen(
        [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(helper_script_path),
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )


def update_installation_readme() -> str:
    return (
        "1. Sul PC sorgente apri il file 'C:\\OratorioCarloAcutis\\Crea pacchetto aggiornamento per altri PC.bat'.\n"
        "2. In alternativa usa il comando PowerShell esterno riportato in questa pagina.\n"
        "3. Attendi la creazione dello ZIP nella cartella aggiornamenti.\n"
        "4. Copia lo ZIP sul PC da aggiornare.\n"
        "5. Sul PC di destinazione estrai lo ZIP ed esegui 'Installa aggiornamento guidato.bat'.\n"
        "6. Il database locale del PC di destinazione resta invariato, quindi dati e ricevute non vengono persi.\n"
        "7. Se il pacchetto contiene il file VERSION.txt, la versione viene solo mostrata come informazione e non blocca l'installazione."
    )


def associati_template_filename() -> str:
    return f"modello_importazione_associati_v{app_version_safe()}.xlsx"


def build_associati_template_xlsx() -> tuple[str, bytes]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Istruzioni"
    instructions["A1"] = f"{APP_NAME} - Modello importazione associati"
    instructions["A2"] = f"Versione gestionale: {APP_VERSION}"
    instructions["A4"] = "Compila il foglio 'Associati' senza cambiare le intestazioni della riga 1."
    instructions["A5"] = "Campi minimi consigliati: Nome, Cognome. Se Data prima iscrizione e vuota verra usata la data odierna."
    instructions["A6"] = "Duplicati: il gestionale salta i record con stesso codice fiscale oppure stesso Nome+Cognome+Data nascita."
    instructions["A7"] = "Per i minorenni sono obbligatori solo Nome, Cognome e Cellulare del Genitore/Tutore."
    instructions["A8"] = "Valori guidati: Sesso = M/F, Stato associato = Attivo/Sospeso/Concluso, Liberatoria Video = Si/No."
    instructions["A9"] = "Tipo documento: Carta d'identitÃ  oppure Patente di guida."
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A1"].fill = PatternFill("solid", fgColor="FFF1E4")
    instructions.column_dimensions["A"].width = 110
    for cell in instructions["A4:A9"]:
        for item in cell:
            item.alignment = Alignment(wrap_text=True, vertical="top")

    sheet = workbook.create_sheet("Associati")
    headers = [label for _, label in ASSOCIATI_IMPORT_FIELDS]
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="F6E4D4")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        width = max(18, min(42, len(header) + 4))
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    return associati_template_filename(), output.getvalue()


def normalize_import_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_import_document_type(value: str) -> str:
    raw = plain_text(value).strip().lower()
    if raw.startswith("patente"):
        return "Patente di guida"
    return "Carta d'identitÃ "


def normalize_import_yes_no(value: str, default: str = "Si") -> str:
    raw = plain_text(value).strip().lower()
    if raw in {"no", "n", "false", "0"}:
        return "No"
    if raw in {"si", "sÃ¬", "s", "true", "1"}:
        return "Si"
    return default


def normalize_import_sesso(value: str) -> str:
    raw = plain_text(value).strip().lower()
    if raw.startswith("f"):
        return "F"
    return "M"


def normalize_import_document_type(value: str) -> str:
    raw = plain_text(value).strip().lower()
    if raw.startswith("patente"):
        return "Patente di guida"
    return DEFAULT_DOCUMENT_TYPE


def normalize_import_yes_no(value: str, default: str = "Si") -> str:
    raw = plain_text(value).strip().lower()
    if raw in {"no", "n", "false", "0"}:
        return "No"
    if raw in {"si", "sÃ¬", "s", "true", "1"}:
        return "Si"
    return default


def import_associati_from_excel(file_bytes: bytes, current_user: dict[str, object] | None = None) -> dict[str, object]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook["Associati"] if "Associati" in workbook.sheetnames else workbook.worksheets[0]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_map = {
        normalize_import_text(label): key
        for key, label in ASSOCIATI_IMPORT_FIELDS
    }
    column_keys: list[str | None] = [header_map.get(normalize_import_text(cell)) for cell in header_row]
    inserted = 0
    duplicates = 0
    errors: list[str] = []
    today_value = date.today().isoformat()

    with get_connection() as connection:
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_payload: dict[str, str] = {}
            for key, value in zip(column_keys, values):
                if not key:
                    continue
                row_payload[key] = normalize_import_text(value)

            if not any((value or "").strip() for value in row_payload.values()):
                continue

            nome = row_payload.get("nome", "").strip()
            cognome = row_payload.get("cognome", "").strip()
            if not nome or not cognome:
                errors.append(f"Riga {row_index}: Nome e Cognome sono obbligatori.")
                continue

            data_nascita = row_payload.get("data_nascita", "").strip()
            codice_fiscale = row_payload.get("codice_fiscale", "").strip().upper()
            duplicate_row = None
            if codice_fiscale:
                duplicate_row = connection.execute(
                    "SELECT id FROM associati WHERE UPPER(TRIM(COALESCE(codice_fiscale, ''))) = ?",
                    (codice_fiscale,),
                ).fetchone()
            if duplicate_row is None:
                duplicate_row = connection.execute(
                    """
                    SELECT id
                    FROM associati
                    WHERE LOWER(TRIM(COALESCE(nome, ''))) = ?
                      AND LOWER(TRIM(COALESCE(cognome, ''))) = ?
                      AND COALESCE(data_nascita, '') = ?
                    """,
                    (nome.strip().lower(), cognome.strip().lower(), data_nascita),
                ).fetchone()
            if duplicate_row is not None:
                duplicates += 1
                continue

            if data_nascita:
                age = calculate_age(data_nascita)
            else:
                age = None
            is_minor = age is not None and age < 18
            if is_minor:
                if not row_payload.get("genitore_tutore_nome", "").strip():
                    errors.append(f"Riga {row_index}: per i minorenni il Nome del Genitore/Tutore e obbligatorio.")
                    continue
                if not row_payload.get("genitore_tutore_cognome", "").strip():
                    errors.append(f"Riga {row_index}: per i minorenni il Cognome del Genitore/Tutore e obbligatorio.")
                    continue
                if not row_payload.get("genitore_tutore_cellulare", "").strip():
                    errors.append(f"Riga {row_index}: per i minorenni il Cellulare del Genitore/Tutore e obbligatorio.")
                    continue

            progressive_number = reserve_progressive_number(connection, "associati")
            associati_insert_placeholders = ", ".join(["?"] * 40)
            connection.execute(
                f"""
                INSERT INTO associati (
                    numero_progressivo, codice_associato, nome, cognome, codice_fiscale, data_nascita,
                    sesso, comune_nascita, provincia_nascita, carica, email, telefono, indirizzo, cap, citta, provincia, impiego,
                    data_prima_iscrizione, stato_associato, liberatoria_video, patologie,
                    genitore_tutore_cognome, genitore_tutore_nome, genitore_tutore_cellulare, genitore_tutore_email, genitore_tutore_impiego, genitore_tutore_tipo_documento, genitore_tutore_numero_documento,
                    prelievo_altro_genitore_nome, prelievo_altro_genitore_cognome, prelievo_altro_genitore_cellulare, prelievo_altro_genitore_impiego, prelievo_altro_genitore_tipo_documento, prelievo_altro_genitore_numero_documento,
                    prelievo_altra_persona_nome, prelievo_altra_persona_cognome, prelievo_altra_persona_cellulare, prelievo_altra_persona_tipo_documento, prelievo_altra_persona_numero_documento,
                    note
                ) VALUES ({associati_insert_placeholders})
                """,
                (
                    progressive_number,
                    format_progressive_code("associati", progressive_number),
                    nome,
                    cognome,
                    codice_fiscale,
                    data_nascita,
                    normalize_import_sesso(row_payload.get("sesso", "M")),
                    row_payload.get("comune_nascita", "").strip(),
                    row_payload.get("provincia_nascita", "").strip().upper(),
                    resolved_carica_value({"carica": row_payload.get("carica", "Associato")}, current_user, existing_value="Associato"),
                    row_payload.get("email", "").strip(),
                    row_payload.get("telefono", "").strip(),
                    row_payload.get("indirizzo", "").strip(),
                    row_payload.get("cap", "").strip(),
                    row_payload.get("citta", "").strip(),
                    row_payload.get("provincia", "").strip().upper(),
                    row_payload.get("impiego", "").strip(),
                    row_payload.get("data_prima_iscrizione", "").strip() or today_value,
                    row_payload.get("stato_associato", "").strip() or "Attivo",
                    normalize_import_yes_no(row_payload.get("liberatoria_video", "Si"), "Si"),
                    row_payload.get("patologie", "").strip(),
                    row_payload.get("genitore_tutore_cognome", "").strip(),
                    row_payload.get("genitore_tutore_nome", "").strip(),
                    row_payload.get("genitore_tutore_cellulare", "").strip(),
                    row_payload.get("genitore_tutore_email", "").strip(),
                    row_payload.get("genitore_tutore_impiego", "").strip(),
                    normalize_import_document_type(row_payload.get("genitore_tutore_tipo_documento", DEFAULT_DOCUMENT_TYPE)),
                    row_payload.get("genitore_tutore_numero_documento", "").strip(),
                    row_payload.get("prelievo_altro_genitore_nome", "").strip(),
                    row_payload.get("prelievo_altro_genitore_cognome", "").strip(),
                    row_payload.get("prelievo_altro_genitore_cellulare", "").strip(),
                    row_payload.get("prelievo_altro_genitore_impiego", "").strip(),
                    normalize_import_document_type(row_payload.get("prelievo_altro_genitore_tipo_documento", DEFAULT_DOCUMENT_TYPE)),
                    row_payload.get("prelievo_altro_genitore_numero_documento", "").strip(),
                    row_payload.get("prelievo_altra_persona_nome", "").strip(),
                    row_payload.get("prelievo_altra_persona_cognome", "").strip(),
                    row_payload.get("prelievo_altra_persona_cellulare", "").strip(),
                    normalize_import_document_type(row_payload.get("prelievo_altra_persona_tipo_documento", DEFAULT_DOCUMENT_TYPE)),
                    row_payload.get("prelievo_altra_persona_numero_documento", "").strip(),
                    row_payload.get("note", "").strip(),
                ),
            )
            inserted += 1
        connection.commit()
    return {"inserted": inserted, "duplicates": duplicates, "errors": errors}


def table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    row = connection.execute(
        """
        SELECT name
        FROM sqlite_master
        WHERE type = 'table' AND name = ?
        """,
        (table_name,),
    ).fetchone()
    return row is not None


def report_requires_admin(report_key: str) -> bool:
    return report_key in ADMIN_ONLY_REPORT_KEYS


def latest_generated_course_month(work_year: int) -> tuple[int, int] | None:
    row = fetch_one(
        """
        SELECT anno, mese
        FROM rate_corsi_mensili
        WHERE anno = ?
        ORDER BY anno DESC, mese DESC
        LIMIT 1
        """,
        (work_year,),
    )
    if row is None:
        return None
    return int(row["anno"]), int(row["mese"])


def increment_year_month(year: int, month: int) -> tuple[int, int]:
    if month >= 12:
        return year + 1, 1
    return year, month + 1


def parse_year_month_value(value: str, field_label: str = "Competenza") -> tuple[int, int]:
    raw_value = (value or "").strip()
    match = re.fullmatch(r"(\d{4})-(\d{2})", raw_value)
    if not match:
        raise ValueError(f"{field_label} non valida.")
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError(f"{field_label} non valida.")
    return year, month


def default_mass_rate_due_date(year: int, month: int) -> str:
    next_year, next_month = increment_year_month(year, month)
    return f"{next_year:04d}-{next_month:02d}-14"


def pending_mass_generation_cutoff(today_value: date | None = None) -> tuple[int, int] | None:
    today_value = today_value or date.today()
    if today_value.day > 14:
        return today_value.year, today_value.month
    if today_value.month == 1:
        return today_value.year - 1, 12
    return today_value.year, today_value.month - 1


def pending_course_monthly_generations(
    work_year: int,
    today_value: date | None = None,
) -> dict[str, object] | None:
    today_value = today_value or date.today()
    if work_year != today_value.year:
        return None

    cutoff = pending_mass_generation_cutoff(today_value)
    if cutoff is None:
        return None
    cutoff_year, cutoff_month = cutoff
    if cutoff_year != work_year:
        return None

    latest = latest_generated_course_month(work_year)
    start_year, start_month = (work_year, 1) if latest is None else increment_year_month(*latest)
    if (start_year, start_month) > (cutoff_year, cutoff_month):
        return None

    missing_months: list[tuple[int, int]] = []
    current_year, current_month = start_year, start_month
    while (current_year, current_month) <= (cutoff_year, cutoff_month):
        missing_months.append((current_year, current_month))
        current_year, current_month = increment_year_month(current_year, current_month)

    if not missing_months:
        return None

    first_missing_year, first_missing_month = missing_months[0]
    cutoff_last_day = f"{cutoff_year:04d}-{cutoff_month:02d}-{monthrange(cutoff_year, cutoff_month)[1]:02d}"
    first_missing_day = f"{first_missing_year:04d}-{first_missing_month:02d}-01"
    active_count = int(
        fetch_scalar(
            """
            SELECT COUNT(*)
            FROM iscrizioni_corsi
            WHERE stato_iscrizione = 'Attiva'
              AND COALESCE(NULLIF(data_inizio, ''), data_iscrizione, ?) <= ?
              AND (data_fine IS NULL OR data_fine = '' OR data_fine >= ?)
            """,
            (cutoff_last_day, cutoff_last_day, first_missing_day),
        )
        or 0
    )
    if active_count <= 0:
        return None

    missing_labels = [f"{month_label(month)} {year}" for year, month in missing_months]
    latest_label = (
        f"{month_label(latest[1])} {latest[0]}"
        if latest is not None
        else "nessuna generazione ancora presente"
    )
    return {
        "work_year": work_year,
        "latest_label": latest_label,
        "missing_months": missing_months,
        "missing_labels": missing_labels,
        "missing_labels_text": ", ".join(missing_labels),
        "message": (
            f"L'ultima generazione massiva registrata e {latest_label}. "
            f"Risultano ancora da generare: {', '.join(missing_labels)}. "
            "Vuoi eseguire ora la generazione automatica dei mesi mancanti?"
        ),
    }


def parse_iso_date_required(value: str, label: str) -> date:
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f"{label} non valida.") from exc


def validate_course_date_range(data_inizio_value: str, data_fine_value: str) -> tuple[str, str]:
    data_inizio = parse_iso_date_required(data_inizio_value, "La data inizio del corso")
    data_fine = parse_iso_date_required(data_fine_value, "La data fine del corso")
    if data_inizio > data_fine:
        raise ValueError("La data fine del corso deve essere uguale o successiva alla data inizio.")
    return data_inizio.isoformat(), data_fine.isoformat()


def compute_course_enrollment_effective_window(
    enrollment_start_value: str,
    enrollment_end_value: str | None,
    course_start_value: str | None,
    course_end_value: str | None,
) -> tuple[str, str | None]:
    start_candidates = [value for value in [enrollment_start_value, course_start_value or ""] if value]
    if not start_candidates:
        raise ValueError("Decorrenza del corso non valida.")
    effective_start = max(start_candidates)
    end_candidates = [value for value in [enrollment_end_value or "", course_end_value or ""] if value]
    effective_end = min(end_candidates) if end_candidates else None
    if effective_end and effective_start > effective_end:
        raise ValueError("L'intervallo del corso non consente l'iscrizione nelle date indicate.")
    return effective_start, effective_end


def resolve_course_enrollment_window(
    connection: sqlite3.Connection,
    corso_id: int,
    data_iscrizione_value: str,
    requested_start_value: str | None = None,
    requested_end_value: str | None = None,
) -> tuple[str, str | None]:
    data_iscrizione_iso = parse_iso_date_required(data_iscrizione_value, "La data iscrizione del corso").isoformat()
    requested_start_iso = (
        parse_iso_date_required(requested_start_value, "La data inizio del tesserato").isoformat()
        if requested_start_value
        else data_iscrizione_iso
    )
    requested_end_iso = (
        parse_iso_date_required(requested_end_value, "La data fine del tesserato").isoformat()
        if requested_end_value
        else None
    )
    corso = connection.execute(
        """
        SELECT
            COALESCE(data_inizio, '') AS data_inizio,
            COALESCE(data_fine, '') AS data_fine
        FROM corsi
        WHERE id = ?
        """,
        (corso_id,),
    ).fetchone()
    if corso is None:
        raise ValueError("Corso non trovato.")
    course_start_iso = (
        parse_iso_date_required(str(corso["data_inizio"]), "La data inizio del corso").isoformat()
        if corso["data_inizio"]
        else None
    )
    course_end_iso = (
        parse_iso_date_required(str(corso["data_fine"]), "La data fine del corso").isoformat()
        if corso["data_fine"]
        else None
    )
    return compute_course_enrollment_effective_window(
        requested_start_iso,
        requested_end_iso,
        course_start_iso,
        course_end_iso,
    )


def generate_course_rates_for_month(
    connection: sqlite3.Connection,
    anno: int,
    mese: int,
    *,
    data_scadenza: str = "",
    note: str = "",
) -> tuple[int, int]:
    if mese < 1 or mese > 12:
        raise ValueError("Il mese deve essere compreso tra 1 e 12.")

    first_day = f"{anno}-{mese:02d}-01"
    last_day = f"{anno}-{mese:02d}-{monthrange(anno, mese)[1]:02d}"
    inserted = 0
    skipped = 0
    iscrizioni_attive = connection.execute(
        """
        SELECT
            ic.id,
            ic.quota_mensile,
            COALESCE(NULLIF(ic.data_inizio, ''), ic.data_iscrizione, ?) AS data_inizio_iscrizione,
            COALESCE(NULLIF(ic.data_fine, ''), '') AS data_fine_iscrizione,
            COALESCE(NULLIF(c.data_inizio, ''), '') AS data_inizio_corso,
            COALESCE(NULLIF(c.data_fine, ''), '') AS data_fine_corso
        FROM iscrizioni_corsi ic
        JOIN corsi c ON c.id = ic.corso_id
        WHERE ic.stato_iscrizione = 'Attiva'
        ORDER BY ic.id
        """,
        (last_day,),
    ).fetchall()

    for iscrizione in iscrizioni_attive:
        effective_start, effective_end = compute_course_enrollment_effective_window(
            str(iscrizione["data_inizio_iscrizione"] or ""),
            str(iscrizione["data_fine_iscrizione"] or ""),
            str(iscrizione["data_inizio_corso"] or ""),
            str(iscrizione["data_fine_corso"] or ""),
        )
        if effective_start > last_day:
            skipped += 1
            continue
        if effective_end and effective_end < first_day:
            skipped += 1
            continue
        cursor = connection.execute(
            """
            INSERT OR IGNORE INTO rate_corsi_mensili (
                iscrizione_corso_id, anno, mese, importo_dovuto, data_scadenza, note
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                iscrizione["id"],
                anno,
                mese,
                iscrizione["quota_mensile"],
                data_scadenza or None,
                note or None,
            ),
        )
        if cursor.rowcount == 1:
            inserted += 1
        else:
            skipped += 1
    return inserted, skipped


def ensure_course_rate_for_enrollment(
    connection: sqlite3.Connection,
    iscrizione_corso_id: int,
    anno: int,
    mese: int,
    *,
    note: str = "",
) -> int:
    if mese < 1 or mese > 12:
        raise ValueError("Il mese deve essere compreso tra 1 e 12.")

    first_day = f"{anno}-{mese:02d}-01"
    last_day = f"{anno}-{mese:02d}-{monthrange(anno, mese)[1]:02d}"
    iscrizione = connection.execute(
        """
        SELECT
            ic.id,
            ic.quota_mensile,
            COALESCE(NULLIF(ic.data_inizio, ''), ic.data_iscrizione, ?) AS data_attivazione,
            COALESCE(NULLIF(ic.data_fine, ''), '') AS data_fine_iscrizione,
            COALESCE(NULLIF(c.data_inizio, ''), '') AS data_inizio_corso,
            COALESCE(NULLIF(c.data_fine, ''), '') AS data_fine_corso
        FROM iscrizioni_corsi ic
        JOIN corsi c ON c.id = ic.corso_id
        WHERE ic.id = ?
        """,
        (last_day, iscrizione_corso_id),
    ).fetchone()
    if iscrizione is None:
        raise ValueError("Iscrizione corso non trovata.")

    effective_start, effective_end = compute_course_enrollment_effective_window(
        str(iscrizione["data_attivazione"] or ""),
        str(iscrizione["data_fine_iscrizione"] or ""),
        str(iscrizione["data_inizio_corso"] or ""),
        str(iscrizione["data_fine_corso"] or ""),
    )
    if effective_start and effective_start > last_day:
        raise ValueError("La quota del mese di iscrizione non puo essere generata prima della data di attivazione.")
    if effective_end and effective_end < first_day:
        raise ValueError("L'iscrizione corso risulta gia conclusa per il mese selezionato.")

    data_scadenza = last_day
    connection.execute(
        """
        INSERT OR IGNORE INTO rate_corsi_mensili (
            iscrizione_corso_id, anno, mese, importo_dovuto, data_scadenza, note
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            iscrizione_corso_id,
            anno,
            mese,
            iscrizione["quota_mensile"],
            data_scadenza,
            note or None,
        ),
    )
    row = connection.execute(
        """
        SELECT id
        FROM rate_corsi_mensili
        WHERE iscrizione_corso_id = ? AND anno = ? AND mese = ?
        """,
        (iscrizione_corso_id, anno, mese),
    ).fetchone()
    if row is None:
        raise ValueError("Impossibile generare la quota mensile del mese di iscrizione.")
    return int(row["id"])


def ensure_course_rates_for_enrollment_range(
    connection: sqlite3.Connection,
    iscrizione_corso_id: int,
    start_year: int,
    start_month: int,
    end_year: int,
    end_month: int,
    *,
    note: str = "",
) -> list[int]:
    if start_month < 1 or start_month > 12 or end_month < 1 or end_month > 12:
        raise ValueError("Il mese deve essere compreso tra 1 e 12.")
    if (end_year, end_month) < (start_year, start_month):
        raise ValueError("Il mese finale deve essere uguale o successivo al mese iniziale.")

    rate_ids: list[int] = []
    current_year = start_year
    current_month = start_month
    while (current_year, current_month) <= (end_year, end_month):
        rate_ids.append(
            ensure_course_rate_for_enrollment(
                connection,
                iscrizione_corso_id,
                current_year,
                current_month,
                note=note,
            )
        )
        if current_month == 12:
            current_year += 1
            current_month = 1
        else:
            current_month += 1
    return rate_ids


def auto_generate_course_rates_on_startup(today_value: date | None = None) -> dict[str, object]:
    today_value = today_value or date.today()
    work_year = int(today_value.year)
    generated_months: list[str] = []
    inserted_total = 0
    skipped_total = 0
    with get_connection() as connection:
        for mese in range(1, today_value.month + 1):
            inserted, skipped = generate_course_rates_for_month(
                connection,
                work_year,
                mese,
                data_scadenza=f"{work_year:04d}-{mese:02d}-{monthrange(work_year, mese)[1]:02d}",
                note="Generazione automatica all'avvio del software",
            )
            inserted_total += inserted
            skipped_total += skipped
            if inserted:
                generated_months.append(f"{month_label(mese)} {work_year}")
        connection.commit()
    return {
        "work_year": work_year,
        "generated_months": generated_months,
        "inserted_total": inserted_total,
        "skipped_total": skipped_total,
    }


def export_activity_log_xls() -> None:
    ensure_output_dir()
    with get_connection() as connection:
        if not table_exists(connection, "registro_attivita"):
            return
        prune_old_activity_log_rows(connection)
        connection.commit()
        rows = connection.execute(
            """
            SELECT
                id,
                data_ora,
                substr(data_ora, 1, 10) AS data,
                substr(data_ora, 12, 8) AS ora,
                COALESCE(username, '') AS username,
                COALESCE(associato_codice, '') AS associato_codice,
                COALESCE(associato_nominativo, '') AS associato_nominativo,
                COALESCE(nome_pc, '') AS nome_pc,
                COALESCE(ip_client, '') AS ip_client,
                COALESCE(metodo_http, '') AS metodo_http,
                COALESCE(percorso, '') AS percorso,
                COALESCE(categoria, '') AS categoria,
                COALESCE(descrizione_attivita, '') AS descrizione_attivita,
                COALESCE(dettaglio, '') AS dettaglio,
                COALESCE(esito, '') AS esito,
                COALESCE(anno_lavoro, '') AS anno_lavoro,
                COALESCE(user_agent, '') AS user_agent
            FROM registro_attivita
            ORDER BY data_ora DESC, id DESC
            """
        ).fetchall()

    headers = [
        "ID",
        "Data",
        "Ora",
        "DataOra",
        "Utente",
        "Codice",
        "Associato",
        "Nome PC",
        "IP client",
        "Metodo",
        "Percorso",
        "Categoria",
        "Attivita",
        "Dettaglio",
        "Esito",
        "Anno lavoro",
        "User agent",
    ]
    workbook_rows = [headers]
    for row in rows:
        workbook_rows.append(
            [
                str(row["id"]),
                row["data"],
                row["ora"],
                row["data_ora"],
                row["username"],
                row["associato_codice"],
                row["associato_nominativo"],
                row["nome_pc"],
                row["ip_client"],
                row["metodo_http"],
                row["percorso"],
                row["categoria"],
                row["descrizione_attivita"],
                row["dettaglio"],
                row["esito"],
                str(row["anno_lavoro"]),
                row["user_agent"],
            ]
        )

    def xml_cell(value: object) -> str:
        safe_value = html.escape("" if value is None else str(value))
        return f'<Cell><Data ss:Type="String">{safe_value}</Data></Cell>'

    xml_rows = [
        "<Row>" + "".join(xml_cell(cell) for cell in row_values) + "</Row>"
        for row_values in workbook_rows
    ]
    workbook = (
        '<?xml version="1.0"?>\n'
        '<?mso-application progid="Excel.Sheet"?>\n'
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"\n'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel"\n'
        ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\n'
        ' <Worksheet ss:Name="Registro attivita">\n'
        "  <Table>\n"
        f"   {''.join(xml_rows)}\n"
        "  </Table>\n"
        " </Worksheet>\n"
        "</Workbook>\n"
    )
    ACTIVITY_LOG_XLS_PATH.write_text(workbook, encoding="utf-8")


def record_activity(
    *,
    username: str,
    path: str,
    method: str,
    description: str,
    category: str,
    outcome: str,
    work_year: int | None = None,
    detail: str = "",
    associato_id: int | None = None,
    associato_codice: str = "",
    associato_nominativo: str = "",
    ip_client: str = "",
    user_agent: str = "",
) -> None:
    try:
        with get_connection() as connection:
            if not table_exists(connection, "registro_attivita"):
                return
            prune_old_activity_log_rows(connection)
            connection.execute(
                """
                INSERT INTO registro_attivita (
                    data_ora,
                    username,
                    associato_id,
                    associato_codice,
                    associato_nominativo,
                    nome_pc,
                    ip_client,
                    metodo_http,
                    percorso,
                    categoria,
                    descrizione_attivita,
                    dettaglio,
                    esito,
                    anno_lavoro,
                    user_agent
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    current_timestamp(),
                    username,
                    associato_id,
                    associato_codice,
                    associato_nominativo,
                    os.environ.get("COMPUTERNAME") or os.environ.get("HOSTNAME") or "",
                    ip_client,
                    method,
                    path,
                    category,
                    description,
                    detail,
                    outcome,
                    work_year,
                    user_agent,
                ),
            )
            connection.commit()
        export_activity_log_xls()
    except Exception:
        pass


def iscrizione_corso_year_relevance_sql(alias: str = "ic") -> str:
    return f"""
        (
            substr(COALESCE(NULLIF({alias}.data_inizio, ''), {alias}.data_iscrizione, ''), 1, 4) = ?
            OR EXISTS (
                SELECT 1
                FROM rate_corsi_mensili r
                WHERE r.iscrizione_corso_id = {alias}.id
                  AND r.anno = ?
            )
        )
    """


def iscrizione_corso_year_relevance_params(work_year: int) -> tuple[object, ...]:
    return (str(work_year), work_year)


def associato_year_relevance_sql(alias: str = "a") -> str:
    return f"""
        (
            substr(COALESCE({alias}.data_prima_iscrizione, ''), 1, 4) = ?
            OR EXISTS (
                SELECT 1
                FROM tesseramenti_annuali t
                WHERE t.associato_id = {alias}.id
                  AND t.anno_sociale = ?
            )
            OR EXISTS (
                SELECT 1
                FROM iscrizioni_corsi ic
                WHERE ic.associato_id = {alias}.id
                  AND {iscrizione_corso_year_relevance_sql('ic')}
            )
            OR EXISTS (
                SELECT 1
                FROM iscrizioni_campi_estivi ice
                JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
                WHERE ice.associato_id = {alias}.id
                  AND ce.anno = ?
            )
            OR EXISTS (
                SELECT 1
                FROM iscrizioni_eventi ie
                JOIN eventi e ON e.id = ie.evento_id
                WHERE ie.associato_id = {alias}.id
                  AND substr(COALESCE(e.data_evento, ''), 1, 4) = ?
            )
        )
    """


def associato_year_relevance_params(work_year: int) -> tuple[object, ...]:
    return (
        str(work_year),
        work_year,
        *iscrizione_corso_year_relevance_params(work_year),
        work_year,
        str(work_year),
    )


def corso_year_relevance_sql(alias: str = "c") -> str:
    return f"""
        (
            substr(COALESCE({alias}.creato_il, ''), 1, 4) = ?
            OR EXISTS (
                SELECT 1
                FROM iscrizioni_corsi ic
                WHERE ic.corso_id = {alias}.id
                  AND {iscrizione_corso_year_relevance_sql('ic')}
            )
        )
    """


def corso_year_relevance_params(work_year: int) -> tuple[object, ...]:
    return (str(work_year), *iscrizione_corso_year_relevance_params(work_year))


def available_work_years(selected_year: int) -> list[int]:
    rows = fetch_all(
        """
        WITH anni AS (
            SELECT anno_sociale AS anno FROM tesseramenti_annuali
            UNION
            SELECT anno FROM rate_corsi_mensili
            UNION
            SELECT anno FROM campi_estivi
            UNION
            SELECT CAST(strftime('%Y', data_evento) AS INTEGER) FROM eventi
            UNION
            SELECT CAST(strftime('%Y', data_prima_iscrizione) AS INTEGER) FROM associati
        )
        SELECT DISTINCT anno
        FROM anni
        WHERE anno IS NOT NULL
        ORDER BY anno DESC
        """
    )
    values = {selected_year, date.today().year}
    values.update(int(row["anno"]) for row in rows if row["anno"])
    return sorted(values, reverse=True)


def month_label(month_number: int) -> str:
    return MONTH_NAMES.get(month_number, str(month_number))


def clean_phone_number(value: str | None) -> str:
    digits = re.sub(r"\D", "", value or "")
    if digits.startswith("00"):
        digits = digits[2:]
    if not digits:
        return ""
    if digits.startswith("39"):
        return digits
    if digits.startswith("0"):
        return f"39{digits.lstrip('0')}"
    if digits.startswith("3"):
        return f"39{digits}"
    return digits


def plain_text(value: object) -> str:
    return re.sub(r"<[^>]+>", "", "" if value is None else str(value)).strip()


def normalize_searchable_text(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", plain_text(value))
    ascii_only = "".join(character for character in normalized if not unicodedata.combining(character))
    return ascii_only.lower().strip()


def normalize_codice_fiscale(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (value or "").upper())[:16]


def normalize_lookup_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", (value or "").strip())
    ascii_only = "".join(character for character in normalized if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]", "", ascii_only.upper())


def calculate_age(birth_date_value: str | None, *, today: date | None = None) -> int | None:
    if not birth_date_value:
        return None
    try:
        born = date.fromisoformat(birth_date_value)
    except ValueError:
        return None
    reference = today or date.today()
    years = reference.year - born.year
    if (reference.month, reference.day) < (born.month, born.day):
        years -= 1
    return max(years, 0)


def label_with_age(label: str, birth_date_value: str | None) -> str:
    age = calculate_age(birth_date_value)
    if age is None:
        return label
    return f"{label} ({age} anni)"


def is_minor_birth_date(birth_date_value: str | None, *, today: date | None = None) -> bool:
    age = calculate_age(birth_date_value, today=today)
    return age is not None and age < 18


def resolve_minor_contact_channels(data: dict[str, object]) -> dict[str, str]:
    is_minor = is_minor_birth_date(str(data.get("data_nascita") or ""))
    email = str(data.get("email") or "").strip()
    telefono = str(data.get("telefono") or "").strip()
    guardian_email = str(data.get("genitore_tutore_email") or "").strip()
    guardian_phone = str(data.get("genitore_tutore_cellulare") or "").strip()
    if is_minor:
        return {
            "email": guardian_email or email,
            "telefono": guardian_phone or telefono,
        }
    return {
        "email": email,
        "telefono": telefono,
    }


@lru_cache(maxsize=1)
def comuni_records() -> tuple[dict[str, object], ...]:
    if not COMUNI_JSON_PATH.exists():
        return ()
    try:
        data = json.loads(COMUNI_JSON_PATH.read_text(encoding="utf-8"))
    except Exception:
        return ()

    records: list[dict[str, object]] = []
    for item in data:
        nome = str(item.get("nome") or "").strip()
        sigla = str(item.get("sigla") or "").strip()
        codice_catastale = str(item.get("codiceCatastale") or "").upper().strip()
        caps = [str(cap).strip() for cap in (item.get("cap") or []) if str(cap).strip()]
        if not nome:
            continue
        records.append(
            {
                "nome": nome,
                "sigla": sigla,
                "codice_catastale": codice_catastale,
                "caps": caps,
                "key": normalize_lookup_key(nome),
            }
        )
    return tuple(records)


@lru_cache(maxsize=1)
def comuni_lookup_by_catastale() -> dict[str, tuple[str, str]]:
    lookup: dict[str, tuple[str, str]] = {}
    for item in comuni_records():
        code = str(item["codice_catastale"])
        if not code:
            continue
        comune = str(item["nome"])
        provincia = str(item["sigla"])
        lookup[code] = (comune, provincia)
    return lookup


@lru_cache(maxsize=1)
def comuni_lookup_by_name() -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {}
    for item in comuni_records():
        key = str(item["key"])
        if key and key not in lookup:
            lookup[key] = dict(item)
    return lookup


@lru_cache(maxsize=1)
def comuni_lookup_by_cap() -> dict[str, list[dict[str, object]]]:
    lookup: dict[str, list[dict[str, object]]] = {}
    for item in comuni_records():
        for cap in item["caps"]:
            lookup.setdefault(cap, []).append(dict(item))
    for cap, items in lookup.items():
        items.sort(key=lambda row: (str(row["nome"]), str(row["sigla"])))
    return lookup


def lookup_comune_details(value: str) -> dict[str, str] | None:
    key = normalize_lookup_key(value)
    if not key:
        return None
    item = comuni_lookup_by_name().get(key)
    if item is None:
        partial_matches = [record for record in comuni_records() if str(record["key"]).startswith(key)]
        if len(partial_matches) == 1:
            item = dict(partial_matches[0])
    if item is None:
        return None
    caps = [str(cap).strip() for cap in (item.get("caps") or []) if str(cap).strip()]
    return {
        "comune": str(item["nome"]),
        "provincia": str(item["sigla"]),
        "codice_catastale": str(item["codice_catastale"]),
        "cap": caps[0] if caps else "",
    }


def lookup_cap_details(value: str) -> dict[str, object] | None:
    cap = re.sub(r"\D", "", value or "")[:5]
    if len(cap) != 5:
        return None
    matches = comuni_lookup_by_cap().get(cap, [])
    if not matches:
        return None
    first = matches[0]
    province_values = sorted({str(item["sigla"]) for item in matches if str(item["sigla"])})
    city_values = sorted({str(item["nome"]) for item in matches if str(item["nome"])})
    return {
        "cap": cap,
        "citta": str(first["nome"]),
        "provincia": str(first["sigla"]),
        "ambiguous": len(city_values) > 1,
        "candidates": city_values[:10],
        "province_candidates": province_values[:10],
    }


def resolve_birth_year(two_digits: int, month: int, day: int) -> int | None:
    reference = date.today()
    candidates = [1900 + two_digits, 2000 + two_digits]
    valid_candidates: list[tuple[int, int]] = []
    for year_value in candidates:
        try:
            candidate_date = date(year_value, month, day)
        except ValueError:
            continue
        if candidate_date > reference:
            continue
        age = calculate_age(candidate_date.isoformat())
        if age is None:
            continue
        if 0 <= age <= 115:
            valid_candidates.append((age, year_value))

    if valid_candidates:
        valid_candidates.sort(key=lambda item: (item[0], item[1]))
        return valid_candidates[0][1]

    for year_value in candidates:
        try:
            date(year_value, month, day)
            return year_value
        except ValueError:
            continue
    return None


def decode_codice_fiscale_birth_data(value: str) -> dict[str, str] | None:
    codice_fiscale = normalize_codice_fiscale(value)
    if len(codice_fiscale) != 16 or not re.fullmatch(r"[A-Z0-9]{16}", codice_fiscale):
        return None

    month = CODICE_FISCALE_MONTHS.get(codice_fiscale[8])
    if month is None:
        return None

    try:
        two_digits_year = int(codice_fiscale[6:8])
        encoded_day = int(codice_fiscale[9:11])
    except ValueError:
        return None

    is_female = encoded_day > 40
    day = encoded_day - 40 if is_female else encoded_day
    year_value = resolve_birth_year(two_digits_year, month, day)
    if year_value is None:
        return None

    try:
        birth_date_value = date(year_value, month, day).isoformat()
    except ValueError:
        return None

    comune_nascita = ""
    provincia_nascita = ""
    birth_place = comuni_lookup_by_catastale().get(codice_fiscale[11:15])
    if birth_place:
        comune_nascita, provincia_nascita = birth_place

    return {
        "codice_fiscale": codice_fiscale,
        "data_nascita": birth_date_value,
        "comune_nascita": comune_nascita,
        "provincia_nascita": provincia_nascita,
        "sesso": "F" if is_female else "M",
    }


def codice_fiscale_token_from_text(value: str, *, is_name: bool = False) -> str:
    cleaned = re.sub(r"[^A-Z]", "", normalize_lookup_key(value))
    consonants = [character for character in cleaned if character not in "AEIOU"]
    vowels = [character for character in cleaned if character in "AEIOU"]
    if is_name and len(consonants) >= 4:
        code = consonants[0] + consonants[2] + consonants[3]
    else:
        code = "".join(consonants[:3])
    if len(code) < 3:
        code += "".join(vowels[: 3 - len(code)])
    return (code + "XXX")[:3]


def codice_fiscale_control_char(partial_code: str) -> str:
    total = 0
    for index, character in enumerate(partial_code.upper(), start=1):
        if index % 2 == 1:
            total += CODICE_FISCALE_ODD_MAP[character]
        else:
            total += CODICE_FISCALE_EVEN_MAP[character]
    return CODICE_FISCALE_CONTROL_CHARS[total % 26]


def calculate_codice_fiscale(
    *,
    nome: str,
    cognome: str,
    data_nascita: str,
    sesso: str,
    comune_nascita: str,
) -> dict[str, str] | None:
    if not all([nome.strip(), cognome.strip(), data_nascita.strip(), sesso.strip(), comune_nascita.strip()]):
        return None
    try:
        born = date.fromisoformat(data_nascita)
    except ValueError:
        return None

    comune = lookup_comune_details(comune_nascita)
    if comune is None or not comune["codice_catastale"]:
        return None

    gender = (sesso or "").strip().upper()
    if gender not in {"M", "F"}:
        return None

    surname_code = codice_fiscale_token_from_text(cognome)
    name_code = codice_fiscale_token_from_text(nome, is_name=True)
    year_code = f"{born.year % 100:02d}"
    month_code = CODICE_FISCALE_MONTH_CODES[born.month]
    day_number = born.day + (40 if gender == "F" else 0)
    day_code = f"{day_number:02d}"
    partial_code = f"{surname_code}{name_code}{year_code}{month_code}{day_code}{comune['codice_catastale']}"
    control = codice_fiscale_control_char(partial_code)
    return {
        "codice_fiscale": partial_code + control,
        "comune_nascita": comune["comune"],
        "provincia_nascita": comune["provincia"],
    }


def json_response(start_response, payload: dict, *, status: str = "200 OK") -> list[bytes]:
    content = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    start_response(
        status,
        [
            ("Content-Type", "application/json; charset=utf-8"),
            ("Content-Length", str(len(content))),
        ],
    )
    return [content]


def normalize_username(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").strip()).lower()


def validate_username(value: str) -> str:
    username = normalize_username(value)
    if not re.fullmatch(r"[a-z0-9._-]{3,32}", username):
        raise ValueError("Lo username deve contenere da 3 a 32 caratteri: lettere, numeri, punto, trattino o underscore.")
    return username


def validate_password(password: str, confirmation: str) -> str:
    if len(password) < 8:
        raise ValueError("La password deve contenere almeno 8 caratteri.")
    if password != confirmation:
        raise ValueError("La conferma password non corrisponde.")
    return password


def hash_password(password: str, *, salt: bytes | None = None, iterations: int = PASSWORD_HASH_ITERATIONS) -> tuple[str, str, int]:
    current_salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), current_salt, iterations)
    return current_salt.hex(), digest.hex(), iterations


def verify_password(password: str, row: sqlite3.Row) -> bool:
    salt = bytes.fromhex(row["password_salt"])
    iterations = int(row["password_iterations"] or PASSWORD_HASH_ITERATIONS)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations).hex()
    return hmac.compare_digest(digest, row["password_hash"])


def current_timestamp() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def years_ago_timestamp(years: int) -> str:
    reference = datetime.now().replace(microsecond=0)
    try:
        target = reference.replace(year=reference.year - years)
    except ValueError:
        # Gestisce il 29 febbraio negli anni non bisestili.
        target = reference.replace(year=reference.year - years, day=28)
    return target.isoformat(sep=" ")


def prune_old_activity_log_rows(connection: sqlite3.Connection) -> int:
    if not table_exists(connection, "registro_attivita"):
        return 0
    cursor = connection.execute(
        """
        DELETE FROM registro_attivita
        WHERE data_ora < ?
        """,
        (years_ago_timestamp(ACTIVITY_LOG_RETENTION_YEARS),),
    )
    return max(int(cursor.rowcount or 0), 0)


def build_cookie_header(name: str, value: str, *, max_age: int | None = None) -> tuple[str, str]:
    cookie = SimpleCookie()
    cookie[name] = value
    morsel = cookie[name]
    morsel["path"] = "/"
    morsel["httponly"] = True
    morsel["samesite"] = "Lax"
    if max_age is not None:
        morsel["max-age"] = str(max_age)
    return "Set-Cookie", morsel.OutputString()


def session_cookie_header(token: str) -> tuple[str, str]:
    return build_cookie_header(
        SESSION_COOKIE_NAME,
        token,
        max_age=SESSION_DURATION_DAYS * 24 * 60 * 60,
    )


def clear_session_cookie_header() -> tuple[str, str]:
    return build_cookie_header(SESSION_COOKIE_NAME, "", max_age=0)


def get_connection() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def fetch_all(query: str, params: tuple | list = ()) -> list[sqlite3.Row]:
    with get_connection() as connection:
        return connection.execute(query, params).fetchall()


def fetch_one(query: str, params: tuple | list = ()) -> sqlite3.Row | None:
    with get_connection() as connection:
        return connection.execute(query, params).fetchone()


def fetch_scalar(query: str, params: tuple | list = ()) -> object:
    row = fetch_one(query, params)
    if row is None:
        return None
    return row[0]


def execute(query: str, params: tuple | list = ()) -> None:
    with get_connection() as connection:
        connection.execute(query, params)
        connection.commit()

UploadedFile = dict[str, object]
ParsedRequest = tuple[str, str, dict[str, str], dict[str, str], dict[str, UploadedFile], dict[str, str]]


def parse_multipart_form(payload: bytes, content_type: str) -> tuple[dict[str, str], dict[str, UploadedFile]]:
    match = re.search(r'boundary="?([^";]+)"?', content_type, re.IGNORECASE)
    if not match:
        return {}, {}

    boundary = match.group(1).encode("utf-8", "ignore")
    delimiter = b"--" + boundary
    form_data: dict[str, str] = {}
    form_files: dict[str, UploadedFile] = {}

    for raw_part in payload.split(delimiter):
        part = raw_part.strip()
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2]
        part = part.strip(b"\r\n")
        if not part:
            continue
        header_blob, separator, body = part.partition(b"\r\n\r\n")
        if not separator:
            continue
        headers: dict[str, str] = {}
        for line in header_blob.decode("utf-8", "ignore").split("\r\n"):
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            headers[key.strip().lower()] = value.strip()
        disposition = headers.get("content-disposition", "")
        if not disposition:
            continue
        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue
        field_name = name_match.group(1).strip()
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        if filename_match:
            filename = Path(filename_match.group(1).strip()).name
            if not filename:
                continue
            form_files[field_name] = {
                "filename": filename,
                "content_type": headers.get("content-type", "application/octet-stream"),
                "content": body,
                "size": len(body),
            }
            continue
        value = body.decode("utf-8", "ignore")
        if field_name in form_data and form_data[field_name]:
            form_data[field_name] = f"{form_data[field_name]},{value}"
        else:
            form_data[field_name] = value

    return form_data, form_files


def parse_request(environ: dict) -> ParsedRequest:
    path = environ.get("PATH_INFO") or "/"
    method = (environ.get("REQUEST_METHOD") or "GET").upper()
    query_params = {
        key: ",".join(values)
        for key, values in parse_qs(
            environ.get("QUERY_STRING", ""),
            keep_blank_values=True,
        ).items()
    }

    form_data: dict[str, str] = {}
    form_files: dict[str, UploadedFile] = {}
    if method == "POST":
        size = int(environ.get("CONTENT_LENGTH") or 0)
        payload = environ["wsgi.input"].read(size)
        content_type = (environ.get("CONTENT_TYPE") or "").strip()
        if content_type.lower().startswith("multipart/form-data"):
            form_data, form_files = parse_multipart_form(payload, content_type)
        else:
            form_data = {
                key: ",".join(values)
                for key, values in parse_qs(payload.decode("utf-8"), keep_blank_values=True).items()
            }

    cookie_jar = SimpleCookie()
    cookie_jar.load(environ.get("HTTP_COOKIE", ""))
    cookies = {key: morsel.value for key, morsel in cookie_jar.items()}

    return path, method, query_params, form_data, form_files, cookies


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value))


def normalized(form_data: dict[str, str], key: str, default: str = "") -> str:
    return (form_data.get(key, default) or "").strip()


def multi_values(form_data: dict[str, str], key: str) -> list[str]:
    raw_value = form_data.get(key, "") or ""
    if not raw_value:
        return []
    return [item.strip() for item in raw_value.split(",") if item.strip()]


def required(form_data: dict[str, str], key: str, label: str) -> str:
    value = normalized(form_data, key)
    if not value:
        raise ValueError(f"Il campo '{label}' e obbligatorio.")
    return value


def optional(form_data: dict[str, str], key: str) -> str | None:
    value = normalized(form_data, key)
    return value or None


def money(value: object) -> str:
    if value in (None, ""):
        return ""
    amount = float(value)
    rendered = f"{amount:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{rendered} EUR"


def decimal_amount(value: object, *, minimum: str = "0.00") -> Decimal:
    amount = Decimal(str(value or "0")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    floor = Decimal(minimum).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if amount < floor:
        raise ValueError("L'importo indicato non e valido.")
    return amount


def decimal_input(value: object) -> str:
    return format(decimal_amount(value), ".2f")


def decimal_or_none(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip()).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return None


def summary_label_index(columns: list[tuple]) -> int:
    for index, column in enumerate(columns):
        label = (column[1] or "").lower()
        if label not in {"ricevuta", "azioni"}:
            return index
    return 0


def summary_total_indexes(columns: list[tuple]) -> list[int]:
    total_hints = ("importo", "dovuto", "pagato", "residuo", "quota", "saldo")
    indexes: list[int] = []
    for index, column in enumerate(columns):
        label = str(column[1] or "").lower()
        if any(hint in label for hint in total_hints):
            indexes.append(index)
    return indexes


def summary_rows_for_table(rows: list[sqlite3.Row], columns: list[tuple]) -> list[list[str]]:
    if not rows:
        return []

    total_indexes = summary_total_indexes(columns)
    totals: dict[int, str] = {}
    for index in total_indexes:
        column = columns[index]
        total = Decimal("0.00")
        found = False
        for row in rows:
            amount = decimal_or_none(row[column[0]])
            if amount is None:
                continue
            total += amount
            found = True
        if found:
            totals[index] = money(total)

    if not totals:
        return []

    summary_row = [""] * len(columns)
    summary_row[summary_label_index(columns)] = "Totali"
    for index, value in totals.items():
        summary_row[index] = value
    return [summary_row]


def option_data_attrs(row: sqlite3.Row, data_keys: list[str] | None = None) -> str:
    if not data_keys:
        return ""
    row_keys = set(row.keys())
    attrs: list[str] = []
    for key in data_keys:
        if key not in row_keys:
            continue
        value = row[key]
        if value in (None, ""):
            continue
        html_key = key.replace("_", "-")
        attrs.append(f' data-{esc(html_key)}="{esc(value)}"')
    return "".join(attrs)


def render_associato_options(
    rows: list[sqlite3.Row],
    selected: str | None = None,
    *,
    extra_data_keys: list[str] | None = None,
) -> str:
    data_keys = ["search_text", "autocomplete_label"]
    for key in extra_data_keys or []:
        if key not in data_keys:
            data_keys.append(key)
    return render_select_options(rows, selected, data_keys=data_keys)


def render_select_options(
    rows: list[sqlite3.Row],
    selected: str | None = None,
    blank_label: str = "Seleziona...",
    data_keys: list[str] | None = None,
) -> str:
    items = [f'<option value="">{esc(blank_label)}</option>']
    selected_value = "" if selected is None else str(selected)
    for row in rows:
        value = str(row["id"])
        label = row["label"]
        selected_attr = ' selected="selected"' if value == selected_value else ""
        items.append(
            f'<option value="{esc(value)}"{selected_attr}{option_data_attrs(row, data_keys)}>{esc(label)}</option>'
        )
    return "".join(items)


def render_select_options_multi(
    rows: list[sqlite3.Row],
    selected_values: list[str] | None = None,
    data_keys: list[str] | None = None,
) -> str:
    selected = set(selected_values or [])
    items = []
    for row in rows:
        value = str(row["id"])
        label = row["label"]
        selected_attr = ' selected="selected"' if value in selected else ""
        items.append(
            f'<option value="{esc(value)}"{selected_attr}{option_data_attrs(row, data_keys)}>{esc(label)}</option>'
        )
    return "".join(items)


def render_static_options(
    options: list[tuple[str, str]],
    selected: str | None = None,
    *,
    blank_label: str | None = None,
) -> str:
    items: list[str] = []
    if blank_label is not None:
        items.append(f'<option value="">{esc(blank_label)}</option>')

    selected_value = "" if selected is None else str(selected)
    for value, label in options:
        selected_attr = ' selected="selected"' if str(value) == selected_value else ""
        items.append(f'<option value="{esc(value)}"{selected_attr}>{esc(label)}</option>')
    return "".join(items)


def hidden_fields_html(hidden_fields: dict[str, str] | None = None) -> str:
    if not hidden_fields:
        return ""
    return "".join(
        f'<input type="hidden" name="{esc(key)}" value="{esc(value)}">'
        for key, value in hidden_fields.items()
    )


def input_field(
    label: str,
    name: str,
    *,
    input_type: str = "text",
    value: str = "",
    required_field: bool = False,
    step: str | None = None,
    minimum: str | None = None,
    placeholder: str = "",
    wide: bool = False,
    element_id: str | None = None,
    attrs: dict[str, str] | None = None,
    revealable: bool = False,
    wrapper_attrs: dict[str, str] | None = None,
) -> str:
    attr_parts = [
        f'type="{esc(input_type)}"',
        f'name="{esc(name)}"',
        'class="control"',
    ]
    if input_type != "file":
        attr_parts.insert(2, f'value="{esc(value)}"')
    if element_id:
        attr_parts.append(f'id="{esc(element_id)}"')
    if required_field:
        attr_parts.append("required")
    if step is not None:
        attr_parts.append(f'step="{esc(step)}"')
    if minimum is not None:
        attr_parts.append(f'min="{esc(minimum)}"')
    if placeholder:
        attr_parts.append(f'placeholder="{esc(placeholder)}"')
    if attrs is not None:
        for key, attr_value in attrs.items():
            attr_parts.append(f'{esc(key)}="{esc(attr_value)}"')
    class_name = "field wide" if wide else "field"
    wrapper_extra_attrs = ""
    if wrapper_attrs is not None:
        wrapper_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in wrapper_attrs.items())
    input_html = f"<input {' '.join(attr_parts)}>"
    if revealable and input_type == "password":
        input_html = (
            '<div class="password-field-wrap">'
            f"{input_html}"
            '<button type="button" class="password-toggle" onclick="togglePasswordVisibility(this)">Mostra</button>'
            "</div>"
        )
    return (
        f'<label class="{class_name}"{wrapper_extra_attrs}><span>{esc(label)}</span>'
        f"{input_html}</label>"
    )


def readonly_field(
    label: str,
    value: str,
    *,
    wide: bool = False,
    element_id: str | None = None,
    wrapper_attrs: dict[str, str] | None = None,
) -> str:
    class_name = "field wide" if wide else "field"
    id_attr = f' id="{esc(element_id)}"' if element_id else ""
    wrapper_extra_attrs = ""
    if wrapper_attrs is not None:
        wrapper_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in wrapper_attrs.items())
    return (
        f'<label class="{class_name}"{wrapper_extra_attrs}><span>{esc(label)}</span>'
        f'<div{id_attr} class="control readonly-control">{esc(value)}</div></label>'
    )


def select_field(
    label: str,
    name: str,
    options_html: str,
    *,
    required_field: bool = False,
    wide: bool = False,
    element_id: str | None = None,
    attrs: dict[str, str] | None = None,
    control_class: str = "control",
    searchable: bool = False,
    search_placeholder: str = "Cerca per codice, nome, telefono o email...",
    wrapper_attrs: dict[str, str] | None = None,
) -> str:
    class_name = "field wide" if wide else "field"
    required_attr = " required" if required_field and not searchable else ""
    search_required_attr = " required" if required_field and searchable else ""
    select_id = element_id or (f"{name}-{uuid.uuid4().hex[:8]}" if searchable else None)
    panel_id = f"{select_id}-autocomplete" if searchable and select_id else None
    extra_attrs = ""
    if select_id:
        extra_attrs += f' id="{esc(select_id)}"'
    if attrs:
        extra_attrs += "".join(f' {esc(key)}="{esc(value)}"' for key, value in attrs.items())
    wrapper_extra_attrs = ""
    if wrapper_attrs is not None:
        wrapper_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in wrapper_attrs.items())
    select_class = control_class if not searchable else f"{control_class} searchable-select-source".strip()
    search_html = ""
    if searchable and select_id:
        search_html = (
            f'<div class="select-autocomplete">'
            f'<input type="search" class="control select-search-control" '
            f'placeholder="{esc(search_placeholder)}" data-select-search-target="{esc(select_id)}" '
            f'data-select-search-panel="{esc(panel_id or "")}" '
            f'oninput="handleSelectSearch(this)" onfocus="openSelectAutocomplete(this)" '
            f'onkeydown="handleSelectSearchKeydown(event, this)" onblur="closeSelectAutocompleteLater(this)" '
            f'autocomplete="off"{search_required_attr}>'
            f'<div id="{esc(panel_id or "")}" class="select-autocomplete-panel" hidden></div>'
            "</div>"
        )
    return (
        f'<label class="{class_name}"{wrapper_extra_attrs}><span>{esc(label)}</span>'
        f"{search_html}"
        f'<select name="{esc(name)}" class="{esc(select_class)}"{required_attr}{extra_attrs}>{options_html}</select>'
        "</label>"
    )


def multi_select_field(
    label: str,
    name: str,
    options_html: str,
    *,
    required_field: bool = False,
    wide: bool = True,
    size: int = 8,
    element_id: str | None = None,
    attrs: dict[str, str] | None = None,
) -> str:
    class_name = "field wide" if wide else "field"
    required_attr = " required" if required_field else ""
    extra_attrs = ""
    if element_id:
        extra_attrs += f' id="{esc(element_id)}"'
    if attrs:
        extra_attrs += "".join(f' {esc(key)}="{esc(value)}"' for key, value in attrs.items())
    return (
        f'<label class="{class_name}"><span>{esc(label)}</span>'
        f'<select name="{esc(name)}" class="control multi-control" multiple size="{size}"{required_attr}{extra_attrs}>{options_html}</select>'
        "</label>"
    )


def textarea_field(
    label: str,
    name: str,
    *,
    value: str = "",
    rows: int = 4,
    wide: bool = True,
    attrs: dict[str, str] | None = None,
    wrapper_attrs: dict[str, str] | None = None,
    ) -> str:
    class_name = "field wide" if wide else "field"
    extra_attrs = ""
    if attrs is not None:
        extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in attrs.items())
    wrapper_extra_attrs = ""
    if wrapper_attrs is not None:
        wrapper_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in wrapper_attrs.items())
    return (
        f'<label class="{class_name}"{wrapper_extra_attrs}><span>{esc(label)}</span>'
        f'<textarea name="{esc(name)}" rows="{rows}" class="control"{extra_attrs}>{esc(value)}</textarea>'
        "</label>"
    )


def inline_fields_row(
    fields_html: list[str],
    *,
    row_class: str = "two-up",
    wrapper_attrs: dict[str, str] | None = None,
) -> str:
    wrapper_extra_attrs = ""
    if wrapper_attrs is not None:
        wrapper_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in wrapper_attrs.items())
    return (
        f'<div class="form-inline-row {esc(row_class)}"{wrapper_extra_attrs}>'
        f'{"".join(fields_html)}'
        "</div>"
    )


def form_section_block(
    title: str,
    subtitle: str = "",
    *,
    attrs: dict[str, str] | None = None,
) -> str:
    extra_attrs = ""
    if attrs is not None:
        extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in attrs.items())
    subtitle_html = f"<p>{esc(subtitle)}</p>" if subtitle else ""
    return (
        f'<div class="form-section-block field wide"{extra_attrs}>'
        f'<div class="form-section-head"><strong>{esc(title)}</strong>{subtitle_html}</div>'
        "</div>"
    )


def form_card(
    title: str,
    subtitle: str,
    action: str,
    fields_html: str,
    button_label: str,
    *,
    hidden_fields: dict[str, str] | None = None,
    card_class: str = "",
    form_attrs: dict[str, str] | None = None,
) -> str:
    section_class = f"card {card_class}".strip()
    form_extra_attrs = ""
    if form_attrs:
        form_extra_attrs = "".join(f' {esc(key)}="{esc(value)}"' for key, value in form_attrs.items())
    head_html = ""
    if title or subtitle:
        head_html = f"""
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(title)}</h2>
          <p>{esc(subtitle)}</p>
        </div>
      </div>
      """
    return f"""
    <section class="{section_class}">
      {head_html}
      <form method="post" action="{esc(action)}" class="form-grid"{form_extra_attrs}>
        {hidden_fields_html(hidden_fields)}
        {fields_html}
        <div class="form-actions">
          <button type="submit" class="button">{esc(button_label)}</button>
        </div>
      </form>
    </section>
    """


def table_card(
    title: str,
    subtitle: str,
    rows: list[sqlite3.Row],
    columns: list[tuple],
    *,
    empty_message: str = "Nessun dato disponibile.",
    table_class: str = "",
    summary_rows: list[list[str]] | None = None,
    head_actions_html: str = "",
    table_id: str | None = None,
    column_filters_in_header: bool = True,
    draggable_columns: bool = True,
) -> str:
    table_id = table_id or f"table-{slugify(title)}-{hashlib.md5('|'.join(str(column[0] or column[1] or '') for column in columns).encode('utf-8')).hexdigest()[:10]}"
    has_filterable_columns = any(
        str(column[0] or "").strip() and str(column[1] or "").lower() not in {"azioni", "ricevuta"}
        for column in columns
    )
    default_head_actions = ""
    if column_filters_in_header and table_id and has_filterable_columns:
        default_head_actions = (
            f'<button type="button" class="button action" onclick="clearTableColumnFilters(\'{esc(table_id)}\')">'
            f"Annulla filtri</button>"
        )
    combined_head_actions = "".join(part for part in [head_actions_html, default_head_actions] if part)
    subtitle_html = f"<p>{esc(subtitle)}</p>" if subtitle else ""
    actions_html = f'<div class="card-head-actions screen-only">{combined_head_actions}</div>' if combined_head_actions else ""
    return f"""
    <section class="card">
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(title)}</h2>
          {subtitle_html}
        </div>
        {actions_html}
      </div>
      {render_table(rows, columns, empty_message=empty_message, table_class=table_class, summary_rows=summary_rows, table_id=table_id, column_filters_in_header=column_filters_in_header, draggable_columns=draggable_columns)}
    </section>
    """


def column_sort_type(column: tuple, index: int, total_index_set: set[int]) -> str:
    key = str(column[0] or "")
    label = str(column[1] or "").lower()
    if index in total_index_set:
        return "number"
    if key in {"id", "numero_progressivo", "numero_progressivo_anno", "anno", "anno_sociale", "mese", "eta"}:
        return "number"
    if any(token in label for token in {"importo", "dovuto", "pagato", "residuo", "quota", "saldo", "eta", "numero", "anno"}):
        return "number"
    return "text"


def cell_sort_value(value: object, column: tuple, index: int, total_index_set: set[int]) -> str:
    if value is None:
        return ""
    if index in total_index_set:
        amount = decimal_or_none(value)
        if amount is not None:
            return format(amount, ".2f")
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    return plain_text(str(value))


def render_table(
    rows: list[sqlite3.Row],
    columns: list[tuple],
    *,
    empty_message: str = "Nessun dato disponibile.",
    table_class: str = "",
    summary_rows: list[list[str]] | None = None,
    table_id: str | None = None,
    column_filters_in_header: bool = False,
    draggable_columns: bool = False,
) -> str:
    if not rows:
        return f'<div class="empty-state">{esc(empty_message)}</div>'

    total_indexes = summary_total_indexes(columns)
    total_index_set = set(total_indexes)
    head = []
    filter_head = []
    body = []
    for index, column in enumerate(columns):
        key = str(column[0] or "")
        label = str(column[1] or "")
        sortable = label.lower() not in {"azioni", "ricevuta"}
        column_key_attr = f' data-column-key="{esc(key)}"' if key else ""
        draggable_attr = ' draggable="true" data-draggable-column="1"' if draggable_columns and key and label.lower() not in {"azioni", "ricevuta"} else ""
        if sortable:
            head.append(
                f'<th scope="col"{column_key_attr}{draggable_attr}>'
                f'<button type="button" class="table-sort-button" data-sort-index="{index}" data-sort-type="{column_sort_type(column, index, total_index_set)}">'
                f'<span>{esc(label)}</span><span class="table-sort-glyph" aria-hidden="true"></span>'
                f"</button></th>"
            )
        else:
            head.append(f"<th scope=\"col\"{column_key_attr}{draggable_attr}>{esc(label)}</th>")
        if column_filters_in_header:
            if key and label.lower() not in {"azioni", "ricevuta"} and table_id:
                filter_head.append(
                    f'<th scope="col"{column_key_attr}><input type="search" class="control table-column-filter" data-table-filter data-target-table="{esc(table_id)}" data-column-key="{esc(key)}" placeholder="Filtra..." oninput="handleTableColumnFilter(this)"></th>'
                )
            else:
                filter_head.append(f"<th scope=\"col\"{column_key_attr}></th>")

    for row in rows:
        cells = []
        for index, column in enumerate(columns):
            key = column[0]
            formatter = column[2] if len(column) > 2 else None
            value = row[key]
            rendered = formatter(value, row) if callable(formatter) else esc(value)
            sort_value = cell_sort_value(value, column, index, total_index_set)
            cell_attrs = f' data-column-key="{esc(str(key))}" data-sort-value="{esc(sort_value)}"'
            if index in total_index_set:
                amount = decimal_or_none(value)
                if amount is not None:
                    cell_attrs += f' data-sum-value="{format(amount, ".2f")}"'
            cells.append(f"<td{cell_attrs}>{rendered}</td>")
        body.append(f"<tr>{''.join(cells)}</tr>")

    footer_html = ""
    if summary_rows:
        footer_rows = []
        for summary_row in summary_rows:
            footer_cells = [
                f'<td data-column-key="{esc(str(columns[index][0]))}">{esc(cell)}</td>'
                for index, cell in enumerate(summary_row)
            ]
            footer_rows.append(f"<tr>{''.join(footer_cells)}</tr>")
        footer_html = f"<tfoot>{''.join(footer_rows)}</tfoot>"

    table_attrs = ""
    if summary_rows and total_indexes:
        table_attrs = (
            f' data-summary-columns="{esc(",".join(str(index) for index in total_indexes))}"'
            f' data-summary-label-index="{summary_label_index(columns)}"'
        )
    if table_id:
        table_attrs += f' id="{esc(table_id)}"'
    if draggable_columns:
        table_attrs += ' data-column-reorder="1"'

    return (
        f'<div class="table-wrap"><table class="data-table {esc(table_class)}"{table_attrs}><thead><tr>'
        + "".join(head)
        + "</tr>"
        + (f'<tr class="table-filter-row">{"".join(filter_head)}</tr>' if column_filters_in_header else "")
        + "</thead><tbody>"
        + "".join(body)
        + f"</tbody>{footer_html}</table></div>"
    )


def stat_card(label: str, value: object, link: str) -> str:
    return f"""
    <a class="stat-card" href="{esc(link)}">
      <span class="stat-label">{esc(label)}</span>
      <strong class="stat-value">{esc(value)}</strong>
    </a>
    """


def message_banner(query_params: dict[str, str]) -> str:
    ok_message = query_params.get("ok")
    error_message = query_params.get("err")
    items = []
    if ok_message:
        items.append(f'<div class="notice success">{esc(ok_message)}</div>')
    if error_message:
        items.append(f'<div class="notice error">{esc(error_message)}</div>')
    return "".join(items)


def work_year_selector(query_params: dict[str, str], *, action: str = "/") -> str:
    selected_year = current_work_year(query_params)
    options = render_static_options(
        [(str(year), str(year)) for year in available_work_years(selected_year)],
        str(selected_year),
    )
    return f"""
    <section class="card compact screen-only work-year-card">
      <div class="card-head">
        <h2>Anno di lavoro</h2>
        <p>Seleziona l'anno operativo da usare in dashboard, report e filtri predefiniti.</p>
      </div>
      <form method="get" action="{esc(action)}" class="form-grid">
        {select_field("Anno di lavoro", "anno_lavoro", options, required_field=True)}
        <div class="form-actions">
          <button type="submit" class="button">Applica anno</button>
        </div>
      </form>
    </section>
    """


def header_work_year_selector(query_params: dict[str, str], *, action: str = "/") -> str:
    selected_year = current_work_year(query_params)
    options = render_static_options(
        [(str(year), str(year)) for year in available_work_years(selected_year)],
        str(selected_year),
        blank_label=None,
    )
    return f"""
    <form method="get" action="{esc(action)}" class="header-year-form screen-only">
      <label class="header-year-field">
        <span>Anno di lavoro</span>
        <select name="anno_lavoro" class="control">{options}</select>
      </label>
      <button type="submit" class="button">Applica</button>
    </form>
    """


def data_view_query(query_params: dict[str, str]) -> dict[str, str]:
    params = work_year_query(query_params)
    params["vista"] = "dati"
    return params


def insert_view_query(query_params: dict[str, str]) -> dict[str, str]:
    return work_year_query(query_params)


def current_page_query(query_params: dict[str, str]) -> dict[str, str]:
    if normalized(query_params, "vista", "") == "dati":
        return data_view_query(query_params)
    return work_year_query(query_params)


def view_mode_switch(current_path: str, query_params: dict[str, str], data_label: str) -> str:
    data_only = normalized(query_params, "vista", "") == "dati"
    primary_href = with_query(current_path, insert_view_query(query_params) if data_only else data_view_query(query_params))
    primary_label = "Torna a inserimento" if data_only else data_label
    secondary_href = with_query(current_path, data_view_query(query_params) if data_only else insert_view_query(query_params))
    secondary_label = data_label if data_only else "Vista inserimento"
    return f"""
    <section class="report-toolbar screen-only view-mode-toolbar">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Vista area</span>
        <p>Passa dalla maschera di inserimento alla schermata dei dati gia registrati.</p>
      </div>
      <div class="report-toolbar-actions view-mode-toolbar-actions">
        <div class="report-toolbar-action-row report-toolbar-action-row-inline-start">
          <a class="button" href="{esc(primary_href)}">{esc(primary_label)}</a>
          <a class="button secondary" href="{esc(secondary_href)}">{esc(secondary_label)}</a>
        </div>
      </div>
    </section>
    """


def data_view_search_toolbar(table_id: str | None = None) -> str:
    target_attr = f' data-target-table="{esc(table_id)}"' if table_id else ""
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Ricerca dati</span>
        <p>Filtra rapidamente tutte le colonne presenti nelle tabelle dei dati gia registrati.</p>
      </div>
      <div class="report-toolbar-actions">
        <label class="report-search">
          <span>Cerca</span>
          <input
            type="search"
            class="control"
            placeholder="Filtra tutte le colonne..."
            oninput="handleDataSearch(this)"{target_attr}
          >
        </label>
      </div>
    </section>
    """


def dashboard_associati_search_toolbar() -> str:
    return """
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Posizione tesserati</span>
        <p>Cerca rapidamente un tesserato nella tabella della dashboard.</p>
      </div>
      <div class="report-toolbar-actions">
        <label class="report-search">
          <span>Cerca</span>
          <input
            type="search"
            class="control"
            placeholder="Filtra per nome, codice o stato..."
            oninput="handleDataSearch(this)"
          >
        </label>
      </div>
    </section>
    """


def column_visibility_toolbar(
    title: str,
    subtitle: str,
    table_id: str,
    columns: list[tuple],
    *,
    default_visible_keys: set[str] | None = None,
    allow_reorder: bool = False,
) -> str:
    visible_keys = default_visible_keys or {str(column[0]) for column in columns}
    toggle_items = []
    for column in columns:
        key = str(column[0] or "")
        label = str(column[1] or "")
        if not key or label.lower() in {"azioni", "ricevuta"}:
            continue
        checked = " checked" if key in visible_keys else ""
        order_controls = ""
        if allow_reorder:
            order_controls = f"""
            <span class="column-toggle-order">
              <button type="button" class="column-toggle-move" data-column-move="left" data-target-table="{esc(table_id)}" data-column-key="{esc(key)}" aria-label="Sposta {esc(label)} a sinistra">â€¹</button>
              <button type="button" class="column-toggle-move" data-column-move="right" data-target-table="{esc(table_id)}" data-column-key="{esc(key)}" aria-label="Sposta {esc(label)} a destra">â€º</button>
            </span>
            """
        toggle_items.append(
            f"""
            <label class="column-toggle-pill" data-column-toggle-chip data-target-table="{esc(table_id)}" data-column-key="{esc(key)}">
              <input type="checkbox" data-column-toggle data-target-table="{esc(table_id)}" data-column-key="{esc(key)}"{checked}>
              <span>{esc(label)}</span>
              {order_controls}
            </label>
            """
        )
    return f"""
    <section class="report-toolbar screen-only column-visibility-toolbar">
      <div class="report-toolbar-copy">
        <span class="eyebrow">{esc(title)}</span>
        <p>{esc(subtitle)}</p>
      </div>
      <div class="column-visibility-actions">
        <button type="button" class="button action" data-column-select-all data-target-table="{esc(table_id)}">Seleziona tutti</button>
        <button type="button" class="button action" data-column-deselect-all data-target-table="{esc(table_id)}">Deseleziona tutti</button>
      </div>
      <div class="column-toggle-grid" data-column-toggle-group="{esc(table_id)}">
        {''.join(toggle_items)}
      </div>
    </section>
    """


def column_filters_toolbar(
    title: str,
    subtitle: str,
    table_id: str,
    columns: list[tuple],
) -> str:
    filter_items = []
    for column in columns:
        key = str(column[0] or "")
        label = str(column[1] or "")
        if not key or label.lower() in {"azioni", "ricevuta"}:
            continue
        filter_items.append(
            f"""
            <label class="report-search column-filter-field">
              <span>{esc(label)}</span>
              <input
                type="search"
                class="control"
                data-table-filter
                data-target-table="{esc(table_id)}"
                data-column-key="{esc(key)}"
                placeholder="Filtra {esc(label.lower())}..."
                oninput="handleTableColumnFilter(this)"
              >
            </label>
            """
        )
    return f"""
    <section class="report-toolbar screen-only column-filter-toolbar">
      <div class="report-toolbar-copy">
        <span class="eyebrow">{esc(title)}</span>
        <p>{esc(subtitle)}</p>
      </div>
      <div class="column-filter-grid">
        {''.join(filter_items)}
      </div>
    </section>
    """


def column_visibility_toolbar(
    title: str,
    subtitle: str,
    table_id: str,
    columns: list[tuple],
    *,
    default_visible_keys: set[str] | None = None,
    allow_reorder: bool = False,
) -> str:
    visible_keys = default_visible_keys or {str(column[0]) for column in columns}
    toggle_items = []
    for column in columns:
        key = str(column[0] or "")
        label = str(column[1] or "")
        if not key or label.lower() in {"azioni", "ricevuta"}:
            continue
        checked = " checked" if key in visible_keys else ""
        toggle_items.append(
            f"""
            <label class="column-picker-option" data-column-toggle-chip data-target-table="{esc(table_id)}" data-column-key="{esc(key)}">
              <input type="checkbox" data-column-toggle data-target-table="{esc(table_id)}" data-column-key="{esc(key)}"{checked}>
              <span>{esc(label)}</span>
            </label>
            """
        )
    reorder_hint = '<p class="column-picker-hint">Trascina con il mouse le intestazioni della tabella per riordinare le colonne.</p>' if allow_reorder else ""
    return f"""
    <section class="report-toolbar screen-only column-visibility-toolbar">
      <div class="report-toolbar-copy">
        <span class="eyebrow">{esc(title)}</span>
        <p>{esc(subtitle)}</p>
      </div>
      <details class="column-picker" data-column-picker="{esc(table_id)}">
        <summary class="button action column-picker-summary">Seleziona colonne</summary>
        <div class="column-picker-panel">
          <label class="column-picker-search">
            <span>Cerca campo</span>
            <input
              type="search"
              class="control"
              data-column-picker-search
              data-target-table="{esc(table_id)}"
              placeholder="Digita per filtrare i campi..."
              oninput="handleColumnPickerSearch(this)"
            >
          </label>
          <div class="column-visibility-actions">
            <button type="button" class="button action" data-column-select-all data-target-table="{esc(table_id)}">Seleziona tutti</button>
            <button type="button" class="button action" data-column-deselect-all data-target-table="{esc(table_id)}">Deseleziona tutti</button>
          </div>
          <div class="column-picker-list" data-column-toggle-group="{esc(table_id)}">
            {''.join(toggle_items)}
          </div>
          {reorder_hint}
        </div>
      </details>
    </section>
    """


def table_card(
    title: str,
    subtitle: str,
    rows: list[sqlite3.Row],
    columns: list[tuple],
    *,
    empty_message: str = "Nessun dato disponibile.",
    table_class: str = "",
    summary_rows: list[list[str]] | None = None,
    head_actions_html: str = "",
    table_id: str | None = None,
    column_filters_in_header: bool = True,
    draggable_columns: bool = True,
) -> str:
    table_id = table_id or f"table-{slugify(title)}-{hashlib.md5('|'.join(str(column[0] or column[1] or '') for column in columns).encode('utf-8')).hexdigest()[:10]}"
    has_filterable_columns = any(
        str(column[0] or "").strip() and str(column[1] or "").lower() not in {"azioni", "ricevuta"}
        for column in columns
    )
    default_head_actions = ""
    if column_filters_in_header and table_id and has_filterable_columns:
        default_head_actions = (
            f'<button type="button" class="button action" onclick="clearTableColumnFilters(\'{esc(table_id)}\')">'
            f"Annulla filtri</button>"
        )
    combined_head_actions = "".join(part for part in [head_actions_html, default_head_actions] if part)
    subtitle_html = f"<p>{esc(subtitle)}</p>" if subtitle else ""
    actions_html = f'<div class="card-head-actions screen-only">{combined_head_actions}</div>' if combined_head_actions else ""
    return f"""
    <section class="card">
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(title)}</h2>
          {subtitle_html}
        </div>
        {actions_html}
      </div>
      {render_table(rows, columns, empty_message=empty_message, table_class=table_class, summary_rows=summary_rows, table_id=table_id, column_filters_in_header=column_filters_in_header, draggable_columns=draggable_columns)}
    </section>
    """


def render_table(
    rows: list[sqlite3.Row],
    columns: list[tuple],
    *,
    empty_message: str = "Nessun dato disponibile.",
    table_class: str = "",
    summary_rows: list[list[str]] | None = None,
    table_id: str | None = None,
    column_filters_in_header: bool = False,
    draggable_columns: bool = False,
) -> str:
    if not rows:
        return f'<div class="empty-state">{esc(empty_message)}</div>'

    total_indexes = summary_total_indexes(columns)
    total_index_set = set(total_indexes)
    head = []
    filter_head = []
    body = []
    for index, column in enumerate(columns):
        key = str(column[0] or "")
        label = str(column[1] or "")
        sortable = label.lower() not in {"azioni", "ricevuta"}
        column_key_attr = f' data-column-key="{esc(key)}"' if key else ""
        draggable_attr = ' draggable="true" data-draggable-column="1"' if draggable_columns and key and label.lower() not in {"azioni", "ricevuta"} else ""
        if sortable:
            head.append(
                f'<th scope="col"{column_key_attr}{draggable_attr}>'
                f'<button type="button" class="table-sort-button" data-sort-index="{index}" data-sort-type="{column_sort_type(column, index, total_index_set)}">'
                f'<span>{esc(label)}</span><span class="table-sort-glyph" aria-hidden="true"></span>'
                f"</button></th>"
            )
        else:
            head.append(f"<th scope=\"col\"{column_key_attr}{draggable_attr}>{esc(label)}</th>")
        if column_filters_in_header:
            if key and label.lower() not in {"azioni", "ricevuta"} and table_id:
                filter_head.append(
                    f'<th scope="col"{column_key_attr}><input type="search" class="control table-column-filter" data-table-filter data-target-table="{esc(table_id)}" data-column-key="{esc(key)}" placeholder="Filtra..." oninput="handleTableColumnFilter(this)"></th>'
                )
            else:
                filter_head.append(f"<th scope=\"col\"{column_key_attr}></th>")

    for row in rows:
        cells = []
        for index, column in enumerate(columns):
            key = column[0]
            formatter = column[2] if len(column) > 2 else None
            value = row[key]
            rendered = formatter(value, row) if callable(formatter) else esc(value)
            sort_value = cell_sort_value(value, column, index, total_index_set)
            cell_attrs = f' data-column-key="{esc(str(key))}" data-sort-value="{esc(sort_value)}"'
            if index in total_index_set:
                amount = decimal_or_none(value)
                if amount is not None:
                    cell_attrs += f' data-sum-value="{format(amount, ".2f")}"'
            cells.append(f"<td{cell_attrs}>{rendered}</td>")
        body.append(f"<tr>{''.join(cells)}</tr>")

    footer_html = ""
    if summary_rows:
        footer_rows = []
        for summary_row in summary_rows:
            footer_cells = [
                f'<td data-column-key="{esc(str(columns[index][0]))}">{esc(cell)}</td>'
                for index, cell in enumerate(summary_row)
            ]
            footer_rows.append(f"<tr>{''.join(footer_cells)}</tr>")
        footer_html = f"<tfoot>{''.join(footer_rows)}</tfoot>"

    table_attrs = ""
    if summary_rows and total_indexes:
        table_attrs = (
            f' data-summary-columns="{esc(",".join(str(index) for index in total_indexes))}"'
            f' data-summary-label-index="{summary_label_index(columns)}"'
        )
    if table_id:
        table_attrs += f' id="{esc(table_id)}"'
    if draggable_columns:
        table_attrs += ' data-column-reorder="1"'

    return (
        f'<div class="table-wrap"><table class="data-table {esc(table_class)}"{table_attrs}><thead><tr>'
        + "".join(head)
        + "</tr>"
        + (f'<tr class="table-filter-row">{"".join(filter_head)}</tr>' if column_filters_in_header else "")
        + "</thead><tbody>"
        + "".join(body)
        + f"</tbody>{footer_html}</table></div>"
    )


def posizione_associati_params(work_year: int) -> tuple[object, ...]:
    return (
        work_year,
        work_year,
        work_year,
        work_year,
        str(work_year),
        *associato_year_relevance_params(work_year),
    )


def posizione_associati_query(*, limit: int | None = None) -> str:
    limit_sql = f"\n        LIMIT {int(limit)}" if limit else ""
    return f"""
        WITH movimenti AS (
            SELECT
                associato_id,
                importo_dovuto,
                importo_pagato,
                saldo_residuo
            FROM v_tesseramenti_saldo
            WHERE anno_sociale = ?

            UNION ALL

            SELECT
                associato_id,
                importo_dovuto,
                importo_pagato,
                saldo_residuo
            FROM v_rate_corsi_saldo
            WHERE anno = ?

            UNION ALL

            SELECT
                associato_id,
                importo_dovuto,
                importo_pagato,
                saldo_residuo
            FROM v_campi_estivi_saldo
            WHERE anno = ?

            UNION ALL

            SELECT
                associato_id,
                importo_dovuto,
                importo_pagato,
                saldo_residuo
            FROM v_eventi_saldo
            WHERE substr(COALESCE(data_evento, ''), 1, 4) = ?
        ),
        posizioni AS (
            SELECT
                associato_id,
                COALESCE(SUM(importo_dovuto), 0) AS totale_dovuto,
                COALESCE(SUM(importo_pagato), 0) AS totale_pagato,
                COALESCE(SUM(saldo_residuo), 0) AS saldo_residuo
            FROM movimenti
            GROUP BY associato_id
        )
        SELECT
            a.id AS associato_id,
            COALESCE(t.codice_tesseramento, '') AS codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            CASE
                WHEN COALESCE(vts.saldo_residuo, t.importo_dovuto, 0) <= 0 THEN 'Concluso'
                ELSE 'Attivo'
            END AS stato_associato,
            COALESCE(posizioni.totale_dovuto, 0) AS totale_dovuto,
            COALESCE(posizioni.totale_pagato, 0) AS totale_pagato,
            COALESCE(posizioni.saldo_residuo, 0) AS saldo_residuo
        FROM associati a
        LEFT JOIN posizioni ON posizioni.associato_id = a.id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = a.id AND t.anno_sociale = ?
        LEFT JOIN v_tesseramenti_saldo vts ON vts.id = t.id
        WHERE {associato_year_relevance_sql('a')}
          AND t.id IS NOT NULL
          AND COALESCE(posizioni.totale_dovuto, 0) > 0
        ORDER BY associato{limit_sql}
    """


def delete_action_form(action: str, prompt: str, *, extra_fields: dict[str, str] | None = None) -> str:
    fields = dict(extra_fields or {})
    fields.setdefault("vista", "")
    return f"""
    <form
      method="post"
      action="{esc(action)}"
      class="inline-form"
      data-confirm-dialog="true"
      data-confirm-title="Conferma eliminazione"
      data-confirm-message="{esc(prompt)}"
      data-confirm-button="Elimina">
      {hidden_fields_html(fields)}
      <button type="submit" class="table-action danger">Elimina</button>
    </form>
    """


def action_links_html(
    *,
    edit_href: str | None = None,
    delete_action: str | None = None,
    delete_prompt: str | None = None,
    extra_links: list[tuple[str, str]] | None = None,
    extra_fields: dict[str, str] | None = None,
) -> str:
    items: list[str] = []
    if edit_href:
        items.append(f'<a class="table-action" href="{esc(edit_href)}">Modifica</a>')
    if extra_links:
        for href, label in extra_links:
            items.append(f'<a class="table-action" href="{esc(href)}">{esc(label)}</a>')
    if delete_action and delete_prompt:
        items.append(delete_action_form(delete_action, delete_prompt, extra_fields=extra_fields))
    return f'<div class="table-actions">{ "".join(items) }</div>'


def nav_link(current_path: str, href: str, label: str) -> str:
    active = " active" if current_path == href.split("?", 1)[0] else ""
    return f'<a class="nav-link{active}" href="{esc(href)}">{esc(label)}</a>'


def access_role_label(current_user: dict[str, object] | None) -> str:
    if not current_user:
        return ""
    return "Amministratore" if current_user.get("is_admin") else "Utente"


def render_navigation(current_path: str, query_params: dict[str, str], current_user: dict[str, object] | None = None) -> str:
    kicker_html = f'<span class="brand-kicker">{esc(APP_SUBTITLE)}</span>' if APP_SUBTITLE else ""
    copy_class = "brand-copy with-kicker" if APP_SUBTITLE else "brand-copy"
    admin_group = ""
    if current_user and current_user.get("is_admin"):
        admin_group = f"""
      <div class="nav-group">
        <span class="nav-group-title">Amministrazione</span>
        {nav_link(current_path, with_query("/maschere/utenti", work_year_query(query_params)), "Utenti")}
        {nav_link(current_path, with_query("/report/registro-attivita", work_year_query(query_params)), "Registro attivita")}
        {nav_link(current_path, with_query("/maschere/backup", work_year_query(query_params)), "Backup")}
        {nav_link(current_path, with_query("/maschere/aggiornamenti", work_year_query(query_params)), "Aggiornamenti")}
        {nav_link(current_path, with_query("/maschere/importa-associati", work_year_query(query_params)), "Importa associati")}
      </div>
      """
    user_block = ""
    if current_user:
        user_block = f"""
      <div class="sidebar-user-card">
        <div class="sidebar-user-inline">
          <strong class="sidebar-user-name">{esc(current_user.get("username", ""))}</strong>
        </div>
        <form method="post" action="/azioni/accesso/logout" class="sidebar-user-form">
          {hidden_fields_html(work_year_query(query_params))}
          <button type="submit" class="sidebar-user-button">Logout</button>
        </form>
      </div>
      """
    return f"""
    <nav class="sidebar">
      <div class="brand">
        <div class="brand-mark">
          <div class="brand-logo-wrap">
            <img class="brand-logo" src="{LOGO_URL}" alt="Logo {APP_NAME}">
          </div>
          <div class="{copy_class}">
            {kicker_html}
            <strong class="brand-title">{APP_NAME}</strong>
          </div>
        </div>
      </div>
      {user_block}
      <div class="sidebar-nav-scroll">
      <div class="nav-group">
        <span class="nav-group-title">Panoramica</span>
        {nav_link(current_path, with_query("/", work_year_query(query_params)), "Dashboard")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">STRUTTURA SOCIALE</span>
        {nav_link(current_path, with_query("/maschere/consiglio-direttivo", work_year_query(query_params)), "Consiglio Direttivo")}
        {nav_link(current_path, with_query("/maschere/tesserati", work_year_query(query_params)), "Tesserati")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">ISCRIZIONI</span>
        {nav_link(current_path, with_query("/maschere/associati", work_year_query(query_params)), "Nuovo associato")}
        {nav_link(current_path, with_query("/maschere/tesseramenti", work_year_query(query_params)), "Rinnovo tesseramento")}
        {nav_link(current_path, with_query("/maschere/oratorio", work_year_query(query_params)), ORATORIO_LABEL)}
        {nav_link(current_path, with_query("/maschere/corsi", work_year_query(query_params)), "Corsi")}
        {nav_link(current_path, with_query("/maschere/campi-estivi", work_year_query(query_params)), ESTATE_LABEL)}
        {nav_link(current_path, with_query("/maschere/eventi", work_year_query(query_params)), "Eventi")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">GESTIONE ECONOMICA</span>
        {nav_link(current_path, with_query("/maschere/pagamenti-multi-area", work_year_query(query_params)), "Pagamenti")}
        {nav_link(current_path, with_query("/report/scadenze", work_year_query(query_params)), "Scadenze da incassare")}
        {nav_link(current_path, with_query("/report/incassi", work_year_query(query_params)), "Incassi totali")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">Report</span>
        {nav_link(current_path, with_query("/report/associati", work_year_query(query_params)), "Posizione tesserati")}
        {nav_link(current_path, with_query("/report/partecipanti", work_year_query(query_params)), "Partecipanti attivita")}
        {nav_link(current_path, with_query("/report/tesseramenti", work_year_query(query_params)), "Situazione tesseramenti")}
        {nav_link(current_path, with_query("/report/corsi", work_year_query(query_params)), "Situazione corsi")}
        {nav_link(current_path, with_query("/report/oratorio", work_year_query(query_params)), "Situazione oratorio")}
        {nav_link(current_path, with_query("/report/campi-estivi", work_year_query(query_params)), "Situazione campo estivo")}
        {nav_link(current_path, with_query("/report/eventi", work_year_query(query_params)), "Situazione eventi")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">DATI STORICI</span>
        {nav_link(current_path, with_query("/report/storico-tesseramenti", work_year_query(query_params)), "Storico tesseramenti")}
      </div>
      <div class="nav-group">
        <span class="nav-group-title">Accesso</span>
        {nav_link(current_path, with_query("/maschere/accesso", work_year_query(query_params)), "Profilo accesso")}
        {nav_link(current_path, with_query("/maschere/guida", work_year_query(query_params)), "Tutorial")}
      </div>
      {admin_group}
      </div>
      <div class="sidebar-version-footer">Versione {esc(APP_VERSION)}</div>
    </nav>
    """


def app_dialog_markup() -> str:
    return """
    <section id="app-dialog-overlay" class="app-dialog-overlay" hidden>
      <div class="app-dialog-backdrop"></div>
      <div class="app-dialog-card" data-variant="info" role="alertdialog" aria-modal="true" aria-labelledby="app-dialog-title" aria-describedby="app-dialog-message">
        <div class="app-dialog-accent">
          <span id="app-dialog-badge" class="app-dialog-badge">Notifica</span>
          <h2 id="app-dialog-title">Notifica</h2>
          <p id="app-dialog-message" class="app-dialog-message"></p>
          <div id="app-dialog-extra" class="app-dialog-extra" hidden></div>
        </div>
        <div class="app-dialog-actions">
          <button type="button" id="app-dialog-cancel" class="button ghost">Annulla</button>
          <button type="button" id="app-dialog-confirm" class="button">Conferma</button>
        </div>
      </div>
    </section>
    """


def shared_dialog_script() -> str:
    return """
      let activeAppDialog = null;

      function togglePasswordVisibility(button) {
        const wrapper = button.closest('.password-field-wrap');
        if (!wrapper) {
          return;
        }
        const input = wrapper.querySelector('input');
        if (!input) {
          return;
        }
        const isHidden = input.type === 'password';
        input.type = isHidden ? 'text' : 'password';
        button.textContent = isHidden ? 'Nascondi' : 'Mostra';
      }

      function syncModalOpenState() {
        const reminderVisible = !!document.querySelector('[data-course-generation-reminder="true"]:not([hidden])');
        const dialogOverlay = document.getElementById('app-dialog-overlay');
        const dialogVisible = !!dialogOverlay && !dialogOverlay.hidden;
        document.body.classList.toggle('modal-open', reminderVisible || dialogVisible);
      }

      function appDialogElements() {
        const overlay = document.getElementById('app-dialog-overlay');
        if (!overlay) {
          return null;
        }
        return {
          overlay,
          backdrop: overlay.querySelector('.app-dialog-backdrop'),
          card: overlay.querySelector('.app-dialog-card'),
          badge: document.getElementById('app-dialog-badge'),
          title: document.getElementById('app-dialog-title'),
          message: document.getElementById('app-dialog-message'),
          extra: document.getElementById('app-dialog-extra'),
          cancelButton: document.getElementById('app-dialog-cancel'),
          confirmButton: document.getElementById('app-dialog-confirm'),
        };
      }

      function closeAppDialog(result = false) {
        const elements = appDialogElements();
        if (!elements || elements.overlay.hidden) {
          activeAppDialog = null;
          return;
        }
        const resolver = activeAppDialog && typeof activeAppDialog.resolve === 'function'
          ? activeAppDialog.resolve
          : null;
        elements.overlay.hidden = true;
        if (elements.card) {
          elements.card.dataset.variant = 'info';
        }
        if (elements.extra) {
          elements.extra.hidden = true;
          elements.extra.innerHTML = '';
        }
        if (elements.confirmButton) {
          elements.confirmButton.classList.remove('danger');
        }
        activeAppDialog = null;
        syncModalOpenState();
        if (resolver) {
          resolver(result);
        }
      }

      function showAppDialog(options = {}) {
        const elements = appDialogElements();
        if (!elements) {
          if (options.cancelLabel) {
            return Promise.resolve(window.confirm(options.message || 'Confermare l\\'operazione?'));
          }
          window.alert(options.message || 'Operazione non disponibile.');
          return Promise.resolve(true);
        }

        if (activeAppDialog && typeof activeAppDialog.resolve === 'function') {
          activeAppDialog.resolve(false);
        }

        const variant = options.variant || 'info';
        const dismissible = options.dismissible !== false;
        const confirmLabel = options.confirmLabel || 'Chiudi';
        const cancelLabel = options.cancelLabel || '';
        const badgeText = options.badge || (
          variant === 'danger' ? 'Attenzione'
            : variant === 'warning' ? 'Conferma'
            : 'Notifica'
        );

        elements.overlay.hidden = false;
        if (elements.card) {
          elements.card.dataset.variant = variant;
        }
        if (elements.badge) {
          elements.badge.textContent = badgeText;
        }
        if (elements.title) {
          elements.title.textContent = options.title || 'Notifica';
        }
        if (elements.message) {
          elements.message.textContent = options.message || '';
        }
        if (elements.extra) {
          if (options.html) {
            elements.extra.hidden = false;
            elements.extra.innerHTML = options.html;
          } else {
            elements.extra.hidden = true;
            elements.extra.innerHTML = '';
          }
        }
        if (elements.cancelButton) {
          if (cancelLabel) {
            elements.cancelButton.hidden = false;
            elements.cancelButton.textContent = cancelLabel;
          } else {
            elements.cancelButton.hidden = true;
          }
        }
        if (elements.confirmButton) {
          elements.confirmButton.textContent = confirmLabel;
          elements.confirmButton.classList.toggle('danger', variant === 'danger');
        }

        syncModalOpenState();

        return new Promise((resolve) => {
          activeAppDialog = {
            resolve,
            dismissible,
            resolveValue: typeof options.resolveValue === 'function' ? options.resolveValue : null,
          };

          if (elements.confirmButton) {
            elements.confirmButton.onclick = () => {
              if (activeAppDialog && typeof activeAppDialog.resolveValue === 'function') {
                const resolved = activeAppDialog.resolveValue(elements);
                if (resolved === false || typeof resolved === 'undefined') {
                  return;
                }
                closeAppDialog(resolved);
                return;
              }
              closeAppDialog(true);
            };
          }
          if (elements.cancelButton) {
            elements.cancelButton.onclick = () => closeAppDialog(false);
          }
          if (elements.backdrop) {
            elements.backdrop.onclick = () => {
              if (activeAppDialog && activeAppDialog.dismissible) {
                closeAppDialog(false);
              }
            };
          }

          window.setTimeout(() => {
            const target = (!cancelLabel || variant === 'danger') && elements.confirmButton
              ? elements.confirmButton
              : elements.cancelButton || elements.confirmButton;
            if (target) {
              target.focus();
            }
          }, 30);
        });
      }

      async function appAlert(message, options = {}) {
        return showAppDialog({
          title: options.title || 'Notifica',
          message,
          badge: options.badge || 'Notifica',
          variant: options.variant || 'info',
          confirmLabel: options.confirmLabel || 'Chiudi',
          dismissible: options.dismissible !== false,
        });
      }

      async function appConfirm(message, options = {}) {
        return showAppDialog({
          title: options.title || 'Conferma operazione',
          message,
          html: options.html || '',
          badge: options.badge || 'Conferma',
          variant: options.variant || 'warning',
          confirmLabel: options.confirmLabel || 'Conferma',
          cancelLabel: options.cancelLabel || 'Annulla',
          dismissible: options.dismissible !== false,
        });
      }
    """


def public_page(title: str, content: str, query_params: dict[str, str]) -> bytes:
    document = f"""<!doctype html>
<html lang="it">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{esc(title)} - {esc(APP_NAME)}</title>
    <link rel="icon" type="image/x-icon" href="/static/logo-ca.ico">
    <link rel="shortcut icon" href="/static/logo-ca.ico">
    <link rel="stylesheet" href="/static/style.css">
  </head>
  <body class="auth-body">
    <div class="auth-shell">
      <section class="auth-brand-panel">
        <div class="auth-brand-card">
          <div class="auth-brand-mark">
            <div class="auth-brand-logo-wrap">
              <img class="auth-brand-logo" src="{LOGO_URL}" alt="Logo {APP_NAME}">
            </div>
            <div>
              <span class="eyebrow">Accesso riservato</span>
              <h1>{esc(APP_NAME)}</h1>
            </div>
          </div>
          <p class="auth-brand-text">Un unico spazio per anagrafiche, tesseramenti, corsi, campo estivo, eventi, incassi e ricevute.</p>
          <div class="auth-feature-list">
            <div class="auth-feature-item">Login protetto con username e password</div>
            <div class="auth-feature-item">Cruscotto unico per quote, scadenze, incassi e ricevute</div>
            <div class="auth-feature-item">Interfaccia chiara, pronta per il lavoro quotidiano</div>
          </div>
        </div>
      </section>
      <main class="auth-content">
        {message_banner(query_params)}
        {content}
      </main>
    </div>
    {app_dialog_markup()}
    <script>
      {shared_dialog_script()}
    </script>
  </body>
</html>"""
    return document.encode("utf-8")


def page(
    title: str,
    current_path: str,
    content: str,
    query_params: dict[str, str],
    current_user: dict[str, object] | None = None,
) -> bytes:
    header_class = "page-header home-header" if current_path == "/" else "page-header"
    header_tools = header_work_year_selector(query_params, action="/") if current_path == "/" else ""
    payment_methods_json = json.dumps(
        [{"id": str(row["id"]), "label": plain_text(row["label"])} for row in metodi_options()],
        ensure_ascii=False,
    )
    sidebar_toggle_button = """
    <button
      type="button"
      class="sidebar-toggle-button screen-only"
      data-sidebar-toggle="true"
      data-open-label="Apri menu"
      data-close-label="Nascondi menu"
      aria-expanded="true"
      onclick="toggleSidebar()">
      <span class="sidebar-toggle-icon" aria-hidden="true">&#9776;</span>
      <span class="sidebar-toggle-text">Nascondi menu</span>
    </button>
    """
    document = f"""<!doctype html>
<html lang="it">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{esc(title)} - {esc(APP_NAME)}</title>
    <link rel="icon" type="image/x-icon" href="/static/logo-ca.ico">
    <link rel="shortcut icon" href="/static/logo-ca.ico">
    <link rel="stylesheet" href="/static/style.css">
  </head>
  <body>
    <div class="app-shell">
      {render_navigation(current_path, query_params, current_user)}
      <main class="content">
        <header class="{header_class}">
          <div class="page-leading">
            {sidebar_toggle_button}
            <div class="page-title-block">
              <span class="eyebrow">{esc(APP_NAME)}</span>
              <h1>{esc(title)}</h1>
            </div>
          </div>
          {header_tools}
        </header>
        {message_banner(query_params)}
        {content}
      </main>
    </div>
    {app_dialog_markup()}
    <script>
      {shared_dialog_script()}
      const paymentMethodChoices = {payment_methods_json};
      const sidebarStateStorageKey = 'oratorio:sidebar-collapsed';

      function applySidebarState(collapsed) {{
        document.body.classList.toggle('sidebar-collapsed', !!collapsed);
        document.querySelectorAll('[data-sidebar-toggle="true"]').forEach((button) => {{
          const isCollapsed = !!collapsed;
          const label = isCollapsed
            ? (button.dataset.openLabel || 'Apri menu')
            : (button.dataset.closeLabel || 'Nascondi menu');
          button.setAttribute('aria-expanded', isCollapsed ? 'false' : 'true');
          const text = button.querySelector('.sidebar-toggle-text');
          if (text) {{
            text.textContent = label;
          }}
        }});
      }}

      function toggleSidebar(forceCollapsed = null) {{
        const nextState = forceCollapsed === null
          ? !document.body.classList.contains('sidebar-collapsed')
          : !!forceCollapsed;
        applySidebarState(nextState);
        try {{
          window.localStorage.setItem(sidebarStateStorageKey, nextState ? '1' : '0');
        }} catch (_) {{
          // Ignora gli errori di persistenza UI.
        }}
      }}

      function restoreSidebarState() {{
        let collapsed = false;
        try {{
          collapsed = window.localStorage.getItem(sidebarStateStorageKey) === '1';
        }} catch (_) {{
          collapsed = false;
        }}
        applySidebarState(collapsed);
      }}

      function setFormHiddenValue(form, fieldName, value) {{
        if (!form || !fieldName) {{
          return;
        }}
        let hidden = form.querySelector(`input[name="${{fieldName}}"]`);
        if (!hidden) {{
          hidden = document.createElement('input');
          hidden.type = 'hidden';
          hidden.name = fieldName;
          form.appendChild(hidden);
        }}
        hidden.value = value;
      }}

      function paymentMethodOptionsHtml(selectedValue) {{
        const selected = selectedValue || '';
        return paymentMethodChoices.map((choice) => {{
          const isSelected = String(choice.id) === String(selected) ? ' selected' : '';
          return `<option value="${{choice.id}}"${{isSelected}}>${{choice.label}}</option>`;
        }}).join('');
      }}

      function escapeHtml(value) {{
        return String(value || '')
          .replaceAll('&', '&amp;')
          .replaceAll('<', '&lt;')
          .replaceAll('>', '&gt;')
          .replaceAll('"', '&quot;')
          .replaceAll("'", '&#39;');
      }}

      function paymentHighlightCodeHtml(code, label = 'Numero tesseramento assegnato') {{
        const trimmed = String(code || '').trim();
        if (!trimmed) {{
          return '';
        }}
        return `
          <div class="payment-flow-highlight-code">
            <span>${{escapeHtml(label)}}</span>
            <strong>${{escapeHtml(trimmed)}}</strong>
          </div>
        `;
      }}

      function enrollmentExtraScadenzeHtml() {{
        return `
          <label class="field wide">
            <span>Altre scadenze aperte</span>
            <select class="control" name="dialog_include_other_scadenze">
              <option value="0">No, paga solo questa iscrizione</option>
              <option value="1">Si, seleziona altre scadenze aperte</option>
            </select>
          </label>
          <div class="field wide payment-flow-extra-wrap" hidden>
            <span>Scadenze aperte aggiuntive</span>
            <div class="payment-flow-extra-list" data-extra-scadenze-list></div>
            <p class="payment-flow-extra-note" data-extra-scadenze-note>
              Se selezioni altre scadenze, il gestionale proporra l'importo complessivo ma restera sempre modificabile.
            </p>
          </div>
        `;
      }}

      function parseCourseMonthValue(value) {{
        const raw = String(value || '').trim();
        const match = raw.match(/^(\\d{{4}})-(\\d{{2}})(?:-\\d{{2}})?$/);
        if (!match) {{
          return null;
        }}
        const year = Number.parseInt(match[1], 10);
        const month = Number.parseInt(match[2], 10);
        if (!Number.isInteger(year) || !Number.isInteger(month) || month < 1 || month > 12) {{
          return null;
        }}
        return {{ year, month }};
      }}

      function formatCourseMonthLabel(year, month) {{
        const safeYear = Number.parseInt(year, 10);
        const safeMonth = Number.parseInt(month, 10);
        if (!Number.isInteger(safeYear) || !Number.isInteger(safeMonth) || safeMonth < 1 || safeMonth > 12) {{
          return '';
        }}
        try {{
          const formatter = new Intl.DateTimeFormat('it-IT', {{ month: 'long', year: 'numeric' }});
          const base = formatter.format(new Date(safeYear, safeMonth - 1, 1));
          return base.charAt(0).toUpperCase() + base.slice(1);
        }} catch (_) {{
          return `${{String(safeMonth).padStart(2, '0')}}/${{safeYear}}`;
        }}
      }}

      function courseMonthSpan(startYear, startMonth, endYear, endMonth) {{
        return ((endYear - startYear) * 12) + (endMonth - startMonth) + 1;
      }}

      function enrollmentPaymentWorkYear(form) {{
        const field = form ? form.querySelector('[name="anno_lavoro"]') : null;
        const value = field ? Number.parseInt(String(field.value || '').trim(), 10) : Number.NaN;
        return Number.isInteger(value) ? value : new Date().getFullYear();
      }}

      function enrollmentPaymentAssociatoId(form) {{
        const field = form ? form.querySelector('[name="associato_id"]') : null;
        return field ? String(field.value || '').trim() : '';
      }}

      function selectedEnrollmentExtraTokens(extra) {{
        if (!extra) {{
          return [];
        }}
        return Array.from(extra.querySelectorAll('.payment-flow-extra-checkbox:checked'))
          .map((input) => String(input.value || '').trim())
          .filter(Boolean);
      }}

      function selectedEnrollmentExtraTotal(extra) {{
        if (!extra) {{
          return 0;
        }}
        return Array.from(extra.querySelectorAll('.payment-flow-extra-checkbox:checked')).reduce((total, input) => {{
          const amount = Number.parseFloat(String(input.dataset.residuo || '0').replace(',', '.'));
          return total + (Number.isFinite(amount) ? amount : 0);
        }}, 0);
      }}

      function populateEnrollmentExtraScadenzeList(extra, options = []) {{
        const list = extra ? extra.querySelector('[data-extra-scadenze-list]') : null;
        if (!list) {{
          return;
        }}
        list.innerHTML = '';
        if (!Array.isArray(options) || !options.length) {{
          list.dataset.loaded = '1';
          list.dataset.loading = '0';
          return;
        }}
        options.forEach((option) => {{
          const item = document.createElement('label');
          item.className = 'payment-flow-extra-item';
          const checkbox = document.createElement('input');
          checkbox.type = 'checkbox';
          checkbox.className = 'payment-flow-extra-checkbox';
          checkbox.value = String(option.id || '');
          checkbox.dataset.residuo = String(option.residuo || '0');
          const textWrap = document.createElement('span');
          textWrap.className = 'payment-flow-extra-item-text';
          const label = document.createElement('strong');
          label.className = 'payment-flow-extra-item-label';
          label.textContent = String(option.label || '');
          const detail = document.createElement('span');
          detail.className = 'payment-flow-extra-item-detail';
          detail.textContent = `Residuo ${{formatSummaryMoney(Number.parseFloat(String(option.residuo || '0').replace(',', '.')))}}`;
          textWrap.append(label, detail);
          item.append(checkbox, textWrap);
          list.appendChild(item);
        }});
        list.dataset.loaded = '1';
        list.dataset.loading = '0';
      }}

      function syncEnrollmentExtraScadenzeState(extra, options = {{}}) {{
        if (!extra) {{
          return;
        }}
        const includeField = extra.querySelector('[name="dialog_include_other_scadenze"]');
        const wrap = extra.querySelector('.payment-flow-extra-wrap');
        const list = extra.querySelector('[data-extra-scadenze-list]');
        const note = extra.querySelector('[data-extra-scadenze-note]');
        const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
        const baseAmountRaw = options.baseAmount ?? extra.dataset.baseAmount ?? '0';
        const baseAmount = Number.parseFloat(String(baseAmountRaw).replace(',', '.'));
        const safeBaseAmount = Number.isFinite(baseAmount) ? baseAmount : 0;
        extra.dataset.baseAmount = safeBaseAmount.toFixed(2);
        const showExtras = includeField && includeField.value === '1';
        if (wrap) {{
          wrap.hidden = !showExtras;
        }}

        let extraTotal = 0;
        let extraCount = 0;
        if (showExtras) {{
          extraTotal = selectedEnrollmentExtraTotal(extra);
          extraCount = selectedEnrollmentExtraTokens(extra).length;
        }}

        const suggestedAmount = safeBaseAmount + extraTotal;
        if (amountField) {{
          if (amountField.dataset.userEdited !== '1' || !String(amountField.value || '').trim()) {{
            amountField.value = suggestedAmount > 0 ? suggestedAmount.toFixed(2) : '';
          }}
          amountField.dataset.suggestedValue = suggestedAmount > 0 ? suggestedAmount.toFixed(2) : '';
        }}

        if (!note) {{
          return;
        }}
        if (!showExtras) {{
          note.textContent = 'VerrÃ  registrato solo il pagamento relativo all\\'iscrizione che stai effettuando.';
          return;
        }}
        if (list && list.dataset.loading === '1') {{
          note.textContent = 'Caricamento delle altre scadenze aperte in corso...';
          return;
        }}
        if (!list || list.children.length === 0) {{
          note.textContent = 'Non risultano altre scadenze aperte selezionabili per questo tesserato nell\\'anno di lavoro corrente.';
          return;
        }}
        if (extraCount === 0) {{
          note.textContent = 'Seleziona eventuali altre scadenze aperte da includere nello stesso pagamento.';
          return;
        }}
        note.textContent = `${{extraCount}} scadenze aggiuntive selezionate per un residuo complessivo di ${{formatSummaryMoney(extraTotal)}}.`;
      }}

      async function loadEnrollmentExtraScadenze(extra, form) {{
        if (!extra || !form) {{
          return;
        }}
        const list = extra.querySelector('[data-extra-scadenze-list]');
        if (!list) {{
          return;
        }}
        const associatoId = enrollmentPaymentAssociatoId(form);
        const workYear = enrollmentPaymentWorkYear(form);
        if (!associatoId) {{
          list.innerHTML = '';
          list.dataset.loaded = '1';
          list.dataset.loading = '0';
          syncEnrollmentExtraScadenzeState(extra);
          return;
        }}
        if (list.dataset.loadedFor === `${{associatoId}}:${{workYear}}`) {{
          syncEnrollmentExtraScadenzeState(extra);
          return;
        }}
        list.dataset.loading = '1';
        list.dataset.loaded = '0';
        list.dataset.loadedFor = `${{associatoId}}:${{workYear}}`;
        list.innerHTML = '<p class="payment-flow-extra-placeholder">Caricamento scadenze aperte...</p>';
        syncEnrollmentExtraScadenzeState(extra);
        try {{
          const params = new URLSearchParams({{
            associato_id: associatoId,
            anno_lavoro: String(workYear),
          }});
          const payload = await fetchJson(`/api/pagamenti-multi-area/scadenze-aperte?${{params.toString()}}`);
          populateEnrollmentExtraScadenzeList(extra, payload && Array.isArray(payload.options) ? payload.options : []);
        }} catch (error) {{
          console.warn('Caricamento scadenze aperte non disponibile', error);
          populateEnrollmentExtraScadenzeList(extra, []);
        }}
        syncEnrollmentExtraScadenzeState(extra);
      }}

      function bindEnrollmentExtraScadenze(extra, form, options = {{}}) {{
        if (!extra) {{
          return;
        }}
        const includeField = extra.querySelector('[name="dialog_include_other_scadenze"]');
        const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
        const list = extra.querySelector('[data-extra-scadenze-list]');
        if (amountField && amountField.dataset.extraBound !== '1') {{
          amountField.dataset.extraBound = '1';
          amountField.addEventListener('input', () => {{
            amountField.dataset.userEdited = '1';
          }});
        }}
        if (includeField && includeField.dataset.extraBound !== '1') {{
          includeField.dataset.extraBound = '1';
          includeField.addEventListener('change', async () => {{
            if (amountField) {{
              amountField.dataset.userEdited = '0';
            }}
            if (includeField.value === '1') {{
              await loadEnrollmentExtraScadenze(extra, form);
            }} else if (list) {{
              list.innerHTML = '';
              list.dataset.loaded = '0';
              list.dataset.loading = '0';
              list.dataset.loadedFor = '';
            }}
            syncEnrollmentExtraScadenzeState(extra, options);
          }});
        }}
        if (list && list.dataset.extraBound !== '1') {{
          list.dataset.extraBound = '1';
          list.addEventListener('change', (event) => {{
            if (!event.target || !event.target.classList.contains('payment-flow-extra-checkbox')) {{
              return;
            }}
            if (amountField) {{
              amountField.dataset.userEdited = '0';
            }}
            syncEnrollmentExtraScadenzeState(extra, options);
          }});
        }}
        syncEnrollmentExtraScadenzeState(extra, options);
      }}

      async function appPromptEnrollmentPayment(form, options = {{}}) {{
        const amountValue = options.defaultAmount ? String(options.defaultAmount) : '';
        const defaultMethodId = options.defaultMethodId || '';
        const allowExtraScadenze = options.allowExtraScadenze !== false;
        const dialogPromise = showAppDialog({{
          title: options.title || 'Registrazione pagamento',
          message: options.message || 'Conferma i dati del pagamento da registrare.',
          badge: 'Pagamento',
          variant: 'warning',
          confirmLabel: options.confirmLabel || 'Registra pagamento',
          cancelLabel: 'Annulla',
          dismissible: false,
          html: `
            <div class="payment-flow-grid">
              ${{options.extraIntroHtml || ''}}
              <label class="field wide">
                <span>Metodo</span>
                <select class="control" name="dialog_metodo_pagamento_id">
                  ${{paymentMethodOptionsHtml(defaultMethodId)}}
                </select>
              </label>
              ${{allowExtraScadenze ? enrollmentExtraScadenzeHtml() : ''}}
              <label class="field wide">
                <span>Importo pagato</span>
                <input class="control" type="number" name="dialog_importo_pagato" step="0.01" min="0.01" value="${{amountValue}}">
              </label>
              <p class="payment-flow-error" hidden></p>
            </div>
          `,
          resolveValue: (elements) => {{
            const extra = elements.extra;
            if (!extra) {{
              return false;
            }}
            const methodField = extra.querySelector('[name="dialog_metodo_pagamento_id"]');
            const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
            const includeField = extra.querySelector('[name="dialog_include_other_scadenze"]');
            const extraList = extra.querySelector('[data-extra-scadenze-list]');
            const errorField = extra.querySelector('.payment-flow-error');
            const methodValue = methodField ? String(methodField.value || '').trim() : '';
            const amountValueRaw = amountField ? String(amountField.value || '').trim() : '';
            const amountNumber = Number.parseFloat(amountValueRaw.replace(',', '.'));
            const includeExtras = includeField && includeField.value === '1';
            const extraTokens = includeExtras ? selectedEnrollmentExtraTokens(extra) : [];
            let errorMessage = '';
            if (!methodValue) {{
              errorMessage = 'Seleziona un metodo di pagamento.';
            }} else if (!Number.isFinite(amountNumber) || amountNumber <= 0) {{
              errorMessage = 'Indica un importo pagato valido.';
            }} else if (includeExtras && extraList && extraList.dataset.loading === '1') {{
              errorMessage = 'Attendi il caricamento delle altre scadenze aperte.';
            }}
            if (errorField) {{
              if (errorMessage) {{
                errorField.hidden = false;
                errorField.textContent = errorMessage;
              }} else {{
                errorField.hidden = true;
                errorField.textContent = '';
              }}
            }}
            if (errorMessage) {{
              return false;
            }}
            return {{
              confirmed: true,
              methodId: methodValue,
              importo: amountNumber.toFixed(2),
              extraTokens,
            }};
          }},
        }});
        window.setTimeout(() => {{
          const elements = appDialogElements();
          const extra = elements.extra;
          if (!extra) {{
            return;
          }}
          if (allowExtraScadenze) {{
            bindEnrollmentExtraScadenze(extra, form, {{ baseAmount: amountValue }});
          }}
        }}, 0);
        return dialogPromise;
      }}

      function syncCourseEnrollmentPaymentDialog(extra, options = {{}}) {{
        if (!extra) {{
          return;
        }}
        const scopeField = extra.querySelector('[name="dialog_payment_scope"]');
        const untilWrap = extra.querySelector('.payment-flow-until');
        const untilField = extra.querySelector('[name="dialog_fino_competenza"]');
        const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
        const summaryField = extra.querySelector('[data-payment-summary]');
        const errorField = extra.querySelector('.payment-flow-error');
        if (!scopeField || !untilWrap || !untilField || !amountField || !summaryField) {{
          return;
        }}
        const monthlyAmount = Number.parseFloat(String(options.monthlyAmount || '0').replace(',', '.'));
        const startYear = Number.parseInt(options.startYear, 10);
        const startMonth = Number.parseInt(options.startMonth, 10);
        const startLabel = formatCourseMonthLabel(startYear, startMonth);
        const scopeValue = scopeField.value === 'mensilita-future' ? 'mensilita-future' : 'mese-iscrizione';
        untilWrap.hidden = scopeValue !== 'mensilita-future';
        const endPeriod = parseCourseMonthValue(untilField.value) || {{ year: startYear, month: startMonth }};
        const span = courseMonthSpan(startYear, startMonth, endPeriod.year, endPeriod.month);
        const safeSpan = span > 0 ? span : 1;
        const suggestedAmount = Number.isFinite(monthlyAmount)
          ? (monthlyAmount * safeSpan).toFixed(2)
          : '';
        syncEnrollmentExtraScadenzeState(extra, {{ baseAmount: suggestedAmount }});
        if (errorField) {{
          errorField.hidden = true;
          errorField.textContent = '';
        }}
        if (scopeValue === 'mensilita-future') {{
          if (span < 1) {{
            summaryField.innerHTML = `Seleziona un mese finale uguale o successivo a <strong>${{startLabel}}</strong>.`;
          }} else {{
            const endLabel = formatCourseMonthLabel(endPeriod.year, endPeriod.month);
            summaryField.innerHTML = `Verranno generate automaticamente <strong>${{safeSpan}}</strong> mensilita da <strong>${{startLabel}}</strong> a <strong>${{endLabel}}</strong>.`;
          }}
        }} else {{
          summaryField.innerHTML = `VerrÃ  generata e proposta la quota del mese di iscrizione: <strong>${{startLabel}}</strong>.`;
        }}
      }}

      async function appPromptCourseEnrollmentPayment(form, options = {{}}) {{
        const enrollmentField = form ? form.querySelector('[name="data_iscrizione"]') : null;
        const enrollmentPeriod = parseCourseMonthValue(enrollmentField ? enrollmentField.value : '') || parseCourseMonthValue(new Date().toISOString().slice(0, 10));
        const startYear = enrollmentPeriod ? enrollmentPeriod.year : new Date().getFullYear();
        const startMonth = enrollmentPeriod ? enrollmentPeriod.month : (new Date().getMonth() + 1);
        const startMonthValue = `${{startYear}}-${{String(startMonth).padStart(2, '0')}}`;
        const monthlyAmount = Number.parseFloat(String(options.defaultAmount || '0').replace(',', '.'));
        const dialogPromise = showAppDialog({{
          title: options.title || 'Pagamento quote corso',
          message: options.message || 'Scegli se saldare il solo mese di iscrizione oppure anche mensilita future.',
          badge: 'Corso',
          variant: 'warning',
          confirmLabel: options.confirmLabel || 'Registra pagamento',
          cancelLabel: 'Annulla',
          dismissible: false,
          html: `
            <div class="payment-flow-grid">
              ${{options.extraIntroHtml || ''}}
              <label class="field wide">
                <span>Mensilita da pagare</span>
                <select class="control" name="dialog_payment_scope">
                  <option value="mese-iscrizione">Solo mese di iscrizione</option>
                  <option value="mensilita-future">Anche mensilita future</option>
                </select>
              </label>
              <label class="field wide payment-flow-until" hidden>
                <span>Fino al mese di competenza</span>
                <input class="control" type="month" name="dialog_fino_competenza" value="${{startMonthValue}}" min="${{startMonthValue}}">
              </label>
              <div class="payment-flow-note" data-payment-summary></div>
              <label class="field wide">
                <span>Metodo</span>
                <select class="control" name="dialog_metodo_pagamento_id">
                  ${{paymentMethodOptionsHtml(options.defaultMethodId || '')}}
                </select>
              </label>
              ${{enrollmentExtraScadenzeHtml()}}
              <label class="field wide">
                <span>Importo pagato</span>
                <input class="control" type="number" name="dialog_importo_pagato" step="0.01" min="0.01" value="${{Number.isFinite(monthlyAmount) ? monthlyAmount.toFixed(2) : ''}}">
              </label>
              <p class="payment-flow-error" hidden></p>
            </div>
          `,
          resolveValue: (elements) => {{
            const extra = elements.extra;
            if (!extra) {{
              return false;
            }}
            const scopeField = extra.querySelector('[name="dialog_payment_scope"]');
            const untilField = extra.querySelector('[name="dialog_fino_competenza"]');
            const methodField = extra.querySelector('[name="dialog_metodo_pagamento_id"]');
            const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
            const includeField = extra.querySelector('[name="dialog_include_other_scadenze"]');
            const extraList = extra.querySelector('[data-extra-scadenze-list]');
            const errorField = extra.querySelector('.payment-flow-error');
            const scopeValue = scopeField && scopeField.value === 'mensilita-future' ? 'mensilita-future' : 'mese-iscrizione';
            const methodValue = methodField ? String(methodField.value || '').trim() : '';
            const amountValueRaw = amountField ? String(amountField.value || '').trim() : '';
            const amountNumber = Number.parseFloat(amountValueRaw.replace(',', '.'));
            const includeExtras = includeField && includeField.value === '1';
            const extraTokens = includeExtras ? selectedEnrollmentExtraTokens(extra) : [];
            const endPeriod = scopeValue === 'mensilita-future'
              ? parseCourseMonthValue(untilField ? untilField.value : '')
              : {{ year: startYear, month: startMonth }};
            let errorMessage = '';
            if (!methodValue) {{
              errorMessage = 'Seleziona un metodo di pagamento.';
            }} else if (!Number.isFinite(amountNumber) || amountNumber <= 0) {{
              errorMessage = 'Indica un importo pagato valido.';
            }} else if (!endPeriod) {{
              errorMessage = 'Seleziona il mese finale delle quote da generare.';
            }} else if (courseMonthSpan(startYear, startMonth, endPeriod.year, endPeriod.month) < 1) {{
              errorMessage = 'Il mese finale deve essere uguale o successivo al mese di iscrizione.';
            }} else if (includeExtras && extraList && extraList.dataset.loading === '1') {{
              errorMessage = 'Attendi il caricamento delle altre scadenze aperte.';
            }}
            if (errorField) {{
              if (errorMessage) {{
                errorField.hidden = false;
                errorField.textContent = errorMessage;
              }} else {{
                errorField.hidden = true;
                errorField.textContent = '';
              }}
            }}
            if (errorMessage) {{
              return false;
            }}
            const untilCompetenza = `${{endPeriod.year}}-${{String(endPeriod.month).padStart(2, '0')}}`;
            return {{
              confirmed: true,
              methodId: methodValue,
              importo: amountNumber.toFixed(2),
              scope: scopeValue,
              untilCompetenza,
              extraTokens,
            }};
          }},
        }});

        window.setTimeout(() => {{
          const elements = appDialogElements();
          const extra = elements.extra;
          if (!extra) {{
            return;
          }}
          const scopeField = extra.querySelector('[name="dialog_payment_scope"]');
          const untilField = extra.querySelector('[name="dialog_fino_competenza"]');
          const amountField = extra.querySelector('[name="dialog_importo_pagato"]');
          if (!scopeField || !untilField || !amountField) {{
            return;
          }}
          bindEnrollmentExtraScadenze(extra, form, {{ baseAmount: Number.isFinite(monthlyAmount) ? monthlyAmount.toFixed(2) : '' }});
          const syncState = () => syncCourseEnrollmentPaymentDialog(extra, {{
            startYear,
            startMonth,
            monthlyAmount,
          }});
          scopeField.addEventListener('change', () => {{
            amountField.dataset.userEdited = '0';
            syncState();
          }});
          untilField.addEventListener('change', () => {{
            amountField.dataset.userEdited = '0';
            syncState();
          }});
          syncState();
        }}, 0);

        return dialogPromise;
      }}

      async function handleEnrollmentPaymentFlow(event) {{
        const form = event.currentTarget;
        if (!form || form.dataset.paymentFlowSubmitting === '1') {{
          return;
        }}
        event.preventDefault();
        const amountFieldName = form.dataset.paymentAmountField || '';
        const amountField = amountFieldName ? form.querySelector(`[name="${{amountFieldName}}"]`) : null;
        const defaultAmount = amountField ? amountField.value : (form.dataset.paymentDefaultAmount || '');
        const wantsPayment = await appConfirm(
          form.dataset.paymentPromptMessage || 'Vuoi procedere anche al pagamento?',
          {{
            title: form.dataset.paymentPromptTitle || 'Conferma iscrizione',
            badge: 'Iscrizione',
            confirmLabel: form.dataset.paymentPromptYes || 'Si, procedi',
            cancelLabel: form.dataset.paymentPromptNo || 'No, solo iscrizione',
            html: paymentHighlightCodeHtml(form.dataset.paymentPromptCode || '', form.dataset.paymentPromptCodeLabel || 'Numero tesseramento assegnato'),
          }}
        );
        if (!wantsPayment) {{
          setFormHiddenValue(form, 'procedi_pagamento', '0');
          setFormHiddenValue(form, 'pagamento_metodo_id', '');
          setFormHiddenValue(form, 'pagamento_importo', '');
          setFormHiddenValue(form, 'pagamento_scope', '');
          setFormHiddenValue(form, 'pagamento_competenza_fine', '');
          setFormHiddenValue(form, 'pagamento_scadenze_aggiuntive', '');
          form.dataset.paymentFlowSubmitting = '1';
          form.submit();
          return;
        }}

        const paymentFlowKind = String(form.dataset.paymentFlow || '').trim();
        const paymentResult = paymentFlowKind === 'corso'
          ? await appPromptCourseEnrollmentPayment(form, {{
              title: form.dataset.paymentDialogTitle || 'Pagamento quote corso',
              message: form.dataset.paymentDialogMessage || 'Scegli se pagare solo il mese di iscrizione oppure anche mensilita future.',
              confirmLabel: form.dataset.paymentDialogConfirm || 'Registra pagamento',
              defaultAmount,
              defaultMethodId: form.dataset.paymentMethodDefault || '',
              extraIntroHtml: paymentHighlightCodeHtml(form.dataset.paymentDialogCode || '', form.dataset.paymentDialogCodeLabel || 'Numero tesseramento assegnato'),
            }})
          : await appPromptEnrollmentPayment(form, {{
              title: form.dataset.paymentDialogTitle || 'Pagamento',
              message: form.dataset.paymentDialogMessage || 'Conferma il pagamento da registrare.',
              confirmLabel: form.dataset.paymentDialogConfirm || 'Registra pagamento',
              defaultAmount,
              defaultMethodId: form.dataset.paymentMethodDefault || '',
              allowExtraScadenze: form.dataset.paymentAllowExtraScadenze !== '0',
              extraIntroHtml: paymentHighlightCodeHtml(form.dataset.paymentDialogCode || '', form.dataset.paymentDialogCodeLabel || 'Numero tesseramento assegnato'),
            }});
        if (!paymentResult || paymentResult.confirmed !== true) {{
          return;
        }}

        setFormHiddenValue(form, 'procedi_pagamento', '1');
        setFormHiddenValue(form, 'pagamento_metodo_id', paymentResult.methodId);
        setFormHiddenValue(form, 'pagamento_importo', paymentResult.importo);
        setFormHiddenValue(form, 'pagamento_scope', paymentResult.scope || '');
        setFormHiddenValue(form, 'pagamento_competenza_fine', paymentResult.untilCompetenza || '');
        setFormHiddenValue(form, 'pagamento_scadenze_aggiuntive', (paymentResult.extraTokens || []).join(','));
        form.dataset.paymentFlowSubmitting = '1';
        form.submit();
      }}

      function bindEnrollmentPaymentFlows() {{
        document.querySelectorAll('form[data-payment-flow]').forEach((form) => {{
          if (form.dataset.paymentFlowBound === '1') {{
            return;
          }}
          form.dataset.paymentFlowBound = '1';
          form.addEventListener('submit', handleEnrollmentPaymentFlow);
        }});
      }}

      function syncDeleteFormView(form) {{
        if (!form) {{
          return;
        }}
        const vistaField = form.querySelector('input[name="vista"]');
        if (!vistaField) {{
          return;
        }}
        const params = new URLSearchParams(window.location.search);
        vistaField.value = params.get('vista') || '';
      }}

      function applyCombinedTableFilters(table) {{
        if (!table) {{
          return;
        }}
        const globalNeedle = normalizeSearchValue(table.dataset.globalFilter || '');
        const columnFilters = Array.from(
          document.querySelectorAll(`[data-table-filter][data-target-table="${{table.id}}"]`)
        )
          .map((input) => {{
            const needle = normalizeSearchValue(input.value || '');
            return {{
              key: input.dataset.columnKey || '',
              needle,
            }};
          }})
          .filter((entry) => entry.key && entry.needle);
        table.querySelectorAll('tbody tr').forEach((row) => {{
          const rowMatchesGlobal = !globalNeedle || normalizeSearchValue(row.textContent || '').includes(globalNeedle);
          const rowMatchesColumns = columnFilters.every((entry) => {{
            const cell = row.querySelector(`[data-column-key="${{entry.key}}"]`);
            return !!cell && normalizeSearchValue(cell.textContent || '').includes(entry.needle);
          }});
          row.style.display = rowMatchesGlobal && rowMatchesColumns ? '' : 'none';
        }});
        updateVisibleTableTotals(table);
      }}

      function filterRowsBySelector(input, selector) {{
        const needle = normalizeSearchValue(input.value || '');
        const targetTableId = input.dataset.targetTable || '';
        if (targetTableId) {{
          const table = document.getElementById(targetTableId);
          if (!table) {{
            return;
          }}
          table.dataset.globalFilter = needle;
          applyCombinedTableFilters(table);
          return;
        }}
        document.querySelectorAll(selector).forEach((table) => {{
          table.dataset.globalFilter = needle;
          applyCombinedTableFilters(table);
        }});
      }}

      function formatSummaryMoney(value) {{
        const amount = Number.isFinite(value) ? value : 0;
        return amount.toLocaleString('it-IT', {{
          minimumFractionDigits: 2,
          maximumFractionDigits: 2,
        }}) + ' EUR';
      }}

      function updateVisibleTableTotals(table) {{
        if (!table) {{
          return;
        }}
        const summaryColumns = (table.dataset.summaryColumns || '')
          .split(',')
          .map((value) => Number.parseInt(value, 10))
          .filter((value) => Number.isInteger(value));
        if (!summaryColumns.length) {{
          return;
        }}
        const footerRow = table.querySelector('tfoot tr');
        if (!footerRow) {{
          return;
        }}
        const labelIndex = Number.parseInt(table.dataset.summaryLabelIndex || '0', 10);
        const footerCells = Array.from(footerRow.children);
        footerCells.forEach((cell, index) => {{
          if (index === labelIndex) {{
            cell.textContent = 'Totali';
            return;
          }}
          if (!summaryColumns.includes(index)) {{
            cell.textContent = '';
          }}
        }});
        summaryColumns.forEach((columnIndex) => {{
          let total = 0;
          let found = false;
          table.querySelectorAll('tbody tr').forEach((row) => {{
            if (row.style.display === 'none') {{
              return;
            }}
            const cell = row.children[columnIndex];
            if (!cell || !cell.dataset.sumValue) {{
              return;
            }}
            const amount = Number.parseFloat(cell.dataset.sumValue);
            if (!Number.isFinite(amount)) {{
              return;
            }}
            total += amount;
            found = true;
          }});
          if (footerCells[columnIndex]) {{
            footerCells[columnIndex].textContent = found ? formatSummaryMoney(total) : '';
          }}
        }});
      }}

      function syncToolbarSearchLinks(input) {{
        const toolbar = input.closest('.report-toolbar');
        if (!toolbar) {{
          return;
        }}
        toolbar.querySelectorAll('[data-search-link]').forEach((link) => {{
          const url = new URL(link.href, window.location.origin);
          if (input.value) {{
            url.searchParams.set('search', input.value);
          }} else {{
            url.searchParams.delete('search');
          }}
          link.href = url.pathname + url.search;
        }});
      }}

      function handleReportSearch(input) {{
        filterRowsBySelector(input, '.report-table');
        syncToolbarSearchLinks(input);
        const tableId = input.dataset.targetTable || '';
        if (tableId) {{
          persistTableState(tableId);
        }}
      }}

      function handleDataSearch(input) {{
        filterRowsBySelector(input, '.search-table');
        const tableId = input.dataset.targetTable || '';
        if (tableId) {{
          persistTableState(tableId);
        }}
      }}

      function handleTableColumnFilter(input) {{
        const tableId = input.dataset.targetTable || '';
        if (!tableId) {{
          return;
        }}
        const table = document.getElementById(tableId);
        if (!table) {{
          return;
        }}
        applyCombinedTableFilters(table);
        persistTableState(tableId);
      }}

      function clearTableColumnFilters(tableId) {{
        const normalizedTableId = String(tableId || '').trim();
        if (!normalizedTableId) {{
          return false;
        }}
        document.querySelectorAll(`[data-table-filter][data-target-table="${{normalizedTableId}}"]`).forEach((input) => {{
          input.value = '';
        }});
        const table = document.getElementById(normalizedTableId);
        if (table) {{
          applyCombinedTableFilters(table);
        }}
        persistTableState(normalizedTableId);
        return false;
      }}

      function tableCellSortValue(cell) {{
        if (!cell) {{
          return '';
        }}
        return String(cell.dataset.sortValue || cell.textContent || '').trim();
      }}

      function setTableColumnVisibility(tableId, columnKey, visible) {{
        const table = document.getElementById(tableId);
        if (!table) {{
          return;
        }}
        table.querySelectorAll('[data-column-key]').forEach((cell) => {{
          if ((cell.dataset.columnKey || '') !== columnKey) {{
            return;
          }}
          cell.classList.toggle('column-hidden', !visible);
        }});
      }}

      function refreshTableSortIndexes(table) {{
        const headerCells = Array.from(table.querySelectorAll('thead th'));
        headerCells.forEach((cell, index) => {{
          const button = cell.querySelector('.table-sort-button');
          if (button) {{
            button.dataset.sortIndex = String(index);
          }}
        }});
      }}

      function shouldPersistTableState(tableId) {{
        return !!String(tableId || '').trim();
      }}

      function tableStateStorageKey(tableId) {{
        return `oratorio-table-state:${{window.location.pathname}}:${{tableId}}`;
      }}

      function getTableColumnOrder(table) {{
        if (!table) {{
          return [];
        }}
        const headerRow = table.querySelector('thead tr');
        if (!headerRow) {{
          return [];
        }}
        return Array.from(headerRow.children)
          .map((cell) => String(cell.dataset.columnKey || '').trim())
          .filter((value) => value);
      }}

      function reorderColumnPickerOptions(tableId, orderedKeys) {{
        const group = document.querySelector(`[data-column-toggle-group="${{tableId}}"]`);
        if (!group || !Array.isArray(orderedKeys) || !orderedKeys.length) {{
          return;
        }}
        const optionsByKey = new Map(
          Array.from(group.querySelectorAll('[data-column-toggle-chip]')).map((option) => [
            String(option.dataset.columnKey || '').trim(),
            option,
          ])
        );
        orderedKeys.forEach((key) => {{
          const option = optionsByKey.get(String(key || '').trim());
          if (option) {{
            group.appendChild(option);
          }}
        }});
      }}

      function readStoredTableState(tableId) {{
        if (!window.localStorage || !shouldPersistTableState(tableId)) {{
          return null;
        }}
        try {{
          const raw = window.localStorage.getItem(tableStateStorageKey(tableId));
          return raw ? JSON.parse(raw) : null;
        }} catch (error) {{
          console.warn('Impostazioni tabella non leggibili', error);
          return null;
        }}
      }}

      function persistTableState(tableId) {{
        if (!window.localStorage || !shouldPersistTableState(tableId)) {{
          return;
        }}
        const table = document.getElementById(tableId);
        if (!table) {{
          return;
        }}
        const searchInput = document.querySelector(`input[data-target-table="${{tableId}}"]:not([data-table-filter])`);
        const state = {{
          visibleColumns: Array.from(
            document.querySelectorAll(`[data-column-toggle][data-target-table="${{tableId}}"]`)
          )
            .filter((input) => input.checked)
            .map((input) => String(input.dataset.columnKey || '').trim())
            .filter((value) => value),
          columnOrder: getTableColumnOrder(table),
          sortColumnKey: (() => {{
            const sortIndex = Number.parseInt(table.dataset.sortIndex || '-1', 10);
            if (!Number.isInteger(sortIndex) || sortIndex < 0) {{
              return '';
            }}
            const headerCell = table.querySelectorAll('thead tr:first-child th')[sortIndex];
            return headerCell ? String(headerCell.dataset.columnKey || '').trim() : '';
          }})(),
          sortDirection: String(table.dataset.sortDirection || ''),
          columnFilters: Object.fromEntries(
            Array.from(document.querySelectorAll(`[data-table-filter][data-target-table="${{tableId}}"]`))
              .map((input) => [String(input.dataset.columnKey || '').trim(), String(input.value || '')])
              .filter((entry) => entry[0])
          ),
          globalSearch: searchInput ? String(searchInput.value || '') : '',
        }};
        try {{
          window.localStorage.setItem(tableStateStorageKey(tableId), JSON.stringify(state));
        }} catch (error) {{
          console.warn('Impostazioni tabella non salvabili', error);
        }}
      }}

      function applyStoredTableColumnOrder(table, orderedKeys) {{
        if (!table || !Array.isArray(orderedKeys) || !orderedKeys.length) {{
          return;
        }}
        orderedKeys.forEach((key, targetIndex) => {{
          const headerRow = table.querySelector('thead tr');
          if (!headerRow) {{
            return;
          }}
          const currentHeaders = Array.from(headerRow.children);
          const currentIndex = currentHeaders.findIndex((cell) => (cell.dataset.columnKey || '') === key);
          if (currentIndex < 0 || currentIndex === targetIndex || targetIndex >= currentHeaders.length) {{
            return;
          }}
          moveTableColumnByIndex(table, currentIndex, targetIndex);
        }});
        reorderColumnPickerOptions(table.id || '', getTableColumnOrder(table));
      }}

      function restorePersistedTableState(tableId) {{
        const state = readStoredTableState(tableId);
        if (!state) {{
          return;
        }}
        const table = document.getElementById(tableId);
        if (!table) {{
          return;
        }}
        if (Array.isArray(state.columnOrder) && state.columnOrder.length) {{
          applyStoredTableColumnOrder(table, state.columnOrder);
        }}
        if (Array.isArray(state.visibleColumns)) {{
          const visibleSet = new Set(state.visibleColumns.map((value) => String(value || '').trim()));
          document.querySelectorAll(`[data-column-toggle][data-target-table="${{tableId}}"]`).forEach((input) => {{
            const key = String(input.dataset.columnKey || '').trim();
            input.checked = visibleSet.has(key);
            setTableColumnVisibility(tableId, key, input.checked);
          }});
          updateColumnPickerSummary(tableId);
        }}
        if (state.sortColumnKey) {{
          const sortHeader = Array.from(table.querySelectorAll('thead tr:first-child th')).find(
            (cell) => String(cell.dataset.columnKey || '').trim() === String(state.sortColumnKey || '').trim()
          );
          if (sortHeader) {{
            sortDataTable(table, sortHeader.cellIndex, state.sortDirection || 'asc');
          }}
        }}
        if (state.columnFilters && typeof state.columnFilters === 'object') {{
          document.querySelectorAll(`[data-table-filter][data-target-table="${{tableId}}"]`).forEach((input) => {{
            const key = String(input.dataset.columnKey || '').trim();
            input.value = typeof state.columnFilters[key] === 'string' ? state.columnFilters[key] : '';
          }});
        }}
        const searchInput = document.querySelector(`input[data-target-table="${{tableId}}"]:not([data-table-filter])`);
        if (searchInput && typeof state.globalSearch === 'string') {{
          searchInput.value = state.globalSearch;
          table.dataset.globalFilter = normalizeSearchValue(state.globalSearch);
        }}
        applyCombinedTableFilters(table);
      }}

      function restorePersistedTableStates() {{
        document.querySelectorAll('table[id]').forEach((table) => {{
          if (shouldPersistTableState(table.id || '')) {{
            restorePersistedTableState(table.id || '');
          }}
        }});
      }}

      function setColumnToggleGroupState(tableId, visible) {{
        document.querySelectorAll(`[data-column-toggle][data-target-table="${{tableId}}"]`).forEach((input) => {{
          input.checked = visible;
          setTableColumnVisibility(tableId, input.dataset.columnKey || '', visible);
        }});
        updateColumnPickerSummary(tableId);
        persistTableState(tableId);
      }}

      function handleColumnPickerSearch(input) {{
        const tableId = String(input.dataset.targetTable || '').trim();
        const needle = normalizeSearchValue(String(input.value || ''));
        if (!tableId) {{
          return;
        }}
        document.querySelectorAll(`[data-column-toggle-chip][data-target-table="${{tableId}}"]`).forEach((option) => {{
          const text = normalizeSearchValue(option.textContent || '');
          option.classList.toggle('column-picker-option-hidden', !!needle && !text.includes(needle));
        }});
      }}

      function updateColumnPickerSummary(tableId) {{
        const picker = document.querySelector(`[data-column-picker="${{tableId}}"]`);
        if (!picker) {{
          return;
        }}
        const summary = picker.querySelector('.column-picker-summary');
        const inputs = Array.from(picker.querySelectorAll('[data-column-toggle]'));
        if (!summary || !inputs.length) {{
          return;
        }}
        const selected = inputs.filter((input) => input.checked).length;
        summary.textContent = `Seleziona colonne (${{selected}}/${{inputs.length}})`;
      }}

      function moveColumnToggleChip(tableId, columnKey, direction) {{
        const chip = document.querySelector(`[data-column-toggle-chip][data-target-table="${{tableId}}"][data-column-key="${{columnKey}}"]`);
        if (!chip || !chip.parentElement) {{
          return;
        }}
        const sibling = direction === 'left' ? chip.previousElementSibling : chip.nextElementSibling;
        if (!sibling) {{
          return;
        }}
        if (direction === 'left') {{
          chip.parentElement.insertBefore(chip, sibling);
        }} else {{
          chip.parentElement.insertBefore(chip, sibling.nextElementSibling);
        }}
      }}

      function moveTableColumn(tableId, columnKey, direction) {{
        const table = document.getElementById(tableId);
        if (!table) {{
          return;
        }}
        const headerCells = Array.from(table.querySelectorAll('thead th'));
        const currentIndex = headerCells.findIndex((cell) => (cell.dataset.columnKey || '') === columnKey);
        if (currentIndex < 0) {{
          return;
        }}
        const targetIndex = direction === 'left' ? currentIndex - 1 : currentIndex + 1;
        if (targetIndex < 0 || targetIndex >= headerCells.length) {{
          return;
        }}
        table.querySelectorAll('tr').forEach((row) => {{
          const cells = Array.from(row.children);
          const moving = cells[currentIndex];
          if (!moving) {{
            return;
          }}
          if (direction === 'left') {{
            row.insertBefore(moving, cells[targetIndex]);
          }} else {{
            row.insertBefore(moving, cells[targetIndex].nextElementSibling);
          }}
        }});
        refreshTableSortIndexes(table);
        updateVisibleTableTotals(table);
        reorderColumnPickerOptions(tableId, getTableColumnOrder(table));
        persistTableState(tableId);
      }}

      function applyInitialColumnVisibility() {{
        document.querySelectorAll('[data-column-toggle]').forEach((input) => {{
          setTableColumnVisibility(input.dataset.targetTable || '', input.dataset.columnKey || '', input.checked);
          input.addEventListener('change', () => {{
            setTableColumnVisibility(input.dataset.targetTable || '', input.dataset.columnKey || '', input.checked);
            updateColumnPickerSummary(input.dataset.targetTable || '');
            persistTableState(input.dataset.targetTable || '');
          }});
        }});
        document.querySelectorAll('[data-column-picker]').forEach((picker) => {{
          updateColumnPickerSummary(picker.dataset.columnPicker || '');
        }});
        document.querySelectorAll('[data-column-select-all]').forEach((button) => {{
          button.addEventListener('click', () => {{
            setColumnToggleGroupState(button.dataset.targetTable || '', true);
          }});
        }});
        document.querySelectorAll('[data-column-deselect-all]').forEach((button) => {{
          button.addEventListener('click', () => {{
            setColumnToggleGroupState(button.dataset.targetTable || '', false);
          }});
        }});
        document.querySelectorAll('[data-column-move]').forEach((button) => {{
          button.addEventListener('click', (event) => {{
            event.preventDefault();
            event.stopPropagation();
            const tableId = button.dataset.targetTable || '';
            const columnKey = button.dataset.columnKey || '';
            const direction = button.dataset.columnMove || 'left';
            moveTableColumn(tableId, columnKey, direction);
            moveColumnToggleChip(tableId, columnKey, direction);
          }});
        }});
      }}

      function moveTableColumnByIndex(table, fromIndex, toIndex) {{
        if (!table || fromIndex === toIndex || fromIndex < 0 || toIndex < 0) {{
          return;
        }}
        table.querySelectorAll('tr').forEach((row) => {{
          const cells = Array.from(row.children);
          const moving = cells[fromIndex];
          const target = cells[toIndex];
          if (!moving || !target) {{
            return;
          }}
          row.insertBefore(moving, fromIndex < toIndex ? target.nextElementSibling : target);
        }});
        refreshTableSortIndexes(table);
        updateVisibleTableTotals(table);
        if (table.id) {{
          reorderColumnPickerOptions(table.id, getTableColumnOrder(table));
          persistTableState(table.id);
        }}
      }}

      function initializeDraggableTableColumns() {{
        document.querySelectorAll('table[data-column-reorder="1"]').forEach((table) => {{
          const headerRow = table.querySelector('thead tr');
          if (!headerRow) {{
            return;
          }}
          let dragIndex = -1;
          Array.from(headerRow.children).forEach((cell) => {{
            if (!cell.dataset.columnKey || cell.dataset.draggableColumnBound === '1') {{
              return;
            }}
            cell.dataset.draggableColumnBound = '1';
            cell.addEventListener('dragstart', (event) => {{
              dragIndex = cell.cellIndex;
              cell.classList.add('column-dragging');
              if (event.dataTransfer) {{
                event.dataTransfer.effectAllowed = 'move';
                event.dataTransfer.setData('text/plain', cell.dataset.columnKey || '');
              }}
            }});
            cell.addEventListener('dragend', () => {{
              dragIndex = -1;
              table.querySelectorAll('.column-drag-over, .column-dragging').forEach((element) => {{
                element.classList.remove('column-drag-over', 'column-dragging');
              }});
            }});
            cell.addEventListener('dragover', (event) => {{
              if (dragIndex < 0 || dragIndex === cell.cellIndex) {{
                return;
              }}
              event.preventDefault();
              cell.classList.add('column-drag-over');
            }});
            cell.addEventListener('dragleave', () => {{
              cell.classList.remove('column-drag-over');
            }});
            cell.addEventListener('drop', (event) => {{
              if (dragIndex < 0 || dragIndex === cell.cellIndex) {{
                return;
              }}
              event.preventDefault();
              cell.classList.remove('column-drag-over');
              moveTableColumnByIndex(table, dragIndex, cell.cellIndex);
            }});
          }});
        }});
      }}

      function updateSortButtons(table, columnIndex, direction) {{
        table.querySelectorAll('.table-sort-button').forEach((button) => {{
          const isActive = Number.parseInt(button.dataset.sortIndex || '-1', 10) === columnIndex;
          button.classList.toggle('is-active', isActive);
          button.dataset.sortDirection = isActive ? direction : '';
        }});
      }}

      function sortDataTable(table, columnIndex, explicitDirection) {{
        const tbody = table.tBodies && table.tBodies[0];
        if (!tbody) {{
          return;
        }}
        const rows = Array.from(tbody.rows);
        if (!rows.length) {{
          return;
        }}
        const headerButton = table.querySelector(`.table-sort-button[data-sort-index="${{columnIndex}}"]`);
        const sortType = headerButton ? (headerButton.dataset.sortType || 'text') : 'text';
        const previousIndex = Number.parseInt(table.dataset.sortIndex || '-1', 10);
        const previousDirection = table.dataset.sortDirection || 'asc';
        const direction = explicitDirection || (previousIndex === columnIndex && previousDirection === 'asc' ? 'desc' : 'asc');
        rows.sort((rowA, rowB) => {{
          const valueA = tableCellSortValue(rowA.cells[columnIndex]);
          const valueB = tableCellSortValue(rowB.cells[columnIndex]);
          if (sortType === 'number') {{
            const numberA = Number.parseFloat(String(valueA).replace(',', '.'));
            const numberB = Number.parseFloat(String(valueB).replace(',', '.'));
            const safeA = Number.isFinite(numberA) ? numberA : Number.NEGATIVE_INFINITY;
            const safeB = Number.isFinite(numberB) ? numberB : Number.NEGATIVE_INFINITY;
            return direction === 'asc' ? safeA - safeB : safeB - safeA;
          }}
          const compare = String(valueA).localeCompare(String(valueB), 'it', {{ numeric: true, sensitivity: 'base' }});
          return direction === 'asc' ? compare : -compare;
        }});
        rows.forEach((row) => tbody.appendChild(row));
        table.dataset.sortIndex = String(columnIndex);
        table.dataset.sortDirection = direction;
        updateSortButtons(table, columnIndex, direction);
        updateVisibleTableTotals(table);
        if (table.id) {{
          persistTableState(table.id);
        }}
      }}

      function isMinorBirthDate(value) {{
        if (!value) {{
          return false;
        }}
        const birthDate = new Date(value);
        if (Number.isNaN(birthDate.getTime())) {{
          return false;
        }}
        const today = new Date();
        let age = today.getFullYear() - birthDate.getFullYear();
        const monthDiff = today.getMonth() - birthDate.getMonth();
        if (monthDiff < 0 || (monthDiff === 0 && today.getDate() < birthDate.getDate())) {{
          age -= 1;
        }}
        return age < 18;
      }}

      function syncMinorGuardianSections(form) {{
        if (!form) {{
          return;
        }}
        const birthDateInput = form.querySelector('[name="data_nascita"]');
        const isMinor = birthDateInput ? isMinorBirthDate(String(birthDateInput.value || '').trim()) : false;
        form.querySelectorAll('[data-minor-only="true"]').forEach((element) => {{
          element.classList.toggle('minor-only-hidden', !isMinor);
          element.querySelectorAll('[data-minor-required]').forEach((control) => {{
            control.required = isMinor;
          }});
        }});
      }}

      function bindMinorGuardianForms() {{
        document.querySelectorAll('form[data-minor-guardian-form="1"]').forEach((form) => {{
          if (form.dataset.minorGuardianBound === '1') {{
            syncMinorGuardianSections(form);
            return;
          }}
          form.dataset.minorGuardianBound = '1';
          const birthDateInput = form.querySelector('[name="data_nascita"]');
          if (birthDateInput) {{
            birthDateInput.addEventListener('input', () => syncMinorGuardianSections(form));
            birthDateInput.addEventListener('change', () => syncMinorGuardianSections(form));
          }}
          syncMinorGuardianSections(form);
        }});
      }}

      function initializeSortableTables() {{
        document.querySelectorAll('table.data-table').forEach((table) => {{
          table.querySelectorAll('.table-sort-button').forEach((button) => {{
            button.addEventListener('click', () => {{
              const columnIndex = Number.parseInt(button.dataset.sortIndex || '-1', 10);
              if (Number.isInteger(columnIndex) && columnIndex >= 0) {{
                sortDataTable(table, columnIndex);
              }}
            }});
          }});
        }});
      }}

      function normalizeSearchValue(value) {{
        return (value || '').toLowerCase().trim();
      }}

      function persistVisibleTableStates() {{
        document.querySelectorAll('table[id]').forEach((table) => {{
          if (shouldPersistTableState(table.id || '')) {{
            persistTableState(table.id || '');
          }}
        }});
      }}

      function applyTableExportStateToLink(link) {{
        const tableId = String(link.dataset.exportTable || '').trim();
        if (!tableId) {{
          return true;
        }}
        const url = new URL(link.getAttribute('href') || link.href, window.location.origin);
        const table = document.getElementById(tableId);
        if (!table) {{
          link.href = url.pathname + url.search;
          return true;
        }}

        const visibleColumns = Array.from(
          document.querySelectorAll(`[data-column-toggle][data-target-table="${{tableId}}"]`)
        )
          .filter((input) => input.checked)
          .map((input) => String(input.dataset.columnKey || '').trim())
          .filter((value) => value);
        const columnOrder = getTableColumnOrder(table);
        const searchInput = document.querySelector(`input[data-target-table="${{tableId}}"]:not([data-table-filter])`);
        const globalSearch = searchInput ? String(searchInput.value || '').trim() : '';
        const sortIndex = Number.parseInt(table.dataset.sortIndex || '-1', 10);
        const sortHeader = Number.isInteger(sortIndex) && sortIndex >= 0
          ? table.querySelectorAll('thead tr:first-child th')[sortIndex]
          : null;
        const sortColumn = sortHeader ? String(sortHeader.dataset.columnKey || '').trim() : '';
        const sortDirection = String(table.dataset.sortDirection || '').trim();

        ['columns', 'column_order', 'search', 'sort_column', 'sort_direction'].forEach((key) => {{
          url.searchParams.delete(key);
        }});
        Array.from(url.searchParams.keys()).forEach((key) => {{
          if (String(key || '').startsWith('cf_')) {{
            url.searchParams.delete(key);
          }}
        }});

        if (visibleColumns.length) {{
          url.searchParams.set('columns', visibleColumns.join(','));
        }}
        if (columnOrder.length) {{
          url.searchParams.set('column_order', columnOrder.join(','));
        }}
        if (globalSearch) {{
          url.searchParams.set('search', globalSearch);
        }}
        if (sortColumn) {{
          url.searchParams.set('sort_column', sortColumn);
        }}
        if (sortDirection) {{
          url.searchParams.set('sort_direction', sortDirection);
        }}
        document.querySelectorAll(`[data-table-filter][data-target-table="${{tableId}}"]`).forEach((input) => {{
          const columnKey = String(input.dataset.columnKey || '').trim();
          const filterValue = String(input.value || '').trim();
          if (columnKey && filterValue) {{
            url.searchParams.set(`cf_${{columnKey}}`, filterValue);
          }}
        }});

        link.href = url.pathname + url.search;
        return true;
      }}

      function getAutocompleteElements(input) {{
        const targetId = input.dataset.selectSearchTarget || '';
        const panelId = input.dataset.selectSearchPanel || '';
        return {{
          select: targetId ? document.getElementById(targetId) : null,
          panel: panelId ? document.getElementById(panelId) : null,
        }};
      }}

      function autocompleteOptions(select) {{
        return Array.from(select.options).filter((option) => !option.disabled && option.value);
      }}

      function autocompleteDisplayLabel(option) {{
        return option.dataset.autocompleteLabel || option.textContent || '';
      }}

      function autocompleteSearchText(option) {{
        return normalizeSearchValue(option.dataset.searchText || option.textContent || '');
      }}

      function autocompleteMatches(option, query) {{
        const visibleText = normalizeSearchValue(option.textContent || '');
        const searchText = autocompleteSearchText(option);
        const displayText = normalizeSearchValue(autocompleteDisplayLabel(option));
        return (
          visibleText.startsWith(query)
          || searchText.startsWith(query)
          || displayText.startsWith(query)
          || visibleText.includes(query)
          || searchText.includes(query)
          || displayText.includes(query)
        );
      }}

      function findAutocompleteOption(select, value) {{
        return autocompleteOptions(select).find((option) => option.value === value) || null;
      }}

      function autocompleteItems(panel) {{
        return Array.from(panel.querySelectorAll('.select-autocomplete-item'));
      }}

      function setActiveAutocompleteItem(panel, nextIndex) {{
        const items = autocompleteItems(panel);
        if (!items.length) {{
          panel.dataset.activeIndex = '-1';
          return;
        }}
        const boundedIndex = Math.max(0, Math.min(nextIndex, items.length - 1));
        items.forEach((item, index) => {{
          item.classList.toggle('active', index === boundedIndex);
        }});
        panel.dataset.activeIndex = String(boundedIndex);
      }}

      function closeSelectAutocomplete(input) {{
        const {{ panel }} = getAutocompleteElements(input);
        if (!panel) {{
          return;
        }}
        panel.hidden = true;
        panel.replaceChildren();
        panel.dataset.activeIndex = '-1';
      }}

      function selectAutocompleteValue(input, optionValue) {{
        const {{ select }} = getAutocompleteElements(input);
        if (!select) {{
          return false;
        }}
        const option = findAutocompleteOption(select, optionValue);
        if (!option) {{
          return false;
        }}
        select.value = option.value;
        input.value = autocompleteDisplayLabel(option);
        input.setCustomValidity('');
        select.dispatchEvent(new Event('change', {{ bubbles: true }}));
        closeSelectAutocomplete(input);
        return true;
      }}

      function clearAutocompleteValue(input) {{
        const {{ select }} = getAutocompleteElements(input);
        if (!select) {{
          return;
        }}
        select.value = '';
      }}

      function renderSelectAutocomplete(input, showAllOnEmpty = false) {{
        const {{ select, panel }} = getAutocompleteElements(input);
        if (!select || !panel) {{
          return [];
        }}

        const query = normalizeSearchValue(input.value);
        if (!query && !showAllOnEmpty) {{
          closeSelectAutocomplete(input);
          return [];
        }}

        const matches = autocompleteOptions(select)
          .filter((option) => !query || autocompleteMatches(option, query))
          .slice(0, 8);

        panel.replaceChildren();
        if (!matches.length) {{
          const emptyState = document.createElement('div');
          emptyState.className = 'select-autocomplete-empty';
          emptyState.textContent = 'Nessun associato trovato';
          panel.appendChild(emptyState);
          panel.hidden = false;
          panel.dataset.activeIndex = '-1';
          return [];
        }}

        matches.forEach((option, index) => {{
          const item = document.createElement('button');
          item.type = 'button';
          item.className = 'select-autocomplete-item';
          item.dataset.optionValue = option.value;
          item.textContent = autocompleteDisplayLabel(option);
          item.addEventListener('mousedown', (event) => {{
            event.preventDefault();
          }});
          item.addEventListener('click', () => {{
            selectAutocompleteValue(input, option.value);
          }});
          panel.appendChild(item);
          if (index === 0) {{
            item.classList.add('active');
          }}
        }});

        panel.hidden = false;
        panel.dataset.activeIndex = matches.length ? '0' : '-1';
        return matches;
      }}

      function syncAutocompleteInput(input) {{
        const {{ select }} = getAutocompleteElements(input);
        if (!select) {{
          return;
        }}
        const option = select.options[select.selectedIndex];
        input.value = option && option.value ? autocompleteDisplayLabel(option) : '';
      }}

      function commitSelectAutocomplete(input, strictRequired = false) {{
        const {{ select }} = getAutocompleteElements(input);
        if (!select) {{
          return true;
        }}

        const query = normalizeSearchValue(input.value);
        input.setCustomValidity('');
        if (!query) {{
          clearAutocompleteValue(input);
          closeSelectAutocomplete(input);
          if (strictRequired && input.required) {{
            input.setCustomValidity('Seleziona un associato valido.');
            return false;
          }}
          return true;
        }}

        const current = select.options[select.selectedIndex];
        if (current && current.value && autocompleteMatches(current, query)) {{
          input.value = autocompleteDisplayLabel(current);
          closeSelectAutocomplete(input);
          return true;
        }}

        const matches = renderSelectAutocomplete(input);
        if (matches.length) {{
          return selectAutocompleteValue(input, matches[0].value);
        }}

        clearAutocompleteValue(input);
        closeSelectAutocomplete(input);
        if (strictRequired && input.required) {{
          input.setCustomValidity('Seleziona un associato valido.');
          return false;
        }}
        return true;
      }}

      function openSelectAutocomplete(input) {{
        renderSelectAutocomplete(input, true);
      }}

      function handleSelectSearch(input) {{
        input.setCustomValidity('');
        if (!normalizeSearchValue(input.value)) {{
          clearAutocompleteValue(input);
          closeSelectAutocomplete(input);
          return;
        }}
        renderSelectAutocomplete(input);
      }}

      function handleSelectSearchKeydown(event, input) {{
        const {{ panel }} = getAutocompleteElements(input);
        if (!panel) {{
          return;
        }}

        if (event.key === 'ArrowDown' || event.key === 'ArrowUp') {{
          event.preventDefault();
          if (panel.hidden) {{
            renderSelectAutocomplete(input, true);
          }}
          const items = autocompleteItems(panel);
          if (!items.length) {{
            return;
          }}
          const currentIndex = parseInt(panel.dataset.activeIndex || '0', 10);
          const delta = event.key === 'ArrowDown' ? 1 : -1;
          setActiveAutocompleteItem(panel, currentIndex + delta);
          return;
        }}

        if (event.key === 'Enter') {{
          const items = autocompleteItems(panel);
          if (!panel.hidden && items.length) {{
            event.preventDefault();
            const activeIndex = parseInt(panel.dataset.activeIndex || '0', 10);
            const activeItem = items[Math.max(0, Math.min(activeIndex, items.length - 1))];
            if (activeItem) {{
              selectAutocompleteValue(input, activeItem.dataset.optionValue || '');
            }}
          }}
          return;
        }}

        if (event.key === 'Escape') {{
          closeSelectAutocomplete(input);
        }}
      }}

      function closeSelectAutocompleteLater(input) {{
        window.setTimeout(() => {{
          commitSelectAutocomplete(input);
        }}, 120);
      }}

      function setInputValue(inputId, value) {{
        const target = document.getElementById(inputId);
        if (!target) {{
          return;
        }}
        target.value = value || '';
      }}

      function syncResidualInput(select) {{
        const targetId = select.dataset.residualTarget || '';
        if (!targetId) {{
          return;
        }}
        const option = select.options[select.selectedIndex];
        const residuo = option && option.dataset ? option.dataset.residuo || '' : '';
        setInputValue(targetId, residuo);
      }}

      function syncStandardQuota(select) {{
        const targetId = select.dataset.standardTarget || '';
        if (!targetId) {{
          return;
        }}
        const option = select.options[select.selectedIndex];
        const quota = option && option.dataset ? option.dataset.quotaStandard || '' : '';
        setInputValue(targetId, quota);
      }}

      function syncAmountProposal(select) {{
        const targetId = select.dataset.amountTarget || '';
        if (!targetId) {{
          return;
        }}
        const option = select.options[select.selectedIndex];
        const amount = option && option.dataset ? option.dataset.importo || '' : '';
        setInputValue(targetId, amount);
      }}

      function syncMirroredValue(source) {{
        const targetId = source.dataset.copyValueTarget || '';
        if (!targetId) {{
          return;
        }}
        const target = document.getElementById(targetId);
        if (!target) {{
          return;
        }}
        const sourceValue = source.value || '';
        const lastSourceValue = target.dataset.lastMirroredSourceValue || '';
        const isAutoManaged = target.dataset.autoManagedValue !== 'false';
        if (!target.value || isAutoManaged || target.value === lastSourceValue) {{
          target.value = sourceValue;
          target.dataset.lastMirroredSourceValue = sourceValue;
          target.dataset.autoManagedValue = 'true';
        }}
      }}

      function syncYearEndDate(input) {{
        const targetId = input.dataset.yearEndTarget || '';
        if (!targetId) {{
          return;
        }}
        const normalized = (input.value || '').replace(/\\D/g, '').slice(0, 4);
        input.value = normalized;
        if (normalized.length !== 4) {{
          return;
        }}
        setInputValue(targetId, `${{normalized}}-12-31`);
      }}

      async function syncTesseramentoCodePreview(input) {{
        const targetId = input.dataset.tesseramentoCodeTarget || '';
        if (!targetId) {{
          return;
        }}
        const target = document.getElementById(targetId);
        if (!target) {{
          return;
        }}
        const normalized = (input.value || '').replace(/\\D/g, '').slice(0, 4);
        if (normalized.length !== 4) {{
          target.textContent = '--';
          return;
        }}
        try {{
          const payload = await fetchJson(`/api/tesseramenti/codice-anteprima?anno_sociale=${{encodeURIComponent(normalized)}}`);
          if (payload && payload.ok && payload.codice_tesseramento) {{
            target.textContent = String(payload.codice_tesseramento);
            return;
          }}
        }} catch (error) {{
          console.warn('Anteprima codice tesseramento non disponibile', error);
        }}
        target.textContent = `${{normalized.slice(-2)}}/?`;
      }}

      function syncMonthEndDate(control) {{
        const targetId = control.dataset.monthEndTarget || '';
        const yearSourceId = control.dataset.monthEndYearSource || '';
        const monthSourceId = control.dataset.monthEndMonthSource || '';
        if (!targetId || !yearSourceId || !monthSourceId) {{
          return;
        }}
        const yearSource = document.getElementById(yearSourceId);
        const monthSource = document.getElementById(monthSourceId);
        if (!yearSource || !monthSource) {{
          return;
        }}
        const yearValue = (yearSource.value || '').replace(/\D/g, '').slice(0, 4);
        yearSource.value = yearValue;
        const monthValue = Number.parseInt(monthSource.value || '', 10);
        if (yearValue.length !== 4 || !Number.isInteger(monthValue) || monthValue < 1 || monthValue > 12) {{
          return;
        }}
        const lastDay = new Date(Number.parseInt(yearValue, 10), monthValue, 0).getDate();
        setInputValue(targetId, `${{yearValue}}-${{String(monthValue).padStart(2, '0')}}-${{String(lastDay).padStart(2, '0')}}`);
      }}

      function enrollmentTesseramentoRedirectUrl(select) {{
        const workYear = select.dataset.requireTesseramentoYear || '';
        const associatoId = select.value || '';
        const areaLabel = select.dataset.requireTesseramentoArea || 'questa area';
        const url = new URL('/maschere/tesseramenti', window.location.origin);
        if (workYear) {{
          url.searchParams.set('anno_lavoro', workYear);
        }}
        if (associatoId) {{
          url.searchParams.set('associato_id', associatoId);
        }}
        if (workYear) {{
          url.searchParams.set('err', `Per registrare l'iscrizione a ${{areaLabel}} devi prima inserire il tesseramento dell'anno ${{workYear}}.`);
        }}
        return url;
      }}

      async function checkEnrollmentTesseramento(select) {{
        if (!select || !select.value) {{
          return true;
        }}
        const option = select.options[select.selectedIndex];
        if (!option || (option.dataset.hasTesseramento || '') === '1') {{
          return true;
        }}
        const workYear = select.dataset.requireTesseramentoYear || '';
        const areaLabel = select.dataset.requireTesseramentoArea || 'questa area';
        await appAlert(
          `Per registrare l'iscrizione a ${{areaLabel}} devi prima inserire il tesseramento dell'anno ${{workYear}}.`,
          {{
            title: 'Tesseramento richiesto',
            variant: 'warning',
            badge: 'Controllo dati',
            confirmLabel: 'Apri tesseramento',
            dismissible: false,
          }},
        );
        const redirectUrl = enrollmentTesseramentoRedirectUrl(select);
        window.location.href = redirectUrl.pathname + redirectUrl.search;
        return false;
      }}

      function findTypeaheadMatch(options, buffer) {{
        const normalized = (buffer || '').toLowerCase().trim();
        if (!normalized) {{
          return null;
        }}
        return (
          options.find((option) => {{
            const visibleText = (option.textContent || '').toLowerCase();
            const searchText = (option.dataset.searchText || '').toLowerCase();
            return visibleText.startsWith(normalized) || searchText.startsWith(normalized);
          }})
          || options.find((option) => {{
            const visibleText = (option.textContent || '').toLowerCase();
            const searchText = (option.dataset.searchText || '').toLowerCase();
            return visibleText.includes(normalized) || searchText.includes(normalized);
          }})
          || null
        );
      }}

      function installSelectTypeahead(select) {{
        let buffer = '';
        let resetTimer = null;
        select.addEventListener('keydown', (event) => {{
          if (event.key.length !== 1 || event.altKey || event.ctrlKey || event.metaKey) {{
            return;
          }}
          buffer += event.key.toLowerCase();
          window.clearTimeout(resetTimer);
          resetTimer = window.setTimeout(() => {{
            buffer = '';
          }}, 700);

          const options = Array.from(select.options).filter((option) => !option.disabled && !option.hidden && option.value);
          const match = findTypeaheadMatch(options, buffer)
            || findTypeaheadMatch(options, event.key.toLowerCase());
          if (!match) {{
            return;
          }}

          event.preventDefault();
          if (select.multiple) {{
            Array.from(select.options).forEach((option) => {{
              if (!option.hidden && !option.disabled) {{
                option.selected = false;
              }}
            }});
            match.selected = true;
          }} else {{
            select.value = match.value;
          }}
          match.scrollIntoView({{ block: 'nearest' }});
          select.dispatchEvent(new Event('change', {{ bubbles: true }}));
        }});
        select.addEventListener('blur', () => {{
          buffer = '';
          window.clearTimeout(resetTimer);
        }});
      }}

      function updateMonthlyRates() {{
        const associatoSelect = document.getElementById('monthly-payment-associato');
        const rateSelect = document.getElementById('monthly-rate-select');
        const amountInput = document.getElementById('monthly-payment-amount');
        if (!associatoSelect || !rateSelect || !amountInput) {{
          return;
        }}

        const associatoId = associatoSelect.value;
        let total = 0;
        let visibleCount = 0;
        Array.from(rateSelect.options).forEach((option) => {{
          const visible = !!associatoId && option.dataset.associatoId === associatoId;
          option.hidden = !visible;
          option.disabled = !visible;
          option.selected = visible;
          if (visible) {{
            total += parseFloat(option.dataset.residuo || '0');
            visibleCount += 1;
          }}
        }});
        amountInput.value = visibleCount ? total.toFixed(2) : '';
      }}

      function syncMonthlySelectedAmount() {{
        const rateSelect = document.getElementById('monthly-rate-select');
        const amountInput = document.getElementById('monthly-payment-amount');
        if (!rateSelect || !amountInput) {{
          return;
        }}
        const total = Array.from(rateSelect.selectedOptions)
          .filter((option) => !option.hidden)
          .reduce((sum, option) => sum + parseFloat(option.dataset.residuo || '0'), 0);
        amountInput.value = total ? total.toFixed(2) : '';
      }}

      function updateMultiAreaScadenze(selectedValues = null) {{
        const associatoSelect = document.getElementById('multi-area-associato');
        const scadenzeSelect = document.getElementById('multi-area-scadenze');
        const amountInput = document.getElementById('multi-area-amount');
        if (!associatoSelect || !scadenzeSelect || !amountInput) {{
          return;
        }}

        const associatoId = associatoSelect.value;
        const selectedSet = new Set(Array.isArray(selectedValues) ? selectedValues : []);
        const hasSelectionOverride = Array.isArray(selectedValues);
        let total = 0;
        Array.from(scadenzeSelect.options).forEach((option) => {{
          const visible = !!associatoId && option.dataset.associatoId === associatoId;
          option.hidden = !visible;
          option.disabled = !visible;
          if (!visible) {{
            option.selected = false;
            return;
          }}
          if (hasSelectionOverride) {{
            option.selected = selectedSet.has(option.value);
          }}
          if (option.selected) {{
            total += parseFloat(option.dataset.residuo || '0');
          }}
        }});
        amountInput.value = total ? total.toFixed(2) : '';
        renderMultiAreaScadenzeVisual();
      }}

      function syncMultiAreaSelectedAmount() {{
        const scadenzeSelect = document.getElementById('multi-area-scadenze');
        const amountInput = document.getElementById('multi-area-amount');
        if (!scadenzeSelect || !amountInput) {{
          return;
        }}
        const total = Array.from(scadenzeSelect.selectedOptions)
          .filter((option) => !option.hidden)
          .reduce((sum, option) => sum + parseFloat(option.dataset.residuo || '0'), 0);
        amountInput.value = total ? total.toFixed(2) : '';
        renderMultiAreaScadenzeVisual();
      }}

      function multiAreaSelectedScadenze() {{
        const scadenzeSelect = document.getElementById('multi-area-scadenze');
        if (!scadenzeSelect) {{
          return [];
        }}
        return Array.from(scadenzeSelect.selectedOptions)
          .filter((option) => !option.hidden)
          .map((option) => option.value);
      }}

      function renderMultiAreaScadenzeVisual() {{
        const associatoSelect = document.getElementById('multi-area-associato');
        const scadenzeSelect = document.getElementById('multi-area-scadenze');
        const container = document.getElementById('multi-area-scadenze-visual');
        const note = document.getElementById('multi-area-scadenze-note');
        if (!associatoSelect || !scadenzeSelect || !container) {{
          return;
        }}

        const associatoId = associatoSelect.value;
        const visibleOptions = Array.from(scadenzeSelect.options).filter((option) => !option.hidden && !option.disabled);
        container.innerHTML = '';

        if (!associatoId) {{
          if (note) {{
            note.textContent = 'Seleziona prima un tesserato per visualizzare le scadenze aperte disponibili.';
          }}
          return;
        }}

        if (!visibleOptions.length) {{
          if (note) {{
            note.textContent = 'Non risultano scadenze aperte disponibili per il tesserato selezionato.';
          }}
          return;
        }}

        visibleOptions.forEach((option) => {{
          const item = document.createElement('label');
          item.className = 'payment-flow-extra-item';
          const checkbox = document.createElement('input');
          checkbox.type = 'checkbox';
          checkbox.className = 'payment-flow-extra-checkbox';
          checkbox.checked = option.selected;
          checkbox.dataset.optionValue = option.value;
          checkbox.dataset.residuo = String(option.dataset.residuo || '0');
          checkbox.addEventListener('change', () => {{
            option.selected = checkbox.checked;
            syncMultiAreaSelectedAmount();
          }});
          const textWrap = document.createElement('span');
          textWrap.className = 'payment-flow-extra-item-text';
          const label = document.createElement('strong');
          label.className = 'payment-flow-extra-item-label';
          label.textContent = String(option.textContent || '');
          const detail = document.createElement('span');
          detail.className = 'payment-flow-extra-item-detail';
          detail.textContent = `Residuo ${{formatSummaryMoney(Number.parseFloat(String(option.dataset.residuo || '0').replace(',', '.')))}}`;
          textWrap.append(label, detail);
          item.append(checkbox, textWrap);
          container.appendChild(item);
        }});

        if (note) {{
          const selectedOptions = visibleOptions.filter((option) => option.selected);
          if (!selectedOptions.length) {{
            note.textContent = 'Seleziona le scadenze da includere nel pagamento unico.';
          }} else {{
            const total = selectedOptions.reduce((sum, option) => sum + parseFloat(option.dataset.residuo || '0'), 0);
            note.textContent = `${{selectedOptions.length}} scadenze selezionate per un residuo complessivo di ${{formatSummaryMoney(total)}}.`;
          }}
        }}
      }}

      function replaceMultiAreaScadenzeOptions(rows, selectedValues = []) {{
        const scadenzeSelect = document.getElementById('multi-area-scadenze');
        if (!scadenzeSelect) {{
          return;
        }}
        scadenzeSelect.innerHTML = '';
        (rows || []).forEach((row) => {{
          const option = document.createElement('option');
          option.value = String(row.id || '');
          option.textContent = String(row.label || '');
          option.dataset.associatoId = String(row.associato_id || '');
          option.dataset.residuo = String(row.residuo || '0');
          scadenzeSelect.appendChild(option);
        }});
        updateMultiAreaScadenze(selectedValues);
      }}

      function syncMultiAreaFutureCourseControls() {{
        const associatoSelect = document.getElementById('multi-area-associato');
        const corsoSelect = document.getElementById('multi-area-course-enrollment');
        const untilInput = document.getElementById('multi-area-course-until');
        const actionButton = document.getElementById('multi-area-course-generate');
        const feedback = document.getElementById('multi-area-course-feedback');
        if (!associatoSelect || !corsoSelect || !untilInput || !actionButton) {{
          return;
        }}

        const associatoId = associatoSelect.value;
        let firstVisibleValue = '';
        Array.from(corsoSelect.options).forEach((option) => {{
          if (!option.value) {{
            option.hidden = false;
            option.disabled = false;
            return;
          }}
          const visible = !!associatoId && option.dataset.associatoId === associatoId;
          option.hidden = !visible;
          option.disabled = !visible;
          if (visible && !firstVisibleValue) {{
            firstVisibleValue = option.value;
          }}
        }});

        const selectedOption = corsoSelect.selectedOptions.length ? corsoSelect.selectedOptions[0] : null;
        if (!selectedOption || selectedOption.hidden || selectedOption.disabled) {{
          corsoSelect.value = firstVisibleValue;
        }}

        const activeOption = corsoSelect.selectedOptions.length ? corsoSelect.selectedOptions[0] : null;
        const startCompetenza = activeOption ? String(activeOption.dataset.startCompetenza || '') : '';
        const workYear = String(corsoSelect.dataset.workYear || '');
        const workYearStart = workYear ? `${{workYear}}-01` : '';
        const maxCompetenza = workYear ? `${{workYear}}-12` : '';
        untilInput.disabled = !activeOption || !activeOption.value;
        actionButton.disabled = !associatoId || !activeOption || !activeOption.value || !untilInput.value;
        const effectiveMin = startCompetenza && workYearStart
          ? (startCompetenza > workYearStart ? startCompetenza : workYearStart)
          : (startCompetenza || workYearStart);
        if (effectiveMin) {{
          untilInput.min = effectiveMin;
          if (!untilInput.value || untilInput.value < effectiveMin) {{
            untilInput.value = effectiveMin;
          }}
        }} else {{
          untilInput.min = '';
        }}
        if (maxCompetenza) {{
          untilInput.max = maxCompetenza;
          if (untilInput.value && untilInput.value > maxCompetenza) {{
            untilInput.value = maxCompetenza;
          }}
        }} else {{
          untilInput.max = '';
        }}
        actionButton.disabled = !associatoId || !activeOption || !activeOption.value || !untilInput.value;
        if (feedback && !feedback.dataset.preserveMessage) {{
          feedback.textContent = activeOption && activeOption.value
            ? 'Le mensilita del corso fino al mese indicato verranno selezionate automaticamente in basso.'
            : 'Seleziona prima tesserato e corso per aggiungere quote future.';
        }}
      }}

      async function generateFutureCourseMultiAreaScadenze() {{
        const associatoSelect = document.getElementById('multi-area-associato');
        const corsoSelect = document.getElementById('multi-area-course-enrollment');
        const untilInput = document.getElementById('multi-area-course-until');
        const actionButton = document.getElementById('multi-area-course-generate');
        const feedback = document.getElementById('multi-area-course-feedback');
        if (!associatoSelect || !corsoSelect || !untilInput || !actionButton) {{
          return false;
        }}

        const associatoId = String(associatoSelect.value || '').trim();
        const iscrizioneCorsoId = String(corsoSelect.value || '').trim();
        const fineCompetenza = String(untilInput.value || '').trim();
        if (!associatoId) {{
          await appAlert('Seleziona prima un tesserato.', {{ title: 'Quote future corso', badge: 'Pagamenti' }});
          return false;
        }}
        if (!iscrizioneCorsoId) {{
          await appAlert('Seleziona il corso per cui vuoi generare le mensilita future.', {{ title: 'Quote future corso', badge: 'Pagamenti' }});
          return false;
        }}
        if (!fineCompetenza) {{
          await appAlert('Indica fino a quale mensilita vuoi arrivare.', {{ title: 'Quote future corso', badge: 'Pagamenti' }});
          return false;
        }}

        const previousSelected = multiAreaSelectedScadenze();
        const params = new URLSearchParams({{
          anno_lavoro: String(corsoSelect.dataset.workYear || ''),
          associato_id: associatoId,
          iscrizione_corso_id: iscrizioneCorsoId,
          fine_competenza: fineCompetenza,
        }});

        actionButton.disabled = true;
        if (feedback) {{
          feedback.dataset.preserveMessage = '1';
          feedback.textContent = 'Generazione quote in corso...';
        }}
        try {{
          const payload = await fetchJson(`/api/pagamenti-multi-area/quote-corso-future?${{params.toString()}}`);
          if (!payload || payload.ok !== true) {{
            throw new Error(payload && payload.error ? payload.error : 'Impossibile generare le quote future del corso.');
          }}
          const selectedValues = Array.from(new Set([...(previousSelected || []), ...((payload.selected_scadenze || []).map(String))]));
          replaceMultiAreaScadenzeOptions(payload.options || [], selectedValues);
          if (feedback) {{
            feedback.textContent = payload.message || 'Quote corso aggiornate e selezionate.';
          }}
          return false;
        }} catch (error) {{
          if (feedback) {{
            feedback.textContent = error && error.message ? error.message : 'Impossibile generare le quote future del corso.';
          }}
          await appAlert(
            feedback ? feedback.textContent : 'Impossibile generare le quote future del corso.',
            {{ title: 'Quote future corso', badge: 'Pagamenti', variant: 'danger' }}
          );
          return false;
        }} finally {{
          if (feedback) {{
            feedback.dataset.preserveMessage = '';
          }}
          actionButton.disabled = false;
          syncMultiAreaFutureCourseControls();
        }}
      }}

      const codiceFiscaleLookupTimers = new WeakMap();
      const comuneLookupTimers = new WeakMap();
      const capLookupTimers = new WeakMap();
      const cittaLookupTimers = new WeakMap();
      const codiceFiscaleGenerationTimers = new WeakMap();

      function assignFieldValue(form, fieldName, value) {{
        const field = form.querySelector(`[name="${{fieldName}}"]`);
        if (!field) {{
          return;
        }}
        field.value = value || '';
      }}

      function applyCodiceFiscalePayload(input, payload) {{
        const form = input.closest('form');
        if (!form || !payload) {{
          return;
        }}
        assignFieldValue(form, 'data_nascita', payload.data_nascita || '');
        assignFieldValue(form, 'sesso', payload.sesso || '');
        assignFieldValue(form, 'comune_nascita', payload.comune_nascita || '');
        assignFieldValue(form, 'provincia_nascita', payload.provincia_nascita || '');
      }}

      async function fetchJson(url, options = {{}}) {{
        const response = await fetch(url, {{
          ...options,
          headers: {{ 'Accept': 'application/json', ...(options.headers || {{}}) }},
          cache: options.cache || 'no-store',
        }});
        if (!response.ok) {{
          return null;
        }}
        return response.json();
      }}

      function associatoFormField(form, fieldName) {{
        return form ? form.querySelector(`[name="${{fieldName}}"]`) : null;
      }}

      async function lookupCodiceFiscale(input) {{
        const normalized = (input.value || '').toUpperCase().replace(/[^A-Z0-9]/g, '').slice(0, 16);
        input.value = normalized;
        if (normalized.length !== 16) {{
          return;
        }}
        try {{
          const payload = await fetchJson(`/api/codice-fiscale?value=${{encodeURIComponent(normalized)}}`);
          if (!payload || !payload.found) {{
            return;
          }}
          applyCodiceFiscalePayload(input, payload);
        }} catch (error) {{
          console.warn('Lookup codice fiscale non disponibile', error);
        }}
      }}

      function scheduleCodiceFiscaleLookup(input) {{
        const normalized = (input.value || '').toUpperCase().replace(/[^A-Z0-9]/g, '').slice(0, 16);
        input.value = normalized;
        const previousTimer = codiceFiscaleLookupTimers.get(input);
        if (previousTimer) {{
          window.clearTimeout(previousTimer);
        }}
        if (normalized.length !== 16) {{
          return;
        }}
        const timer = window.setTimeout(() => {{
          lookupCodiceFiscale(input);
        }}, 180);
        codiceFiscaleLookupTimers.set(input, timer);
      }}

      async function lookupComuneNascita(input) {{
        const form = input.closest('form');
        const comuneField = associatoFormField(form, 'comune_nascita');
        if (!form || !comuneField) {{
          return null;
        }}
        const comune = (comuneField.value || '').trim();
        if (!comune) {{
          return null;
        }}
        try {{
          const payload = await fetchJson(`/api/comuni?nome=${{encodeURIComponent(comune)}}`);
          if (!payload || !payload.found) {{
            return null;
          }}
          assignFieldValue(form, 'comune_nascita', payload.comune || comune);
          assignFieldValue(form, 'provincia_nascita', payload.provincia || '');
          return payload;
        }} catch (error) {{
          console.warn('Lookup comune di nascita non disponibile', error);
          return null;
        }}
      }}

      function scheduleComuneNascitaLookup(input) {{
        const previousTimer = comuneLookupTimers.get(input);
        if (previousTimer) {{
          window.clearTimeout(previousTimer);
        }}
        const timer = window.setTimeout(async () => {{
          await lookupComuneNascita(input);
          await generateCodiceFiscaleFromFields(input.closest('form'));
        }}, 180);
        comuneLookupTimers.set(input, timer);
      }}

      async function lookupCittaDetails(input) {{
        const form = input.closest('form');
        const cittaField = associatoFormField(form, 'citta');
        if (!form || !cittaField) {{
          return null;
        }}
        const citta = (cittaField.value || '').trim();
        if (!citta) {{
          return null;
        }}
        try {{
          const payload = await fetchJson(`/api/comuni?nome=${{encodeURIComponent(citta)}}`);
          if (!payload || !payload.found) {{
            return null;
          }}
          assignFieldValue(form, 'citta', payload.comune || citta);
          assignFieldValue(form, 'provincia', payload.provincia || '');
          if (payload.cap) {{
            assignFieldValue(form, 'cap', payload.cap);
          }}
          return payload;
        }} catch (error) {{
          console.warn('Lookup citta non disponibile', error);
          return null;
        }}
      }}

      function scheduleCittaLookup(input) {{
        const previousTimer = cittaLookupTimers.get(input);
        if (previousTimer) {{
          window.clearTimeout(previousTimer);
        }}
        const timer = window.setTimeout(() => {{
          lookupCittaDetails(input);
        }}, 180);
        cittaLookupTimers.set(input, timer);
      }}

      async function lookupCapDetails(input) {{
        const form = input.closest('form');
        const capField = associatoFormField(form, 'cap');
        if (!form || !capField) {{
          return null;
        }}
        const cap = (capField.value || '').replace(/\\D/g, '').slice(0, 5);
        capField.value = cap;
        if (cap.length !== 5) {{
          return null;
        }}
        try {{
          const payload = await fetchJson(`/api/cap?value=${{encodeURIComponent(cap)}}`);
          if (!payload || !payload.found) {{
            return null;
          }}
          assignFieldValue(form, 'citta', payload.citta || '');
          assignFieldValue(form, 'provincia', payload.provincia || '');
          return payload;
        }} catch (error) {{
          console.warn('Lookup CAP non disponibile', error);
          return null;
        }}
      }}

      function scheduleCapLookup(input) {{
        const previousTimer = capLookupTimers.get(input);
        if (previousTimer) {{
          window.clearTimeout(previousTimer);
        }}
        const timer = window.setTimeout(() => {{
          lookupCapDetails(input);
        }}, 180);
        capLookupTimers.set(input, timer);
      }}

      async function generateCodiceFiscaleFromFields(form) {{
        if (!form) {{
          return null;
        }}
        const nome = associatoFormField(form, 'nome');
        const cognome = associatoFormField(form, 'cognome');
        const dataNascita = associatoFormField(form, 'data_nascita');
        const sesso = associatoFormField(form, 'sesso');
        const comuneNascita = associatoFormField(form, 'comune_nascita');
        const codiceFiscale = associatoFormField(form, 'codice_fiscale');
        if (!nome || !cognome || !dataNascita || !sesso || !comuneNascita || !codiceFiscale) {{
          return null;
        }}

        if (!(nome.value || '').trim() || !(cognome.value || '').trim() || !(dataNascita.value || '').trim() || !(sesso.value || '').trim() || !(comuneNascita.value || '').trim()) {{
          return null;
        }}

        try {{
          const url = new URL('/api/codice-fiscale/calcola', window.location.origin);
          url.searchParams.set('nome', nome.value || '');
          url.searchParams.set('cognome', cognome.value || '');
          url.searchParams.set('data_nascita', dataNascita.value || '');
          url.searchParams.set('sesso', sesso.value || '');
          url.searchParams.set('comune_nascita', comuneNascita.value || '');
          const payload = await fetchJson(url.pathname + url.search);
          if (!payload || !payload.found) {{
            return null;
          }}
          assignFieldValue(form, 'codice_fiscale', payload.codice_fiscale || '');
          assignFieldValue(form, 'comune_nascita', payload.comune_nascita || comuneNascita.value || '');
          assignFieldValue(form, 'provincia_nascita', payload.provincia_nascita || '');
          return payload;
        }} catch (error) {{
          console.warn('Calcolo codice fiscale non disponibile', error);
          return null;
        }}
      }}

      function scheduleCodiceFiscaleGeneration(trigger) {{
        const form = trigger ? trigger.closest('form') : null;
        if (!form) {{
          return;
        }}
        const previousTimer = codiceFiscaleGenerationTimers.get(form);
        if (previousTimer) {{
          window.clearTimeout(previousTimer);
        }}
        const timer = window.setTimeout(() => {{
          generateCodiceFiscaleFromFields(form);
        }}, 180);
        codiceFiscaleGenerationTimers.set(form, timer);
      }}

      let whatsappShareWindow = null;
      let activeCourseReminder = null;

      function closeWhatsappShareWindow() {{
        if (!whatsappShareWindow) {{
          return;
        }}
        try {{
          if (!whatsappShareWindow.closed) {{
            whatsappShareWindow.close();
          }}
        }} catch (error) {{
          console.warn('Chiusura finestra WhatsApp non disponibile', error);
        }}
        whatsappShareWindow = null;
      }}

      function openCourseGenerationReminder() {{
        const reminder = document.querySelector('[data-course-generation-reminder="true"]');
        if (!reminder) {{
          return;
        }}
        const reminderKey = reminder.dataset.reminderKey || '';
        if (!reminderKey) {{
          return;
        }}
        const storageKey = `course-generation-reminder:${{reminderKey}}`;
        if (window.sessionStorage && window.sessionStorage.getItem(storageKey) === 'done') {{
          return;
        }}
        activeCourseReminder = {{ element: reminder, storageKey }};
        window.setTimeout(() => {{
          reminder.hidden = false;
          syncModalOpenState();
          const submitButton = reminder.querySelector('button[type=\"submit\"]');
          if (submitButton) {{
            submitButton.focus();
          }}
        }}, 180);
      }}

      function closeCourseGenerationReminder(markDone = true) {{
        const reminderElement = activeCourseReminder && activeCourseReminder.element
          ? activeCourseReminder.element
          : document.querySelector('[data-course-generation-reminder="true"]');
        if (!reminderElement) {{
          return;
        }}
        const reminderKey = activeCourseReminder && activeCourseReminder.storageKey
          ? activeCourseReminder.storageKey
          : (reminderElement.dataset.reminderKey ? `course-generation-reminder:${{reminderElement.dataset.reminderKey}}` : '');
        reminderElement.hidden = true;
        syncModalOpenState();
        if (markDone && window.sessionStorage && reminderKey) {{
          window.sessionStorage.setItem(reminderKey, 'done');
        }}
        activeCourseReminder = null;
      }}

      document.addEventListener('DOMContentLoaded', () => {{
        restoreSidebarState();
        document.querySelectorAll('table[data-summary-columns]').forEach((table) => updateVisibleTableTotals(table));
        initializeSortableTables();
        applyInitialColumnVisibility();
        restorePersistedTableStates();
        initializeDraggableTableColumns();
        bindMinorGuardianForms();

        document.querySelectorAll('.select-search-control').forEach((input) => {{
          const {{ select }} = getAutocompleteElements(input);
          if (!select) {{
            return;
          }}
          syncAutocompleteInput(input);
          select.addEventListener('change', () => syncAutocompleteInput(input));
        }});

        document.querySelectorAll('form').forEach((form) => {{
          form.addEventListener('submit', (event) => {{
            for (const input of form.querySelectorAll('.select-search-control')) {{
              if (!commitSelectAutocomplete(input, input.required)) {{
                event.preventDefault();
                input.reportValidity();
                input.focus();
                return;
              }}
            }}
          }});
        }});

        document.querySelectorAll('form[data-confirm-dialog=\"true\"]').forEach((form) => {{
          form.addEventListener('submit', async (event) => {{
            event.preventDefault();
            syncDeleteFormView(form);
            const confirmed = await appConfirm(form.dataset.confirmMessage || 'Confermare l\\'eliminazione del record selezionato?', {{
              title: form.dataset.confirmTitle || 'Conferma eliminazione',
              confirmLabel: form.dataset.confirmButton || 'Elimina',
              cancelLabel: 'Annulla',
              variant: 'danger',
              badge: 'Eliminazione',
            }});
            if (confirmed) {{
              form.submit();
            }}
          }});
        }});

        document.querySelectorAll('select.control:not(.searchable-select-source)').forEach((select) => {{
          installSelectTypeahead(select);
        }});

        document.querySelectorAll('select[data-residual-target]').forEach((select) => {{
          select.addEventListener('change', () => syncResidualInput(select));
          syncResidualInput(select);
        }});

        document.querySelectorAll('select[data-standard-target]').forEach((select) => {{
          select.addEventListener('change', () => syncStandardQuota(select));
          syncStandardQuota(select);
        }});

        document.querySelectorAll('select[data-amount-target]').forEach((select) => {{
          select.addEventListener('change', () => syncAmountProposal(select));
          syncAmountProposal(select);
        }});

        document.querySelectorAll('input[data-copy-value-target]').forEach((input) => {{
          input.addEventListener('input', () => syncMirroredValue(input));
          input.addEventListener('change', () => syncMirroredValue(input));
          syncMirroredValue(input);
        }});

        document.querySelectorAll('input[data-auto-managed-value]').forEach((input) => {{
          input.addEventListener('input', () => {{
            input.dataset.autoManagedValue = 'false';
          }});
        }});

        document.querySelectorAll('select[data-require-tesseramento=\"true\"]').forEach((select) => {{
          select.addEventListener('change', async () => {{
            await checkEnrollmentTesseramento(select);
          }});
        }});

        document.querySelectorAll('input[data-year-end-target]').forEach((input) => {{
          input.addEventListener('input', () => syncYearEndDate(input));
          input.addEventListener('blur', () => syncYearEndDate(input));
          syncYearEndDate(input);
        }});

        document.querySelectorAll('input[data-tesseramento-code-target]').forEach((input) => {{
          input.addEventListener('input', () => {{
            syncYearEndDate(input);
            syncTesseramentoCodePreview(input);
          }});
          input.addEventListener('blur', () => syncTesseramentoCodePreview(input));
          syncTesseramentoCodePreview(input);
        }});

        document.querySelectorAll('[data-month-end-target]').forEach((control) => {{
          const eventName = control.tagName === 'SELECT' ? 'change' : 'input';
          control.addEventListener(eventName, () => syncMonthEndDate(control));
          if (eventName !== 'change') {{
            control.addEventListener('blur', () => syncMonthEndDate(control));
          }}
          syncMonthEndDate(control);
        }});

        document.querySelectorAll('input[data-codice-fiscale]').forEach((input) => {{
          input.addEventListener('input', () => scheduleCodiceFiscaleLookup(input));
          input.addEventListener('blur', () => lookupCodiceFiscale(input));
          if ((input.value || '').trim().length === 16) {{
            lookupCodiceFiscale(input);
          }}
        }});

        document.querySelectorAll('form').forEach((form) => {{
          ['nome', 'cognome', 'data_nascita'].forEach((fieldName) => {{
            const field = associatoFormField(form, fieldName);
            if (!field) {{
              return;
            }}
            field.addEventListener('input', () => scheduleCodiceFiscaleGeneration(field));
            field.addEventListener('blur', () => generateCodiceFiscaleFromFields(form));
          }});

          const sessoField = associatoFormField(form, 'sesso');
          if (sessoField) {{
            sessoField.addEventListener('change', () => generateCodiceFiscaleFromFields(form));
          }}

          const comuneField = associatoFormField(form, 'comune_nascita');
          if (comuneField) {{
            comuneField.addEventListener('input', () => scheduleComuneNascitaLookup(comuneField));
            comuneField.addEventListener('blur', async () => {{
              await lookupComuneNascita(comuneField);
              await generateCodiceFiscaleFromFields(form);
            }});
          }}

          const capField = associatoFormField(form, 'cap');
          if (capField) {{
            capField.addEventListener('input', () => scheduleCapLookup(capField));
            capField.addEventListener('blur', () => lookupCapDetails(capField));
          }}

          const cittaField = associatoFormField(form, 'citta');
          if (cittaField) {{
            cittaField.addEventListener('input', () => scheduleCittaLookup(cittaField));
            cittaField.addEventListener('blur', () => lookupCittaDetails(cittaField));
          }}
        }});

        const associatoSelect = document.getElementById('monthly-payment-associato');
        const rateSelect = document.getElementById('monthly-rate-select');
        if (associatoSelect && rateSelect) {{
          associatoSelect.addEventListener('change', updateMonthlyRates);
          rateSelect.addEventListener('change', syncMonthlySelectedAmount);
          updateMonthlyRates();
        }}

        const multiAssociatoSelect = document.getElementById('multi-area-associato');
        const multiScadenzeSelect = document.getElementById('multi-area-scadenze');
        if (multiAssociatoSelect && multiScadenzeSelect) {{
          multiAssociatoSelect.addEventListener('change', () => {{
            syncMultiAreaFutureCourseControls();
            updateMultiAreaScadenze([]);
          }});
          multiScadenzeSelect.addEventListener('change', syncMultiAreaSelectedAmount);
          updateMultiAreaScadenze([]);
        }}

        const multiCourseSelect = document.getElementById('multi-area-course-enrollment');
        const multiCourseUntil = document.getElementById('multi-area-course-until');
        const multiCourseButton = document.getElementById('multi-area-course-generate');
        if (multiCourseSelect && multiCourseUntil) {{
          multiCourseSelect.addEventListener('change', syncMultiAreaFutureCourseControls);
          multiCourseUntil.addEventListener('change', syncMultiAreaFutureCourseControls);
          syncMultiAreaFutureCourseControls();
        }}
        if (multiCourseButton) {{
          multiCourseButton.addEventListener('click', generateFutureCourseMultiAreaScadenze);
        }}

        bindEnrollmentPaymentFlows();
        window.addEventListener('pagehide', persistVisibleTableStates);
        window.addEventListener('beforeunload', persistVisibleTableStates);
        document.addEventListener('visibilitychange', () => {{
          if (document.visibilityState === 'hidden') {{
            persistVisibleTableStates();
          }}
        }});

        window.addEventListener('focus', () => {{
          closeWhatsappShareWindow();
        }});

        document.querySelectorAll('[data-reminder-dismiss=\"true\"]').forEach((element) => {{
          element.addEventListener('click', () => closeCourseGenerationReminder(true));
        }});

        document.addEventListener('keydown', (event) => {{
          if (event.key === 'Escape' && activeAppDialog && activeAppDialog.dismissible) {{
            closeAppDialog(false);
            return;
          }}
          if (event.key === 'Escape' && activeCourseReminder) {{
            closeCourseGenerationReminder(true);
          }}
        }});

        const reminderForm = document.getElementById('course-generation-reminder-form');
        if (reminderForm) {{
          reminderForm.addEventListener('submit', () => {{
            if (activeCourseReminder && window.sessionStorage && activeCourseReminder.storageKey) {{
              window.sessionStorage.setItem(activeCourseReminder.storageKey, 'done');
            }}
          }});
        }}

        openCourseGenerationReminder();
      }});

      function shareReceipt(button) {{
        const channel = button.dataset.channel || '';
        if (channel === 'email' && button.dataset.mailto) {{
          window.location.href = button.dataset.mailto;
          return false;
        }}
        if (channel === 'whatsapp' && button.dataset.whatsapp) {{
          closeWhatsappShareWindow();
          whatsappShareWindow = window.open(button.dataset.whatsapp, 'oratorioCarloAcutisWhatsappShare');
          if (whatsappShareWindow) {{
            try {{
              whatsappShareWindow.focus();
            }} catch (error) {{
              console.warn('Focus finestra WhatsApp non disponibile', error);
            }}
          }}
          return false;
        }}
        return false;
      }}

      async function shareReport(button) {{
        const channel = button.dataset.channel || '';
        const reportKey = button.dataset.reportKey || '';
        const recipientSelectId = button.dataset.recipientSelectId || '';
        const recipientSelect = recipientSelectId ? document.getElementById(recipientSelectId) : null;
        const opensWhatsappWindow = channel === 'whatsapp' || channel === 'whatsapp-group';
        const needsRecipient = !!recipientSelectId && (channel === 'email' || channel === 'whatsapp');
        if (!channel || !reportKey) {{
          await appAlert('Invio report non disponibile.', {{
            title: 'Invio report',
            variant: 'danger',
            badge: 'Errore',
          }});
          return false;
        }}
        if (needsRecipient && (!recipientSelect || !recipientSelect.value)) {{
          await appAlert('Seleziona prima un destinatario valido.', {{
            title: 'Invio report',
            variant: 'warning',
            badge: 'Controllo dati',
          }});
          return false;
        }}

        let pendingWhatsappWindow = null;
        let popupBlocked = false;
        if (opensWhatsappWindow) {{
          closeWhatsappShareWindow();
          pendingWhatsappWindow = window.open('', 'oratorioCarloAcutisWhatsappShare');
          if (pendingWhatsappWindow) {{
            whatsappShareWindow = pendingWhatsappWindow;
            try {{
              pendingWhatsappWindow.document.write('<!doctype html><title>Preparazione WhatsApp</title><body style="font-family:Arial,sans-serif;padding:24px;">Preparazione messaggio WhatsApp...</body>');
            }} catch (error) {{
              console.warn('Finestra temporanea WhatsApp non disponibile', error);
            }}
          }}
        }}
        if (opensWhatsappWindow && !pendingWhatsappWindow) {{
          popupBlocked = true;
        }}

        const url = new URL('/api/report-share', window.location.origin);
        const params = new URLSearchParams(window.location.search);
        const toolbar = button.closest('.report-toolbar');
        const searchInput = toolbar ? toolbar.querySelector('.report-search input') : null;
        if (searchInput) {{
          if (searchInput.value) {{
            params.set('search', searchInput.value);
          }} else {{
            params.delete('search');
          }}
        }}
        params.set('report_key', reportKey);
        if (recipientSelect && recipientSelect.value) {{
          params.set('recipient_id', recipientSelect.value);
        }} else {{
          params.delete('recipient_id');
        }}
        params.set('channel', channel);
        url.search = params.toString();

        try {{
          const response = await fetch(url.toString(), {{
            headers: {{ 'Accept': 'application/json' }}
          }});
          const payload = await response.json();
          if (!response.ok || !payload || !payload.ok || !payload.url) {{
            if (pendingWhatsappWindow) {{
              try {{
                pendingWhatsappWindow.close();
              }} catch (closeError) {{
                console.warn('Chiusura finestra WhatsApp non disponibile', closeError);
              }}
            if (whatsappShareWindow === pendingWhatsappWindow) {{
                whatsappShareWindow = null;
              }}
            }}
            await appAlert((payload && payload.error) || 'Invio report non disponibile.', {{
              title: 'Invio report',
              variant: 'danger',
              badge: 'Errore',
            }});
            return false;
          }}

          if (payload.channel === 'email') {{
            window.location.href = payload.url;
            return false;
          }}

          if (payload.channel === 'whatsapp-group') {{
            if (payload.message) {{
              try {{
                if (navigator.clipboard && navigator.clipboard.writeText) {{
                  await navigator.clipboard.writeText(payload.message);
                }}
              }} catch (error) {{
                console.warn('Copia del messaggio gruppo WhatsApp non disponibile', error);
              }}
            }}
            whatsappShareWindow = pendingWhatsappWindow || whatsappShareWindow;
            if (whatsappShareWindow) {{
              whatsappShareWindow.location.href = payload.url;
            }} else if (popupBlocked) {{
              window.location.href = payload.url;
            }} else {{
              whatsappShareWindow = window.open(payload.url, 'oratorioCarloAcutisWhatsappShare');
            }}
            if (whatsappShareWindow) {{
              try {{
                whatsappShareWindow.focus();
              }} catch (error) {{
                console.warn('Focus finestra WhatsApp non disponibile', error);
              }}
            }}
            return false;
          }}

          whatsappShareWindow = pendingWhatsappWindow || whatsappShareWindow;
          if (whatsappShareWindow) {{
            whatsappShareWindow.location.href = payload.url;
          }} else if (popupBlocked) {{
            window.location.href = payload.url;
          }} else {{
            closeWhatsappShareWindow();
            whatsappShareWindow = window.open(payload.url, 'oratorioCarloAcutisWhatsappShare');
          }}
          if (whatsappShareWindow) {{
            try {{
              whatsappShareWindow.focus();
            }} catch (error) {{
              console.warn('Focus finestra WhatsApp non disponibile', error);
            }}
          }}
          return false;
        }} catch (error) {{
          if (pendingWhatsappWindow) {{
            try {{
              pendingWhatsappWindow.close();
            }} catch (closeError) {{
              console.warn('Chiusura finestra WhatsApp non disponibile', closeError);
            }}
            if (whatsappShareWindow === pendingWhatsappWindow) {{
              whatsappShareWindow = null;
            }}
          }}
          console.warn('Invio report non disponibile', error);
          await appAlert('Invio report non disponibile.', {{
            title: 'Invio report',
            variant: 'danger',
            badge: 'Errore',
          }});
          return false;
        }}
      }}
    </script>
  </body>
</html>"""
    return document.encode("utf-8")


def redirect(
    start_response,
    path: str,
    *,
    ok: str | None = None,
    err: str | None = None,
    extra_query: dict[str, str] | None = None,
    extra_headers: list[tuple[str, str]] | None = None,
):
    params = dict(extra_query or {})
    if ok:
        params["ok"] = ok
    if err:
        params["err"] = err
    parsed = urlsplit(path)
    existing_query = {
        key: ",".join(values)
        for key, values in parse_qs(parsed.query, keep_blank_values=True).items()
    }
    existing_query.update(params)
    query_string = urlencode(existing_query)
    location = urlunsplit(("", "", parsed.path, query_string, parsed.fragment))
    headers = [("Location", location)]
    if extra_headers:
        headers.extend(extra_headers)
    start_response("303 See Other", headers)
    return [b""]


def friendly_db_error(error: Exception) -> str:
    message = str(error)
    replacements = {
        "UNIQUE constraint failed: associati.codice_associato": "Esiste gia un associato con questo codice.",
        "UNIQUE constraint failed: associati.codice_fiscale": "Esiste gia un associato con questo codice fiscale.",
        "UNIQUE constraint failed: tesseramenti_annuali.associato_id, tesseramenti_annuali.anno_sociale": "Questo tesserato ha gia un tesseramento per l'anno indicato.",
        "UNIQUE constraint failed: tesseramenti_annuali.codice_tesseramento": "Esiste gia un tesseramento con questo codice.",
        "UNIQUE constraint failed: tesseramenti_annuali.anno_sociale, tesseramenti_annuali.numero_progressivo_anno": "Esiste gia un numero progressivo tesseramento per l'anno indicato.",
        "UNIQUE constraint failed: pagamenti_campi_estivi.iscrizione_campo_id": "Per questa iscrizione al Campo estivo esiste gia un pagamento una tantum.",
        "UNIQUE constraint failed: pagamenti_oratorio.iscrizione_oratorio_id": "Per questa iscrizione a Oratorio esiste gia un pagamento una tantum.",
        "UNIQUE constraint failed: pagamenti_eventi.iscrizione_evento_id": "Per questa iscrizione evento esiste gia un pagamento una tantum.",
        "UNIQUE constraint failed: iscrizioni_campi_estivi.associato_id, iscrizioni_campi_estivi.campo_estivo_id": "Questo tesserato risulta gia iscritto al Campo estivo dell'anno selezionato.",
        "UNIQUE constraint failed: iscrizioni_oratorio.associato_id, iscrizioni_oratorio.oratorio_id": "Questo tesserato risulta gia iscritto a Oratorio per l'anno selezionato.",
        "UNIQUE constraint failed: iscrizioni_eventi.associato_id, iscrizioni_eventi.evento_id": "Questo tesserato risulta gia iscritto all'evento selezionato.",
        "UNIQUE constraint failed: rate_corsi_mensili.iscrizione_corso_id, rate_corsi_mensili.anno, rate_corsi_mensili.mese": "Per questa iscrizione corso esiste gia una quota mensile per anno e mese indicati.",
        "UNIQUE constraint failed: corsi.codice_corso": "Esiste gia un corso con questo codice.",
        "UNIQUE constraint failed: eventi.codice_evento": "Esiste gia un evento con questo codice.",
        "UNIQUE constraint failed: campi_estivi.codice_campo": "Esiste gia un Campo estivo con questo codice.",
        "UNIQUE constraint failed: oratorio.codice_oratorio": "Esiste gia una scheda Oratorio con questo codice.",
        "UNIQUE constraint failed: tipologie_corsi.codice_tipologia": "Esiste gia una tipologia corso con questo codice.",
        "UNIQUE constraint failed: tipologie_corsi.nome": "Esiste gia una tipologia corso con questo nome.",
        "UNIQUE constraint failed: utenti_accesso.username": "Esiste gia un utente con questo username.",
        "FOREIGN KEY constraint failed": "Operazione non possibile: esistono dati collegati che devono essere gestiti prima.",
    }
    for needle, replacement in replacements.items():
        if needle in message:
            return replacement
    return message


def associati_options() -> list[sqlite3.Row]:
    rows = fetch_all(
        """
        SELECT
            associati.id,
            associati.codice_associato,
            nome,
            cognome,
            data_nascita,
            COALESCE(associati.telefono, '') AS telefono,
            COALESCE(associati.email, '') AS email
        FROM associati
        ORDER BY associati.cognome, associati.nome
        """
    )
    options = []
    for row in rows:
        full_name = plain_text(f"{row['cognome']} {row['nome']}")
        named_label = label_with_age(full_name, row["data_nascita"])
        contact_parts = []
        if row["telefono"]:
            contact_parts.append(row["telefono"])
        if row["email"]:
            contact_parts.append(row["email"])
        contact_tail = f" - {' - '.join(contact_parts)}" if contact_parts else ""
        options.append(
            {
                "id": row["id"],
                "label": f"{row['codice_associato']} - {named_label}",
                "autocomplete_label": f"{row['codice_associato']} - {named_label}{contact_tail}",
                "search_text": plain_text(
                    f"{row['codice_associato']} {row['cognome']} {row['nome']} {row['telefono']} {row['email']}"
                ).lower(),
            }
        )
    return options


def associati_options_for_enrollment(work_year: int) -> list[dict[str, object]]:
    rows = fetch_all(
        """
        SELECT
            a.id,
            t.codice_tesseramento,
            a.nome,
            a.cognome,
            a.data_nascita,
            COALESCE(a.telefono, '') AS telefono,
            COALESCE(a.email, '') AS email
        FROM tesseramenti_annuali t
        JOIN associati a ON a.id = t.associato_id
        WHERE t.anno_sociale = ?
        ORDER BY t.numero_progressivo_anno, a.cognome, a.nome
        """,
        (work_year,),
    )
    options: list[dict[str, object]] = []
    for row in rows:
        full_name = plain_text(f"{row['cognome']} {row['nome']}")
        named_label = label_with_age(full_name, row["data_nascita"])
        contact_parts = []
        if row["telefono"]:
            contact_parts.append(row["telefono"])
        if row["email"]:
            contact_parts.append(row["email"])
        contact_tail = f" - {' - '.join(contact_parts)}" if contact_parts else ""
        codice = plain_text(row["codice_tesseramento"] or "")
        options.append(
            {
                "id": row["id"],
                "label": f"{codice} - {named_label}",
                "autocomplete_label": f"{codice} - {named_label}{contact_tail}",
                "search_text": plain_text(
                    f"{codice} {row['cognome']} {row['nome']} {row['telefono']} {row['email']}"
                ).lower(),
                "has_tesseramento": "1",
            }
        )
    return options


def report_recipient_rows() -> list[dict[str, str]]:
    placeholders = ", ".join("?" for _ in REPORT_RECIPIENT_CARICHE)
    rows = fetch_all(
        f"""
        SELECT
            id,
            nome,
            cognome,
            data_nascita,
            carica,
            COALESCE(associati.email, '') AS email,
            COALESCE(associati.telefono, '') AS telefono
        FROM associati
        WHERE carica IN ({placeholders})
        ORDER BY associati.cognome, associati.nome
        """,
        tuple(REPORT_RECIPIENT_CARICHE),
    )
    role_order = {value: index for index, value in enumerate(REPORT_RECIPIENT_CARICHE)}
    recipients: list[dict[str, str]] = []
    for row in rows:
        full_name = plain_text(f"{row['cognome']} {row['nome']}")
        named_label = label_with_age(full_name, row["data_nascita"])
        contact_parts = []
        if row["email"]:
            contact_parts.append(row["email"])
        if row["telefono"]:
            contact_parts.append(row["telefono"])
        contact_tail = f" - {' - '.join(contact_parts)}" if contact_parts else ""
        recipients.append(
            {
                "id": str(row["id"]),
                "label": f"{row['carica']} - {named_label}",
                "autocomplete_label": f"{row['carica']} - {named_label}{contact_tail}",
                "search_text": plain_text(
                    f"{row['carica']} {row['cognome']} {row['nome']} {row['email']} {row['telefono']}"
                ).lower(),
                "email": row["email"],
                "whatsapp_phone": clean_phone_number(row["telefono"]),
                "carica": row["carica"],
            }
        )
    recipients.sort(key=lambda item: (role_order.get(item["carica"], 999), item["label"]))
    return recipients


def report_recipient_options(selected: str | None = None) -> str:
    return render_select_options(
        report_recipient_rows(),
        selected,
        blank_label="Seleziona destinatario...",
        data_keys=["search_text", "autocomplete_label", "email", "whatsapp_phone", "carica"],
    )


def activity_log_user_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT DISTINCT
            username AS id,
            username AS label
        FROM registro_attivita
        WHERE COALESCE(username, '') <> ''
        ORDER BY username
        """
    )


def activity_log_associato_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT DISTINCT
            associato_id AS id,
            trim(COALESCE(associato_codice, '') || CASE
                WHEN COALESCE(associato_codice, '') <> '' AND COALESCE(associato_nominativo, '') <> '' THEN ' - '
                ELSE ''
            END || COALESCE(associato_nominativo, '')) AS label,
            lower(trim(COALESCE(associato_codice, '') || ' ' || COALESCE(associato_nominativo, ''))) AS search_text,
            trim(COALESCE(associato_codice, '') || CASE
                WHEN COALESCE(associato_codice, '') <> '' AND COALESCE(associato_nominativo, '') <> '' THEN ' - '
                ELSE ''
            END || COALESCE(associato_nominativo, '')) AS autocomplete_label
        FROM registro_attivita
        WHERE associato_id IS NOT NULL
          AND COALESCE(associato_nominativo, '') <> ''
        ORDER BY label
        """
    )


def activity_log_attivita_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT DISTINCT
            descrizione_attivita AS id,
            descrizione_attivita AS label
        FROM registro_attivita
        WHERE COALESCE(descrizione_attivita, '') <> ''
        ORDER BY descrizione_attivita
        """
    )


def can_manage_cariche(current_user: dict[str, object] | None) -> bool:
    return bool(current_user and current_user.get("is_admin"))


def available_cariche(current_user: dict[str, object] | None) -> tuple[str, ...]:
    return CARICA_VALUES if can_manage_cariche(current_user) else ("Associato",)


def resolved_carica_value(
    form_data: dict[str, str],
    current_user: dict[str, object] | None,
    *,
    existing_value: str = "Associato",
) -> str:
    if not can_manage_cariche(current_user):
        return existing_value if existing_value in CARICA_VALUES else "Associato"
    selected = normalized(form_data, "carica", existing_value) or "Associato"
    if selected not in CARICA_VALUES:
        raise ValueError("La carica selezionata non e valida.")
    return selected


def associato_carica_field(selected: str, current_user: dict[str, object] | None) -> str:
    current_value = selected if selected in CARICA_VALUES else "Associato"
    if can_manage_cariche(current_user):
        return select_field("Carica", "carica", carica_options(current_value, current_user))
    return readonly_field("Carica", current_value)


def associato_has_tesseramento_for_year(associato_id: int | str, work_year: int) -> bool:
    try:
        associato_id_int = int(associato_id)
    except (TypeError, ValueError):
        return False
    row = fetch_one(
        """
        SELECT id
        FROM tesseramenti_annuali
        WHERE associato_id = ? AND anno_sociale = ?
        LIMIT 1
        """,
        (associato_id_int, work_year),
    )
    return row is not None


def redirect_missing_tesseramento(
    start_response,
    *,
    associato_id: int | str,
    work_year: int,
    area_label: str,
    query_params: dict[str, str] | None = None,
):
    extra_query = work_year_query(query_params or {"anno_lavoro": str(work_year)})
    extra_query["associato_id"] = str(associato_id)
    return redirect(
        start_response,
        "/maschere/tesseramenti",
        err=f"Per registrare l'iscrizione a {area_label} devi prima inserire il tesseramento dell'anno {work_year}.",
        extra_query=extra_query,
    )


def metodi_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT id, nome AS label
        FROM metodi_pagamento
        WHERE attivo = 1
        ORDER BY nome
        """
    )


def preferred_metodo_pagamento_id(metodi: list[sqlite3.Row]) -> str:
    for row in metodi:
        if plain_text(row["label"]).lower() == "contanti":
            return str(row["id"])
    return str(metodi[0]["id"]) if metodi else ""


def quote_predefinite_options(area: str) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            id,
            replace(printf('%.2f', importo), ',', '.') AS importo,
            descrizione || ' - ' || replace(printf('%.2f', importo), '.', ',') || ' EUR' AS label
        FROM quote_predefinite
        WHERE area = ? AND attiva = 1
        ORDER BY descrizione, id
        """,
        (area,),
    )


def quote_predefinite_rows(area: str) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT id, descrizione, importo, attiva, COALESCE(note, '') AS note
        FROM quote_predefinite
        WHERE area = ?
        ORDER BY descrizione, id
        """,
        (area,),
    )


def default_tesseramento_quota_importo() -> str:
    rows = quote_predefinite_options("tesseramenti")
    if not rows:
        return "0.00"
    return str(rows[0]["importo"] or "0.00")


def tipologie_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT id, codice_tipologia || ' - ' || nome AS label
        FROM tipologie_corsi
        WHERE attiva = 1
        ORDER BY nome
        """
    )


def corsi_options() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            c.id,
            c.nome AS label,
            replace(printf('%.2f', c.quota_mensile_standard), ',', '.') AS importo
        FROM corsi c
        WHERE c.attivo = 1
        ORDER BY c.nome
        """
    )


def pagamenti_multi_area_course_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            ic.id,
            ic.associato_id,
            substr(COALESCE(NULLIF(ic.data_inizio, ''), ic.data_iscrizione, ''), 1, 7) AS start_competenza,
            c.nome || ' - decorrenza ' || substr(COALESCE(NULLIF(ic.data_inizio, ''), ic.data_iscrizione, ''), 1, 7) AS label
        FROM iscrizioni_corsi ic
        JOIN corsi c ON c.id = ic.corso_id
        WHERE ic.stato_iscrizione = 'Attiva'
          AND {iscrizione_corso_year_relevance_sql('ic')}
        ORDER BY c.nome, start_competenza, ic.id
        """,
        iscrizione_corso_year_relevance_params(work_year),
    )


def corsi_report_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            c.id,
            c.nome AS label
        FROM corsi c
        WHERE c.attivo = 1
          AND {corso_year_relevance_sql('c')}
        ORDER BY c.nome
        """,
        corso_year_relevance_params(work_year),
    )


def iscrizioni_corsi_options() -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            ic.id,
            {associato_display_sql('a')} || ' - ' || c.nome AS label
        FROM iscrizioni_corsi ic
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        ORDER BY ic.id DESC
        """
    )


def rate_corsi_options() -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            r.id,
            {associato_display_sql('a')} || ' - ' || c.nome || ' ' || printf('%04d-%02d', r.anno, r.mese) AS label
        FROM rate_corsi_mensili r
        JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        ORDER BY r.anno DESC, r.mese DESC, a.cognome, a.nome
        """
    )


def rate_corsi_aperte_options(work_year: int, associato_id: str | None = None) -> list[sqlite3.Row]:
    params: list[object] = [work_year]
    where_clause = "WHERE v.anno = ? AND v.saldo_residuo > 0"
    if associato_id and associato_id.isdigit():
        where_clause += " AND v.associato_id = ?"
        params.append(int(associato_id))
    return fetch_all(
        f"""
        SELECT
            v.id,
            v.associato_id,
            replace(printf('%.2f', v.saldo_residuo), ',', '.') AS residuo,
            COALESCE(v.codice_tesseramento || ' - ', '') || v.associato || ' - ' || v.corso || ' - ' || v.competenza || ' - residuo ' ||
            replace(printf('%.2f', v.saldo_residuo), '.', ',') || ' EUR' AS label
        FROM v_rate_corsi_saldo v
        {where_clause}
        ORDER BY v.associato, v.corso, v.anno, v.mese
        """,
        tuple(params),
    )


def tesseramenti_options() -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            t.id,
            COALESCE(t.codice_tesseramento || ' - ', '') || {associato_display_sql('a')} || ' - anno ' || t.anno_sociale AS label
        FROM tesseramenti_annuali t
        JOIN associati a ON a.id = t.associato_id
        ORDER BY t.anno_sociale DESC, a.cognome, a.nome
        """
    )


def tesseramenti_aperti_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            v.id,
            replace(printf('%.2f', v.saldo_residuo), ',', '.') AS residuo,
            COALESCE(v.codice_tesseramento || ' - ', '') || v.associato || ' - anno ' || v.anno_sociale || ' - residuo ' ||
            replace(printf('%.2f', v.saldo_residuo), '.', ',') || ' EUR' AS label
        FROM v_tesseramenti_saldo v
        WHERE v.anno_sociale = ? AND v.saldo_residuo > 0
        ORDER BY v.associato
        """,
        (work_year,),
    )


def campi_estivi_options() -> list[sqlite3.Row]:
    return campi_estivi_options_for_year(None)


def campi_estivi_options_for_year(work_year: int | None) -> list[sqlite3.Row]:
    params: tuple[object, ...] = ()
    where_clause = "WHERE attivo = 1"
    if work_year is not None:
        where_clause += " AND anno = ?"
        params = (work_year,)
    return fetch_all(
        f"""
        SELECT
            id,
            anno,
            replace(printf('%.2f', quota_partecipazione_standard), ',', '.') AS quota_standard,
            'Campo estivo ' || anno AS label
        FROM campi_estivi
        {where_clause}
        ORDER BY anno DESC, id DESC
        """,
        params,
    )


def estate_record_for_year(work_year: int) -> sqlite3.Row | None:
    return fetch_one(
        """
        SELECT id, anno, quota_partecipazione_standard, attivo
        FROM campi_estivi
        WHERE anno = ?
        ORDER BY id
        LIMIT 1
        """,
        (work_year,),
    )


def ensure_estate_record(connection: sqlite3.Connection, work_year: int, standard_fee: str | None = None) -> int:
    existing_id = connection.execute(
        "SELECT id FROM campi_estivi WHERE anno = ? ORDER BY id LIMIT 1",
        (work_year,),
    ).fetchone()
    if existing_id:
        if standard_fee is not None:
            connection.execute(
                "UPDATE campi_estivi SET quota_partecipazione_standard = ?, attivo = 1 WHERE id = ?",
                (standard_fee, existing_id["id"]),
            )
        return int(existing_id["id"])

    progressive_number = reserve_progressive_number(connection, "campi_estivi")
    connection.execute(
        """
        INSERT INTO campi_estivi (
            numero_progressivo, codice_campo, nome, anno, data_inizio, data_fine, quota_partecipazione_standard, attivo
        ) VALUES (?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (
            progressive_number,
            format_progressive_code("campi_estivi", progressive_number),
            f"{ESTATE_LABEL} {work_year}",
            work_year,
            f"{work_year}-06-01",
            f"{work_year}-06-30",
            standard_fee or "0",
        ),
    )
    return int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])


def oratorio_options() -> list[sqlite3.Row]:
    return oratorio_options_for_year(None)


def oratorio_options_for_year(work_year: int | None) -> list[sqlite3.Row]:
    params: tuple[object, ...] = ()
    where_clause = "WHERE attivo = 1"
    if work_year is not None:
        where_clause += " AND anno = ?"
        params = (work_year,)
    return fetch_all(
        f"""
        SELECT
            id,
            anno,
            replace(printf('%.2f', quota_partecipazione_standard), ',', '.') AS quota_standard,
            '{ORATORIO_LABEL} ' || anno AS label
        FROM oratorio
        {where_clause}
        ORDER BY anno DESC, id DESC
        """,
        params,
    )


def ensure_oratorio_record(connection: sqlite3.Connection, work_year: int, standard_fee: str | None = None) -> int:
    existing_id = connection.execute(
        "SELECT id FROM oratorio WHERE anno = ? ORDER BY id LIMIT 1",
        (work_year,),
    ).fetchone()
    if existing_id:
        if standard_fee is not None:
            connection.execute(
                "UPDATE oratorio SET quota_partecipazione_standard = ?, attivo = 1 WHERE id = ?",
                (standard_fee, existing_id["id"]),
            )
        return int(existing_id["id"])

    progressive_number = reserve_progressive_number(connection, "oratorio")
    connection.execute(
        """
        INSERT INTO oratorio (
            numero_progressivo, codice_oratorio, nome, anno, data_inizio, data_fine, quota_partecipazione_standard, attivo
        ) VALUES (?, ?, ?, ?, ?, ?, ?, 1)
        """,
        (
            progressive_number,
            format_progressive_code("oratorio", progressive_number),
            f"{ORATORIO_LABEL} {work_year}",
            work_year,
            f"{work_year}-01-01",
            f"{work_year}-12-31",
            standard_fee or "0",
        ),
    )
    return int(connection.execute("SELECT last_insert_rowid()").fetchone()[0])


def iscrizioni_campi_options() -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            ice.id,
            COALESCE(t.codice_tesseramento || ' - ', '') || {associato_display_sql('a')} || ' - ' || ESTATE_LABEL || ' ' || ce.anno AS label
        FROM iscrizioni_campi_estivi ice
        JOIN associati a ON a.id = ice.associato_id
        JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
        WHERE ce.anno = ?
        ORDER BY ice.id DESC
        """,
        (work_year,),
    )


def iscrizioni_campi_aperte_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            v.id,
            replace(printf('%.2f', v.saldo_residuo), ',', '.') AS residuo,
            COALESCE(v.codice_tesseramento || ' - ', '') || v.associato || ' - anno ' || v.anno || ' - residuo ' ||
            replace(printf('%.2f', v.saldo_residuo), '.', ',') || ' EUR' AS label
        FROM v_campi_estivi_saldo v
        WHERE v.anno = ? AND v.saldo_residuo > 0
        ORDER BY v.associato
        """,
        (work_year,),
    )


def iscrizioni_oratorio_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            io.id,
            COALESCE(t.codice_tesseramento || ' - ', '') || {associato_display_sql('a')} || ' - ' || '{ORATORIO_LABEL} ' || o.anno AS label
        FROM iscrizioni_oratorio io
        JOIN associati a ON a.id = io.associato_id
        JOIN oratorio o ON o.id = io.oratorio_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
        WHERE o.anno = ?
        ORDER BY io.id DESC
        """,
        (work_year,),
    )


def iscrizioni_oratorio_aperte_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            v.id,
            replace(printf('%.2f', v.saldo_residuo), ',', '.') AS residuo,
            COALESCE(v.codice_tesseramento || ' - ', '') || v.associato || ' - anno ' || v.anno || ' - residuo ' ||
            replace(printf('%.2f', v.saldo_residuo), '.', ',') || ' EUR' AS label
        FROM v_oratorio_saldo v
        WHERE v.anno = ? AND v.saldo_residuo > 0
        ORDER BY v.associato
        """,
        (work_year,),
    )


def eventi_options(work_year: int | None = None) -> list[sqlite3.Row]:
    params: tuple[object, ...] = ()
    where_clause = "WHERE attivo = 1"
    if work_year is not None:
        where_clause += " AND substr(COALESCE(data_evento, ''), 1, 4) = ?"
        params = (str(work_year),)
    return fetch_all(
        f"""
        SELECT
            id,
            codice_evento || ' - ' || nome AS label,
            replace(printf('%.2f', quota_partecipazione_standard), ',', '.') AS importo
        FROM eventi
        {where_clause}
        ORDER BY data_evento DESC, nome
        """,
        params,
    )


def iscrizioni_eventi_options() -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        SELECT
            ie.id,
            COALESCE(t.codice_tesseramento || ' - ', '') || {associato_display_sql('a')} || ' - ' || e.nome AS label
        FROM iscrizioni_eventi ie
        JOIN associati a ON a.id = ie.associato_id
        JOIN eventi e ON e.id = ie.evento_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
        ORDER BY ie.id DESC
        """
    )


def iscrizioni_eventi_aperte_options(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            v.id,
            replace(printf('%.2f', v.saldo_residuo), ',', '.') AS residuo,
            COALESCE(v.codice_tesseramento || ' - ', '') || v.associato || ' - ' || v.evento || ' - residuo ' ||
            replace(printf('%.2f', v.saldo_residuo), '.', ',') || ' EUR' AS label
        FROM v_eventi_saldo v
        WHERE substr(v.data_evento, 1, 4) = ? AND v.saldo_residuo > 0
        ORDER BY v.evento, v.associato
        """,
        (str(work_year),),
    )


def scadenze_multi_area_options(work_year: int, associato_id: int | None = None) -> list[sqlite3.Row]:
    clauses = ["1 = 1"]
    params: list[object] = [work_year, work_year, work_year, work_year, str(work_year)]
    if associato_id is not None:
        clauses.append("associato_id = ?")
        params.append(int(associato_id))
    return fetch_all(
        f"""
        SELECT
            kind || ':' || source_id AS id,
            associato_id,
            area,
            riferimento,
            scadenza,
            replace(printf('%.2f', saldo_residuo), ',', '.') AS residuo,
            area || ' - ' || riferimento || ' - ' || COALESCE(scadenza, '') || ' - residuo ' ||
            replace(printf('%.2f', saldo_residuo), '.', ',') || ' EUR' AS label
        FROM (
            SELECT
                'tesseramenti' AS kind,
                id AS source_id,
                associato_id,
                'Tesseramento annuale' AS area,
                'Anno ' || anno_sociale AS riferimento,
                COALESCE(data_scadenza, data_tesseramento) AS scadenza,
                saldo_residuo
            FROM v_tesseramenti_saldo
            WHERE anno_sociale = ? AND saldo_residuo > 0

            UNION ALL

            SELECT
                'corsi-rate' AS kind,
                id AS source_id,
                associato_id,
                'Corso - quota mensile' AS area,
                corso || ' ' || competenza AS riferimento,
                COALESCE(data_scadenza, printf('%04d-%02d-01', anno, mese)) AS scadenza,
                saldo_residuo
            FROM v_rate_corsi_saldo
            WHERE anno = ? AND saldo_residuo > 0

            UNION ALL

            SELECT
                'campi-estivi' AS kind,
                id AS source_id,
                associato_id,
                'Campo estivo' AS area,
                campo_estivo AS riferimento,
                COALESCE(data_inizio, data_iscrizione) AS scadenza,
                saldo_residuo
            FROM v_campi_estivi_saldo
            WHERE anno = ? AND saldo_residuo > 0

            UNION ALL

            SELECT
                'oratorio' AS kind,
                id AS source_id,
                associato_id,
                'Oratorio' AS area,
                oratorio AS riferimento,
                COALESCE(data_inizio, data_iscrizione) AS scadenza,
                saldo_residuo
            FROM v_oratorio_saldo
            WHERE anno = ? AND saldo_residuo > 0

            UNION ALL

            SELECT
                'eventi' AS kind,
                id AS source_id,
                associato_id,
                'Evento' AS area,
                evento AS riferimento,
                data_evento AS scadenza,
                saldo_residuo
            FROM v_eventi_saldo
            WHERE substr(data_evento, 1, 4) = ? AND saldo_residuo > 0
        ) scadenze
        WHERE {' AND '.join(clauses)}
        ORDER BY scadenza, area, riferimento
        """,
        tuple(params),
    )


def scadenze_multi_area_payload_rows(work_year: int, associato_id: int | None = None) -> list[dict[str, str]]:
    return [
        {
            "id": str(row["id"]),
            "associato_id": str(row["associato_id"]),
            "label": plain_text(row["label"]),
            "residuo": str(row["residuo"]),
        }
        for row in scadenze_multi_area_options(work_year, associato_id)
    ]


def build_manual_scadenza_row(
    *,
    kind: str,
    source_id: int,
    associato_id: int,
    area: str,
    riferimento: str,
    scadenza: str,
    importo_dovuto: object,
    importo_pagato: object = "0.00",
    is_current: bool = False,
) -> dict[str, object]:
    dovuto = decimal_amount(importo_dovuto, minimum="0.00")
    pagato = decimal_amount(importo_pagato, minimum="0.00")
    residuo = (dovuto - pagato).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return {
        "kind": kind,
        "source_id": int(source_id),
        "associato_id": int(associato_id),
        "area": area,
        "riferimento": riferimento,
        "scadenza": scadenza,
        "importo_dovuto": format(dovuto, ".2f"),
        "importo_pagato": format(pagato, ".2f"),
        "saldo_residuo": format(residuo, ".2f"),
        "is_current": bool(is_current),
    }


def popup_payment_extra_tokens(form_data: dict[str, str]) -> list[str]:
    raw_value = normalized(form_data, "pagamento_scadenze_aggiuntive", "")
    tokens: list[str] = []
    seen: set[str] = set()
    for chunk in raw_value.split(","):
        token = chunk.strip()
        if not token or token in seen:
            continue
        parse_scadenza_multi_area_token(token)
        seen.add(token)
        tokens.append(token)
    return tokens


def generate_multi_area_future_course_payload(
    work_year: int,
    associato_id: int,
    iscrizione_corso_id: int,
    fine_competenza: str,
) -> dict[str, object]:
    end_year, end_month = parse_year_month_value(fine_competenza, "Competenza finale")
    if end_year != work_year:
        raise ValueError("Puoi generare quote future del corso solo entro l'anno di lavoro selezionato.")

    iscrizione = fetch_one(
        """
        SELECT
            ic.id,
            ic.associato_id,
            c.nome AS corso,
            COALESCE(NULLIF(ic.data_inizio, ''), ic.data_iscrizione, '') AS data_inizio_iscrizione,
            COALESCE(NULLIF(ic.data_fine, ''), '') AS data_fine_iscrizione,
            COALESCE(NULLIF(c.data_inizio, ''), '') AS data_inizio_corso,
            COALESCE(NULLIF(c.data_fine, ''), '') AS data_fine_corso
        FROM iscrizioni_corsi ic
        JOIN corsi c ON c.id = ic.corso_id
        WHERE ic.id = ?
          AND ic.associato_id = ?
          AND ic.stato_iscrizione = 'Attiva'
        """,
        (iscrizione_corso_id, associato_id),
    )
    if iscrizione is None:
        raise ValueError("Iscrizione corso non valida per il tesserato selezionato.")

    effective_start, _effective_end = compute_course_enrollment_effective_window(
        str(iscrizione["data_inizio_iscrizione"] or ""),
        str(iscrizione["data_fine_iscrizione"] or ""),
        str(iscrizione["data_inizio_corso"] or ""),
        str(iscrizione["data_fine_corso"] or ""),
    )
    start_year, start_month = parse_year_month_value(str(effective_start)[:7], "Decorrenza corso")
    effective_start_year, effective_start_month = (start_year, start_month)
    if (effective_start_year, effective_start_month) < (work_year, 1):
        effective_start_year, effective_start_month = work_year, 1
    if (end_year, end_month) < (effective_start_year, effective_start_month):
        raise ValueError("La competenza finale deve essere uguale o successiva alla decorrenza del corso.")

    with get_connection() as connection:
        ensure_course_rates_for_enrollment_range(
            connection,
            iscrizione_corso_id,
            effective_start_year,
            effective_start_month,
            end_year,
            end_month,
            note="Quota generata automaticamente da Pagamenti multi-area",
        )
        selected_rows = connection.execute(
            """
            SELECT
                'corsi-rate:' || r.id AS token
            FROM rate_corsi_mensili r
            LEFT JOIN (
                SELECT rata_corso_id, SUM(importo) AS totale_pagato
                FROM pagamenti_rate_corsi
                GROUP BY rata_corso_id
            ) pagamenti ON pagamenti.rata_corso_id = r.id
            WHERE r.iscrizione_corso_id = ?
              AND r.anno = ?
              AND (r.anno < ? OR (r.anno = ? AND r.mese <= ?))
              AND (r.importo_dovuto - COALESCE(pagamenti.totale_pagato, 0)) > 0
            ORDER BY r.anno, r.mese, r.id
            """,
            (iscrizione_corso_id, work_year, end_year, end_year, end_month),
        ).fetchall()
        connection.commit()

    selected_tokens = [str(row["token"]) for row in selected_rows]
    target_label = f"{month_label(end_month)} {end_year}"
    return {
        "ok": True,
        "message": (
            f"Quote del corso {plain_text(iscrizione['corso'])} aggiornate fino a {target_label}. "
            "Le mensilita aperte del corso sono state selezionate automaticamente."
        ),
        "selected_scadenze": selected_tokens,
        "options": scadenze_multi_area_payload_rows(work_year, associato_id),
    }


def parse_scadenza_multi_area_token(token: str) -> tuple[str, int]:
    if ":" not in token:
        raise ValueError("Selezione scadenza non valida.")
    kind, raw_id = token.split(":", 1)
    if kind not in {"tesseramenti", "corsi-rate", "campi-estivi", "oratorio", "eventi"} or not raw_id.isdigit():
        raise ValueError("Selezione scadenza non valida.")
    return kind, int(raw_id)


def load_multi_area_scadenze(tokens: list[str]) -> list[dict]:
    parsed = [parse_scadenza_multi_area_token(token) for token in tokens]
    grouped_ids: dict[str, list[int]] = {}
    for kind, source_id in parsed:
        grouped_ids.setdefault(kind, []).append(source_id)

    resolved: dict[str, dict] = {}

    def store_rows(rows: list[sqlite3.Row], kind: str) -> None:
        for row in rows:
            token = f"{kind}:{row['source_id']}"
            resolved[token] = dict(row)

    for kind, ids in grouped_ids.items():
        placeholders = ",".join("?" for _ in ids)
        if kind == "tesseramenti":
            rows = fetch_all(
                f"""
                SELECT
                    id AS source_id,
                    associato_id,
                    'Tesseramento annuale' AS area,
                    'Anno ' || anno_sociale AS riferimento,
                    COALESCE(data_scadenza, data_tesseramento) AS scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo
                FROM v_tesseramenti_saldo
                WHERE id IN ({placeholders})
                """,
                tuple(ids),
            )
        elif kind == "corsi-rate":
            rows = fetch_all(
                f"""
                SELECT
                    id AS source_id,
                    associato_id,
                    'Corso - quota mensile' AS area,
                    corso || ' ' || competenza AS riferimento,
                    COALESCE(data_scadenza, printf('%04d-%02d-01', anno, mese)) AS scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo
                FROM v_rate_corsi_saldo
                WHERE id IN ({placeholders})
                """,
                tuple(ids),
            )
        elif kind == "campi-estivi":
            rows = fetch_all(
                f"""
                SELECT
                    id AS source_id,
                    associato_id,
                    'Campo estivo' AS area,
                    campo_estivo AS riferimento,
                    COALESCE(data_inizio, data_iscrizione) AS scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo
                FROM v_campi_estivi_saldo
                WHERE id IN ({placeholders})
                """,
                tuple(ids),
            )
        elif kind == "oratorio":
            rows = fetch_all(
                f"""
                SELECT
                    id AS source_id,
                    associato_id,
                    'Oratorio' AS area,
                    oratorio AS riferimento,
                    COALESCE(data_inizio, data_iscrizione) AS scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo
                FROM v_oratorio_saldo
                WHERE id IN ({placeholders})
                """,
                tuple(ids),
            )
        else:
            rows = fetch_all(
                f"""
                SELECT
                    id AS source_id,
                    associato_id,
                    'Evento' AS area,
                    evento AS riferimento,
                    data_evento AS scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo
                FROM v_eventi_saldo
                WHERE id IN ({placeholders})
                """,
                tuple(ids),
            )
        store_rows(rows, kind)

    ordered_rows: list[dict] = []
    for token in tokens:
        row = resolved.get(token)
        if row is None:
            raise ValueError("Una o piu scadenze selezionate non sono piu disponibili.")
        row["kind"], row["source_id"] = parse_scadenza_multi_area_token(token)
        ordered_rows.append(row)
    return ordered_rows


def multi_area_receipts_rows(work_year: int) -> list[sqlite3.Row]:
    return fetch_all(
        f"""
        WITH pagamenti AS (
            SELECT
                pt.gruppo_ricevuta,
                pt.data_pagamento,
                pt.importo,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato
            FROM pagamenti_tesseramenti pt
            JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
            JOIN associati a ON a.id = t.associato_id
            WHERE pt.gruppo_ricevuta LIKE 'MGR-%' AND t.anno_sociale = ?

            UNION ALL

            SELECT
                prc.gruppo_ricevuta,
                prc.data_pagamento,
                prc.importo,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato
            FROM pagamenti_rate_corsi prc
            JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
            JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
            JOIN associati a ON a.id = ic.associato_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
            WHERE prc.gruppo_ricevuta LIKE 'MGR-%' AND r.anno = ?

            UNION ALL

            SELECT
                pce.gruppo_ricevuta,
                pce.data_pagamento,
                pce.importo,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato
            FROM pagamenti_campi_estivi pce
            JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
            JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
            JOIN associati a ON a.id = ice.associato_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
            WHERE pce.gruppo_ricevuta LIKE 'MGR-%' AND ce.anno = ?

            UNION ALL

            SELECT
                po.gruppo_ricevuta,
                po.data_pagamento,
                po.importo,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato
            FROM pagamenti_oratorio po
            JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
            JOIN oratorio o ON o.id = io.oratorio_id
            JOIN associati a ON a.id = io.associato_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
            WHERE po.gruppo_ricevuta LIKE 'MGR-%' AND o.anno = ?

            UNION ALL

            SELECT
                pe.gruppo_ricevuta,
                pe.data_pagamento,
                pe.importo,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato
            FROM pagamenti_eventi pe
            JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
            JOIN eventi e ON e.id = ie.evento_id
            JOIN associati a ON a.id = ie.associato_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
            WHERE pe.gruppo_ricevuta LIKE 'MGR-%' AND substr(COALESCE(e.data_evento, ''), 1, 4) = ?
        )
        SELECT
            gruppo_ricevuta,
            codice_tesseramento,
            associato,
            data_pagamento,
            COUNT(*) AS numero_scadenze,
            COALESCE(SUM(importo), 0) AS importo_totale
        FROM pagamenti
        GROUP BY gruppo_ricevuta, codice_tesseramento, associato, data_pagamento
        ORDER BY data_pagamento DESC, gruppo_ricevuta DESC
        """,
        (work_year, work_year, work_year, work_year, str(work_year)),
    )


def delete_multi_area_group(group_code: str) -> None:
    if not group_code:
        raise ValueError("Pagamento multi-area non valido.")
    with get_connection() as connection:
        connection.execute("DELETE FROM pagamenti_tesseramenti WHERE gruppo_ricevuta = ?", (group_code,))
        connection.execute("DELETE FROM pagamenti_rate_corsi WHERE gruppo_ricevuta = ?", (group_code,))
        connection.execute("DELETE FROM pagamenti_campi_estivi WHERE gruppo_ricevuta = ?", (group_code,))
        connection.execute("DELETE FROM pagamenti_oratorio WHERE gruppo_ricevuta = ?", (group_code,))
        connection.execute("DELETE FROM pagamenti_eventi WHERE gruppo_ricevuta = ?", (group_code,))
        connection.commit()


def register_grouped_multi_area_payment(
    connection: sqlite3.Connection,
    rows: list[dict[str, object]],
    *,
    importo_totale: Decimal,
    data_pagamento: str,
    metodo_pagamento_id: str,
    riferimento: str = "",
    note: str = "",
    group_code: str | None = None,
) -> str:
    if not rows:
        raise ValueError("Seleziona almeno una scadenza da saldare.")

    one_time_kinds = {"campi-estivi", "oratorio", "eventi"}
    selected_associati = {str(row["associato_id"]) for row in rows}
    if len(selected_associati) != 1:
        raise ValueError("Le scadenze selezionate devono appartenere tutte allo stesso associato.")

    if any(decimal_amount(row["saldo_residuo"]) <= Decimal("0.00") for row in rows):
        raise ValueError("Una o piu scadenze selezionate risultano gia saldate.")

    for row in rows:
        if row["kind"] in one_time_kinds and not row.get("is_current") and decimal_amount(row["importo_pagato"]) > Decimal("0.00"):
            raise ValueError(
                f"La scadenza {row['area']} - {row['riferimento']} ha gia un pagamento parziale e non puo essere inclusa nel pagamento multi-area."
            )

    totale_residuo = sum(decimal_amount(row["saldo_residuo"]) for row in rows)
    if importo_totale > totale_residuo:
        raise ValueError("L'importo inserito supera il residuo totale delle scadenze selezionate.")

    ordered_rows: list[tuple[int, int, dict[str, object]]] = []
    for index, row in enumerate(rows):
        if row.get("is_current"):
            priority = 0
        elif row["kind"] in one_time_kinds:
            priority = 2
        else:
            priority = 1
        ordered_rows.append((priority, index, row))
    ordered_rows.sort(key=lambda item: (item[0], item[1]))

    effective_group_code = group_code or generate_multi_area_group_code()
    remaining = importo_totale
    inserted = 0
    for _, _, row in ordered_rows:
        residuo = decimal_amount(row["saldo_residuo"])
        if remaining <= Decimal("0.00"):
            break

        if row["kind"] in one_time_kinds and not row.get("is_current") and remaining < residuo:
            raise ValueError(
                f"La scadenza {row['area']} - {row['riferimento']} puo essere inclusa solo a saldo completo nel pagamento multi-area."
            )

        importo_riga = min(remaining, residuo).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if importo_riga <= Decimal("0.00"):
            continue

        common_params = (
            data_pagamento,
            format(importo_riga, ".2f"),
            metodo_pagamento_id,
            riferimento,
            note,
            effective_group_code,
        )

        if row["kind"] == "tesseramenti":
            connection.execute(
                """
                INSERT INTO pagamenti_tesseramenti (
                    tesseramento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (row["source_id"], *common_params),
            )
        elif row["kind"] == "corsi-rate":
            connection.execute(
                """
                INSERT INTO pagamenti_rate_corsi (
                    rata_corso_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (row["source_id"], *common_params),
            )
        elif row["kind"] == "campi-estivi":
            connection.execute(
                """
                INSERT INTO pagamenti_campi_estivi (
                    iscrizione_campo_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (row["source_id"], *common_params),
            )
        elif row["kind"] == "oratorio":
            connection.execute(
                """
                INSERT INTO pagamenti_oratorio (
                    iscrizione_oratorio_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (row["source_id"], *common_params),
            )
        else:
            connection.execute(
                """
                INSERT INTO pagamenti_eventi (
                    iscrizione_evento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (row["source_id"], *common_params),
            )

        inserted += 1
        remaining = (remaining - importo_riga).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if inserted == 0:
        raise ValueError("Nessun pagamento multi-area e stato registrato.")

    return effective_group_code


def boolean_options(selected: str = "1") -> str:
    return render_static_options([("1", "Si"), ("0", "No")], selected, blank_label=None)


def associato_status_options(selected: str = "Attivo") -> str:
    return render_static_options(
        [("Attivo", "Attivo"), ("Sospeso", "Sospeso"), ("Dimesso", "Dimesso")],
        selected,
        blank_label=None,
    )


def liberatoria_video_options(selected: str = "Si") -> str:
    return render_static_options(
        [("Si", "Si"), ("No", "No")],
        selected,
        blank_label=None,
    )


def document_type_options(selected: str = DEFAULT_DOCUMENT_TYPE) -> str:
    return render_static_options(
        [(value, value) for value in DOCUMENT_TYPE_VALUES],
        selected,
        blank_label=None,
    )


def sesso_options(selected: str = "M") -> str:
    return render_static_options(
        [("M", "Maschio"), ("F", "Femmina")],
        selected,
        blank_label=None,
    )


def carica_options(selected: str = "Associato", current_user: dict[str, object] | None = None) -> str:
    return render_static_options(
        [(value, value) for value in available_cariche(current_user)],
        selected,
        blank_label=None,
    )


def corso_enrollment_status_options(selected: str = "Attiva") -> str:
    return render_static_options(
        [("Attiva", "Attiva"), ("Sospesa", "Sospesa"), ("Chiusa", "Chiusa")],
        selected,
        blank_label=None,
    )


def camp_enrollment_status_options(selected: str = "Iscritto") -> str:
    return render_static_options(
        [("Iscritto", "Iscritto"), ("Lista attesa", "Lista attesa"), ("Annullato", "Annullato")],
        selected,
        blank_label=None,
    )


def event_enrollment_status_options(selected: str = "Iscritto") -> str:
    return render_static_options(
        [("Iscritto", "Iscritto"), ("Confermato", "Confermato"), ("Annullato", "Annullato")],
        selected,
        blank_label=None,
    )


def edit_path(entity_key: str, record_id: object, query_params: dict[str, str]) -> str:
    return with_query(f"/modifica/{entity_key}/{record_id}", current_page_query(query_params))


def render_associato_edit_fields(row: sqlite3.Row, current_user: dict[str, object] | None) -> list[str]:
    minor_section_attrs = {"data-minor-only": "true"}
    minor_required_input = {"data-minor-required": "true"}
    return [
        input_field("Codice associato", "codice_associato", value=row["codice_associato"] or "", required_field=True),
        input_field("Cognome", "cognome", value=row["cognome"] or "", required_field=True),
        input_field("Nome", "nome", value=row["nome"] or "", required_field=True),
        input_field(
            "Codice fiscale",
            "codice_fiscale",
            value=row["codice_fiscale"] or "",
            attrs={
                "maxlength": "16",
                "data-codice-fiscale": "true",
                "autocomplete": "off",
                "autocapitalize": "characters",
                "spellcheck": "false",
            },
        ),
        input_field("Data nascita", "data_nascita", input_type="date", value=row["data_nascita"] or ""),
        select_field("Sesso", "sesso", sesso_options(row["sesso"] or "M")),
        input_field("Comune di nascita", "comune_nascita", value=row["comune_nascita"] or ""),
        input_field("Provincia di nascita", "provincia_nascita", value=row["provincia_nascita"] or "", attrs={"maxlength": "2", "autocapitalize": "characters"}),
        input_field("Cellulare", "telefono", value=row["telefono"] or ""),
        input_field("Email", "email", input_type="email", value=row["email"] or ""),
        input_field("Indirizzo", "indirizzo", value=row["indirizzo"] or "", wide=True),
        input_field("CAP", "cap", value=row["cap"] or "", attrs={"maxlength": "5", "inputmode": "numeric"}),
        input_field("CittÃ ", "citta", value=row["citta"] or ""),
        input_field("Provincia", "provincia", value=row["provincia"] or "", attrs={"maxlength": "2", "autocapitalize": "characters"}),
        input_field("Impiego", "impiego", value=row["impiego"] or ""),
        input_field(
            "Data prima iscrizione",
            "data_prima_iscrizione",
            input_type="date",
            value=row["data_prima_iscrizione"] or "",
            required_field=True,
        ),
        associato_carica_field(row["carica"] or "Associato", current_user),
        select_field(
            "Liberatoria Video",
            "liberatoria_video",
            liberatoria_video_options(row["liberatoria_video"] or "Si"),
        ),
        textarea_field(
            "Patologie, allergie, intolleranze alimentari ed eventuali terapie in corso",
            "patologie",
            value=row["patologie"] or "",
        ),
        textarea_field("Note", "note", value=row["note"] or ""),
        form_section_block(
            "Genitore/Tutore",
            "Compila questi dati solo se il tesserato e minorenne.",
            attrs=minor_section_attrs,
        ),
        input_field("Nome", "genitore_tutore_nome", value=row["genitore_tutore_nome"] or "", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
        input_field("Cognome", "genitore_tutore_cognome", value=row["genitore_tutore_cognome"] or "", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
        input_field("Cellulare", "genitore_tutore_cellulare", value=row["genitore_tutore_cellulare"] or "", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
        input_field("Email", "genitore_tutore_email", input_type="email", value=row["genitore_tutore_email"] or "", wrapper_attrs=minor_section_attrs),
        select_field(
            "Tipo documento",
            "genitore_tutore_tipo_documento",
            document_type_options(row["genitore_tutore_tipo_documento"] or DEFAULT_DOCUMENT_TYPE),
            wrapper_attrs=minor_section_attrs,
        ),
        input_field("Numero documento", "genitore_tutore_numero_documento", value=row["genitore_tutore_numero_documento"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Impiego", "genitore_tutore_impiego", value=row["genitore_tutore_impiego"] or "", wrapper_attrs=minor_section_attrs),
        form_section_block(
            "Altri autorizzati al prelievo all'uscita",
            "Indica i soggetti autorizzati al prelievo del minore.",
            attrs=minor_section_attrs,
        ),
        form_section_block("Altro genitore", attrs=minor_section_attrs),
        input_field("Nome", "prelievo_altro_genitore_nome", value=row["prelievo_altro_genitore_nome"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Cognome", "prelievo_altro_genitore_cognome", value=row["prelievo_altro_genitore_cognome"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Cellulare", "prelievo_altro_genitore_cellulare", value=row["prelievo_altro_genitore_cellulare"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Impiego", "prelievo_altro_genitore_impiego", value=row["prelievo_altro_genitore_impiego"] or "", wrapper_attrs=minor_section_attrs),
        select_field(
            "Tipo documento",
            "prelievo_altro_genitore_tipo_documento",
            document_type_options(row["prelievo_altro_genitore_tipo_documento"] or DEFAULT_DOCUMENT_TYPE),
            wrapper_attrs=minor_section_attrs,
        ),
        input_field("Numero documento", "prelievo_altro_genitore_numero_documento", value=row["prelievo_altro_genitore_numero_documento"] or "", wrapper_attrs=minor_section_attrs),
        form_section_block("Altra persona", attrs=minor_section_attrs),
        input_field("Nome", "prelievo_altra_persona_nome", value=row["prelievo_altra_persona_nome"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Cognome", "prelievo_altra_persona_cognome", value=row["prelievo_altra_persona_cognome"] or "", wrapper_attrs=minor_section_attrs),
        input_field("Cellulare", "prelievo_altra_persona_cellulare", value=row["prelievo_altra_persona_cellulare"] or "", wide=True, wrapper_attrs=minor_section_attrs),
        select_field(
            "Tipo documento",
            "prelievo_altra_persona_tipo_documento",
            document_type_options(row["prelievo_altra_persona_tipo_documento"] or DEFAULT_DOCUMENT_TYPE),
            wrapper_attrs=minor_section_attrs,
        ),
        input_field("Numero documento", "prelievo_altra_persona_numero_documento", value=row["prelievo_altra_persona_numero_documento"] or "", wrapper_attrs=minor_section_attrs),
    ]


CRUD_CONFIG = {
    "associati": {
        "page_title": "Modifica associato",
        "page_subtitle": "Aggiorna anagrafica e contatti del socio.",
        "return_path": "/maschere/associati",
        "return_query": {"vista": "dati"},
        "fetch_query": "SELECT * FROM associati WHERE id = ?",
        "update_sql": """
            UPDATE associati
            SET codice_associato = ?, nome = ?, cognome = ?, codice_fiscale = ?, data_nascita = ?,
                sesso = ?, comune_nascita = ?, provincia_nascita = ?, carica = ?, email = ?, telefono = ?, indirizzo = ?, cap = ?, citta = ?, provincia = ?, impiego = ?,
                data_prima_iscrizione = ?, liberatoria_video = ?, patologie = ?,
                genitore_tutore_cognome = ?, genitore_tutore_nome = ?, genitore_tutore_cellulare = ?, genitore_tutore_email = ?, genitore_tutore_impiego = ?, genitore_tutore_tipo_documento = ?, genitore_tutore_numero_documento = ?,
                prelievo_altro_genitore_nome = ?, prelievo_altro_genitore_cognome = ?, prelievo_altro_genitore_cellulare = ?, prelievo_altro_genitore_impiego = ?, prelievo_altro_genitore_tipo_documento = ?, prelievo_altro_genitore_numero_documento = ?,
                prelievo_altra_persona_nome = ?, prelievo_altra_persona_cognome = ?, prelievo_altra_persona_cellulare = ?, prelievo_altra_persona_tipo_documento = ?, prelievo_altra_persona_numero_documento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "codice_associato", "Codice associato"),
            required(form_data, "nome", "Nome"),
            required(form_data, "cognome", "Cognome"),
            optional(form_data, "codice_fiscale"),
            optional(form_data, "data_nascita"),
            normalized(form_data, "sesso", "M") or "M",
            optional(form_data, "comune_nascita"),
            optional(form_data, "provincia_nascita"),
            normalized(form_data, "carica", "Associato") or "Associato",
            optional(form_data, "email"),
            optional(form_data, "telefono"),
            optional(form_data, "indirizzo"),
            optional(form_data, "cap"),
            optional(form_data, "citta"),
            optional(form_data, "provincia"),
            optional(form_data, "impiego"),
            required(form_data, "data_prima_iscrizione", "Data prima iscrizione"),
            normalized(form_data, "liberatoria_video", "Si") or "Si",
            optional(form_data, "patologie"),
            optional(form_data, "genitore_tutore_cognome"),
            optional(form_data, "genitore_tutore_nome"),
            optional(form_data, "genitore_tutore_cellulare"),
            optional(form_data, "genitore_tutore_email"),
            optional(form_data, "genitore_tutore_impiego"),
            optional(form_data, "genitore_tutore_tipo_documento"),
            optional(form_data, "genitore_tutore_numero_documento"),
            optional(form_data, "prelievo_altro_genitore_nome"),
            optional(form_data, "prelievo_altro_genitore_cognome"),
            optional(form_data, "prelievo_altro_genitore_cellulare"),
            optional(form_data, "prelievo_altro_genitore_impiego"),
            optional(form_data, "prelievo_altro_genitore_tipo_documento"),
            optional(form_data, "prelievo_altro_genitore_numero_documento"),
            optional(form_data, "prelievo_altra_persona_nome"),
            optional(form_data, "prelievo_altra_persona_cognome"),
            optional(form_data, "prelievo_altra_persona_cellulare"),
            optional(form_data, "prelievo_altra_persona_tipo_documento"),
            optional(form_data, "prelievo_altra_persona_numero_documento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Codice associato", "codice_associato", value=row["codice_associato"] or "", required_field=True),
            lambda row: input_field("Nome", "nome", value=row["nome"] or "", required_field=True),
            lambda row: input_field("Cognome", "cognome", value=row["cognome"] or "", required_field=True),
            lambda row: input_field(
                "Codice fiscale",
                "codice_fiscale",
                value=row["codice_fiscale"] or "",
                attrs={
                    "maxlength": "16",
                    "data-codice-fiscale": "true",
                    "autocomplete": "off",
                    "autocapitalize": "characters",
                    "spellcheck": "false",
                },
            ),
            lambda row: input_field("Data nascita", "data_nascita", input_type="date", value=row["data_nascita"] or ""),
            lambda row: select_field("Sesso", "sesso", sesso_options(row["sesso"] or "M")),
            lambda row: inline_fields_row([
                input_field("Comune di nascita", "comune_nascita", value=row["comune_nascita"] or ""),
                input_field("Provincia di nascita", "provincia_nascita", value=row["provincia_nascita"] or "", attrs={"maxlength": "2", "autocapitalize": "characters"}),
            ]),
            lambda row: inline_fields_row([
                input_field("Cellulare", "telefono", value=row["telefono"] or ""),
                input_field("Email", "email", input_type="email", value=row["email"] or ""),
            ], row_class="cell-email-row"),
            lambda row: input_field("Indirizzo", "indirizzo", value=row["indirizzo"] or "", wide=True),
            lambda row: input_field("CAP", "cap", value=row["cap"] or "", attrs={"maxlength": "5", "inputmode": "numeric"}),
            lambda row: input_field("CittÃ ", "citta", value=row["citta"] or ""),
            lambda row: input_field("Provincia", "provincia", value=row["provincia"] or "", attrs={"maxlength": "2", "autocapitalize": "characters"}),
            lambda row: input_field("Impiego", "impiego", value=row["impiego"] or ""),
            lambda row: input_field(
                "Data prima iscrizione",
                "data_prima_iscrizione",
                input_type="date",
                value=row["data_prima_iscrizione"] or "",
                required_field=True,
            ),
            lambda row: select_field("Carica", "carica", carica_options(row["carica"] or "Associato")),
            lambda row: select_field(
                "Liberatoria Video",
                "liberatoria_video",
                liberatoria_video_options(row["liberatoria_video"] or "Si"),
            ),
            lambda row: textarea_field(
                "Patologie, allergie, intolleranze alimentari ed eventuali terapie in corso",
                "patologie",
                value=row["patologie"] or "",
            ),
            lambda row: form_section_block(
                "Genitore/Tutore",
                "Compila questi dati solo se il tesserato e minorenne.",
                attrs={"data-minor-only": "true"},
            ),
            lambda row: input_field("Nome", "genitore_tutore_nome", value=row["genitore_tutore_nome"] or "", attrs={"data-minor-required": "true"}, wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cognome", "genitore_tutore_cognome", value=row["genitore_tutore_cognome"] or "", attrs={"data-minor-required": "true"}, wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cellulare", "genitore_tutore_cellulare", value=row["genitore_tutore_cellulare"] or "", attrs={"data-minor-required": "true"}, wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Email", "genitore_tutore_email", input_type="email", value=row["genitore_tutore_email"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Impiego", "genitore_tutore_impiego", value=row["genitore_tutore_impiego"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: select_field("Tipo documento", "genitore_tutore_tipo_documento", document_type_options(row["genitore_tutore_tipo_documento"] or DEFAULT_DOCUMENT_TYPE), wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Numero documento", "genitore_tutore_numero_documento", value=row["genitore_tutore_numero_documento"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: form_section_block(
                "Altri autorizzati al prelievo all'uscita",
                "Indica i soggetti autorizzati al prelievo del minore.",
                attrs={"data-minor-only": "true"},
            ),
            lambda row: form_section_block("Altro genitore", attrs={"data-minor-only": "true"}),
            lambda row: input_field("Nome", "prelievo_altro_genitore_nome", value=row["prelievo_altro_genitore_nome"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cognome", "prelievo_altro_genitore_cognome", value=row["prelievo_altro_genitore_cognome"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cellulare", "prelievo_altro_genitore_cellulare", value=row["prelievo_altro_genitore_cellulare"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Impiego", "prelievo_altro_genitore_impiego", value=row["prelievo_altro_genitore_impiego"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: select_field("Tipo documento", "prelievo_altro_genitore_tipo_documento", document_type_options(row["prelievo_altro_genitore_tipo_documento"] or DEFAULT_DOCUMENT_TYPE), wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Numero documento", "prelievo_altro_genitore_numero_documento", value=row["prelievo_altro_genitore_numero_documento"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: form_section_block("Altra persona", attrs={"data-minor-only": "true"}),
            lambda row: input_field("Nome", "prelievo_altra_persona_nome", value=row["prelievo_altra_persona_nome"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cognome", "prelievo_altra_persona_cognome", value=row["prelievo_altra_persona_cognome"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Cellulare", "prelievo_altra_persona_cellulare", value=row["prelievo_altra_persona_cellulare"] or "", wrapper_attrs={"data-minor-only": "true"}),
            lambda row: select_field("Tipo documento", "prelievo_altra_persona_tipo_documento", document_type_options(row["prelievo_altra_persona_tipo_documento"] or DEFAULT_DOCUMENT_TYPE), wrapper_attrs={"data-minor-only": "true"}),
            lambda row: input_field("Numero documento", "prelievo_altra_persona_numero_documento", value=row["prelievo_altra_persona_numero_documento"] or "", wide=True, wrapper_attrs={"data-minor-only": "true"}),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM associati WHERE id = ?",
        "delete_prompt": "Eliminare questo associato? Verranno eliminati anche tesseramenti, iscrizioni e pagamenti collegati.",
        "success_update": "Associato aggiornato.",
        "success_delete": "Associato eliminato.",
    },
    "tesseramenti_annuali": {
        "page_title": "Modifica tesseramento",
        "page_subtitle": "Aggiorna anno sociale, importi e scadenze.",
        "return_path": "/maschere/tesseramenti",
        "fetch_query": "SELECT * FROM tesseramenti_annuali WHERE id = ?",
        "update_sql": """
            UPDATE tesseramenti_annuali
            SET associato_id = ?, anno_sociale = ?, numero_progressivo_anno = ?, codice_tesseramento = ?, data_tesseramento = ?, importo_dovuto = ?, data_scadenza = ?, note = ?
            WHERE id = ?
        """,
        "fields": [
            lambda row: readonly_field("Codice tesseramento", row["codice_tesseramento"] or "--", wide=True),
            lambda row: select_field(
                "Tesserato",
                "associato_id",
                render_associato_options(
                    associati_options(),
                    str(row["associato_id"]),
                ),
                required_field=True,
                wide=True,
                searchable=True,
            ),
            lambda row: input_field("Anno sociale", "anno_sociale", input_type="number", value=str(row["anno_sociale"] or ""), required_field=True, minimum="2000"),
            lambda row: input_field("Data tesseramento", "data_tesseramento", input_type="date", value=row["data_tesseramento"] or "", required_field=True),
            lambda row: input_field("Importo dovuto", "importo_dovuto", input_type="number", value=str(row["importo_dovuto"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: input_field("Data scadenza", "data_scadenza", input_type="date", value=row["data_scadenza"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM tesseramenti_annuali WHERE id = ?",
        "delete_prompt": "Eliminare questo tesseramento e i pagamenti collegati?",
        "success_update": "Tesseramento aggiornato.",
        "success_delete": "Tesseramento eliminato.",
    },
    "pagamenti_tesseramenti": {
        "page_title": "Modifica pagamento tesseramento",
        "page_subtitle": "Correggi data, importo o metodo del pagamento.",
        "return_path": "/maschere/tesseramenti",
        "fetch_query": "SELECT * FROM pagamenti_tesseramenti WHERE id = ?",
        "update_sql": """
            UPDATE pagamenti_tesseramenti
            SET tesseramento_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "tesseramento_id", "Tesseramento"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Tesseramento", "tesseramento_id", render_select_options(tesseramenti_options(), str(row["tesseramento_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_tesseramenti WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento del tesseramento?",
        "success_update": "Pagamento tesseramento aggiornato.",
        "success_delete": "Pagamento tesseramento eliminato.",
    },
    "tipologie_corsi": {
        "page_title": "Modifica tipologia corso",
        "page_subtitle": "Aggiorna codice, nome e stato della tipologia.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM tipologie_corsi WHERE id = ?",
        "update_sql": """
            UPDATE tipologie_corsi
            SET codice_tipologia = ?, nome = ?, descrizione = ?, attiva = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "codice_tipologia", "Codice tipologia"),
            required(form_data, "nome", "Nome tipologia"),
            optional(form_data, "descrizione"),
            normalized(form_data, "attiva", "1") or "1",
        ),
        "fields": [
            lambda row: input_field("Codice tipologia", "codice_tipologia", value=row["codice_tipologia"] or "", required_field=True),
            lambda row: input_field("Nome tipologia", "nome", value=row["nome"] or "", required_field=True),
            lambda row: select_field("Attiva", "attiva", boolean_options(str(row["attiva"] if row["attiva"] is not None else 1))),
            lambda row: textarea_field("Descrizione", "descrizione", value=row["descrizione"] or ""),
        ],
        "delete_sql": "DELETE FROM tipologie_corsi WHERE id = ?",
        "delete_prompt": "Eliminare questa tipologia corso?",
        "success_update": "Tipologia corso aggiornata.",
        "success_delete": "Tipologia corso eliminata.",
    },
    "quote_tesseramenti": {
        "page_title": "Modifica quota tesseramento",
        "page_subtitle": "Aggiorna descrizione, importo e stato della quota predefinita.",
        "return_path": "/maschere/tesseramenti",
        "return_query": {"vista": "dati"},
        "fetch_query": "SELECT * FROM quote_predefinite WHERE id = ? AND area = 'tesseramenti'",
        "update_sql": """
            UPDATE quote_predefinite
            SET descrizione = ?, importo = ?, attiva = ?, note = ?
            WHERE id = ? AND area = 'tesseramenti'
        """,
        "build_params": lambda form_data: (
            required(form_data, "descrizione", "Descrizione"),
            required(form_data, "importo", "Importo"),
            normalized(form_data, "attiva", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Descrizione", "descrizione", value=row["descrizione"] or "", required_field=True, wide=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Attiva", "attiva", boolean_options(str(row["attiva"] if row["attiva"] is not None else 1))),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM quote_predefinite WHERE id = ? AND area = 'tesseramenti'",
        "delete_prompt": "Eliminare questa quota predefinita di tesseramento?",
        "success_update": "Quota tesseramento aggiornata.",
        "success_delete": "Quota tesseramento eliminata.",
    },
    "quote_campi_estivi": {
        "page_title": "Modifica quota Campo estivo",
        "page_subtitle": "Aggiorna descrizione, importo e stato della quota predefinita.",
        "return_path": "/maschere/campi-estivi",
        "return_query": {"vista": "dati"},
        "fetch_query": "SELECT * FROM quote_predefinite WHERE id = ? AND area = 'campi-estivi'",
        "update_sql": """
            UPDATE quote_predefinite
            SET descrizione = ?, importo = ?, attiva = ?, note = ?
            WHERE id = ? AND area = 'campi-estivi'
        """,
        "build_params": lambda form_data: (
            required(form_data, "descrizione", "Descrizione"),
            required(form_data, "importo", "Importo"),
            normalized(form_data, "attiva", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Descrizione", "descrizione", value=row["descrizione"] or "", required_field=True, wide=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Attiva", "attiva", boolean_options(str(row["attiva"] if row["attiva"] is not None else 1))),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM quote_predefinite WHERE id = ? AND area = 'campi-estivi'",
        "delete_prompt": "Eliminare questa quota predefinita di Campo estivo?",
        "success_update": "Quota Campo estivo aggiornata.",
        "success_delete": "Quota Campo estivo eliminata.",
    },
    "quote_oratorio": {
        "page_title": "Modifica quota Oratorio",
        "page_subtitle": "Aggiorna descrizione, importo e stato della quota predefinita.",
        "return_path": "/maschere/oratorio",
        "return_query": {"vista": "dati"},
        "fetch_query": "SELECT * FROM quote_predefinite WHERE id = ? AND area = 'oratorio'",
        "update_sql": """
            UPDATE quote_predefinite
            SET descrizione = ?, importo = ?, attiva = ?, note = ?
            WHERE id = ? AND area = 'oratorio'
        """,
        "build_params": lambda form_data: (
            required(form_data, "descrizione", "Descrizione"),
            required(form_data, "importo", "Importo"),
            normalized(form_data, "attiva", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Descrizione", "descrizione", value=row["descrizione"] or "", required_field=True, wide=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Attiva", "attiva", boolean_options(str(row["attiva"] if row["attiva"] is not None else 1))),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM quote_predefinite WHERE id = ? AND area = 'oratorio'",
        "delete_prompt": "Eliminare questa quota predefinita di Oratorio?",
        "success_update": "Quota Oratorio aggiornata.",
        "success_delete": "Quota Oratorio eliminata.",
    },
    "corsi": {
        "page_title": "Modifica corso",
        "page_subtitle": "Aggiorna anagrafica corso, quota mensile e organizzazione.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM corsi WHERE id = ?",
        "update_sql": """
            UPDATE corsi
            SET codice_corso = ?, nome = ?, descrizione = ?, quota_iscrizione_standard = 0,
                quota_mensile_standard = ?, data_inizio = ?, data_fine = ?, sede = ?, giorno_settimana = ?, orario = ?, attivo = ?, note = ?,
                tipologia_corso_id = NULL
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "codice_corso", "Codice corso"),
            required(form_data, "nome", "Nome corso"),
            optional(form_data, "descrizione"),
            normalized(form_data, "quota_mensile_standard", "0"),
            required(form_data, "data_inizio", "Data inizio"),
            required(form_data, "data_fine", "Data fine"),
            optional(form_data, "sede"),
            optional(form_data, "giorno_settimana"),
            optional(form_data, "orario"),
            normalized(form_data, "attivo", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Codice corso", "codice_corso", value=row["codice_corso"] or "", required_field=True),
            lambda row: input_field("Nome corso", "nome", value=row["nome"] or "", required_field=True),
            lambda row: input_field("Quota mensile standard", "quota_mensile_standard", input_type="number", value=str(row["quota_mensile_standard"] or ""), step="0.01", minimum="0"),
            lambda row: input_field("Data inizio", "data_inizio", input_type="date", value=row["data_inizio"] or "", required_field=True),
            lambda row: input_field("Data fine", "data_fine", input_type="date", value=row["data_fine"] or "", required_field=True),
            lambda row: input_field("Sede", "sede", value=row["sede"] or ""),
            lambda row: input_field("Giorno settimana", "giorno_settimana", value=row["giorno_settimana"] or ""),
            lambda row: input_field("Orario", "orario", value=row["orario"] or ""),
            lambda row: select_field("Attivo", "attivo", boolean_options(str(row["attivo"] if row["attivo"] is not None else 1))),
            lambda row: textarea_field("Descrizione", "descrizione", value=row["descrizione"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM corsi WHERE id = ?",
        "delete_prompt": "Eliminare questo corso? Le iscrizioni collegate devono essere eliminate prima.",
        "success_update": "Corso aggiornato.",
        "success_delete": "Corso eliminato.",
    },
    "iscrizioni_corsi": {
        "page_title": "Modifica iscrizione corso",
        "page_subtitle": "Aggiorna partecipazione e quota mensile del corso.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM iscrizioni_corsi WHERE id = ?",
        "update_sql": """
            UPDATE iscrizioni_corsi
            SET associato_id = ?, corso_id = ?, data_iscrizione = ?, data_inizio = ?, data_fine = ?,
                quota_mensile = ?, stato_iscrizione = ?, note = ?, quota_iscrizione = 0
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "associato_id", "Associato"),
            required(form_data, "corso_id", "Corso"),
            required(form_data, "data_iscrizione", "Data iscrizione"),
            optional(form_data, "data_inizio"),
            optional(form_data, "data_fine"),
            required(form_data, "quota_mensile", "Quota mensile"),
            normalized(form_data, "stato_iscrizione", "Attiva") or "Attiva",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field(
                "Tesserato",
                "associato_id",
                render_associato_options(
                    associati_options_for_enrollment(int(str(row["data_iscrizione"] or date.today().isoformat())[:4])),
                    str(row["associato_id"]),
                    extra_data_keys=["has_tesseramento"],
                ),
                required_field=True,
                wide=True,
                searchable=True,
            ),
            lambda row: select_field("Corso", "corso_id", render_select_options(corsi_options(), str(row["corso_id"])), required_field=True, wide=True),
            lambda row: input_field(
                "Data iscrizione",
                "data_iscrizione",
                input_type="date",
                value=row["data_iscrizione"] or "",
                required_field=True,
                element_id="modifica-corso-data-iscrizione",
                attrs={"data-copy-value-target": "modifica-corso-data-inizio"},
            ),
            lambda row: input_field(
                "Data inizio",
                "data_inizio",
                input_type="date",
                value=row["data_inizio"] or row["data_iscrizione"] or "",
                element_id="modifica-corso-data-inizio",
                attrs={"data-auto-managed-value": "true"},
            ),
            lambda row: input_field("Data fine", "data_fine", input_type="date", value=row["data_fine"] or ""),
            lambda row: input_field("Quota mensile", "quota_mensile", input_type="number", value=str(row["quota_mensile"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Stato iscrizione", "stato_iscrizione", corso_enrollment_status_options(row["stato_iscrizione"] or "Attiva")),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM iscrizioni_corsi WHERE id = ?",
        "delete_prompt": "Eliminare questa iscrizione corso e tutti i pagamenti collegati?",
        "success_update": "Iscrizione corso aggiornata.",
        "success_delete": "Iscrizione corso eliminata.",
    },
    "pagamenti_iscrizioni_corsi": {
        "page_title": "Modifica pagamento iscrizione corso",
        "page_subtitle": "Aggiorna il versamento iniziale del corso.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM pagamenti_iscrizioni_corsi WHERE id = ?",
        "update_sql": """
            UPDATE pagamenti_iscrizioni_corsi
            SET iscrizione_corso_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "iscrizione_corso_id", "Iscrizione corso"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Iscrizione corso", "iscrizione_corso_id", render_select_options(iscrizioni_corsi_options(), str(row["iscrizione_corso_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_iscrizioni_corsi WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento di iscrizione corso?",
        "success_update": "Pagamento iscrizione corso aggiornato.",
        "success_delete": "Pagamento iscrizione corso eliminato.",
    },
    "rate_corsi_mensili": {
        "page_title": "Modifica quota mensile corso",
        "page_subtitle": "Aggiorna competenza, scadenza e importo della rata.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM rate_corsi_mensili WHERE id = ?",
        "update_sql": """
            UPDATE rate_corsi_mensili
            SET iscrizione_corso_id = ?, anno = ?, mese = ?, importo_dovuto = ?, data_scadenza = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "iscrizione_corso_id", "Iscrizione corso"),
            required(form_data, "anno", "Anno"),
            required(form_data, "mese", "Mese"),
            required(form_data, "importo_dovuto", "Importo dovuto"),
            optional(form_data, "data_scadenza"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Iscrizione corso", "iscrizione_corso_id", render_select_options(iscrizioni_corsi_options(), str(row["iscrizione_corso_id"])), required_field=True, wide=True),
            lambda row: input_field("Anno", "anno", input_type="number", value=str(row["anno"] or ""), required_field=True, minimum="2000"),
            lambda row: input_field("Mese", "mese", input_type="number", value=str(row["mese"] or ""), required_field=True, minimum="1"),
            lambda row: input_field("Importo dovuto", "importo_dovuto", input_type="number", value=str(row["importo_dovuto"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: input_field("Data scadenza", "data_scadenza", input_type="date", value=row["data_scadenza"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM rate_corsi_mensili WHERE id = ?",
        "delete_prompt": "Eliminare questa quota mensile e i pagamenti collegati?",
        "success_update": "Quota mensile aggiornata.",
        "success_delete": "Quota mensile eliminata.",
    },
    "pagamenti_rate_corsi": {
        "page_title": "Modifica pagamento quota mensile",
        "page_subtitle": "Aggiorna il pagamento della rata mensile.",
        "return_path": "/maschere/corsi",
        "fetch_query": "SELECT * FROM pagamenti_rate_corsi WHERE id = ?",
        "update_sql": """
            UPDATE pagamenti_rate_corsi
            SET rata_corso_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "rata_corso_id", "Quota mensile"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Quota mensile", "rata_corso_id", render_select_options(rate_corsi_options(), str(row["rata_corso_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_rate_corsi WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento della quota mensile?",
        "success_update": "Pagamento quota mensile aggiornato.",
        "success_delete": "Pagamento quota mensile eliminato.",
    },
    "campi_estivi": {
        "page_title": "Modifica Campo estivo",
        "page_subtitle": "Aggiorna i dati interni del Campo estivo.",
        "return_path": "/maschere/campi-estivi",
        "fetch_query": "SELECT * FROM campi_estivi WHERE id = ?",
        "update_sql": """
            UPDATE campi_estivi
            SET codice_campo = ?, nome = ?, anno = ?, data_inizio = ?, data_fine = ?, sede = ?,
                quota_partecipazione_standard = ?, posti_massimi = ?, descrizione = ?, attivo = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "codice_campo", "Codice campo"),
            required(form_data, "nome", "Nome"),
            required(form_data, "anno", "Anno"),
            required(form_data, "data_inizio", "Data inizio"),
            required(form_data, "data_fine", "Data fine"),
            optional(form_data, "sede"),
            normalized(form_data, "quota_partecipazione_standard", "0"),
            optional(form_data, "posti_massimi"),
            optional(form_data, "descrizione"),
            normalized(form_data, "attivo", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: readonly_field("Numero progressivo", str(row["numero_progressivo"] or "")),
            lambda row: input_field("Codice campo", "codice_campo", value=row["codice_campo"] or "", required_field=True),
            lambda row: input_field("Nome", "nome", value=row["nome"] or "", required_field=True),
            lambda row: input_field("Anno", "anno", input_type="number", value=str(row["anno"] or ""), required_field=True, minimum="2000"),
            lambda row: input_field("Data inizio", "data_inizio", input_type="date", value=row["data_inizio"] or "", required_field=True),
            lambda row: input_field("Data fine", "data_fine", input_type="date", value=row["data_fine"] or "", required_field=True),
            lambda row: input_field("Sede", "sede", value=row["sede"] or ""),
            lambda row: input_field("Quota standard", "quota_partecipazione_standard", input_type="number", value=str(row["quota_partecipazione_standard"] or ""), step="0.01", minimum="0"),
            lambda row: input_field("Posti massimi", "posti_massimi", input_type="number", value=str(row["posti_massimi"] or "") if row["posti_massimi"] is not None else "", minimum="1"),
            lambda row: select_field("Attivo", "attivo", boolean_options(str(row["attivo"] if row["attivo"] is not None else 1))),
            lambda row: textarea_field("Descrizione", "descrizione", value=row["descrizione"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM campi_estivi WHERE id = ?",
        "delete_prompt": "Eliminare questo Campo estivo e tutti i dati collegati?",
        "success_update": "Campo estivo aggiornato.",
        "success_delete": "Campo estivo eliminato.",
    },
    "iscrizioni_campi_estivi": {
        "page_title": "Modifica iscrizione Campo estivo",
        "page_subtitle": "Aggiorna partecipante, stato e quota di partecipazione.",
        "return_path": "/maschere/campi-estivi",
        "fetch_query": """
            SELECT ice.*, ce.anno AS anno_estate
            FROM iscrizioni_campi_estivi ice
            JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
            WHERE ice.id = ?
        """,
        "update_sql": """
            UPDATE iscrizioni_campi_estivi
            SET associato_id = ?, data_iscrizione = ?, quota_partecipazione = ?, stato_iscrizione = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "associato_id", "Associato"),
            required(form_data, "data_iscrizione", "Data iscrizione"),
            required(form_data, "quota_partecipazione", "Quota partecipazione"),
            normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field(
                "Tesserato",
                "associato_id",
                render_associato_options(
                    associati_options_for_enrollment(int(str(row["data_iscrizione"] or date.today().isoformat())[:4])),
                    str(row["associato_id"]),
                    extra_data_keys=["has_tesseramento"],
                ),
                required_field=True,
                wide=True,
                searchable=True,
            ),
            lambda row: readonly_field("Anno", str(row["anno_estate"] or "")),
            lambda row: input_field("Data iscrizione", "data_iscrizione", input_type="date", value=row["data_iscrizione"] or "", required_field=True),
            lambda row: input_field("Quota partecipazione", "quota_partecipazione", input_type="number", value=str(row["quota_partecipazione"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Stato iscrizione", "stato_iscrizione", camp_enrollment_status_options(row["stato_iscrizione"] or "Iscritto")),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM iscrizioni_campi_estivi WHERE id = ?",
        "delete_prompt": "Eliminare questa iscrizione al Campo estivo e il pagamento collegato?",
        "success_update": "Iscrizione Campo estivo aggiornata.",
        "success_delete": "Iscrizione Campo estivo eliminata.",
    },
    "pagamenti_campi_estivi": {
        "page_title": "Modifica pagamento Campo estivo",
        "page_subtitle": "Aggiorna il pagamento una tantum del Campo estivo.",
        "return_path": "/maschere/campi-estivi",
        "fetch_query": "SELECT * FROM pagamenti_campi_estivi WHERE id = ?",
        "update_sql": """
            UPDATE pagamenti_campi_estivi
            SET iscrizione_campo_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "iscrizione_campo_id", "Iscrizione Campo estivo"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Iscrizione Campo estivo", "iscrizione_campo_id", render_select_options(iscrizioni_campi_options(), str(row["iscrizione_campo_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_campi_estivi WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento del Campo estivo?",
        "success_update": "Pagamento Campo estivo aggiornato.",
        "success_delete": "Pagamento Campo estivo eliminato.",
    },
    "iscrizioni_oratorio": {
        "page_title": "Modifica iscrizione Oratorio",
        "page_subtitle": "Aggiorna partecipante, stato e quota di partecipazione.",
        "return_path": "/maschere/oratorio",
        "fetch_query": """
            SELECT io.*, o.anno AS anno_oratorio
            FROM iscrizioni_oratorio io
            JOIN oratorio o ON o.id = io.oratorio_id
            WHERE io.id = ?
        """,
        "update_sql": """
            UPDATE iscrizioni_oratorio
            SET associato_id = ?, data_iscrizione = ?, quota_partecipazione = ?, stato_iscrizione = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "associato_id", "Associato"),
            required(form_data, "data_iscrizione", "Data iscrizione"),
            required(form_data, "quota_partecipazione", "Quota partecipazione"),
            normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field(
                "Tesserato",
                "associato_id",
                render_associato_options(
                    associati_options_for_enrollment(int(row["anno_oratorio"] or date.today().year)),
                    str(row["associato_id"]),
                    extra_data_keys=["has_tesseramento"],
                ),
                required_field=True,
                wide=True,
                searchable=True,
            ),
            lambda row: readonly_field("Anno", str(row["anno_oratorio"] or "")),
            lambda row: input_field("Data iscrizione", "data_iscrizione", input_type="date", value=row["data_iscrizione"] or "", required_field=True),
            lambda row: input_field("Quota partecipazione", "quota_partecipazione", input_type="number", value=str(row["quota_partecipazione"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Stato iscrizione", "stato_iscrizione", camp_enrollment_status_options(row["stato_iscrizione"] or "Iscritto")),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM iscrizioni_oratorio WHERE id = ?",
        "delete_prompt": "Eliminare questa iscrizione a Oratorio e il pagamento collegato?",
        "success_update": "Iscrizione Oratorio aggiornata.",
        "success_delete": "Iscrizione Oratorio eliminata.",
    },
    "pagamenti_oratorio": {
        "page_title": "Modifica pagamento Oratorio",
        "page_subtitle": "Aggiorna il pagamento una tantum di Oratorio.",
        "return_path": "/maschere/oratorio",
        "fetch_query": """
            SELECT po.*, o.anno AS anno_oratorio
            FROM pagamenti_oratorio po
            JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
            JOIN oratorio o ON o.id = io.oratorio_id
            WHERE po.id = ?
        """,
        "update_sql": """
            UPDATE pagamenti_oratorio
            SET iscrizione_oratorio_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "iscrizione_oratorio_id", "Iscrizione Oratorio"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Iscrizione Oratorio", "iscrizione_oratorio_id", render_select_options(iscrizioni_oratorio_options(int(row["anno_oratorio"] or date.today().year)), str(row["iscrizione_oratorio_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_oratorio WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento di Oratorio?",
        "success_update": "Pagamento Oratorio aggiornato.",
        "success_delete": "Pagamento Oratorio eliminato.",
    },
    "eventi": {
        "page_title": "Modifica evento",
        "page_subtitle": "Aggiorna dati, quota e stato dell'evento.",
        "return_path": "/maschere/eventi",
        "fetch_query": "SELECT * FROM eventi WHERE id = ?",
        "update_sql": """
            UPDATE eventi
            SET codice_evento = ?, nome = ?, tipologia = ?, data_evento = ?, luogo = ?,
                quota_partecipazione_standard = ?, posti_massimi = ?, descrizione = ?, attivo = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "codice_evento", "Codice evento"),
            required(form_data, "nome", "Nome evento"),
            optional(form_data, "tipologia"),
            required(form_data, "data_evento", "Data evento"),
            optional(form_data, "luogo"),
            normalized(form_data, "quota_partecipazione_standard", "0"),
            optional(form_data, "posti_massimi"),
            optional(form_data, "descrizione"),
            normalized(form_data, "attivo", "1") or "1",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: input_field("Codice evento", "codice_evento", value=row["codice_evento"] or "", required_field=True),
            lambda row: input_field("Nome evento", "nome", value=row["nome"] or "", required_field=True),
            lambda row: input_field("Tipologia", "tipologia", value=row["tipologia"] or ""),
            lambda row: input_field("Data evento", "data_evento", input_type="date", value=row["data_evento"] or "", required_field=True),
            lambda row: input_field("Luogo", "luogo", value=row["luogo"] or ""),
            lambda row: input_field("Quota standard", "quota_partecipazione_standard", input_type="number", value=str(row["quota_partecipazione_standard"] or ""), step="0.01", minimum="0"),
            lambda row: input_field("Posti massimi", "posti_massimi", input_type="number", value=str(row["posti_massimi"] or "") if row["posti_massimi"] is not None else "", minimum="1"),
            lambda row: select_field("Attivo", "attivo", boolean_options(str(row["attivo"] if row["attivo"] is not None else 1))),
            lambda row: textarea_field("Descrizione", "descrizione", value=row["descrizione"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM eventi WHERE id = ?",
        "delete_prompt": "Eliminare questo evento e tutte le iscrizioni collegate?",
        "success_update": "Evento aggiornato.",
        "success_delete": "Evento eliminato.",
    },
    "iscrizioni_eventi": {
        "page_title": "Modifica iscrizione evento",
        "page_subtitle": "Aggiorna partecipante, quota e stato dell'evento.",
        "return_path": "/maschere/eventi",
        "fetch_query": "SELECT * FROM iscrizioni_eventi WHERE id = ?",
        "update_sql": """
            UPDATE iscrizioni_eventi
            SET associato_id = ?, evento_id = ?, data_iscrizione = ?, quota_partecipazione = ?, stato_iscrizione = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "associato_id", "Associato"),
            required(form_data, "evento_id", "Evento"),
            required(form_data, "data_iscrizione", "Data iscrizione"),
            required(form_data, "quota_partecipazione", "Quota partecipazione"),
            normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Tesserato", "associato_id", render_associato_options(associati_options(), str(row["associato_id"])), required_field=True, wide=True, searchable=True),
            lambda row: select_field("Evento", "evento_id", render_select_options(eventi_options(), str(row["evento_id"])), required_field=True, wide=True),
            lambda row: input_field("Data iscrizione", "data_iscrizione", input_type="date", value=row["data_iscrizione"] or "", required_field=True),
            lambda row: input_field("Quota partecipazione", "quota_partecipazione", input_type="number", value=str(row["quota_partecipazione"] or ""), step="0.01", minimum="0", required_field=True),
            lambda row: select_field("Stato iscrizione", "stato_iscrizione", event_enrollment_status_options(row["stato_iscrizione"] or "Iscritto")),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM iscrizioni_eventi WHERE id = ?",
        "delete_prompt": "Eliminare questa iscrizione evento e il pagamento collegato?",
        "success_update": "Iscrizione evento aggiornata.",
        "success_delete": "Iscrizione evento eliminata.",
    },
    "pagamenti_eventi": {
        "page_title": "Modifica pagamento evento",
        "page_subtitle": "Aggiorna il pagamento una tantum dell'evento.",
        "return_path": "/maschere/eventi",
        "fetch_query": "SELECT * FROM pagamenti_eventi WHERE id = ?",
        "update_sql": """
            UPDATE pagamenti_eventi
            SET iscrizione_evento_id = ?, data_pagamento = ?, importo = ?, metodo_pagamento_id = ?, riferimento = ?, note = ?
            WHERE id = ?
        """,
        "build_params": lambda form_data: (
            required(form_data, "iscrizione_evento_id", "Iscrizione evento"),
            required(form_data, "data_pagamento", "Data pagamento"),
            required(form_data, "importo", "Importo"),
            optional(form_data, "metodo_pagamento_id"),
            optional(form_data, "riferimento"),
            optional(form_data, "note"),
        ),
        "fields": [
            lambda row: select_field("Iscrizione evento", "iscrizione_evento_id", render_select_options(iscrizioni_eventi_options(), str(row["iscrizione_evento_id"])), required_field=True, wide=True),
            lambda row: input_field("Data pagamento", "data_pagamento", input_type="date", value=row["data_pagamento"] or "", required_field=True),
            lambda row: input_field("Importo", "importo", input_type="number", value=str(row["importo"] or ""), step="0.01", minimum="0.01", required_field=True),
            lambda row: select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi_options(), str(row["metodo_pagamento_id"] or "")), wide=True),
            lambda row: input_field("Riferimento", "riferimento", value=row["riferimento"] or ""),
            lambda row: textarea_field("Note", "note", value=row["note"] or ""),
        ],
        "delete_sql": "DELETE FROM pagamenti_eventi WHERE id = ?",
        "delete_prompt": "Eliminare questo pagamento evento?",
        "success_update": "Pagamento evento aggiornato.",
        "success_delete": "Pagamento evento eliminato.",
    },
}


def get_crud_config(entity_key: str) -> dict:
    config = CRUD_CONFIG.get(entity_key)
    if config is None:
        raise KeyError(entity_key)
    return config


def access_users_rows() -> list[sqlite3.Row]:
    return fetch_all(
        """
        SELECT
            id,
            username,
            COALESCE(email_recupero, '') AS email_recupero,
            CASE WHEN is_admin = 1 THEN 'Amministratore' ELSE 'Utente' END AS profilo,
            CASE WHEN attivo = 1 THEN 'Attivo' ELSE 'Disattivato' END AS stato,
            creato_il,
            COALESCE(ultimo_accesso, '') AS ultimo_accesso
        FROM utenti_accesso
        ORDER BY is_admin DESC, username
        """
    )


def login_page(query_params: dict[str, str]) -> bytes:
    next_target = login_destination(query_params.get("next"))
    first_access = bootstrap_admin_required()
    if first_access:
        title = "Primo accesso"
        subtitle = "Crea l'utente amministratore che potra poi generare gli altri utenti operativi."
        action = "/azioni/accesso/primo-admin"
        button_label = "Crea amministratore"
        intro_html = '<span class="auth-mode">Configurazione iniziale</span>'
        recovery_html = ""
    else:
        title = "Accesso"
        subtitle = "Inserisci username e password per entrare nel gestionale."
        action = "/azioni/accesso/login"
        button_label = "Accedi"
        intro_html = '<span class="auth-mode">Area riservata</span>'
        recovery_html = f"""
        <div class="auth-support-actions">
          <a class="button ghost" href="{esc(with_query('/recupera-password', {'next': next_target} if next_target != '/' else {}))}">Recupera password</a>
        </div>
        """

    form = form_card(
        "",
        "",
        action,
        "".join(
            [
                input_field(
                    "Username",
                    "username",
                    required_field=True,
                    wide=True,
                    attrs={"autocomplete": "username"},
                ),
                input_field(
                    "Password",
                    "password",
                    input_type="password",
                    required_field=True,
                    wide=True,
                    revealable=True,
                    attrs={"autocomplete": "current-password" if not first_access else "new-password"},
                ),
                input_field(
                    "Conferma password",
                    "password_conferma",
                    input_type="password",
                    required_field=first_access,
                    wide=True,
                    revealable=True,
                    attrs={"autocomplete": "new-password"},
                ) if first_access else "",
            ]
        ),
        button_label,
        hidden_fields={"next": next_target},
    )
    content = f"""
    <div class="auth-panel-stack">
      <div class="auth-intro-card">
        {intro_html}
        <h2>{esc(title)}</h2>
        <p>{esc(subtitle)}</p>
      </div>
      {form}
      {recovery_html}
    </div>
    """
    return public_page(title, content, query_params)


def recover_password_page(query_params: dict[str, str]) -> bytes:
    next_target = login_destination(query_params.get("next"))
    form = form_card(
        "Recupera password",
        "Inserisci username ed email di recupero per impostare una nuova password del tuo utente standard.",
        "/azioni/accesso/recupera-password",
        "".join(
            [
                input_field("Username", "username", required_field=True, wide=True, attrs={"autocomplete": "username"}),
                input_field("Email recupero", "email_recupero", input_type="email", required_field=True, wide=True, attrs={"autocomplete": "email"}),
                input_field("Nuova password", "password", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
                input_field("Conferma nuova password", "password_conferma", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
            ]
        ),
        "Aggiorna password",
        hidden_fields={"next": next_target},
    )
    content = f"""
    <div class="auth-panel-stack">
      <div class="auth-intro-card">
        <span class="auth-mode">Recupero credenziali</span>
        <h2>Recupera password</h2>
        <p>Per ragioni di sicurezza la password attuale non viene mostrata: qui puoi solo impostarne una nuova dopo verifica.</p>
      </div>
      {form}
      <div class="auth-support-actions">
        <a class="button ghost" href="{esc(with_query('/login', {'next': next_target} if next_target != '/' else {}))}">Torna al login</a>
      </div>
    </div>
    """
    return public_page("Recupera password", content, query_params)


def access_profile_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    if current_user is None:
        raise KeyError("current_user")
    user_row = access_user_row(int(current_user["id"]))
    if user_row is None:
        raise KeyError("current_user")

    info_card = table_card(
        "Dati accesso",
        "Le password restano protette e non sono visualizzabili. Da questa area puoi aggiornarle in sicurezza.",
        [user_row],
        [
            ("username", "Username"),
            ("email_recupero", "Email recupero"),
            ("ultimo_accesso", "Ultimo accesso"),
            ("aggiornato_il", "Ultimo aggiornamento"),
        ],
    )
    change_password_form = form_card(
        "Cambia password",
        "Inserisci la password attuale e poi la nuova password.",
        "/azioni/accesso/cambia-password",
        "".join(
            [
                input_field("Password attuale", "password_attuale", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "current-password"}),
                input_field("Nuova password", "password", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
                input_field("Conferma nuova password", "password_conferma", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
            ]
        ),
        "Aggiorna password",
        hidden_fields=work_year_query(query_params),
    )
    content = f"""
    <div class="cards-grid">
      {change_password_form}
    </div>
    {info_card}
    """
    return page("Profilo accesso", "/maschere/accesso", content, query_params, current_user)


def utenti_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    users = access_users_rows()
    create_form = form_card(
        "Nuovo utente",
        "Crea un utente operativo. Gli utenti creati da qui non possono amministrare altri account.",
        "/azioni/utenti/crea",
        "".join(
            [
                readonly_field("Profilo", "Utente standard"),
                input_field("Username", "username", required_field=True, attrs={"autocomplete": "off"}),
                input_field("Password", "password", input_type="password", required_field=True, revealable=True, attrs={"autocomplete": "new-password"}),
                input_field("Conferma password", "password_conferma", input_type="password", required_field=True, revealable=True, attrs={"autocomplete": "new-password"}),
                input_field("Email recupero", "email_recupero", input_type="email", required_field=True, wide=True, attrs={"autocomplete": "off"}),
            ]
        ),
        "Salva utente",
        hidden_fields=work_year_query(query_params),
    )
    content = f"""
    <div class="cards-grid">
      {create_form}
    </div>
    {table_card(
        "Utenti registrati",
        "Le password non sono leggibili ne esportabili: l'amministratore puo pero reimpostarle e disattivare gli utenti standard.",
        users,
        [
            ("username", "Username"),
            ("email_recupero", "Email recupero"),
            ("profilo", "Profilo"),
            ("stato", "Stato"),
            ("creato_il", "Creato il"),
            ("ultimo_accesso", "Ultimo accesso"),
            (
                "id",
                "Azioni",
                lambda value, row: ""
                if row["profilo"] == "Amministratore"
                else action_links_html(
                    extra_links=[
                        (
                            with_query(f"/maschere/utenti/gestione/{value}", work_year_query(query_params)),
                            "Gestisci",
                        )
                    ]
                ),
            ),
        ],
    )}
    """
    return page("Utenti", "/maschere/utenti", content, query_params, current_user)


def backup_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    if not current_user or not current_user.get("is_admin"):
        raise KeyError("backup")

    backups = backup_archive_rows(query_params)
    create_form = form_card(
        "Backup completo",
        "Crea un archivio ZIP con applicazione, database, runtime, script e file di ripristino per questo PC o per un altro PC.",
        "/azioni/backup/crea",
        "".join(
            [
                readonly_field("Percorso salvataggio", str(BACKUP_DIR), wide=True),
                readonly_field("Ripristino automatico incluso", "Si", wide=True),
            ]
        ),
        "Crea backup e scarica",
        hidden_fields=work_year_query(query_params),
    )
    backups_table = table_card(
        "Backup disponibili",
        "Puoi riscaricare in qualsiasi momento uno dei backup gia creati dal gestionale.",
        backups,
        [
            ("filename", "File"),
            ("created_at", "Creato il"),
            ("size", "Dimensione"),
            ("path", "Percorso"),
            (
                "download_url",
                "Azioni",
                lambda value, row: action_links_html(
                    extra_links=[(str(value or ""), "Scarica backup")],
                    delete_action=f"/azioni/backup/elimina/{quote(str(row['filename'] or ''))}",
                    delete_prompt="Eliminare definitivamente questo backup?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ],
        empty_message="Nessun backup ancora disponibile.",
    )
    content = f"""
    <div class="cards-grid">
      {create_form}
    </div>
    {backups_table}
    """
    return page("Backup", "/maschere/backup", content, query_params, current_user)


def aggiornamenti_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    if not current_user or not current_user.get("is_admin"):
        raise KeyError("aggiornamenti")

    content = f"""
    <section class="card">
      <div class="card-head">
        <h2>Installa aggiornamento guidato</h2>
      </div>
      <form method="post" action="/azioni/aggiornamenti/installa" class="form-grid" enctype="multipart/form-data">
        {hidden_fields_html(work_year_query(query_params))}
        {input_field("File ZIP pacchetto aggiornamento", "file_zip_aggiornamento", input_type="file", required_field=True, wide=True, attrs={"accept": ".zip"})}
        <div class="form-actions">
          <button type="submit" class="button">Installa aggiornamento</button>
        </div>
      </form>
    </section>
    """
    return page("Aggiornamenti", "/maschere/aggiornamenti", content, query_params, current_user)


def importa_associati_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    if not current_user or not current_user.get("is_admin"):
        raise KeyError("importa-associati")

    model_form = form_card(
        "Genera Modello",
        "Scarica il file Excel con tutti i campi disponibili da compilare. Le intestazioni non devono essere modificate.",
        "/azioni/associati/template-excel",
        "".join(
            [
                readonly_field("Versione modello", APP_VERSION, wide=True),
                readonly_field("Foglio da compilare", "Associati", wide=True),
                readonly_field("Controllo duplicati", "Codice fiscale oppure Nome + Cognome + Data nascita", wide=True),
            ]
        ),
        "Genera modello Excel",
        hidden_fields=work_year_query(query_params),
        card_class="import-associati-card",
    )
    import_form = form_card(
        "Importa Excel",
        "Carica il file Excel compilato. Il gestionale saltera automaticamente i record gia presenti e importera solo quelli nuovi.",
        "/azioni/associati/importa-excel",
        "".join(
            [
                input_field("File Excel", "file_excel_associati", input_type="file", required_field=True, wide=True, attrs={"accept": ".xlsx,.xlsm,.xltx,.xltm"}),
                readonly_field("Se l'associato e minorenne", "Sono obbligatori solo Nome, Cognome e Cellulare del Genitore/Tutore", wide=True),
            ]
        ),
        "Importa dati da Excel",
        hidden_fields=work_year_query(query_params),
        card_class="import-associati-card",
        form_attrs={"enctype": "multipart/form-data"},
    )
    content = f"""
    <div class="cards-grid">
      {model_form}
      {import_form}
    </div>
    """
    return page("Importa associati", "/maschere/importa-associati", content, query_params, current_user)


def load_guide_media_catalog() -> dict[str, dict[str, object]]:
    if not GUIDE_MEDIA_JSON_PATH.exists():
        return {}
    try:
        raw_catalog = json.loads(GUIDE_MEDIA_JSON_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    catalog: dict[str, dict[str, object]] = {}
    for course_key, course_value in raw_catalog.items():
        slides = []
        for index, slide in enumerate(course_value.get("slides", []), start=1):
            slides.append(
                {
                    "title": slide.get("title", ""),
                    "subtitle": slide.get("subtitle", ""),
                    "bullets": slide.get("bullets", []),
                    "audio_text": slide.get("audio_text", ""),
                    "image_url": f"/static/guide-media/images/{course_key}/{index:02d}.png",
                }
            )
        catalog[course_key] = {
            "title": course_value.get("title", course_key.replace("_", " ").title()),
            "subtitle": course_value.get("subtitle", ""),
            "slides": slides,
            "download_url": f"/download/tutorial/{quote(course_key)}.mp4",
        }
    return catalog


def tutorial_video_output_path(course_key: str) -> Path:
    ensure_tutorial_video_dir()
    safe_key = re.sub(r"[^0-9A-Za-z._-]+", "-", str(course_key or "tutorial")).strip("-") or "tutorial"
    return TUTORIAL_VIDEO_DIR / f"{safe_key}.mp4"


def tutorial_video_download_name(course_key: str, course: dict[str, object]) -> str:
    raw_name = str(course.get("title", "") or course_key or "tutorial")
    safe_name = re.sub(r"[^0-9A-Za-z._-]+", "-", raw_name).strip("-") or "tutorial"
    return f"{safe_name}.mp4"


def tutorial_video_is_stale(course_key: str, output_path: Path) -> bool:
    if not output_path.exists():
        return True
    output_mtime = output_path.stat().st_mtime
    dependencies = [GUIDE_MEDIA_JSON_PATH, TUTORIAL_VIDEO_SCRIPT]
    image_dir = STATIC_DIR / "guide-media" / "images" / course_key
    if image_dir.exists():
        dependencies.extend(sorted(image_dir.glob("*.png")))
    for dependency in dependencies:
        if dependency.exists() and dependency.stat().st_mtime > output_mtime:
            return True
    return False


def ensure_tutorial_video_file(course_key: str) -> tuple[Path, str]:
    catalog = load_guide_media_catalog()
    course = catalog.get(course_key)
    if course is None:
        raise KeyError(course_key)
    output_path = tutorial_video_output_path(course_key)
    if tutorial_video_is_stale(course_key, output_path):
        if not TUTORIAL_VIDEO_SCRIPT.exists():
            raise ValueError("Script di generazione tutorial non trovato.")
        command = [
            "powershell.exe",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(TUTORIAL_VIDEO_SCRIPT),
            "-CatalogPath",
            str(GUIDE_MEDIA_JSON_PATH),
            "-CourseKey",
            course_key,
            "-BaseDir",
            str(BASE_DIR),
            "-OutputPath",
            str(output_path),
        ]
        try:
            subprocess.run(
                command,
                cwd=str(BASE_DIR),
                capture_output=True,
                text=True,
                check=True,
                creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
        except subprocess.CalledProcessError as error:
            detail = (error.stderr or error.stdout or "").strip()
            raise ValueError(f"Generazione video tutorial non riuscita. {detail}".strip()) from error
    if not output_path.is_file():
        raise ValueError("Il file video del tutorial non e stato generato.")
    return output_path, tutorial_video_download_name(course_key, course)


def guide_media_card(course_key: str, course: dict[str, object]) -> str:
    slides = course.get("slides", [])
    first_slide = slides[0] if slides else {}
    return f"""
    <section class="card guide-media-card" data-guide-player="{esc(course_key)}">
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(str(course.get("title", "")))}</h2>
          <p>{esc(str(course.get("subtitle", "")))}</p>
        </div>
        <div class="card-head-actions mini-actions">
          <button type="button" class="button ghost" data-guide-action="restart">Ricomincia</button>
          <button type="button" class="button" data-guide-action="toggle">Avvia narrazione</button>
        </div>
      </div>
      <div class="guide-media-layout">
        <div class="guide-media-stage">
          <div class="guide-media-screen">
            <img class="guide-media-image" src="{esc(str(first_slide.get("image_url", "")))}" alt="{esc(str(course.get("title", "")))}">
          </div>
          <div class="guide-media-caption">
            <div class="guide-media-progress">Slide <span data-guide-current>1</span>/<span data-guide-total>{len(slides)}</span></div>
            <h3 data-guide-slide-title>{esc(str(first_slide.get("title", "")))}</h3>
            <p data-guide-slide-subtitle>{esc(str(first_slide.get("subtitle", "")))}</p>
          </div>
        </div>
        <div class="guide-media-track-panel">
          <div class="guide-media-track-head">
            <strong>Percorso</strong>
            <span class="guide-media-note">Immagini e voce del browser</span>
          </div>
          <ol class="guide-media-track-list" data-guide-track-list></ol>
        </div>
      </div>
    </section>
    """


def guida_in_linea_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    sections = [
        ("Accesso", "Accedi con le credenziali fornite. L'amministratore puo gestire utenti, backup, aggiornamenti e importazioni massime da Excel."),
        ("Nuovo associato", "Compila l'anagrafica di base. Al salvataggio il gestionale crea automaticamente anche il tesseramento dell'anno di lavoro e ti chiede se vuoi registrare subito il pagamento."),
        ("Rinnovo Tesseramento", "Usa questa maschera per i rinnovi annuali. Il codice tesseramento viene assegnato in formato anno/progressivo, ad esempio 26/1."),
        ("Iscrizioni", "Per Corsi, Oratorio, Campo estivo ed Eventi il gestionale verifica il tesseramento dell'anno e puo proporti subito il pagamento con ricevuta."),
        ("Pagamenti", "La maschera Pagamenti consente di saldare insieme scadenze provenienti da aree diverse e di includere anche quote future dei corsi."),
        ("Report", "Ogni report puo essere filtrato, ordinato, esportato in Excel o PDF, stampato e condiviso tramite mail o WhatsApp."),
        ("Tesserati e anagrafica storica", "Le tabelle ricordano la tua vista, permettono filtri per colonna, riordino trascinando le intestazioni e scelta dinamica dei campi da mostrare."),
        ("Backup", "Solo l'amministratore puo creare backup completi ZIP con database, applicazione, runtime e file di ripristino."),
        ("Aggiornamenti", "Solo l'amministratore puo installare aggiornamenti guidati dal software. Il pacchetto aggiornamento va creato da comando esterno e mantiene i dati del PC di destinazione."),
        ("Importazione Excel", "Solo l'amministratore puo generare il modello Excel associati e importarlo evitando duplicati gia presenti."),
    ]
    guide_cards = "".join(
        f"""
        <article class="card guide-card" data-guide-topic="{esc(title.lower())} {esc(text.lower())}">
          <div class="card-head">
            <h2>{esc(title)}</h2>
            <p>{esc(text)}</p>
          </div>
        </article>
        """
        for title, text in sections
    )
    assistant_knowledge = json.dumps(
        [
            {"keywords": ["backup", "salvataggio"], "answer": "Apri Amministrazione > Backup e crea lo ZIP completo. Per ripristinarlo estrai il file ed esegui il file di ripristino incluso."},
            {"keywords": ["aggiornamento", "versione"], "answer": "Apri Amministrazione > Aggiornamenti per leggere la procedura guidata. Il pacchetto si crea da comando esterno oppure dal file 'Crea pacchetto aggiornamento per altri PC.bat', poi sul PC di destinazione esegui 'Installa aggiornamento guidato.bat'."},
            {"keywords": ["import", "excel", "modello"], "answer": "Apri Amministrazione > Importa associati, genera il modello Excel, compilalo e poi importalo dallo stesso pannello."},
            {"keywords": ["tesseramento", "codice"], "answer": "Il codice tesseramento viene assegnato automaticamente nel formato anno/progressivo, ad esempio 26/1, 26/2, 26/3."},
            {"keywords": ["pagamento", "ricevuta"], "answer": "Dopo ogni iscrizione puoi scegliere se registrare subito il pagamento. Il gestionale genera automaticamente la relativa ricevuta."},
            {"keywords": ["minorenne", "genitore", "tutore"], "answer": "Per i minorenni sono obbligatori solo Nome, Cognome e Cellulare del Genitore/Tutore. Le ricevute vengono inviate ai recapiti del Genitore/Tutore."},
        ],
        ensure_ascii=False,
    )
    guide_media_catalog = load_guide_media_catalog()
    guide_media_cards = "".join(guide_media_card(course_key, course) for course_key, course in guide_media_catalog.items())
    guide_media_json = json.dumps(guide_media_catalog, ensure_ascii=False)
    content = f"""
    <section class="report-toolbar">
      <div class="report-toolbar-copy">
        <h2>Guida in linea</h2>
        <p>Raccoglie in modo semplice le funzioni principali del gestionale e propone anche un assistente locale in stile chat per orientarti piu rapidamente.</p>
      </div>
      <div class="report-toolbar-actions">
        <div class="field wide">
          <span>Cerca nella guida</span>
          <input type="search" class="control" id="guide-search-input" placeholder="Scrivi ad esempio: backup, tesseramento, import Excel..." oninput="filterGuideCards(this.value)">
        </div>
      </div>
    </section>
    <div class="cards-grid guide-media-grid">
      {guide_media_cards}
    </div>
    <section class="card guide-chat-card">
      <div class="card-head">
        <h2>Assistente guida</h2>
        <p>Risponde alle domande piu frequenti basandosi sulle funzioni gia presenti nel software.</p>
      </div>
      <div id="guide-chat-log" class="guide-chat-log">
        <div class="guide-chat-message assistant">Scrivi una domanda breve, ad esempio: come faccio un backup oppure come aggiorno un altro PC?</div>
      </div>
      <div class="guide-chat-compose">
        <input type="search" class="control" id="guide-chat-input" placeholder="Fai una domanda sulla procedura da seguire..." onkeydown="if (event.key === 'Enter') {{ event.preventDefault(); askGuideAssistant(); }}">
        <button type="button" class="button" onclick="askGuideAssistant()">Invia domanda</button>
      </div>
    </section>
    <div class="cards-grid guide-grid">
      {guide_cards}
    </div>
    <script>
      const guideAssistantKnowledge = {assistant_knowledge};
      function filterGuideCards(rawValue) {{
        const search = String(rawValue || '').trim().toLowerCase();
        document.querySelectorAll('[data-guide-topic]').forEach((card) => {{
          const haystack = String(card.dataset.guideTopic || '');
          card.hidden = !!search && !haystack.includes(search);
        }});
      }}
      function appendGuideChatMessage(role, text) {{
        const log = document.getElementById('guide-chat-log');
        if (!log) {{
          return;
        }}
        const message = document.createElement('div');
        message.className = `guide-chat-message ${{role}}`;
        message.textContent = text;
        log.appendChild(message);
        log.scrollTop = log.scrollHeight;
      }}
      function askGuideAssistant() {{
        const input = document.getElementById('guide-chat-input');
        if (!input) {{
          return;
        }}
        const question = String(input.value || '').trim();
        if (!question) {{
          return;
        }}
        appendGuideChatMessage('user', question);
        input.value = '';
        const normalized = question.toLowerCase();
        const match = guideAssistantKnowledge.find((item) => item.keywords.some((keyword) => normalized.includes(keyword)));
        appendGuideChatMessage('assistant', match ? match.answer : 'Non ho trovato una risposta specifica. Prova a cercare nella guida qui sopra oppure usa parole chiave come backup, aggiornamento, tesseramento o import Excel.');
      }}
      const guideMediaCatalog = {guide_media_json};
      let guideNarrationVoice = null;
      const guideMediaStates = [];
      function resolveGuideVoice() {{
        const voices = window.speechSynthesis ? window.speechSynthesis.getVoices() : [];
        guideNarrationVoice = voices.find((voice) => String(voice.lang || '').toLowerCase().startsWith('it')) || voices[0] || null;
        return guideNarrationVoice;
      }}
      if (window.speechSynthesis && typeof window.speechSynthesis.onvoiceschanged !== 'undefined') {{
        window.speechSynthesis.onvoiceschanged = resolveGuideVoice;
      }}
      resolveGuideVoice();
      function updateGuideToggleButton(state) {{
        if (!state || !state.toggleButton) {{
          return;
        }}
        state.toggleButton.textContent = state.playing ? 'Interrompi narrazione' : 'Avvia narrazione';
      }}
      function renderGuideTrack(state) {{
        if (!state || !state.trackList) {{
          return;
        }}
        state.trackList.innerHTML = '';
        state.slides.forEach((slide, index) => {{
          const item = document.createElement('li');
          item.className = 'guide-media-track-item';
          const button = document.createElement('button');
          button.type = 'button';
          button.className = 'guide-media-track-button';
          if (index === state.index) {{
            button.classList.add('active');
          }}
          button.innerHTML = `<strong>${{index + 1}}. ${{slide.title}}</strong><span>${{slide.subtitle}}</span>`;
          button.addEventListener('click', () => {{
            state.index = index;
            renderGuideSlide(state);
            if (state.playing) {{
              speakGuideSlide(state);
            }}
          }});
          item.appendChild(button);
          state.trackList.appendChild(item);
        }});
      }}
      function renderGuideSlide(state) {{
        if (!state) {{
          return;
        }}
        const slide = state.slides[state.index];
        if (!slide) {{
          return;
        }}
        state.image.src = slide.image_url;
        state.image.alt = slide.title;
        state.currentEl.textContent = String(state.index + 1);
        state.titleEl.textContent = slide.title;
        state.subtitleEl.textContent = slide.subtitle;
        renderGuideTrack(state);
        updateGuideToggleButton(state);
      }}
      function stopGuideNarration(state) {{
        if (!state) {{
          return;
        }}
        state.playing = false;
        if (window.speechSynthesis) {{
          window.speechSynthesis.cancel();
        }}
        updateGuideToggleButton(state);
      }}
      function speakGuideSlide(state) {{
        if (!state || !window.speechSynthesis) {{
          return;
        }}
        const slide = state.slides[state.index];
        if (!slide) {{
          return;
        }}
        window.speechSynthesis.cancel();
        const utterance = new SpeechSynthesisUtterance(slide.audio_text || `${{slide.title}}. ${{slide.subtitle}}`);
        utterance.lang = 'it-IT';
        utterance.rate = 0.97;
        utterance.pitch = 1;
        const chosenVoice = resolveGuideVoice();
        if (chosenVoice) {{
          utterance.voice = chosenVoice;
        }}
        utterance.onend = () => {{
          if (!state.playing) {{
            return;
          }}
          if (state.index < state.slides.length - 1) {{
            state.index += 1;
            renderGuideSlide(state);
            speakGuideSlide(state);
          }} else {{
            stopGuideNarration(state);
          }}
        }};
        utterance.onerror = () => {{
          stopGuideNarration(state);
        }};
        window.speechSynthesis.speak(utterance);
      }}
      function initGuideMediaPlayers() {{
        document.querySelectorAll('[data-guide-player]').forEach((card) => {{
          const courseKey = String(card.dataset.guidePlayer || '');
          const course = guideMediaCatalog[courseKey];
          if (!course || !Array.isArray(course.slides) || !course.slides.length) {{
            return;
          }}
          const state = {{
            slides: course.slides,
            index: 0,
            playing: false,
            image: card.querySelector('.guide-media-image'),
            currentEl: card.querySelector('[data-guide-current]'),
            titleEl: card.querySelector('[data-guide-slide-title]'),
            subtitleEl: card.querySelector('[data-guide-slide-subtitle]'),
            trackList: card.querySelector('[data-guide-track-list]'),
            toggleButton: card.querySelector('[data-guide-action=\"toggle\"]'),
            restartButton: card.querySelector('[data-guide-action=\"restart\"]'),
          }};
          if (state.toggleButton) {{
            state.toggleButton.addEventListener('click', () => {{
              if (state.playing) {{
                stopGuideNarration(state);
              }} else {{
                guideMediaStates.forEach((otherState) => {{
                  if (otherState !== state) {{
                    stopGuideNarration(otherState);
                  }}
                }});
                state.playing = true;
                updateGuideToggleButton(state);
                speakGuideSlide(state);
              }}
            }});
          }}
          if (state.restartButton) {{
            state.restartButton.addEventListener('click', () => {{
              state.index = 0;
              renderGuideSlide(state);
              if (state.playing) {{
                speakGuideSlide(state);
              }}
            }});
          }}
          renderGuideSlide(state);
          guideMediaStates.push(state);
        }});
      }}
      initGuideMediaPlayers();
      window.addEventListener('beforeunload', () => {{
        guideMediaStates.forEach((state) => stopGuideNarration(state));
      }});
      document.addEventListener('visibilitychange', () => {{
        if (document.hidden) {{
          guideMediaStates.forEach((state) => stopGuideNarration(state));
        }}
      }});
    </script>
    """
    return page("Guida in linea", "/maschere/guida", content, query_params, current_user)


def video_tutorial_card(course_key: str, course: dict[str, object]) -> str:
    slides = course.get("slides", [])
    first_slide = slides[0] if slides else {}
    return f"""
    <section class="card guide-media-card" data-video-player="{esc(course_key)}">
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(str(course.get("title", "")))}</h2>
          <p>{esc(str(course.get("subtitle", "")))}</p>
        </div>
        <div class="card-head-actions mini-actions">
          <button type="button" class="button" data-video-action="toggle">Avvia presentazione</button>
          <button type="button" class="button ghost" data-video-action="download">Scarica video</button>
        </div>
      </div>
      <div class="guide-media-layout">
        <div class="guide-media-stage">
          <div class="guide-media-screen">
            <img class="guide-media-image" src="{esc(str(first_slide.get("image_url", "")))}" alt="{esc(str(course.get("title", "")))}">
          </div>
          <div class="guide-media-caption">
            <div class="guide-media-progress">Slide <span data-video-current>1</span>/<span data-video-total>{len(slides)}</span></div>
            <h3 data-video-slide-title>{esc(str(first_slide.get("title", "")))}</h3>
            <p data-video-slide-subtitle>{esc(str(first_slide.get("subtitle", "")))}</p>
          </div>
        </div>
        <div class="guide-media-track-panel">
          <div class="guide-media-track-head">
            <strong>Percorso</strong>
            <span class="guide-media-note">Corso completo con schermate reali del software</span>
          </div>
          <ol class="guide-media-track-list" data-video-track-list></ol>
        </div>
      </div>
    </section>
    """


def video_tutorial_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    guide_media_catalog = load_guide_media_catalog()
    ordered_keys = [key for key in ("promo_course", "full_course") if key in guide_media_catalog]
    ordered_keys.extend(key for key in guide_media_catalog.keys() if key not in ordered_keys)
    video_cards = "".join(video_tutorial_card(course_key, guide_media_catalog[course_key]) for course_key in ordered_keys)
    video_catalog_json = json.dumps(guide_media_catalog, ensure_ascii=False)
    content = f"""
    <div class="cards-grid guide-media-grid">
      {video_cards}
    </div>
    <section id="video-tutorial-overlay" class="guide-player-overlay" hidden>
      <div class="guide-player-backdrop"></div>
      <div class="guide-player-shell">
        <div class="guide-player-topbar">
          <div class="guide-player-topcopy">
            <span class="guide-player-kicker">Tutorial</span>
            <h2 id="video-tutorial-course-title">Presentazione</h2>
            <p id="video-tutorial-course-subtitle"></p>
          </div>
          <div class="guide-player-actions">
            <button type="button" class="button ghost" id="video-tutorial-restart">Ricomincia</button>
            <button type="button" class="button ghost" id="video-tutorial-download">Scarica video</button>
            <button type="button" class="button" id="video-tutorial-toggle">Avvia presentazione</button>
            <button type="button" class="button ghost" id="video-tutorial-close">Chiudi</button>
          </div>
        </div>
        <div class="guide-player-body">
          <div class="guide-player-stage">
            <div class="guide-player-frame">
              <img id="video-tutorial-image" class="guide-player-image" src="" alt="Slide tutorial">
            </div>
            <div class="guide-player-meta">
              <div class="guide-player-progress">Slide <span id="video-tutorial-current">1</span>/<span id="video-tutorial-total">0</span></div>
              <h3 id="video-tutorial-slide-title"></h3>
              <p id="video-tutorial-slide-subtitle"></p>
            </div>
          </div>
          <aside class="guide-player-track-panel">
            <div class="guide-player-track-head">
              <strong>Scaletta</strong>
              <span>Le slide scorrono in modo sincronizzato con la narrazione</span>
            </div>
            <ol id="video-tutorial-track-list" class="guide-player-track-list"></ol>
          </aside>
        </div>
      </div>
    </section>
    <script>
      const videoTutorialCatalog = {video_catalog_json};
      const videoTutorialStates = [];
      const videoTutorialOverlay = document.getElementById('video-tutorial-overlay');
      const videoTutorialCourseTitle = document.getElementById('video-tutorial-course-title');
      const videoTutorialCourseSubtitle = document.getElementById('video-tutorial-course-subtitle');
      const videoTutorialImage = document.getElementById('video-tutorial-image');
      const videoTutorialCurrent = document.getElementById('video-tutorial-current');
      const videoTutorialTotal = document.getElementById('video-tutorial-total');
      const videoTutorialSlideTitle = document.getElementById('video-tutorial-slide-title');
      const videoTutorialSlideSubtitle = document.getElementById('video-tutorial-slide-subtitle');
      const videoTutorialTrackList = document.getElementById('video-tutorial-track-list');
      const videoTutorialRestart = document.getElementById('video-tutorial-restart');
      const videoTutorialDownload = document.getElementById('video-tutorial-download');
      const videoTutorialToggle = document.getElementById('video-tutorial-toggle');
      const videoTutorialClose = document.getElementById('video-tutorial-close');
      let videoTutorialVoice = null;
      let activeVideoTutorialState = null;

      function videoVoiceScore(voice) {{
        const name = String((voice && voice.name) || '').toLowerCase();
        const lang = String((voice && voice.lang) || '').toLowerCase();
        let score = 0;
        if (lang.startsWith('it')) score += 20;
        if (name.includes('natural') || name.includes('online')) score += 12;
        if (name.includes('elsa') || name.includes('isabella') || name.includes('female') || name.includes('femmin')) score += 18;
        if (name.includes('google italiano')) score += 10;
        if (name.includes('diego') || name.includes('luca') || name.includes('male') || name.includes('masch')) score -= 8;
        return score;
      }}

      function resolveVideoTutorialVoice() {{
        const voices = window.speechSynthesis ? window.speechSynthesis.getVoices() : [];
        if (!voices.length) {{
          videoTutorialVoice = null;
          return null;
        }}
        videoTutorialVoice = [...voices].sort((left, right) => videoVoiceScore(right) - videoVoiceScore(left))[0] || voices[0] || null;
        return videoTutorialVoice;
      }}

      if (window.speechSynthesis && typeof window.speechSynthesis.onvoiceschanged !== 'undefined') {{
        window.speechSynthesis.onvoiceschanged = resolveVideoTutorialVoice;
      }}
      resolveVideoTutorialVoice();

      function videoTutorialSlug(value) {{
        return String(value || 'video-tutorial')
          .toLowerCase()
          .normalize('NFD')
          .replace(/[\\u0300-\\u036f]/g, '')
          .replace(/[^a-z0-9]+/g, '-')
          .replace(/^-+|-+$/g, '') || 'video-tutorial';
      }}

      function setVideoTutorialDownloadState(state, exporting, label) {{
        if (!state) {{
          return;
        }}
        const text = label || (exporting ? 'Preparazione video...' : 'Scarica video');
        const buttons = Array.isArray(state.downloadButtons) ? state.downloadButtons : [];
        buttons.forEach((button) => {{
          if (!button) {{
            return;
          }}
          button.disabled = !!exporting;
          button.textContent = text;
        }});
        if (state.toggleButton) {{
          state.toggleButton.disabled = !!exporting;
        }}
        if (activeVideoTutorialState === state && videoTutorialDownload) {{
          videoTutorialDownload.disabled = !!exporting;
          videoTutorialDownload.textContent = text;
        }}
        if (activeVideoTutorialState === state) {{
          if (videoTutorialToggle) {{
            videoTutorialToggle.disabled = !!exporting;
          }}
          if (videoTutorialRestart) {{
            videoTutorialRestart.disabled = !!exporting;
          }}
          if (videoTutorialClose) {{
            videoTutorialClose.disabled = !!exporting;
          }}
        }}
      }}

      async function exportVideoTutorial(state) {{
        if (!state || !state.downloadUrl) {{
          return;
        }}
        window.location.href = state.downloadUrl;
      }}

      function updateVideoTutorialButtons(state) {{
        if (!state) {{
          return;
        }}
        const label = state.playing ? 'Interrompi narrazione' : 'Avvia presentazione';
        if (state.toggleButton) {{
          state.toggleButton.textContent = label;
        }}
        if (activeVideoTutorialState === state && videoTutorialToggle) {{
          videoTutorialToggle.textContent = label;
        }}
      }}

      function renderVideoTutorialTrack(state) {{
        if (!state || !state.trackList) {{
          return;
        }}
        state.trackList.innerHTML = '';
        state.slides.forEach((slide, index) => {{
          const item = document.createElement('li');
          item.className = 'guide-media-track-item';
          const button = document.createElement('button');
          button.type = 'button';
          button.className = 'guide-media-track-button';
          button.disabled = !!state.exporting;
          if (index === state.index) {{
            button.classList.add('active');
          }}
          button.innerHTML = `<strong>${{index + 1}}. ${{slide.title}}</strong><span>${{slide.subtitle}}</span>`;
          button.addEventListener('click', () => {{
            state.index = index;
            renderVideoTutorialSlide(state);
            if (state.playing) {{
              speakVideoTutorialSlide(state);
            }}
          }});
          item.appendChild(button);
          state.trackList.appendChild(item);
        }});
        if (activeVideoTutorialState === state && videoTutorialTrackList) {{
          videoTutorialTrackList.innerHTML = state.trackList.innerHTML;
          Array.from(videoTutorialTrackList.querySelectorAll('button')).forEach((button, index) => {{
            button.disabled = !!state.exporting;
            button.addEventListener('click', () => {{
              state.index = index;
              renderVideoTutorialSlide(state);
              if (state.playing) {{
                speakVideoTutorialSlide(state);
              }}
            }});
            if (index === state.index) {{
              button.scrollIntoView({{ block: 'nearest' }});
            }}
          }});
        }}
      }}

      function renderVideoTutorialSlide(state) {{
        if (!state) {{
          return;
        }}
        const slide = state.slides[state.index];
        if (!slide) {{
          return;
        }}
        state.image.src = slide.image_url;
        state.image.alt = slide.title;
        state.currentEl.textContent = String(state.index + 1);
        state.titleEl.textContent = slide.title;
        state.subtitleEl.textContent = slide.subtitle;
        if (activeVideoTutorialState === state) {{
          videoTutorialCourseTitle.textContent = state.courseTitle;
          videoTutorialCourseSubtitle.textContent = state.courseSubtitle;
          videoTutorialImage.src = slide.image_url;
          videoTutorialImage.alt = slide.title;
          videoTutorialCurrent.textContent = String(state.index + 1);
          videoTutorialTotal.textContent = String(state.slides.length);
          videoTutorialSlideTitle.textContent = slide.title;
          videoTutorialSlideSubtitle.textContent = slide.subtitle;
        }}
        renderVideoTutorialTrack(state);
        updateVideoTutorialButtons(state);
      }}

      function stopVideoTutorialNarration(state) {{
        if (!state) {{
          return;
        }}
        state.playing = false;
        if (window.speechSynthesis) {{
          window.speechSynthesis.cancel();
        }}
        updateVideoTutorialButtons(state);
      }}

      function openVideoTutorialPresentation(state) {{
        if (!state || !videoTutorialOverlay) {{
          return;
        }}
        activeVideoTutorialState = state;
        videoTutorialOverlay.hidden = false;
        document.body.classList.add('guide-player-open');
        renderVideoTutorialSlide(state);
      }}

      function closeVideoTutorialPresentation() {{
        if (activeVideoTutorialState) {{
          stopVideoTutorialNarration(activeVideoTutorialState);
        }}
        activeVideoTutorialState = null;
        if (videoTutorialOverlay) {{
          videoTutorialOverlay.hidden = true;
        }}
        document.body.classList.remove('guide-player-open');
      }}

      function speakVideoTutorialSlide(state) {{
        if (!state || !window.speechSynthesis) {{
          return;
        }}
        const slide = state.slides[state.index];
        if (!slide) {{
          return;
        }}
        window.speechSynthesis.cancel();
        const utterance = new SpeechSynthesisUtterance(slide.audio_text || `${{slide.title}}. ${{slide.subtitle}}`);
        utterance.lang = 'it-IT';
        utterance.rate = 0.94;
        utterance.pitch = 1.02;
        utterance.volume = 1;
        const chosenVoice = resolveVideoTutorialVoice();
        if (chosenVoice) {{
          utterance.voice = chosenVoice;
        }}
        utterance.onend = () => {{
          if (!state.playing) {{
            return;
          }}
          if (state.index < state.slides.length - 1) {{
            state.index += 1;
            renderVideoTutorialSlide(state);
            speakVideoTutorialSlide(state);
          }} else {{
            stopVideoTutorialNarration(state);
          }}
        }};
        utterance.onerror = () => {{
          stopVideoTutorialNarration(state);
        }};
        window.speechSynthesis.speak(utterance);
      }}

      function startVideoTutorialPresentation(state, withNarration) {{
        if (!state) {{
          return;
        }}
        videoTutorialStates.forEach((otherState) => {{
          if (otherState !== state) {{
            stopVideoTutorialNarration(otherState);
          }}
        }});
        openVideoTutorialPresentation(state);
        state.playing = !!withNarration;
        updateVideoTutorialButtons(state);
        if (withNarration) {{
          speakVideoTutorialSlide(state);
        }}
      }}

      function initVideoTutorialPlayers() {{
        document.querySelectorAll('[data-video-player]').forEach((card) => {{
          const courseKey = String(card.dataset.videoPlayer || '');
          const course = videoTutorialCatalog[courseKey];
          if (!course || !Array.isArray(course.slides) || !course.slides.length) {{
            return;
          }}
          const state = {{
            courseTitle: String(course.title || ''),
            courseSubtitle: String(course.subtitle || ''),
            slides: course.slides,
            index: 0,
            playing: false,
            image: card.querySelector('.guide-media-image'),
            currentEl: card.querySelector('[data-video-current]'),
            titleEl: card.querySelector('[data-video-slide-title]'),
            subtitleEl: card.querySelector('[data-video-slide-subtitle]'),
            trackList: card.querySelector('[data-video-track-list]'),
            toggleButton: card.querySelector('[data-video-action=\"toggle\"]'),
            downloadButtons: Array.from(card.querySelectorAll('[data-video-action=\"download\"]')),
            downloadUrl: String(course.download_url || ''),
            exporting: false,
          }};
          if (state.toggleButton) {{
            state.toggleButton.addEventListener('click', () => {{
              if (state.playing) {{
                stopVideoTutorialNarration(state);
              }} else {{
                startVideoTutorialPresentation(state, true);
              }}
            }});
          }}
          state.downloadButtons.forEach((button) => {{
            button.addEventListener('click', () => {{
              exportVideoTutorial(state);
            }});
          }});
          renderVideoTutorialSlide(state);
          videoTutorialStates.push(state);
        }});
      }}

      initVideoTutorialPlayers();

      if (videoTutorialToggle) {{
        videoTutorialToggle.addEventListener('click', () => {{
          if (!activeVideoTutorialState) {{
            return;
          }}
          if (activeVideoTutorialState.playing) {{
            stopVideoTutorialNarration(activeVideoTutorialState);
          }} else {{
            activeVideoTutorialState.playing = true;
            updateVideoTutorialButtons(activeVideoTutorialState);
            speakVideoTutorialSlide(activeVideoTutorialState);
          }}
        }});
      }}

      if (videoTutorialRestart) {{
        videoTutorialRestart.addEventListener('click', () => {{
          if (!activeVideoTutorialState) {{
            return;
          }}
          activeVideoTutorialState.index = 0;
          renderVideoTutorialSlide(activeVideoTutorialState);
          if (activeVideoTutorialState.playing) {{
            speakVideoTutorialSlide(activeVideoTutorialState);
          }}
        }});
      }}

      if (videoTutorialDownload) {{
        videoTutorialDownload.addEventListener('click', () => {{
          if (!activeVideoTutorialState) {{
            return;
          }}
          exportVideoTutorial(activeVideoTutorialState);
        }});
      }}

      if (videoTutorialClose) {{
        videoTutorialClose.addEventListener('click', closeVideoTutorialPresentation);
      }}

      if (videoTutorialOverlay) {{
        videoTutorialOverlay.addEventListener('click', (event) => {{
          if (event.target === videoTutorialOverlay || event.target.classList.contains('guide-player-backdrop')) {{
            closeVideoTutorialPresentation();
          }}
        }});
      }}

      document.addEventListener('keydown', (event) => {{
        if (event.key === 'Escape' && activeVideoTutorialState) {{
          closeVideoTutorialPresentation();
        }}
      }});

      window.addEventListener('beforeunload', () => {{
        videoTutorialStates.forEach((state) => stopVideoTutorialNarration(state));
      }});

      document.addEventListener('visibilitychange', () => {{
        if (document.hidden) {{
          videoTutorialStates.forEach((state) => stopVideoTutorialNarration(state));
        }}
      }});
    </script>
    """
    return page("Tutorial", "/maschere/guida", content, query_params, current_user)


def gestione_utente_page(
    user_id: int,
    query_params: dict[str, str],
    current_user: dict[str, object] | None = None,
) -> bytes:
    if not current_user or not current_user.get("is_admin"):
        raise KeyError(user_id)
    user_row = access_user_row(user_id)
    if user_row is None or user_row["is_admin"]:
        raise KeyError(user_id)

    info_table = table_card(
        "Dati utente",
        "L'account puo essere aggiornato, disattivato o ricevere una nuova password, ma la password attuale resta protetta.",
        [user_row],
        [
            ("username", "Username"),
            ("email_recupero", "Email recupero"),
            ("attivo", "Attivo", lambda value, row: "Si" if int(value or 0) == 1 else "No"),
            ("creato_il", "Creato il"),
            ("ultimo_accesso", "Ultimo accesso"),
        ],
    )
    update_form = form_card(
        "Anagrafica utente",
        "Aggiorna username ed email di recupero del profilo standard.",
        f"/azioni/utenti/aggiorna/{user_id}",
        "".join(
            [
                input_field("Username", "username", required_field=True, value=user_row["username"] or ""),
                input_field("Email recupero", "email_recupero", input_type="email", required_field=True, wide=True, value=user_row["email_recupero"] or ""),
            ]
        ),
        "Salva dati",
        hidden_fields=work_year_query(query_params),
    )
    password_form = form_card(
        "Reimposta password",
        "Inserisci una nuova password per l'utente standard. La password attuale non e visualizzabile.",
        f"/azioni/utenti/password/{user_id}",
        "".join(
            [
                input_field("Nuova password", "password", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
                input_field("Conferma nuova password", "password_conferma", input_type="password", required_field=True, wide=True, revealable=True, attrs={"autocomplete": "new-password"}),
            ]
        ),
        "Aggiorna password",
        hidden_fields=work_year_query(query_params),
    )
    status_action = "riattiva" if int(user_row["attivo"] or 0) == 0 else "disattiva"
    status_label = "Riattiva utente" if status_action == "riattiva" else "Disattiva utente"
    status_form = f"""
    <section class="card">
      <div class="card-head">
        <h2>Stato utente</h2>
        <p>Puoi annullare l'utente standard disattivandolo. Se e disattivato, non potra accedere finche non lo riattivi.</p>
      </div>
      <form method="post" action="{esc(f'/azioni/utenti/stato/{user_id}')}" class="form-grid">
        {hidden_fields_html({**work_year_query(query_params), 'azione_stato': status_action})}
        <div class="form-actions">
          <button type="submit" class="button ghost">{esc(status_label)}</button>
        </div>
      </form>
    </section>
    """
    back_url = with_query("/maschere/utenti", work_year_query(query_params))
    content = f"""
    <section class="hero">
      <div>
        <span class="eyebrow">Gestione utenti</span>
        <h2>{esc(user_row['username'])}</h2>
        <p>Qui puoi modificare solo gli utenti standard. L'account amministratore resta gestito dal proprio profilo accesso.</p>
      </div>
      <div class="hero-actions">
        <a class="button ghost" href="{esc(back_url)}">Torna agli utenti</a>
      </div>
    </section>
    <div class="cards-grid">
      {update_form}
      {password_form}
    </div>
    {status_form}
    {info_table}
    """
    return page("Gestione utente", "/maschere/utenti", content, query_params, current_user)


def render_crud_edit_page(
    entity_key: str,
    record_id: int,
    query_params: dict[str, str],
    current_user: dict[str, object] | None = None,
) -> bytes:
    if entity_key in LOCKED_CRUD_ENTITIES:
        raise KeyError(entity_key)
    config = get_crud_config(entity_key)
    row = fetch_one(config["fetch_query"], (record_id,))
    if row is None:
        raise KeyError(entity_key)

    fields_html = (
        "".join(render_associato_edit_fields(row, current_user))
        if entity_key == "associati"
        else "".join(field(row) for field in config["fields"])
    )

    form = form_card(
        config["page_title"],
        config["page_subtitle"],
        f"/azioni/crud/aggiorna/{entity_key}/{record_id}",
        fields_html,
        "Salva modifiche",
        hidden_fields=current_page_query(query_params),
        form_attrs={"data-minor-guardian-form": "1"} if entity_key == "associati" else None,
    )
    delete_block = f"""
    <section class="card screen-only">
      <div class="card-head">
        <h2>Eliminazione record</h2>
        <p>Usa questa azione solo se vuoi rimuovere definitivamente il dato selezionato.</p>
      </div>
      {delete_action_form(
          f"/azioni/crud/elimina/{entity_key}/{record_id}",
          config["delete_prompt"],
          extra_fields=current_page_query(query_params),
      )}
    </section>
    """
    back_url = with_query(config["return_path"], current_page_query(query_params))
    content = f"""
    <section class="hero">
      <div>
        <span class="eyebrow">Gestione dati</span>
        <h2>{esc(config["page_title"])}</h2>
        <p>Torna alla maschera principale dopo aver salvato o eliminato il record.</p>
      </div>
      <div class="hero-actions">
        <a class="button ghost" href="{esc(back_url)}">Torna alla maschera</a>
      </div>
    </section>
    {form}
    {delete_block}
    """
    return page(config["page_title"], config["return_path"], content, query_params, current_user)


def handle_crud_update(
    entity_key: str,
    record_id: int,
    form_data: dict[str, str],
    start_response,
    current_user: dict[str, object] | None = None,
):
    if entity_key in LOCKED_CRUD_ENTITIES:
        raise ValueError(LOCKED_CRUD_ENTITIES[entity_key])
    config = get_crud_config(entity_key)
    if entity_key == "associati":
        existing_row = fetch_one("SELECT carica FROM associati WHERE id = ?", (record_id,))
        if existing_row is None:
            raise ValueError("Associato non trovato.")
        params = (
            required(form_data, "codice_associato", "Codice associato"),
            required(form_data, "nome", "Nome"),
            required(form_data, "cognome", "Cognome"),
            optional(form_data, "codice_fiscale"),
            optional(form_data, "data_nascita"),
            normalized(form_data, "sesso", "M") or "M",
            optional(form_data, "comune_nascita"),
            optional(form_data, "provincia_nascita"),
            resolved_carica_value(form_data, current_user, existing_value=existing_row["carica"] or "Associato"),
            optional(form_data, "email"),
            optional(form_data, "telefono"),
            optional(form_data, "indirizzo"),
            optional(form_data, "cap"),
            optional(form_data, "citta"),
            optional(form_data, "provincia"),
            optional(form_data, "impiego"),
            required(form_data, "data_prima_iscrizione", "Data prima iscrizione"),
            normalized(form_data, "stato_associato", "Attivo") or "Attivo",
            normalized(form_data, "liberatoria_video", "Si") or "Si",
            optional(form_data, "patologie"),
            optional(form_data, "genitore_tutore_cognome"),
            optional(form_data, "genitore_tutore_nome"),
            optional(form_data, "genitore_tutore_cellulare"),
            optional(form_data, "genitore_tutore_email"),
            optional(form_data, "genitore_tutore_impiego"),
            normalized(form_data, "genitore_tutore_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
            optional(form_data, "genitore_tutore_numero_documento"),
            optional(form_data, "prelievo_altro_genitore_nome"),
            optional(form_data, "prelievo_altro_genitore_cognome"),
            optional(form_data, "prelievo_altro_genitore_cellulare"),
            optional(form_data, "prelievo_altro_genitore_impiego"),
            normalized(form_data, "prelievo_altro_genitore_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
            optional(form_data, "prelievo_altro_genitore_numero_documento"),
            optional(form_data, "prelievo_altra_persona_nome"),
            optional(form_data, "prelievo_altra_persona_cognome"),
            optional(form_data, "prelievo_altra_persona_cellulare"),
            normalized(form_data, "prelievo_altra_persona_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
            optional(form_data, "prelievo_altra_persona_numero_documento"),
            optional(form_data, "note"),
            record_id,
        )
    elif entity_key == "corsi":
        data_inizio, data_fine = validate_course_date_range(
            required(form_data, "data_inizio", "Data inizio"),
            required(form_data, "data_fine", "Data fine"),
        )
        params = (
            required(form_data, "codice_corso", "Codice corso"),
            required(form_data, "nome", "Nome corso"),
            optional(form_data, "descrizione"),
            normalized(form_data, "quota_mensile_standard", "0"),
            data_inizio,
            data_fine,
            optional(form_data, "sede"),
            optional(form_data, "giorno_settimana"),
            optional(form_data, "orario"),
            normalized(form_data, "attivo", "1") or "1",
            optional(form_data, "note"),
            record_id,
        )
    elif entity_key == "iscrizioni_corsi":
        corso_id = int(required(form_data, "corso_id", "Corso"))
        data_iscrizione = required(form_data, "data_iscrizione", "Data iscrizione")
        with get_connection() as connection:
            effective_start, effective_end = resolve_course_enrollment_window(
                connection,
                corso_id,
                data_iscrizione,
                optional(form_data, "data_inizio"),
                optional(form_data, "data_fine"),
            )
            params = (
                required(form_data, "associato_id", "Associato"),
                corso_id,
                data_iscrizione,
                effective_start,
                effective_end,
                required(form_data, "quota_mensile", "Quota mensile"),
                normalized(form_data, "stato_iscrizione", "Attiva") or "Attiva",
                optional(form_data, "note"),
                record_id,
            )
            connection.execute(config["update_sql"], params)
            connection.commit()
        extra_query = work_year_query_from_form(form_data)
        extra_query.update(config.get("return_query", {}))
        return redirect(
            start_response,
            config["return_path"],
            ok=config["success_update"],
            extra_query=extra_query,
        )
    elif entity_key == "tesseramenti_annuali":
        existing_row = fetch_one(
            """
            SELECT anno_sociale, numero_progressivo_anno, codice_tesseramento
            FROM tesseramenti_annuali
            WHERE id = ?
            """,
            (record_id,),
        )
        if existing_row is None:
            raise ValueError("Tesseramento non trovato.")
        anno_sociale = int(required(form_data, "anno_sociale", "Anno sociale"))
        preferred_number = (
            int(existing_row["numero_progressivo_anno"])
            if existing_row["anno_sociale"] == anno_sociale and existing_row["numero_progressivo_anno"]
            else None
        )
        with get_connection() as connection:
            numero_progressivo_anno, codice_tesseramento = assign_tesseramento_identifier(
                connection,
                anno_sociale,
                preferred_number=preferred_number,
                exclude_id=record_id,
            )
            params = (
                required(form_data, "associato_id", "Associato"),
                anno_sociale,
                numero_progressivo_anno,
                codice_tesseramento,
                required(form_data, "data_tesseramento", "Data tesseramento"),
                required(form_data, "importo_dovuto", "Importo dovuto"),
                optional(form_data, "data_scadenza"),
                optional(form_data, "note"),
                record_id,
            )
            connection.execute(config["update_sql"], params)
            connection.commit()
        extra_query = work_year_query_from_form(form_data)
        extra_query.update(config.get("return_query", {}))
        return redirect(
            start_response,
            config["return_path"],
            ok=config["success_update"],
            extra_query=extra_query,
        )
    else:
        params = config["build_params"](form_data) + (record_id,)
    execute(config["update_sql"], params)
    extra_query = work_year_query_from_form(form_data)
    extra_query.update(config.get("return_query", {}))
    return redirect(
        start_response,
        config["return_path"],
        ok=config["success_update"],
        extra_query=extra_query,
    )


def handle_crud_delete(entity_key: str, record_id: int, form_data: dict[str, str], start_response):
    if entity_key in LOCKED_CRUD_ENTITIES:
        raise ValueError(LOCKED_CRUD_ENTITIES[entity_key])
    config = get_crud_config(entity_key)
    execute(config["delete_sql"], (record_id,))
    extra_query = work_year_query_from_form(form_data)
    extra_query.update(config.get("return_query", {}))
    return redirect(
        start_response,
        config["return_path"],
        ok=config["success_delete"],
        extra_query=extra_query,
    )


def course_generation_reminder_block(query_params: dict[str, str]) -> str:
    return ""


def dashboard_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    work_year = current_work_year(query_params)
    date_from, date_to = year_start_end(work_year)
    stats = [
        stat_card(
            f"Tesserati {work_year}",
            fetch_scalar(
                """
                SELECT COUNT(DISTINCT associato_id)
                FROM tesseramenti_annuali
                WHERE anno_sociale = ?
                """,
                (work_year,),
            ) or 0,
            with_query("/maschere/tesserati", work_year_query(query_params)),
        ),
        stat_card(
            f"Oratorio {work_year}",
            fetch_scalar(
                """
                SELECT COUNT(*)
                FROM iscrizioni_oratorio io
                JOIN oratorio o ON o.id = io.oratorio_id
                WHERE o.anno = ?
                """,
                (work_year,),
            ) or 0,
            with_query("/maschere/oratorio", data_view_query(query_params)),
        ),
        stat_card(
            f"Corsi {work_year}",
            fetch_scalar(
                f"""
                SELECT COUNT(*)
                FROM iscrizioni_corsi ic
                WHERE {iscrizione_corso_year_relevance_sql('ic')}
                """,
                iscrizione_corso_year_relevance_params(work_year),
            ) or 0,
            with_query("/maschere/corsi", data_view_query(query_params)),
        ),
        stat_card(
            f"Campo estivo {work_year}",
            fetch_scalar(
                """
                SELECT COUNT(*)
                FROM iscrizioni_campi_estivi ice
                JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
                WHERE ce.anno = ?
                """,
                (work_year,),
            ) or 0,
            with_query("/maschere/campi-estivi", data_view_query(query_params)),
        ),
        stat_card(
            f"Eventi {work_year}",
            fetch_scalar(
                """
                SELECT COUNT(*)
                FROM iscrizioni_eventi ie
                JOIN eventi e ON e.id = ie.evento_id
                WHERE substr(e.data_evento, 1, 4) = ?
                """,
                (str(work_year),),
            ) or 0,
            with_query("/maschere/eventi", data_view_query(query_params)),
        ),
    ]

    associati_summary = fetch_all(
        posizione_associati_query(limit=20),
        posizione_associati_params(work_year),
    )
    associati_summary_columns = tesserato_columns([
        ("codice_tesseramento", "Codice"),
        (
            "associato",
            "Associato",
            lambda value, row: report_link(value, f"/report/associato/{row['associato_id']}", work_year_query(query_params)),
        ),
        ("totale_dovuto", "Totale dovuto", lambda value, _: money(value)),
        ("totale_pagato", "Totale pagato", lambda value, _: money(value)),
        ("saldo_residuo", "Saldo residuo", lambda value, _: money(value)),
    ])

    content = f"""
    <section class="stat-grid">
      {''.join(stats)}
    </section>
    {dashboard_charts(query_params)}
    {dashboard_associati_search_toolbar()}
    {table_card(
        f"Posizione tesserati anno {work_year}",
        "Clicca sul nome per aprire il dettaglio con incassi e scadenze del singolo tesserato.",
        associati_summary,
        associati_summary_columns,
        empty_message="Nessun associato disponibile.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(associati_summary, associati_summary_columns),
    )}
    """
    return page("Dashboard", "/", content, query_params, current_user)


def consiglio_direttivo_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    work_year = current_work_year(query_params)
    placeholders = ", ".join("?" for _ in CONSIGLIO_DIRETTIVO_ORDER)
    rows = fetch_all(
        f"""
        SELECT
            id,
            numero_progressivo,
            codice_associato,
            {associato_base_name_sql('a')} AS associato,
            data_nascita,
            carica,
            COALESCE(a.telefono, '') AS telefono,
            COALESCE(a.email, '') AS email,
            COALESCE(a.citta, '') AS citta,
            stato_associato
        FROM associati a
        WHERE carica IN ({placeholders})
          AND {associato_year_relevance_sql('a')}
        ORDER BY
            CASE carica
                WHEN 'Presidente' THEN 1
                WHEN 'Vice Presidente' THEN 2
                WHEN 'Tesoriere' THEN 3
                WHEN 'Segretario' THEN 4
                WHEN 'Consigliere' THEN 5
                WHEN 'Consigliere spirituale' THEN 6
                ELSE 99
            END,
            cognome,
            nome
        """,
        (*CONSIGLIO_DIRETTIVO_ORDER, *associato_year_relevance_params(work_year)),
    )
    grouped_rows = {role: [] for role in CONSIGLIO_DIRETTIVO_ORDER}
    for row in rows:
        grouped_rows[row["carica"]].append(row)

    cards_html = []
    for role in CONSIGLIO_DIRETTIVO_ORDER:
        role_rows = grouped_rows.get(role, [])
        members_html = ""
        if role_rows:
            member_cards: list[str] = []
            for row in role_rows:
                metadata: list[str] = []
                display_name = esc(label_with_age(row["associato"], row["data_nascita"]))
                recapiti_parts: list[str] = []
                if row["telefono"]:
                    recapiti_parts.append(
                        f'<span class="direttivo-member-inline-item"><strong>Telefono</strong><span>{esc(row["telefono"])}</span></span>'
                    )
                if row["email"]:
                    recapiti_parts.append(
                        f'<span class="direttivo-member-inline-item"><strong>Email</strong><span>{esc(row["email"])}</span></span>'
                    )
                if recapiti_parts:
                    metadata.append(
                        f'<div class="direttivo-member-fact direttivo-member-fact-contacts is-wide"><div class="direttivo-member-inline">{"".join(recapiti_parts)}</div></div>'
                    )
                actions_html = ""
                if can_manage_cariche(current_user):
                    actions_html = (
                        f'<div class="direttivo-actions">'
                        f'{action_links_html(edit_href=edit_path("associati", row["id"], query_params), extra_fields=work_year_query(query_params))}'
                        f"</div>"
                    )
                member_cards.append(
                    f"""
                    <article class="direttivo-member">
                      <div class="direttivo-member-copy">
                        <header class="direttivo-member-header">
                          <h3>{display_name}</h3>
                        </header>
                        <div class="direttivo-member-meta">
                          {''.join(metadata) if metadata else '<div class="direttivo-member-fact direttivo-member-fact-contacts is-empty is-wide"><div class="direttivo-member-empty-note">Nessun recapito inserito</div></div>'}
                        </div>
                        {actions_html}
                      </div>
                    </article>
                    """
                )
            members_html = "".join(member_cards)
        else:
            members_html = '<div class="empty-state">Nessun associato con questa carica.</div>'

        cards_html.append(
            f"""
            <section class="card direttivo-role-card">
              <div class="direttivo-role-head">
                <h2>{esc(CONSIGLIO_DIRETTIVO_LABELS.get(role, role))}</h2>
                <div class="direttivo-role-accent"></div>
              </div>
              <div class="direttivo-role-body">
                {members_html}
              </div>
            </section>
            """
        )

    content = f"""
    <div class="cards-grid consiglio-direttivo-grid">
      {''.join(cards_html)}
    </div>
    """
    return page("Consiglio Direttivo", "/maschere/consiglio-direttivo", content, query_params, current_user)


def tesserati_dataset_rows(work_year: int) -> list[dict[str, object]]:
    source_rows = fetch_all(
        f"""
        SELECT
            t.codice_tesseramento,
            {associato_display_sql('a')} AS tesserato,
            COALESCE(a.codice_fiscale, '') AS codice_fiscale,
            COALESCE(a.data_nascita, '') AS data_nascita,
            COALESCE(a.sesso, '') AS sesso,
            COALESCE(a.comune_nascita, '') AS comune_nascita,
            COALESCE(a.provincia_nascita, '') AS provincia_nascita,
            COALESCE(a.telefono, '') AS telefono,
            COALESCE(a.email, '') AS email,
            COALESCE(a.liberatoria_video, '') AS liberatoria_video,
            COALESCE(a.patologie, '') AS patologie,
            COALESCE(a.genitore_tutore_nome, '') AS genitore_tutore_nome,
            COALESCE(a.genitore_tutore_cognome, '') AS genitore_tutore_cognome,
            COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
            COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
            COALESCE(a.genitore_tutore_impiego, '') AS genitore_tutore_impiego,
            COALESCE(a.genitore_tutore_tipo_documento, 'Carta d''identitÃ ') AS genitore_tutore_tipo_documento,
            COALESCE(a.genitore_tutore_numero_documento, '') AS genitore_tutore_numero_documento,
            COALESCE(a.prelievo_altro_genitore_nome, '') AS prelievo_altro_genitore_nome,
            COALESCE(a.prelievo_altro_genitore_cognome, '') AS prelievo_altro_genitore_cognome,
            COALESCE(a.prelievo_altro_genitore_cellulare, '') AS prelievo_altro_genitore_cellulare,
            COALESCE(a.prelievo_altro_genitore_impiego, '') AS prelievo_altro_genitore_impiego,
            COALESCE(a.prelievo_altro_genitore_tipo_documento, 'Carta d''identitÃ ') AS prelievo_altro_genitore_tipo_documento,
            COALESCE(a.prelievo_altro_genitore_numero_documento, '') AS prelievo_altro_genitore_numero_documento,
            COALESCE(a.prelievo_altra_persona_nome, '') AS prelievo_altra_persona_nome,
            COALESCE(a.prelievo_altra_persona_cognome, '') AS prelievo_altra_persona_cognome,
            COALESCE(a.prelievo_altra_persona_cellulare, '') AS prelievo_altra_persona_cellulare,
            COALESCE(a.prelievo_altra_persona_tipo_documento, 'Carta d''identitÃ ') AS prelievo_altra_persona_tipo_documento,
            COALESCE(a.prelievo_altra_persona_numero_documento, '') AS prelievo_altra_persona_numero_documento,
            COALESCE(a.indirizzo, '') AS indirizzo,
            COALESCE(a.cap, '') AS cap,
            COALESCE(a.citta, '') AS citta,
            COALESCE(a.provincia, '') AS provincia,
            COALESCE(a.impiego, '') AS impiego,
            COALESCE(a.carica, '') AS carica,
            CASE
                WHEN COALESCE(vts.saldo_residuo, t.importo_dovuto, 0) <= 0 THEN 'Concluso'
                ELSE 'Attivo'
            END AS stato_associato,
            COALESCE(t.data_tesseramento, '') AS data_tesseramento,
            COALESCE(t.data_scadenza, '') AS data_scadenza
        FROM tesseramenti_annuali t
        JOIN associati a ON a.id = t.associato_id
        LEFT JOIN v_tesseramenti_saldo vts ON vts.id = t.id
        WHERE t.anno_sociale = ?
        ORDER BY t.numero_progressivo_anno, a.cognome, a.nome
        """,
        (work_year,),
    )
    rows: list[dict[str, object]] = []
    for row in source_rows:
        item = dict(row)
        age = calculate_age(item.get("data_nascita"))
        item["eta"] = "" if age is None else str(age)
        item["eta_legale"] = "" if age is None else ("Maggiorenne" if age >= 18 else "Minorenne")
        rows.append(item)
    return rows


def tesserati_columns_definition() -> list[tuple]:
    return [
        ("codice_tesseramento", "Codice"),
        ("tesserato", "Tesserato"),
        ("eta", "Eta"),
        ("eta_legale", "EtÃ  legale"),
        ("codice_fiscale", "Codice fiscale"),
        ("data_nascita", "Data nascita"),
        ("sesso", "Sesso"),
        ("comune_nascita", "Comune nascita"),
        ("provincia_nascita", "Prov. nascita"),
        ("telefono", "Cellulare"),
        ("email", "Email"),
        ("liberatoria_video", "Liberatoria Video"),
        ("patologie", "Patologie / allergie / terapie"),
        ("genitore_tutore_nome", "Genitore/Tutore nome"),
        ("genitore_tutore_cognome", "Genitore/Tutore cognome"),
        ("genitore_tutore_cellulare", "Genitore/Tutore cellulare"),
        ("genitore_tutore_email", "Genitore/Tutore email"),
        ("genitore_tutore_impiego", "Genitore/Tutore impiego"),
        ("genitore_tutore_tipo_documento", "Genitore/Tutore tipo documento"),
        ("genitore_tutore_numero_documento", "Genitore/Tutore numero documento"),
        ("prelievo_altro_genitore_nome", "Altro genitore nome"),
        ("prelievo_altro_genitore_cognome", "Altro genitore cognome"),
        ("prelievo_altro_genitore_cellulare", "Altro genitore cellulare"),
        ("prelievo_altro_genitore_impiego", "Altro genitore impiego"),
        ("prelievo_altro_genitore_tipo_documento", "Altro genitore tipo documento"),
        ("prelievo_altro_genitore_numero_documento", "Altro genitore numero documento"),
        ("prelievo_altra_persona_nome", "Altra persona nome"),
        ("prelievo_altra_persona_cognome", "Altra persona cognome"),
        ("prelievo_altra_persona_cellulare", "Altra persona cellulare"),
        ("prelievo_altra_persona_tipo_documento", "Altra persona tipo documento"),
        ("prelievo_altra_persona_numero_documento", "Altra persona numero documento"),
        ("indirizzo", "Indirizzo"),
        ("cap", "CAP"),
        ("citta", "CittÃ "),
        ("provincia", "Provincia"),
        ("impiego", "Impiego"),
        ("carica", "Carica"),
        ("stato_associato", "Stato"),
        ("data_tesseramento", "Data tesseramento"),
        ("data_scadenza", "Scadenza"),
    ]


def parse_table_export_state(query_params: dict[str, str]) -> dict[str, object]:
    visible_columns = [
        item.strip()
        for item in (query_params.get("columns") or "").split(",")
        if item.strip()
    ]
    column_order = [
        item.strip()
        for item in (query_params.get("column_order") or "").split(",")
        if item.strip()
    ]
    column_filters = {
        key.removeprefix("cf_"): value.strip()
        for key, value in query_params.items()
        if key.startswith("cf_") and value.strip()
    }
    return {
        "visible_columns": visible_columns,
        "column_order": column_order,
        "sort_column": normalized(query_params, "sort_column", ""),
        "sort_direction": normalized(query_params, "sort_direction", "asc").lower(),
        "global_search": normalized(query_params, "search", ""),
        "column_filters": column_filters,
    }


def apply_table_export_state(
    definition: dict[str, object],
    rows: list[dict[str, object]],
    query_params: dict[str, str],
) -> tuple[dict[str, object], list[dict[str, object]]]:
    state = parse_table_export_state(query_params)
    original_columns = list(definition.get("columns", []))
    if not original_columns:
        return definition, rows

    by_key = {
        str(column[0] or "").strip(): column
        for column in original_columns
        if str(column[0] or "").strip()
    }
    ordered_keys = [key for key in state["column_order"] if key in by_key]
    ordered_columns = [by_key[key] for key in ordered_keys]
    ordered_columns.extend(
        column for column in original_columns if str(column[0] or "").strip() not in ordered_keys
    )

    visible_set = set(state["visible_columns"])
    if visible_set:
        filtered_columns = [
            column for column in ordered_columns if str(column[0] or "").strip() in visible_set
        ]
        if filtered_columns:
            ordered_columns = filtered_columns

    filterable_columns = {
        str(column[0] or "").strip(): column
        for column in ordered_columns
        if len(column) >= 2 and str(column[1] or "").lower() not in {"azioni", "ricevuta"}
    }
    searchable_columns = list(filterable_columns.values())

    def row_text(row: dict[str, object], column: tuple) -> str:
        return normalize_searchable_text(report_display_value(row, column))

    def row_value(row: object, key: str):
        if isinstance(row, dict):
            return row.get(key)
        try:
            return row[key]
        except (KeyError, IndexError, TypeError):
            return None

    filtered_rows = list(rows)
    global_search = normalize_searchable_text(str(state["global_search"] or ""))
    if global_search:
        filtered_rows = [
            row
            for row in filtered_rows
            if any(global_search in row_text(row, column) for column in searchable_columns)
        ]

    for key, value in state["column_filters"].items():
        column = filterable_columns.get(key)
        normalized_filter = normalize_searchable_text(value)
        if column is None or not normalized_filter:
            continue
        filtered_rows = [
            row
            for row in filtered_rows
            if normalized_filter in row_text(row, column)
        ]

    sort_column_key = str(state["sort_column"] or "").strip()
    sort_direction = "desc" if str(state["sort_direction"] or "").lower() == "desc" else "asc"
    if sort_column_key and sort_column_key in by_key:
        sort_column = by_key[sort_column_key]
        sort_index = next(
            (index for index, column in enumerate(ordered_columns) if str(column[0] or "").strip() == sort_column_key),
            0,
        )
        total_index_set = set(summary_total_indexes(ordered_columns))
        sort_type = column_sort_type(sort_column, sort_index, total_index_set)

        def row_sort_key(row: dict[str, object]):
            value = row_value(row, sort_column_key)
            if sort_type == "number":
                amount = decimal_or_none(value)
                if amount is not None:
                    return float(amount)
                text_value = normalize_searchable_text(report_display_value(row, sort_column))
                try:
                    return float(text_value.replace(",", "."))
                except ValueError:
                    return float("-inf")
            return normalize_searchable_text(report_display_value(row, sort_column))

        filtered_rows = sorted(filtered_rows, key=row_sort_key, reverse=(sort_direction == "desc"))

    active_filters = list(definition.get("filters", []))
    if global_search:
        active_filters.append({"label": "Ricerca", "value": str(state["global_search"])})
    for key, value in state["column_filters"].items():
        column = by_key.get(key)
        if column is not None and value:
            active_filters.append({"label": str(column[1]), "value": str(value)})

    return {**definition, "columns": ordered_columns, "filters": active_filters}, filtered_rows


def build_tesserati_export_definition(query_params: dict[str, str]) -> tuple[dict[str, object], list[dict[str, object]]]:
    work_year = current_work_year(query_params)
    definition = {
        "title": "Tesserati",
        "subtitle": f"Anagrafica dei tesserati per l'anno di lavoro {work_year}.",
        "sheet_name": "Tesserati",
        "export_name": f"tesserati_{work_year}.xlsx",
        "filters": [{"label": "Anno di lavoro", "value": str(work_year)}],
        "columns": tesserati_columns_definition(),
        "empty_message": "Nessun tesserato disponibile.",
    }
    return apply_table_export_state(definition, tesserati_dataset_rows(work_year), query_params)


def tesserati_toolbar(query_params: dict[str, str]) -> str:
    excel_url = with_query("/export/excel/tesserati", work_year_query(query_params))
    pdf_url = with_query("/export/pdf/tesserati", work_year_query(query_params))
    table_id = "tesserati-anagrafica-table"
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Azioni tesserati</span>
        <p>Esporta l'anagrafica tesserati in Excel o PDF, oppure stampala direttamente dal browser.</p>
      </div>
      <div class="report-toolbar-actions">
        <div class="report-toolbar-action-row">
          <a class="button action" href="{esc(excel_url)}" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta Excel</a>
          <a class="button action" href="{esc(pdf_url)}" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta PDF</a>
          <button type="button" class="button action" onclick="window.print()">Stampa report</button>
        </div>
      </div>
    </section>
    """


def associati_storici_dataset_rows(work_year: int) -> list[dict[str, object]]:
    source_rows = fetch_all(
        f"""
        SELECT
            associati.id,
            associati.codice_associato,
            {associato_display_sql('associati')} AS associato,
            COALESCE(associati.codice_fiscale, '') AS codice_fiscale,
            COALESCE(associati.data_nascita, '') AS data_nascita,
            COALESCE(associati.sesso, '') AS sesso,
            COALESCE(associati.comune_nascita, '') AS comune_nascita,
            COALESCE(associati.provincia_nascita, '') AS provincia_nascita,
            COALESCE(associati.telefono, '') AS telefono,
            COALESCE(associati.email, '') AS email,
            COALESCE(associati.liberatoria_video, '') AS liberatoria_video,
            COALESCE(associati.patologie, '') AS patologie,
            COALESCE(associati.genitore_tutore_nome, '') AS genitore_tutore_nome,
            COALESCE(associati.genitore_tutore_cognome, '') AS genitore_tutore_cognome,
            COALESCE(associati.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
            COALESCE(associati.genitore_tutore_email, '') AS genitore_tutore_email,
            COALESCE(associati.genitore_tutore_impiego, '') AS genitore_tutore_impiego,
            COALESCE(associati.genitore_tutore_tipo_documento, 'Carta d''identitÃ ') AS genitore_tutore_tipo_documento,
            COALESCE(associati.genitore_tutore_numero_documento, '') AS genitore_tutore_numero_documento,
            COALESCE(associati.prelievo_altro_genitore_nome, '') AS prelievo_altro_genitore_nome,
            COALESCE(associati.prelievo_altro_genitore_cognome, '') AS prelievo_altro_genitore_cognome,
            COALESCE(associati.prelievo_altro_genitore_cellulare, '') AS prelievo_altro_genitore_cellulare,
            COALESCE(associati.prelievo_altro_genitore_impiego, '') AS prelievo_altro_genitore_impiego,
            COALESCE(associati.prelievo_altro_genitore_tipo_documento, 'Carta d''identitÃ ') AS prelievo_altro_genitore_tipo_documento,
            COALESCE(associati.prelievo_altro_genitore_numero_documento, '') AS prelievo_altro_genitore_numero_documento,
            COALESCE(associati.prelievo_altra_persona_nome, '') AS prelievo_altra_persona_nome,
            COALESCE(associati.prelievo_altra_persona_cognome, '') AS prelievo_altra_persona_cognome,
            COALESCE(associati.prelievo_altra_persona_cellulare, '') AS prelievo_altra_persona_cellulare,
            COALESCE(associati.prelievo_altra_persona_tipo_documento, 'Carta d''identitÃ ') AS prelievo_altra_persona_tipo_documento,
            COALESCE(associati.prelievo_altra_persona_numero_documento, '') AS prelievo_altra_persona_numero_documento,
            COALESCE(associati.indirizzo, '') AS indirizzo,
            COALESCE(associati.cap, '') AS cap,
            COALESCE(associati.citta, '') AS citta,
            COALESCE(associati.provincia, '') AS provincia,
            COALESCE(associati.impiego, '') AS impiego,
            COALESCE(associati.carica, '') AS carica,
            CASE
                WHEN t.id IS NULL THEN 'Sospeso'
                WHEN COALESCE(vts.saldo_residuo, t.importo_dovuto, 0) <= 0 THEN 'Concluso'
                ELSE 'Attivo'
            END AS stato_associato,
            COALESCE(associati.data_prima_iscrizione, '') AS data_prima_iscrizione
        FROM associati
        LEFT JOIN tesseramenti_annuali t
          ON t.associato_id = associati.id
         AND t.anno_sociale = ?
        LEFT JOIN v_tesseramenti_saldo vts ON vts.id = t.id
        ORDER BY associati.cognome, associati.nome
        """,
        (work_year,),
    )
    rows: list[dict[str, object]] = []
    for row in source_rows:
        item = dict(row)
        age = calculate_age(item.get("data_nascita"))
        item["eta"] = "" if age is None else str(age)
        item["eta_legale"] = "" if age is None else ("Maggiorenne" if age >= 18 else "Minorenne")
        rows.append(item)
    return rows


def associati_storici_columns_definition(*, include_actions: bool = False, query_params: dict[str, str] | None = None) -> list[tuple]:
    columns: list[tuple] = [
        ("codice_associato", "Codice"),
        ("associato", "Associato"),
        ("eta", "Eta"),
        ("eta_legale", "EtÃ  legale"),
        ("codice_fiscale", "Codice fiscale"),
        ("data_nascita", "Data nascita"),
        ("sesso", "Sesso"),
        ("comune_nascita", "Comune nascita"),
        ("provincia_nascita", "Prov. nascita"),
        ("telefono", "Cellulare"),
        ("email", "Email"),
        ("liberatoria_video", "Liberatoria Video"),
        ("patologie", "Patologie / allergie / terapie"),
        ("genitore_tutore_nome", "Genitore/Tutore nome"),
        ("genitore_tutore_cognome", "Genitore/Tutore cognome"),
        ("genitore_tutore_cellulare", "Genitore/Tutore cellulare"),
        ("genitore_tutore_email", "Genitore/Tutore email"),
        ("genitore_tutore_impiego", "Genitore/Tutore impiego"),
        ("genitore_tutore_tipo_documento", "Genitore/Tutore tipo documento"),
        ("genitore_tutore_numero_documento", "Genitore/Tutore numero documento"),
        ("prelievo_altro_genitore_nome", "Altro genitore nome"),
        ("prelievo_altro_genitore_cognome", "Altro genitore cognome"),
        ("prelievo_altro_genitore_cellulare", "Altro genitore cellulare"),
        ("prelievo_altro_genitore_impiego", "Altro genitore impiego"),
        ("prelievo_altro_genitore_tipo_documento", "Altro genitore tipo documento"),
        ("prelievo_altro_genitore_numero_documento", "Altro genitore numero documento"),
        ("prelievo_altra_persona_nome", "Altra persona nome"),
        ("prelievo_altra_persona_cognome", "Altra persona cognome"),
        ("prelievo_altra_persona_cellulare", "Altra persona cellulare"),
        ("prelievo_altra_persona_tipo_documento", "Altra persona tipo documento"),
        ("prelievo_altra_persona_numero_documento", "Altra persona numero documento"),
        ("indirizzo", "Indirizzo"),
        ("cap", "CAP"),
        ("citta", "CittÃ "),
        ("provincia", "Provincia"),
        ("impiego", "Impiego"),
        ("carica", "Carica"),
        ("stato_associato", "Stato"),
        ("data_prima_iscrizione", "Data prima iscrizione"),
    ]
    if include_actions and query_params is not None:
        columns.append(
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("associati", value, query_params),
                    delete_action=f"/azioni/crud/elimina/associati/{value}",
                    delete_prompt="Eliminare questo associato e tutti i dati collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            )
        )
    return columns


def build_associati_storici_export_definition(query_params: dict[str, str]) -> tuple[dict[str, object], list[dict[str, object]]]:
    work_year = current_work_year(query_params)
    definition = {
        "title": "Anagrafica storica associati",
        "subtitle": "Elenco completo dell'anagrafica associati registrata nel gestionale.",
        "sheet_name": "Associati storici",
        "export_name": "anagrafica_storica_associati.xlsx",
        "filters": [{"label": "Anno di lavoro", "value": str(work_year)}],
        "columns": associati_storici_columns_definition(),
        "empty_message": "Nessun associato disponibile.",
    }
    return apply_table_export_state(definition, associati_storici_dataset_rows(work_year), query_params)


def associati_storici_toolbar(query_params: dict[str, str]) -> str:
    excel_url = with_query("/export/excel/anagrafica-associati", work_year_query(query_params))
    pdf_url = with_query("/export/pdf/anagrafica-associati", work_year_query(query_params))
    table_id = "associati-storici-table"
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Azioni anagrafica</span>
        <p>Esporta l'anagrafica storica associati in Excel o PDF, oppure stampala direttamente dal browser.</p>
      </div>
      <div class="report-toolbar-actions">
        <div class="report-toolbar-action-row">
          <a class="button action" href="{esc(excel_url)}" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta Excel</a>
          <a class="button action" href="{esc(pdf_url)}" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta PDF</a>
          <button type="button" class="button action" onclick="window.print()">Stampa report</button>
        </div>
      </div>
    </section>
    """


def tesserati_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    work_year = current_work_year(query_params)
    rows = tesserati_dataset_rows(work_year)
    columns = tesserati_columns_definition()
    table_id = "tesserati-anagrafica-table"
    content = f"""
    {tesserati_toolbar(query_params)}
    {data_view_search_toolbar(table_id)}
    {column_visibility_toolbar(
        "Colonne tesserati",
        "Scegli quali dati anagrafici del tesserato visualizzare nella tabella.",
        table_id,
        columns,
        default_visible_keys={"codice_tesseramento", "tesserato", "eta", "eta_legale", "telefono", "email", "liberatoria_video", "genitore_tutore_nome", "genitore_tutore_cognome", "genitore_tutore_cellulare", "carica", "stato_associato", "data_tesseramento", "data_scadenza"},
        allow_reorder=True,
    )}
    {table_card(
        f"Tesserati anno {work_year}",
        "Visualizzazione anagrafica dei soggetti tesserati nell'anno di lavoro selezionato.",
        rows,
        columns,
        empty_message="Nessun tesserato disponibile per l'anno selezionato.",
        table_class="search-table",
        table_id=table_id,
        column_filters_in_header=True,
        draggable_columns=True,
    )}
    """
    return page("Tesserati", "/maschere/tesserati", content, query_params, current_user)


def associati_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    tesseramento_auto_importo = default_tesseramento_quota_importo()
    tesseramento_auto_codice = peek_next_tesseramento_code(work_year)
    associati = associati_storici_dataset_rows(work_year)
    minor_section_attrs = {"data-minor-only": "true"}
    minor_required_input = {"data-minor-required": "true"}

    associati_form = form_card(
        "Nuovo associato",
        f"Inserisci l'anagrafica di base del socio. Al salvataggio verra creato automaticamente anche il tesseramento {work_year}.",
        "/azioni/associati/crea",
        "".join(
            [
                readonly_field("Codice associato assegnato", peek_next_progressive_code("associati"), wide=True),
                input_field("Cognome", "cognome", required_field=True),
                input_field("Nome", "nome", required_field=True),
                input_field(
                    "Codice fiscale",
                    "codice_fiscale",
                    attrs={
                        "maxlength": "16",
                        "data-codice-fiscale": "true",
                        "autocomplete": "off",
                        "autocapitalize": "characters",
                        "spellcheck": "false",
                    },
                ),
                input_field("Data nascita", "data_nascita", input_type="date"),
                select_field("Sesso", "sesso", sesso_options("M")),
                inline_fields_row([
                    input_field("Comune di nascita", "comune_nascita"),
                    input_field("Provincia di nascita", "provincia_nascita", attrs={"maxlength": "2", "autocapitalize": "characters"}),
                ]),
                inline_fields_row([
                    input_field("Cellulare", "telefono"),
                    input_field("Email", "email", input_type="email"),
                ], row_class="cell-email-row"),
                input_field("Indirizzo", "indirizzo", wide=True),
                input_field("CAP", "cap", attrs={"maxlength": "5", "inputmode": "numeric"}),
                input_field("CittÃ ", "citta"),
                input_field("Provincia", "provincia", attrs={"maxlength": "2", "autocapitalize": "characters"}),
                input_field("Impiego", "impiego"),
                input_field(
                    "Data prima iscrizione",
                    "data_prima_iscrizione",
                    input_type="date",
                    value=date.today().isoformat(),
                    required_field=True,
                ),
                associato_carica_field("Associato", current_user),
                select_field(
                    "Liberatoria Video",
                    "liberatoria_video",
                    liberatoria_video_options("Si"),
                ),
                textarea_field(
                    "Patologie, allergie, intolleranze alimentari ed eventuali terapie in corso",
                    "patologie",
                ),
                textarea_field("Note", "note"),
                form_section_block(
                    "Genitore/Tutore",
                    "Compila questi dati solo se il tesserato e minorenne.",
                    attrs=minor_section_attrs,
                ),
                input_field("Nome", "genitore_tutore_nome", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
                input_field("Cognome", "genitore_tutore_cognome", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
                input_field("Cellulare", "genitore_tutore_cellulare", attrs=minor_required_input, wrapper_attrs=minor_section_attrs),
                input_field("Email", "genitore_tutore_email", input_type="email", wrapper_attrs=minor_section_attrs),
                select_field(
                    "Tipo documento",
                    "genitore_tutore_tipo_documento",
                    document_type_options(DEFAULT_DOCUMENT_TYPE),
                    wrapper_attrs=minor_section_attrs,
                ),
                input_field("Numero documento", "genitore_tutore_numero_documento", wrapper_attrs=minor_section_attrs),
                input_field("Impiego", "genitore_tutore_impiego", wrapper_attrs=minor_section_attrs),
                form_section_block(
                    "Altri autorizzati al prelievo all'uscita",
                    "Indica i soggetti autorizzati al prelievo del minore.",
                    attrs=minor_section_attrs,
                ),
                form_section_block("Altro genitore", attrs=minor_section_attrs),
                input_field("Nome", "prelievo_altro_genitore_nome", wrapper_attrs=minor_section_attrs),
                input_field("Cognome", "prelievo_altro_genitore_cognome", wrapper_attrs=minor_section_attrs),
                input_field("Cellulare", "prelievo_altro_genitore_cellulare", wrapper_attrs=minor_section_attrs),
                input_field("Impiego", "prelievo_altro_genitore_impiego", wrapper_attrs=minor_section_attrs),
                select_field(
                    "Tipo documento",
                    "prelievo_altro_genitore_tipo_documento",
                    document_type_options(DEFAULT_DOCUMENT_TYPE),
                    wrapper_attrs=minor_section_attrs,
                ),
                input_field("Numero documento", "prelievo_altro_genitore_numero_documento", wrapper_attrs=minor_section_attrs),
                form_section_block("Altra persona", attrs=minor_section_attrs),
                input_field("Nome", "prelievo_altra_persona_nome", wrapper_attrs=minor_section_attrs),
                input_field("Cognome", "prelievo_altra_persona_cognome", wrapper_attrs=minor_section_attrs),
                input_field("Cellulare", "prelievo_altra_persona_cellulare", wide=True, wrapper_attrs=minor_section_attrs),
                select_field(
                    "Tipo documento",
                    "prelievo_altra_persona_tipo_documento",
                    document_type_options(DEFAULT_DOCUMENT_TYPE),
                    wrapper_attrs=minor_section_attrs,
                ),
                input_field("Numero documento", "prelievo_altra_persona_numero_documento", wrapper_attrs=minor_section_attrs),
            ]
        ),
        "Salva associato",
        hidden_fields=work_year_query(query_params),
        form_attrs={
            "data-payment-flow": "associato-tesseramento",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-default-amount": tesseramento_auto_importo,
            "data-payment-allow-extra-scadenze": "0",
            "data-payment-prompt-title": "Conferma nuovo associato",
            "data-payment-prompt-message": f"Dopo il salvataggio verra creato automaticamente anche il tesseramento {work_year}. Vuoi procedere subito anche al pagamento?",
            "data-payment-prompt-code": tesseramento_auto_codice,
            "data-payment-prompt-code-label": "Numero tesseramento assegnato",
            "data-payment-prompt-yes": "Si, registra anche il pagamento",
            "data-payment-prompt-no": "No, solo iscrizione",
            "data-payment-dialog-title": "Pagamento tesseramento",
            "data-payment-dialog-message": f"Conferma i dati del pagamento del tesseramento {work_year}.",
            "data-payment-dialog-code": "",
            "data-payment-dialog-code-label": "",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
            "data-minor-guardian-form": "1",
        },
    )

    forms_html = f"""
    <div class="cards-grid">
      {associati_form}
    </div>
    """ if not data_only else ""

    storico_columns = associati_storici_columns_definition(include_actions=True, query_params=query_params)
    storico_table_id = "associati-storici-table"
    tables_html = f"""
    {associati_storici_toolbar(query_params)}
    {data_view_search_toolbar(storico_table_id)}
    {column_visibility_toolbar(
        "Colonne associati",
        "Scegli quali dati dell'anagrafica storica associati visualizzare nella tabella.",
        storico_table_id,
        storico_columns,
        default_visible_keys={"codice_associato", "associato", "eta", "eta_legale", "telefono", "email", "liberatoria_video", "genitore_tutore_nome", "genitore_tutore_cognome", "genitore_tutore_cellulare", "carica", "stato_associato", "data_prima_iscrizione"},
        allow_reorder=True,
    )}
    {table_card(
        "Anagrafica storica associati",
        "Elenco completo dell'anagrafica storica associati registrata nel gestionale.",
        associati,
        storico_columns,
        table_class="search-table",
        table_id=storico_table_id,
        column_filters_in_header=True,
        draggable_columns=True,
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/associati", query_params, "Apri dati associati")}
    {forms_html}
    {tables_html if data_only else ""}
    """
    return page("Anagrafica storica associati" if data_only else "Associati", "/maschere/associati", content, query_params, current_user)


def tesseramenti_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    year = str(current_work_year(query_params))
    associati = associati_options()
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    tesseramenti = tesseramenti_aperti_options(int(year))
    quote_tesseramenti = quote_predefinite_options("tesseramenti")
    quota_tesseramento_default = quote_tesseramenti[0] if len(quote_tesseramenti) == 1 else None
    quota_tesseramento_default_id = str(quota_tesseramento_default["id"]) if quota_tesseramento_default else None
    quota_tesseramento_default_importo = str(quota_tesseramento_default["importo"]) if quota_tesseramento_default else ""
    tesseramento_scadenza_default = f"{year}-12-31"
    selected_associato_id = normalized(query_params, "associato_id", "")
    quote_tesseramenti_table = quote_predefinite_rows("tesseramenti")
    tesseramenti_table = fetch_all(
        f"""
        SELECT
            t.id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            t.anno_sociale,
            t.data_tesseramento,
            t.importo_dovuto,
            t.data_scadenza
        FROM tesseramenti_annuali t
        JOIN associati a ON a.id = t.associato_id
        WHERE t.anno_sociale = ?
        ORDER BY a.cognome, a.nome
        """,
        (int(year),),
    )
    payments_table = fetch_all(
        f"""
        SELECT
            pt.id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            t.anno_sociale,
            pt.data_pagamento,
            pt.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(pt.riferimento, '') AS riferimento
        FROM pagamenti_tesseramenti pt
        JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
        JOIN associati a ON a.id = t.associato_id
        LEFT JOIN metodi_pagamento mp ON mp.id = pt.metodo_pagamento_id
        WHERE t.anno_sociale = ?
        ORDER BY pt.data_pagamento DESC, associato
        """,
        (int(year),),
    )
    saldo = fetch_all(
        """
        SELECT codice_tesseramento, associato, anno_sociale, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
        FROM v_tesseramenti_saldo
        WHERE anno_sociale = ?
        ORDER BY anno_sociale DESC, associato
        """,
        (int(year),),
    )

    create_form = form_card(
        "Rinnovo Tesseramento",
        "Registra la quota annuale per un tesserato richiamando, se vuoi, una quota predefinita.",
        "/azioni/tesseramenti/crea",
        "".join(
            [
                readonly_field(
                    "Codice tesseramento assegnato",
                    peek_next_tesseramento_code(int(year)),
                    wide=True,
                    element_id="tesseramento-codice-preview",
                ),
                readonly_field("Anno di lavoro", year),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati, selected_associato_id),
                    required_field=True,
                    wide=True,
                    searchable=True,
                ),
                select_field(
                    "Quota tesseramento",
                    "quota_predefinita_id",
                    render_select_options(quote_tesseramenti, selected=quota_tesseramento_default_id, data_keys=["importo"]),
                    wide=True,
                    element_id="tesseramento-quota-select",
                    attrs={"data-amount-target": "tesseramento-importo-dovuto"},
                ),
                input_field("Data tesseramento", "data_tesseramento", input_type="date", value=date.today().isoformat(), required_field=True),
                input_field(
                    "Importo dovuto",
                    "importo_dovuto",
                    input_type="number",
                    step="0.01",
                    minimum="0",
                    required_field=True,
                    value=quota_tesseramento_default_importo,
                    element_id="tesseramento-importo-dovuto",
                ),
                input_field(
                    "Data scadenza",
                    "data_scadenza",
                    input_type="date",
                    value=tesseramento_scadenza_default,
                    element_id="tesseramento-data-scadenza",
                ),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva tesseramento",
        hidden_fields={**work_year_query(query_params), "anno_sociale": year},
        form_attrs={
            "data-payment-flow": "tesseramento",
            "data-payment-amount-field": "importo_dovuto",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-allow-extra-scadenze": "0",
            "data-payment-prompt-title": "Conferma tesseramento",
            "data-payment-prompt-message": "Vuoi procedere subito anche al pagamento del tesseramento?",
            "data-payment-prompt-yes": "Si, registra anche il pagamento",
            "data-payment-prompt-no": "No, solo tesseramento",
            "data-payment-dialog-title": "Pagamento tesseramento",
            "data-payment-dialog-message": "Conferma i dati del pagamento del tesseramento.",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
        },
    )

    quote_form = form_card(
        "Quota tesseramento",
        "Inserisci una quota predefinita con descrizione e importo da richiamare in fase di rinnovo.",
        "/azioni/quote/crea",
        "".join(
            [
                input_field("Descrizione", "descrizione", required_field=True, wide=True),
                input_field("Importo", "importo", input_type="number", step="0.01", minimum="0", required_field=True),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva quota",
        hidden_fields={**work_year_query(query_params), "area": "tesseramenti"},
    )

    forms_html = f"""
    <div class="cards-grid cards-stack">
      {create_form}
      {quote_form}
    </div>
    """ if not data_only else ""

    tables_html = f"""
    {table_card(
        f"Tesseramenti anno {year}",
        "Anagrafica dei tesseramenti inseriti nell'anno di lavoro selezionato.",
        tesseramenti_table,
        with_codice_tesserato([
            ("codice_tesseramento", "Codice"),
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("data_tesseramento", "Data"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("data_scadenza", "Scadenza"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("tesseramenti_annuali", value, query_params),
                    delete_action=f"/azioni/crud/elimina/tesseramenti_annuali/{value}",
                    delete_prompt="Eliminare questo tesseramento e i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ]),
        empty_message="Nessun tesseramento presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(tesseramenti_table, with_codice_tesserato([
            ("codice_tesseramento", "Codice"),
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("data_tesseramento", "Data"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("data_scadenza", "Scadenza"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("tesseramenti_annuali", value, query_params),
                    delete_action=f"/azioni/crud/elimina/tesseramenti_annuali/{value}",
                    delete_prompt="Eliminare questo tesseramento e i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ])),
    )}
    {table_card(
        f"Pagamenti tesseramenti anno {year}",
        "Incassi registrati sulle quote annuali.",
        payments_table,
        with_codice_tesserato([
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("data_pagamento", "Data pagamento"),
            ("importo", "Importo", lambda value, _: money(value)),
            ("metodo_pagamento", "Metodo"),
            ("riferimento", "Riferimento"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("pagamenti_tesseramenti", value, query_params),
                    extra_links=[(receipt_link("tesseramenti", value, query_params), "Ricevuta")],
                    delete_action=f"/azioni/crud/elimina/pagamenti_tesseramenti/{value}",
                    delete_prompt="Eliminare questo pagamento del tesseramento?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ]),
        empty_message="Nessun pagamento tesseramento presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(payments_table, with_codice_tesserato([
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("data_pagamento", "Data pagamento"),
            ("importo", "Importo", lambda value, _: money(value)),
            ("metodo_pagamento", "Metodo"),
            ("riferimento", "Riferimento"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("pagamenti_tesseramenti", value, query_params),
                    extra_links=[(receipt_link("tesseramenti", value, query_params), "Ricevuta")],
                    delete_action=f"/azioni/crud/elimina/pagamenti_tesseramenti/{value}",
                    delete_prompt="Eliminare questo pagamento del tesseramento?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ])),
    )}
    {table_card(
        f"Situazione tesseramenti anno {year}",
        "Vista con dovuto, pagato e saldo residuo.",
        saldo,
        with_codice_tesserato([
            ("codice_tesseramento", "Codice"),
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
        table_class="search-table",
        summary_rows=summary_rows_for_table(saldo, with_codice_tesserato([
            ("codice_tesseramento", "Codice"),
            ("associato", "Tesserato"),
            ("anno_sociale", "Anno"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ])),
    )}
    {table_card(
        "Quota tesseramento",
        "Quota predefinita disponibile per la registrazione del tesseramento.",
        quote_tesseramenti_table,
        [
            ("descrizione", "Descrizione"),
            ("importo", "Importo", lambda value, _: money(value)),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("quote_tesseramenti", value, query_params),
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ],
        empty_message="Nessuna quota tesseramento presente.",
        table_class="search-table",
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/tesseramenti", query_params, "Apri dati tesseramenti")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati tesseramenti" if data_only else "Rinnovo Tesseramento", "/maschere/tesseramenti", content, query_params, current_user)


def corsi_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    associati = associati_options()
    associati_iscrizioni = associati_options_for_enrollment(work_year)
    corsi = corsi_options()
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    open_rate_options = rate_corsi_aperte_options(work_year)
    corsi_table = fetch_all(
        f"""
        SELECT
            c.id,
            c.numero_progressivo,
            c.codice_corso,
            c.nome AS corso,
            c.quota_mensile_standard,
            c.giorno_settimana,
            c.orario,
            c.attivo
        FROM corsi c
        WHERE {corso_year_relevance_sql('c')}
        ORDER BY corso
        """,
        corso_year_relevance_params(work_year),
    )
    iscrizioni_table = fetch_all(
        f"""
        SELECT
            ic.id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            c.nome AS corso,
            ic.data_iscrizione,
            ic.data_inizio,
            ic.data_fine,
            ic.quota_mensile,
            ic.stato_iscrizione
        FROM iscrizioni_corsi ic
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = ?
        WHERE {iscrizione_corso_year_relevance_sql('ic')}
        ORDER BY ic.id DESC
        """,
        (work_year, *iscrizione_corso_year_relevance_params(work_year)),
    )
    rate_table = fetch_all(
        """
        SELECT
            id,
            corso,
            codice_tesseramento,
            associato,
            competenza,
            importo_dovuto,
            importo_pagato,
            saldo_residuo,
            stato_pagamento
        FROM v_rate_corsi_saldo
        WHERE anno = ?
        ORDER BY anno DESC, mese DESC, corso, associato
        """,
        (work_year,),
    )
    rate_payments_table = fetch_all(
        f"""
        SELECT
            prc.id,
            COALESCE(prc.gruppo_ricevuta, '') AS gruppo_ricevuta,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            c.nome AS corso,
            printf('%04d-%02d', r.anno, r.mese) AS competenza,
            prc.data_pagamento,
            prc.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(prc.riferimento, '') AS riferimento
        FROM pagamenti_rate_corsi prc
        JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
        JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id
        WHERE r.anno = ?
        ORDER BY prc.data_pagamento DESC, associato
        """,
        (work_year,),
    )
    create_course_form = form_card(
        "Nuovo corso",
        "Crea un corso e definisci la quota mensile di riferimento.",
        "/azioni/corsi/crea",
        "".join(
            [
                readonly_field("Codice corso assegnato", peek_next_progressive_code("corsi")),
                input_field("Nome corso", "nome", required_field=True),
                input_field("Data inizio", "data_inizio", input_type="date", required_field=True),
                input_field("Data fine", "data_fine", input_type="date", required_field=True),
                input_field("Quota mensile", "quota_mensile_standard", input_type="number", step="0.01", minimum="0"),
                input_field("Giorno settimana", "giorno_settimana", placeholder="Lunedi"),
                input_field("Orario", "orario", placeholder="17:00-18:00"),
                textarea_field("Descrizione", "descrizione"),
            ]
        ),
        "Salva corso",
        hidden_fields=work_year_query(query_params),
    )

    create_enrollment_form = form_card(
        "Iscrizione corso",
        "Collega un tesserato a un corso e definisci la quota mensile da applicare.",
        "/azioni/corsi/iscrizione",
        "".join(
            [
                readonly_field("Anno di lavoro", str(work_year)),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati_iscrizioni, extra_data_keys=["has_tesseramento"]),
                    required_field=True,
                    wide=True,
                    searchable=True,
                    attrs={
                        "data-require-tesseramento": "true",
                        "data-require-tesseramento-year": str(work_year),
                        "data-require-tesseramento-area": "Corsi",
                    },
                ),
                select_field(
                    "Corso",
                    "corso_id",
                    render_select_options(corsi, data_keys=["importo"]),
                    required_field=True,
                    wide=True,
                    element_id="corso-iscrizione-select",
                    attrs={"data-amount-target": "corso-iscrizione-quota"},
                ),
                input_field(
                    "Data iscrizione",
                    "data_iscrizione",
                    input_type="date",
                    value=date.today().isoformat(),
                    required_field=True,
                    element_id="corso-iscrizione-data-iscrizione",
                    attrs={"data-copy-value-target": "corso-iscrizione-data-inizio"},
                ),
                input_field(
                    "Data inizio",
                    "data_inizio",
                    input_type="date",
                    value=date.today().isoformat(),
                    element_id="corso-iscrizione-data-inizio",
                    attrs={"data-auto-managed-value": "true"},
                ),
                input_field(
                    "Quota mensile",
                    "quota_mensile",
                    input_type="number",
                    step="0.01",
                    minimum="0",
                    element_id="corso-iscrizione-quota",
                ),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva iscrizione corso",
        hidden_fields=work_year_query(query_params),
        form_attrs={
            "data-payment-flow": "corso",
            "data-payment-amount-field": "quota_mensile",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-prompt-title": "Conferma iscrizione corso",
            "data-payment-prompt-message": "Dopo il salvataggio verra generata automaticamente la prima quota mensile utile del corso. Vuoi procedere subito anche al pagamento?",
            "data-payment-prompt-yes": "Si, procedi al pagamento",
            "data-payment-prompt-no": "No, solo iscrizione",
            "data-payment-dialog-title": "Pagamento quote corso",
            "data-payment-dialog-message": "Scegli se pagare solo la prima mensilita utile oppure anche le mensilita future.",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
        },
    )

    forms_html = f"""
    <div class="cards-grid cards-stack">
      {create_enrollment_form}
      {create_course_form}
    </div>
    """ if not data_only else ""

    tables_html = f"""
    {table_card(
        "Corsi registrati",
        "Catalogo corsi con quota mensile e organizzazione.",
        corsi_table,
        [
            ("codice_corso", "Codice"),
            ("corso", "Corso"),
            ("quota_mensile_standard", "Mensile", lambda value, _: money(value)),
            ("giorno_settimana", "Giorno"),
            ("orario", "Orario"),
            ("attivo", "Attivo"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("corsi", value, query_params),
                    delete_action=f"/azioni/crud/elimina/corsi/{value}",
                    delete_prompt="Eliminare questo corso? Le iscrizioni collegate devono essere eliminate prima.",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ],
        table_class="search-table",
    )}
    {table_card(
        "Iscrizioni corsi",
        "Elenco delle iscrizioni corsi con quota mensile e stato.",
        iscrizioni_table,
        with_codice_tesserato([
            ("associato", "Tesserato"),
            ("corso", "Corso"),
            ("data_iscrizione", "Data iscrizione"),
            ("data_inizio", "Inizio"),
            ("data_fine", "Fine"),
            ("quota_mensile", "Quota mensile", lambda value, _: money(value)),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("iscrizioni_corsi", value, query_params),
                    delete_action=f"/azioni/crud/elimina/iscrizioni_corsi/{value}",
                    delete_prompt="Eliminare questa iscrizione corso e tutti i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ]),
        table_class="search-table",
        summary_rows=summary_rows_for_table(iscrizioni_table, with_codice_tesserato([
            ("associato", "Tesserato"),
            ("corso", "Corso"),
            ("data_iscrizione", "Data iscrizione"),
            ("data_inizio", "Inizio"),
            ("data_fine", "Fine"),
            ("quota_mensile", "Quota mensile", lambda value, _: money(value)),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("iscrizioni_corsi", value, query_params),
                    delete_action=f"/azioni/crud/elimina/iscrizioni_corsi/{value}",
                    delete_prompt="Eliminare questa iscrizione corso e tutti i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ])),
    )}
    {table_card(
        f"Quote mensili corsi anno {work_year}",
        "Vista operativa per tesserato, competenza e stato pagamento.",
        rate_table,
        with_codice_tesserato([
            ("corso", "Corso"),
            ("associato", "Tesserato"),
            ("competenza", "Competenza"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("rate_corsi_mensili", value, query_params),
                    delete_action=f"/azioni/crud/elimina/rate_corsi_mensili/{value}",
                    delete_prompt="Eliminare questa quota mensile e i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ]),
        empty_message="Nessuna quota mensile presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(rate_table, with_codice_tesserato([
            ("corso", "Corso"),
            ("associato", "Tesserato"),
            ("competenza", "Competenza"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("rate_corsi_mensili", value, query_params),
                    delete_action=f"/azioni/crud/elimina/rate_corsi_mensili/{value}",
                    delete_prompt="Eliminare questa quota mensile e i pagamenti collegati?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ])),
    )}
    {table_card(
        f"Pagamenti quote mensili anno {work_year}",
        "Incassi registrati sulle quote mensili dei corsi.",
        rate_payments_table,
        with_codice_tesserato([
            ("associato", "Tesserato"),
            ("corso", "Corso"),
            ("competenza", "Competenza"),
            ("data_pagamento", "Data pagamento"),
            ("importo", "Importo", lambda value, _: money(value)),
            ("metodo_pagamento", "Metodo"),
            ("riferimento", "Riferimento"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("pagamenti_rate_corsi", value, query_params),
                    extra_links=[(
                        grouped_receipt_link(row["gruppo_ricevuta"], query_params)
                        if row["gruppo_ricevuta"]
                        else receipt_link("corsi-rata", value, query_params),
                        "Ricevuta",
                    )],
                    delete_action=f"/azioni/crud/elimina/pagamenti_rate_corsi/{value}",
                    delete_prompt="Eliminare questo pagamento della quota mensile?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ]),
        empty_message="Nessun pagamento mensile presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(rate_payments_table, with_codice_tesserato([
            ("associato", "Tesserato"),
            ("corso", "Corso"),
            ("competenza", "Competenza"),
            ("data_pagamento", "Data pagamento"),
            ("importo", "Importo", lambda value, _: money(value)),
            ("metodo_pagamento", "Metodo"),
            ("riferimento", "Riferimento"),
            (
                "id",
                "Azioni",
                lambda value, row: action_links_html(
                    edit_href=edit_path("pagamenti_rate_corsi", value, query_params),
                    extra_links=[(
                        grouped_receipt_link(row["gruppo_ricevuta"], query_params)
                        if row["gruppo_ricevuta"]
                        else receipt_link("corsi-rata", value, query_params),
                        "Ricevuta",
                    )],
                    delete_action=f"/azioni/crud/elimina/pagamenti_rate_corsi/{value}",
                    delete_prompt="Eliminare questo pagamento della quota mensile?",
                    extra_fields=work_year_query(query_params),
                ),
            ),
        ])),
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/corsi", query_params, "Apri dati corsi")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati corsi" if data_only else "Corsi", "/maschere/corsi", content, query_params, current_user)


def campi_estivi_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    associati = associati_options()
    associati_iscrizioni = associati_options_for_enrollment(work_year)
    iscrizioni = iscrizioni_campi_aperte_options(work_year)
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    quote_campo = quote_predefinite_options("campi-estivi")
    quote_campo_table = quote_predefinite_rows("campi-estivi")
    iscrizioni_table = fetch_all(
        f"""
        SELECT
            ice.id,
            ce.anno,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            ice.data_iscrizione,
            ice.quota_partecipazione,
            ice.stato_iscrizione
        FROM iscrizioni_campi_estivi ice
        JOIN associati a ON a.id = ice.associato_id
        JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
        WHERE ce.anno = ?
        ORDER BY ice.id DESC
        """,
        (work_year,),
    )
    pagamenti_table = fetch_all(
        f"""
        SELECT
            pce.id,
            ce.anno,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            pce.data_pagamento,
            pce.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(pce.riferimento, '') AS riferimento
        FROM pagamenti_campi_estivi pce
        JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
        JOIN associati a ON a.id = ice.associato_id
        JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = pce.metodo_pagamento_id
        WHERE ce.anno = ?
        ORDER BY pce.data_pagamento DESC, associato
        """,
        (work_year,),
    )
    saldo = fetch_all(
        """
        SELECT anno, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
        FROM v_campi_estivi_saldo
        WHERE anno = ?
        ORDER BY associato
        """,
        (work_year,),
    )

    quote_form = form_card(
        "Nuova quota Campo estivo",
        "Inserisci una quota predefinita con descrizione e importo da richiamare in fase di iscrizione.",
        "/azioni/quote/crea",
        "".join(
            [
                input_field("Descrizione", "descrizione", required_field=True, wide=True),
                input_field("Importo", "importo", input_type="number", step="0.01", minimum="0", required_field=True),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva quota",
        hidden_fields={**work_year_query(query_params), "area": "campi-estivi"},
    )

    enrollment_form = form_card(
        "Iscrizione Campo estivo",
        "Registra il tesserato e proponi la quota scelta, sempre modificabile.",
        "/azioni/campi-estivi/iscrizione",
        "".join(
            [
                readonly_field("Anno di lavoro", str(work_year)),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati_iscrizioni, extra_data_keys=["has_tesseramento"]),
                    required_field=True,
                    wide=True,
                    searchable=True,
                    attrs={
                        "data-require-tesseramento": "true",
                        "data-require-tesseramento-year": str(work_year),
                        "data-require-tesseramento-area": "Campo estivo",
                    },
                ),
                select_field(
                    "Quota predefinita",
                    "quota_predefinita_id",
                    render_select_options(quote_campo, data_keys=["importo"]),
                    wide=True,
                    element_id="campo-quota-select",
                    attrs={"data-amount-target": "campo-estivo-quota"},
                ),
                input_field("Data iscrizione", "data_iscrizione", input_type="date", value=date.today().isoformat(), required_field=True),
                input_field(
                    "Importo",
                    "quota_partecipazione",
                    input_type="number",
                    step="0.01",
                    minimum="0",
                    required_field=True,
                    element_id="campo-estivo-quota",
                    wide=True,
                ),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva iscrizione Campo estivo",
        hidden_fields=work_year_query(query_params),
        form_attrs={
            "data-payment-flow": "campo-estivo",
            "data-payment-amount-field": "quota_partecipazione",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-prompt-title": "Conferma iscrizione Campo estivo",
            "data-payment-prompt-message": "Vuoi procedere subito anche al pagamento dell'iscrizione al Campo estivo?",
            "data-payment-prompt-yes": "Si, registra anche il pagamento",
            "data-payment-prompt-no": "No, solo iscrizione",
            "data-payment-dialog-title": "Pagamento Campo estivo",
            "data-payment-dialog-message": "Conferma i dati del pagamento del Campo estivo.",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
        },
    )

    forms_html = f"""
    <div class="cards-grid cards-stack">
      {enrollment_form}
      {quote_form}
    </div>
    """ if not data_only else ""

    quote_campo_columns = [
        ("descrizione", "Descrizione"),
        ("importo", "Importo", lambda value, _: money(value)),
        ("attiva", "Attiva"),
        ("note", "Note"),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("quote_campi_estivi", value, query_params),
                delete_action=f"/azioni/crud/elimina/quote_campi_estivi/{value}",
                delete_prompt="Eliminare questa quota predefinita di Campo estivo?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ]
    iscrizioni_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("data_iscrizione", "Data iscrizione"),
        ("quota_partecipazione", "Quota", lambda value, _: money(value)),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("iscrizioni_campi_estivi", value, query_params),
                delete_action=f"/azioni/crud/elimina/iscrizioni_campi_estivi/{value}",
                delete_prompt="Eliminare questa iscrizione al Campo estivo e il pagamento collegato?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    pagamenti_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("data_pagamento", "Data"),
        ("importo", "Importo", lambda value, _: money(value)),
        ("metodo_pagamento", "Metodo"),
        ("riferimento", "Riferimento"),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("pagamenti_campi_estivi", value, query_params),
                extra_links=[(receipt_link("campi-estivi", value, query_params), "Ricevuta")],
                delete_action=f"/azioni/crud/elimina/pagamenti_campi_estivi/{value}",
                delete_prompt="Eliminare questo pagamento del Campo estivo?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    saldo_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
        ("importo_pagato", "Pagato", lambda value, _: money(value)),
        ("saldo_residuo", "Residuo", lambda value, _: money(value)),
        ("stato_pagamento", "Stato"),
    ])

    tables_html = f"""
    {table_card(
        "Iscrizioni Campo estivo",
        "Partecipanti registrati al Campo estivo dell'anno selezionato.",
        iscrizioni_table,
        iscrizioni_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(iscrizioni_table, iscrizioni_columns),
    )}
    {table_card(
        "Pagamenti Campo estivo",
        "Pagamenti una tantum registrati per il Campo estivo.",
        pagamenti_table,
        pagamenti_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(pagamenti_table, pagamenti_columns),
    )}
    {table_card(
        f"Situazione Campo estivo anno {work_year}",
        "Elenco iscritti con saldo della quota di partecipazione.",
        saldo,
        saldo_columns,
        empty_message="Nessun movimento presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(saldo, saldo_columns),
    )}
    {table_card(
        "Quote Campo estivo",
        "Elenco delle quote predefinite disponibili per le iscrizioni al Campo estivo.",
        quote_campo_table,
        quote_campo_columns,
        empty_message="Nessuna quota predefinita presente.",
        table_class="search-table",
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/campi-estivi", query_params, "Apri dati Campo estivo")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati Campo estivo" if data_only else ESTATE_LABEL, "/maschere/campi-estivi", content, query_params, current_user)


def oratorio_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    associati_iscrizioni = associati_options_for_enrollment(work_year)
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    quote_oratorio = quote_predefinite_options("oratorio")
    contributo_oratorio_default = quote_oratorio[0] if len(quote_oratorio) == 1 else None
    contributo_oratorio_default_id = str(contributo_oratorio_default["id"]) if contributo_oratorio_default else None
    contributo_oratorio_default_importo = str(contributo_oratorio_default["importo"]) if contributo_oratorio_default else ""
    quote_oratorio_table = quote_predefinite_rows("oratorio")
    iscrizioni_table = fetch_all(
        f"""
        SELECT
            io.id,
            o.anno,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            io.data_iscrizione,
            io.quota_partecipazione,
            io.stato_iscrizione
        FROM iscrizioni_oratorio io
        JOIN associati a ON a.id = io.associato_id
        JOIN oratorio o ON o.id = io.oratorio_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
        WHERE o.anno = ?
        ORDER BY io.id DESC
        """,
        (work_year,),
    )
    pagamenti_table = fetch_all(
        f"""
        SELECT
            po.id,
            o.anno,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            po.data_pagamento,
            po.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(po.riferimento, '') AS riferimento
        FROM pagamenti_oratorio po
        JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
        JOIN associati a ON a.id = io.associato_id
        JOIN oratorio o ON o.id = io.oratorio_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = po.metodo_pagamento_id
        WHERE o.anno = ?
        ORDER BY po.data_pagamento DESC, associato
        """,
        (work_year,),
    )
    saldo = fetch_all(
        """
        SELECT anno, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
        FROM v_oratorio_saldo
        WHERE anno = ?
        ORDER BY associato
        """,
        (work_year,),
    )

    enrollment_form = form_card(
        "Iscrizione Oratorio",
        "Registra il tesserato e proponi la quota scelta, sempre modificabile.",
        "/azioni/oratorio/iscrizione",
        "".join(
            [
                readonly_field("Anno di lavoro", str(work_year)),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati_iscrizioni, extra_data_keys=["has_tesseramento"]),
                    required_field=True,
                    wide=True,
                    searchable=True,
                    attrs={
                        "data-require-tesseramento": "true",
                        "data-require-tesseramento-year": str(work_year),
                        "data-require-tesseramento-area": ORATORIO_LABEL,
                    },
                ),
                select_field(
                    "Contributo Iscrizione",
                    "quota_predefinita_id",
                    render_select_options(quote_oratorio, selected=contributo_oratorio_default_id, data_keys=["importo"]),
                    wide=True,
                    element_id="oratorio-quota-select",
                    attrs={"data-amount-target": "oratorio-quota"},
                ),
                input_field("Data iscrizione", "data_iscrizione", input_type="date", value=date.today().isoformat(), required_field=True, wide=False),
                input_field(
                    "Importo",
                    "quota_partecipazione",
                    input_type="number",
                    step="0.01",
                    minimum="0",
                    required_field=True,
                    element_id="oratorio-quota",
                    value=contributo_oratorio_default_importo,
                    wide=False,
                ),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva iscrizione Oratorio",
        hidden_fields=work_year_query(query_params),
        form_attrs={
            "data-payment-flow": "oratorio",
            "data-payment-amount-field": "quota_partecipazione",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-prompt-title": "Conferma iscrizione Oratorio",
            "data-payment-prompt-message": "Vuoi procedere subito anche al pagamento dell'iscrizione a Oratorio?",
            "data-payment-prompt-yes": "Si, registra anche il pagamento",
            "data-payment-prompt-no": "No, solo iscrizione",
            "data-payment-dialog-title": "Pagamento Oratorio",
            "data-payment-dialog-message": "Conferma i dati del pagamento di Oratorio.",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
        },
    )

    quote_form = form_card(
        "Contributo iscrizione",
        "Inserisci un contributo predefinito con descrizione e importo da richiamare in fase di iscrizione.",
        "/azioni/quote/crea",
        "".join(
            [
                input_field("Descrizione", "descrizione", required_field=True, wide=True),
                input_field("Importo", "importo", input_type="number", step="0.01", minimum="0", required_field=True),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva contributo",
        hidden_fields={**work_year_query(query_params), "area": "oratorio"},
    )

    forms_html = f"""
    <div class="cards-grid cards-stack">
      {enrollment_form}
      {quote_form}
    </div>
    """ if not data_only else ""

    quote_columns = [
        ("descrizione", "Descrizione"),
        ("importo", "Importo", lambda value, _: money(value)),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("quote_oratorio", value, query_params),
                extra_fields=work_year_query(query_params),
            ),
        ),
    ]
    iscrizioni_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("data_iscrizione", "Data iscrizione"),
        ("quota_partecipazione", "Quota", lambda value, _: money(value)),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("iscrizioni_oratorio", value, query_params),
                delete_action=f"/azioni/crud/elimina/iscrizioni_oratorio/{value}",
                delete_prompt="Eliminare questa iscrizione a Oratorio e il pagamento collegato?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    pagamenti_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("data_pagamento", "Data"),
        ("importo", "Importo", lambda value, _: money(value)),
        ("metodo_pagamento", "Metodo"),
        ("riferimento", "Riferimento"),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("pagamenti_oratorio", value, query_params),
                extra_links=[(receipt_link("oratorio", value, query_params), "Ricevuta")],
                delete_action=f"/azioni/crud/elimina/pagamenti_oratorio/{value}",
                delete_prompt="Eliminare questo pagamento di Oratorio?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    saldo_columns = with_codice_tesserato([
        ("associato", "Tesserato"),
        ("anno", "Anno"),
        ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
        ("importo_pagato", "Pagato", lambda value, _: money(value)),
        ("saldo_residuo", "Residuo", lambda value, _: money(value)),
        ("stato_pagamento", "Stato"),
    ])

    tables_html = f"""
    {table_card(
        "Iscrizioni Oratorio",
        "Partecipanti registrati a Oratorio dell'anno selezionato.",
        iscrizioni_table,
        iscrizioni_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(iscrizioni_table, iscrizioni_columns),
    )}
    {table_card(
        "Pagamenti Oratorio",
        "Pagamenti una tantum registrati per Oratorio.",
        pagamenti_table,
        pagamenti_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(pagamenti_table, pagamenti_columns),
    )}
    {table_card(
        f"Situazione Oratorio anno {work_year}",
        "Elenco iscritti con saldo della quota di partecipazione.",
        saldo,
        saldo_columns,
        empty_message="Nessun movimento presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(saldo, saldo_columns),
    )}
    {table_card(
        "Contributo Oratorio",
        "Elenco delle quote predefinite disponibili per le iscrizioni a Oratorio.",
        quote_oratorio_table,
        quote_columns,
        empty_message="Nessuna quota predefinita presente.",
        table_class="search-table",
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/oratorio", query_params, "Apri dati Oratorio")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati Oratorio" if data_only else ORATORIO_LABEL, "/maschere/oratorio", content, query_params, current_user)


def eventi_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    associati = associati_options()
    associati_iscrizioni = associati_options_for_enrollment(work_year)
    eventi = eventi_options(work_year)
    iscrizioni = iscrizioni_eventi_aperte_options(work_year)
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)

    eventi_table = fetch_all(
        """
        SELECT
            id,
            numero_progressivo,
            codice_evento,
            nome,
            tipologia,
            data_evento,
            quota_partecipazione_standard,
            attivo
        FROM eventi
        WHERE substr(COALESCE(data_evento, ''), 1, 4) = ?
        ORDER BY data_evento DESC, nome
        """,
        (str(work_year),),
    )
    iscrizioni_table = fetch_all(
        f"""
        SELECT
            ie.id,
            e.nome AS evento,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            ie.data_iscrizione,
            ie.quota_partecipazione,
            ie.stato_iscrizione
        FROM iscrizioni_eventi ie
        JOIN associati a ON a.id = ie.associato_id
        JOIN eventi e ON e.id = ie.evento_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
        WHERE substr(COALESCE(e.data_evento, ''), 1, 4) = ?
        ORDER BY ie.id DESC
        """,
        (str(work_year),),
    )
    pagamenti_table = fetch_all(
        f"""
        SELECT
            pe.id,
            e.nome AS evento,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            pe.data_pagamento,
            pe.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(pe.riferimento, '') AS riferimento
        FROM pagamenti_eventi pe
        JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
        JOIN associati a ON a.id = ie.associato_id
        JOIN eventi e ON e.id = ie.evento_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
        LEFT JOIN metodi_pagamento mp ON mp.id = pe.metodo_pagamento_id
        WHERE substr(COALESCE(e.data_evento, ''), 1, 4) = ?
        ORDER BY pe.data_pagamento DESC, associato
        """,
        (str(work_year),),
    )
    saldo = fetch_all(
        """
        SELECT evento, tipologia, data_evento, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
        FROM v_eventi_saldo
        WHERE substr(data_evento, 1, 4) = ?
        ORDER BY data_evento DESC, evento, associato
        """,
        (str(work_year),),
    )

    create_event_form = form_card(
        "Nuovo evento",
        "Crea eventi diversi con quota di partecipazione una tantum.",
        "/azioni/eventi/crea",
        "".join(
            [
                readonly_field("Codice evento assegnato", peek_next_progressive_code("eventi")),
                input_field("Nome evento", "nome", required_field=True),
                input_field("Data evento", "data_evento", input_type="date", required_field=True),
                input_field("Luogo", "luogo"),
                input_field("Quota dovuta", "quota_partecipazione_standard", input_type="number", step="0.01", minimum="0"),
                textarea_field("Descrizione", "descrizione"),
            ]
        ),
        "Salva evento",
        hidden_fields=work_year_query(query_params),
    )

    enrollment_form = form_card(
        "Iscrizione evento",
        "Registra i tesserati all'evento e la loro quota.",
        "/azioni/eventi/iscrizione",
        "".join(
            [
                readonly_field("Anno di lavoro", str(work_year)),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati_iscrizioni, extra_data_keys=["has_tesseramento"]),
                    required_field=True,
                    wide=True,
                    searchable=True,
                    attrs={
                        "data-require-tesseramento": "true",
                        "data-require-tesseramento-year": str(work_year),
                        "data-require-tesseramento-area": "Eventi",
                    },
                ),
                select_field(
                    "Evento",
                    "evento_id",
                    render_select_options(eventi, data_keys=["importo"]),
                    required_field=True,
                    wide=True,
                    element_id="evento-iscrizione-select",
                    attrs={"data-amount-target": "evento-iscrizione-quota"},
                ),
                input_field("Data iscrizione", "data_iscrizione", input_type="date", value=date.today().isoformat(), required_field=True),
                input_field(
                    "Quota dovuta",
                    "quota_partecipazione",
                    input_type="number",
                    step="0.01",
                    minimum="0",
                    required_field=True,
                    element_id="evento-iscrizione-quota",
                ),
                textarea_field("Note", "note"),
            ]
        ),
        "Salva iscrizione evento",
        hidden_fields=work_year_query(query_params),
        form_attrs={
            "data-payment-flow": "evento",
            "data-payment-amount-field": "quota_partecipazione",
            "data-payment-method-default": metodo_predefinito,
            "data-payment-prompt-title": "Conferma iscrizione evento",
            "data-payment-prompt-message": "Vuoi procedere subito anche al pagamento dell'iscrizione all'evento?",
            "data-payment-prompt-yes": "Si, registra anche il pagamento",
            "data-payment-prompt-no": "No, solo iscrizione",
            "data-payment-dialog-title": "Pagamento evento",
            "data-payment-dialog-message": "Conferma i dati del pagamento dell'evento.",
            "data-payment-dialog-confirm": "Registra pagamento e genera ricevuta",
        },
    )

    forms_html = f"""
    <div class="cards-grid cards-stack">
      {enrollment_form}
      {create_event_form}
    </div>
    """ if not data_only else ""

    eventi_columns = [
        ("codice_evento", "Codice"),
        ("nome", "Evento"),
        ("data_evento", "Data"),
        ("quota_partecipazione_standard", "Quota dovuta", lambda value, _: money(value)),
        ("attivo", "Attivo"),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("eventi", value, query_params),
                delete_action=f"/azioni/crud/elimina/eventi/{value}",
                delete_prompt="Eliminare questo evento e tutte le iscrizioni collegate?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ]
    iscrizioni_columns = with_codice_tesserato([
        ("evento", "Evento"),
        ("associato", "Tesserato"),
        ("data_iscrizione", "Data iscrizione"),
        ("quota_partecipazione", "Quota dovuta", lambda value, _: money(value)),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("iscrizioni_eventi", value, query_params),
                delete_action=f"/azioni/crud/elimina/iscrizioni_eventi/{value}",
                delete_prompt="Eliminare questa iscrizione evento e il pagamento collegato?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    pagamenti_columns = with_codice_tesserato([
        ("evento", "Evento"),
        ("associato", "Tesserato"),
        ("data_pagamento", "Data"),
        ("importo", "Importo", lambda value, _: money(value)),
        ("metodo_pagamento", "Metodo"),
        ("riferimento", "Riferimento"),
        (
            "id",
            "Azioni",
            lambda value, row: action_links_html(
                edit_href=edit_path("pagamenti_eventi", value, query_params),
                extra_links=[(receipt_link("eventi", value, query_params), "Ricevuta")],
                delete_action=f"/azioni/crud/elimina/pagamenti_eventi/{value}",
                delete_prompt="Eliminare questo pagamento evento?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])
    saldo_columns = with_codice_tesserato([
        ("evento", "Evento"),
        ("data_evento", "Data"),
        ("associato", "Tesserato"),
        ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
        ("importo_pagato", "Pagato", lambda value, _: money(value)),
        ("saldo_residuo", "Residuo", lambda value, _: money(value)),
        ("stato_pagamento", "Stato"),
    ])

    tables_html = f"""
    {table_card(
        "Eventi registrati",
        "Anagrafica completa degli eventi.",
        eventi_table,
        eventi_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(eventi_table, eventi_columns),
    )}
    {table_card(
        "Iscrizioni eventi",
        "Partecipanti registrati agli eventi.",
        iscrizioni_table,
        iscrizioni_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(iscrizioni_table, iscrizioni_columns),
    )}
    {table_card(
        "Pagamenti eventi",
        "Pagamenti una tantum registrati per gli eventi.",
        pagamenti_table,
        pagamenti_columns,
        table_class="search-table",
        summary_rows=summary_rows_for_table(pagamenti_table, pagamenti_columns),
    )}
    {table_card(
        f"Situazione eventi anno {work_year}",
        "Partecipanti, quota una tantum e stato del pagamento.",
        saldo,
        saldo_columns,
        empty_message="Nessun movimento presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(saldo, saldo_columns),
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/eventi", query_params, "Apri dati eventi")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati eventi" if data_only else "Eventi", "/maschere/eventi", content, query_params, current_user)


def pagamenti_multi_area_page(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    data_only = normalized(query_params, "vista", "") == "dati"
    work_year = current_work_year(query_params)
    associati = associati_options()
    metodi = metodi_options()
    metodo_predefinito = preferred_metodo_pagamento_id(metodi)
    scadenze = scadenze_multi_area_options(work_year)
    future_course_options = pagamenti_multi_area_course_options(work_year)
    ricevute_table = multi_area_receipts_rows(work_year)
    default_until_month = (
        f"{work_year}-{date.today().month:02d}"
        if work_year == date.today().year
        else f"{work_year}-12"
    )

    payment_form = form_card(
        "Pagamento multi-area",
        "Seleziona un tesserato e registra in un unico pagamento scadenze aperte di tesseramenti, quote mensili corsi, Oratorio, Campo estivo ed eventi.",
        "/azioni/pagamenti-multi-area/crea",
        "".join(
            [
                readonly_field("Anno di lavoro", str(work_year)),
                select_field(
                    "Tesserato",
                    "associato_id",
                    render_associato_options(associati),
                    required_field=True,
                    wide=True,
                    searchable=True,
                    element_id="multi-area-associato",
                ),
                select_field(
                    "Corso per quote future",
                    "multi_area_iscrizione_corso_id",
                    render_select_options(future_course_options, data_keys=["associato_id", "start_competenza"]),
                    wide=True,
                    element_id="multi-area-course-enrollment",
                    attrs={"data-work-year": str(work_year)},
                ),
                input_field(
                    "Genera fino alla competenza",
                    "multi_area_competenza_fine",
                    input_type="month",
                    value=default_until_month,
                    wide=True,
                    element_id="multi-area-course-until",
                ),
                (
                    '<div class="field wide multi-area-course-actions">'
                    '<span>Quote future corso</span>'
                    '<div class="multi-area-course-actions-row">'
                    '<button type="button" class="button secondary" id="multi-area-course-generate">Aggiungi quote future corso</button>'
                    "</div>"
                    '<p id="multi-area-course-feedback" class="multi-area-course-feedback">'
                    "Seleziona tesserato e corso, indica fino a quale mensilita vuoi arrivare e il gestionale selezionera in basso le relative quote."
                    "</p>"
                    "</div>"
                ),
                multi_select_field(
                    "Scadenze da saldare",
                    "scadenza_id",
                    render_select_options_multi(scadenze, data_keys=["associato_id", "residuo"]),
                    required_field=False,
                    wide=True,
                    size=12,
                    element_id="multi-area-scadenze",
                ),
                (
                    '<div class="field wide multi-area-scadenze-display">'
                    '<div id="multi-area-scadenze-visual" class="payment-flow-extra-list"></div>'
                    '<p id="multi-area-scadenze-note" class="payment-flow-extra-note">'
                    'Seleziona prima un tesserato per visualizzare le scadenze aperte disponibili.'
                    "</p>"
                    "</div>"
                ),
                input_field(
                    "Importo da registrare",
                    "importo",
                    input_type="number",
                    step="0.01",
                    minimum="0.01",
                    required_field=True,
                    element_id="multi-area-amount",
                ),
                select_field("Metodo", "metodo_pagamento_id", render_select_options(metodi, metodo_predefinito), required_field=True),
                input_field("Data pagamento", "data_pagamento", input_type="date", value=date.today().isoformat(), required_field=True),
                input_field("Riferimento", "riferimento"),
                textarea_field("Note", "note"),
            ]
        ),
        "Registra pagamento e ricevuta unica",
        hidden_fields=work_year_query(query_params),
    )

    forms_html = f"""
    <div class="cards-grid">
      {payment_form}
    </div>
    """ if not data_only else ""

    ricevute_columns = with_codice_tesserato([
        ("gruppo_ricevuta", "Codice ricevuta"),
        ("associato", "Tesserato"),
        ("data_pagamento", "Data pagamento"),
        ("numero_scadenze", "Scadenze"),
        ("importo_totale", "Importo", lambda value, _: money(value)),
        (
            "gruppo_ricevuta",
            "Azioni",
            lambda value, row: action_links_html(
                extra_links=[(with_query(f"/ricevute/multi-area-gruppo/{value}", work_year_query(query_params)), "Ricevuta")],
                delete_action=f"/azioni/pagamenti-multi-area/elimina/{value}",
                delete_prompt="Eliminare questo pagamento multi-area e ripristinare le scadenze saldate?",
                extra_fields=work_year_query(query_params),
            ),
        ),
    ])

    tables_html = f"""
    {table_card(
        f"Ricevute multi-area anno {work_year}",
        "Storico dei pagamenti unici registrati su scadenze provenienti da aree diverse.",
        ricevute_table,
        ricevute_columns,
        empty_message="Nessuna ricevuta multi-area presente per l'anno selezionato.",
        table_class="search-table",
        summary_rows=summary_rows_for_table(ricevute_table, ricevute_columns),
    )}
    """

    content = f"""
    {view_mode_switch("/maschere/pagamenti-multi-area", query_params, "Apri dati pagamenti")}
    {forms_html}
    {data_view_search_toolbar() if data_only else ""}
    {tables_html if data_only else ""}
    """
    return page("Dati pagamenti" if data_only else "Pagamenti", "/maschere/pagamenti-multi-area", content, query_params, current_user)


def with_query(path: str, query_params: dict[str, str]) -> str:
    active_params = {key: value for key, value in query_params.items() if value}
    if not active_params:
        return path
    return f"{path}?{urlencode(active_params)}"


def slugify(value: str) -> str:
    cleaned = []
    previous_dash = False
    for character in value.lower():
        if character.isalnum():
            cleaned.append(character)
            previous_dash = False
        elif not previous_dash:
            cleaned.append("-")
            previous_dash = True
    return "".join(cleaned).strip("-") or "report"


def report_display_value(row: sqlite3.Row, column: tuple) -> str:
    key = column[0]
    formatter = column[2] if len(column) > 2 else None
    value = row[key]
    if callable(formatter):
        return plain_text(formatter(value, row))
    return plain_text(value)


def report_share_detail_columns(columns: list[tuple]) -> list[tuple]:
    excluded_labels = {"Azioni", "Ricevuta"}
    excluded_keys = {"id", "payment_id"}
    return [
        column
        for column in columns
        if len(column) >= 2
        and column[1] not in excluded_labels
        and column[0] not in excluded_keys
    ]


def report_share_row_line(row: sqlite3.Row, columns: list[tuple], index: int) -> str:
    parts: list[str] = []
    for column in report_share_detail_columns(columns):
        value = report_display_value(row, column)
        if not value:
            continue
        parts.append(f"{column[1]}: {value}")
    if not parts:
        return f"{index}. Riga {index}"
    return f"{index}. " + " | ".join(parts)


def report_search_term(query_params: dict[str, str]) -> str:
    return normalized(query_params, "search", "")


def definition_with_search(definition: dict, query_params: dict[str, str]) -> dict:
    search_term = report_search_term(query_params)
    if not search_term:
        return definition
    filters = list(definition.get("filters", []))
    filters.append({"label": "Ricerca testo", "value": search_term})
    enriched = dict(definition)
    enriched["filters"] = filters
    return enriched


def ui_tesserato_text(text: str) -> str:
    value = str(text or "")
    replacements = [
        ("Codice associato", "Codice"),
        ("codice associato", "codice"),
        ("Associati", "Tesserati"),
        ("associati", "tesserati"),
        ("Associato", "Tesserato"),
        ("associato", "tesserato"),
    ]
    for source, target in replacements:
        value = value.replace(source, target)
    return value


def tesserato_columns(columns: list[tuple]) -> list[tuple]:
    updated: list[tuple] = []
    for column in columns:
        if len(column) < 2:
            updated.append(column)
            continue
        key = column[0]
        label = column[1]
        if key == "associato":
            label = "Tesserato"
        elif key == "codice_tesseramento":
            label = "Codice"
        elif key in {"codice_associato", "associato_codice"}:
            label = "Codice"
        elif key == "associato_nominativo":
            label = "Tesserato"
        updated.append((key, label, *column[2:]))
    return updated


def tesserato_filters(filters: list[dict[str, str]]) -> list[dict[str, str]]:
    updated: list[dict[str, str]] = []
    for row in filters:
        updated.append(
            {
                **row,
                "label": "Tesserato"
                if row.get("label") == "Associato"
                else (
                    "Codice"
                    if row.get("label") in {"Codice associato", "Codice tesserato"}
                    else row.get("label", "")
                ),
            }
        )
    return updated


def tesserato_definition(definition: dict) -> dict:
    updated = dict(definition)
    updated["columns"] = tesserato_columns(list(definition.get("columns", [])))
    updated["filters"] = tesserato_filters(list(definition.get("filters", [])))
    if "title" in updated:
        updated["title"] = ui_tesserato_text(updated["title"])
    if "subtitle" in updated:
        updated["subtitle"] = ui_tesserato_text(updated["subtitle"])
    return updated


def with_codice_tesserato(columns: list[tuple]) -> list[tuple]:
    updated: list[tuple] = []
    has_code = any(column and column[0] in {"codice_tesseramento", "codice_associato"} for column in columns)
    for column in columns:
        if not column:
            updated.append(column)
            continue
        if column[0] == "associato" and not has_code:
            updated.append(("codice_tesseramento", "Codice"))
        updated.append(column)
    return updated


def filter_report_rows(rows: list[sqlite3.Row], columns: list[tuple], search_term: str) -> list[sqlite3.Row]:
    needle = search_term.lower().strip()
    if not needle:
        return rows
    return [
        row
        for row in rows
        if needle in " ".join(report_display_value(row, column).lower() for column in columns)
    ]


def participants_target_options(area: str, work_year: int) -> list[sqlite3.Row]:
    if area == "corsi":
        return fetch_all(
            f"""
            SELECT c.id, c.nome AS label
            FROM corsi c
            WHERE {corso_year_relevance_sql('c')}
            ORDER BY c.nome
            """,
            corso_year_relevance_params(work_year),
        )
    if area == "campi-estivi":
        return fetch_all(
            """
            SELECT id, 'Campo estivo ' || anno AS label
            FROM campi_estivi
            WHERE anno = ?
            ORDER BY id
            """,
            (work_year,),
        )
    if area == "oratorio":
        return fetch_all(
            """
            SELECT id, 'Oratorio ' || anno AS label
            FROM oratorio
            WHERE anno = ?
            ORDER BY id
            """,
            (work_year,),
        )
    if area == "eventi":
        return fetch_all(
            """
            SELECT id, nome || ' - ' || data_evento AS label
            FROM eventi
            WHERE substr(data_evento, 1, 4) = ?
            ORDER BY data_evento, nome
            """,
            (str(work_year),),
        )
    return []


def build_participants_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    area = normalized(query_params, "area", "corsi")
    target_id = normalized(query_params, "target_id", "")
    area_options = render_static_options(
        [("corsi", "Corso"), ("oratorio", ORATORIO_LABEL), ("campi-estivi", ESTATE_LABEL), ("eventi", "Evento")],
        area,
        blank_label=None,
    )
    target_rows = participants_target_options(area, work_year)
    target_options = render_select_options(target_rows, target_id, "Seleziona...")
    current_query = dict(work_year_query(query_params))
    current_query["area"] = area
    if target_id:
        current_query["target_id"] = target_id

    lead_html = f"""
    <section class="card compact screen-only">
      <div class="card-head">
        <h2>Filtro partecipanti</h2>
        <p>Seleziona il corso, Oratorio, il Campo estivo o l'evento di cui vuoi visualizzare e stampare i partecipanti.</p>
      </div>
      <form method="get" action="/report/partecipanti" class="form-grid">
        <input type="hidden" name="anno_lavoro" value="{esc(work_year)}">
        {select_field("Area", "area", area_options, required_field=True, attrs={"onchange": "const target = this.form.querySelector('[name=target_id]'); if (target) target.value = ''; this.form.submit();"})}
        {select_field("Elemento", "target_id", target_options, required_field=True, wide=True)}
        <div class="form-actions">
          <button type="submit" class="button">Apri partecipanti</button>
        </div>
      </form>
    </section>
    """

    base_definition = {
        "title": "Partecipanti attivitÃ ",
        "current_path": "/report/partecipanti",
        "subtitle": "Visualizzazione e stampa partecipanti per corso, Oratorio, Campo estivo o evento.",
        "query": """
            SELECT '' AS codice_tesseramento, '' AS associato, '' AS telefono, '' AS email, '' AS data_iscrizione, '' AS stato_iscrizione, 0 AS quota
            WHERE 1 = 0
        """,
        "params": (),
        "sheet_name": "Partecipanti attivita",
        "export_name": "partecipanti.xlsx",
        "lead_html": lead_html,
        "filters": [{"label": "Anno di lavoro", "value": str(work_year)}],
        "columns": [
            ("codice_tesseramento", "Codice"),
            ("associato", "Tesserato"),
            ("telefono", "Telefono"),
            ("email", "Email"),
            ("data_iscrizione", "Data iscrizione"),
            ("stato_iscrizione", "Stato"),
            ("quota", "Quota", lambda value, _: money(value)),
        ],
    }

    if not target_id.isdigit():
        return base_definition

    if area == "corsi":
        course = fetch_one(
            f"""
            SELECT nome
            FROM corsi c
            WHERE c.id = ?
              AND {corso_year_relevance_sql('c')}
            """,
            (int(target_id), *corso_year_relevance_params(work_year)),
        )
        if course is None:
            return base_definition
        base_definition.update(
            {
                "subtitle": f"Partecipanti iscritti al corso {course['nome']}.",
                "query": f"""
                    SELECT
                        t.codice_tesseramento,
                        {associato_display_sql('a')} AS associato,
                        COALESCE(a.telefono, '') AS telefono,
                        COALESCE(a.email, '') AS email,
                        ic.data_iscrizione,
                        ic.stato_iscrizione,
                        ic.quota_mensile AS quota
                    FROM iscrizioni_corsi ic
                    JOIN associati a ON a.id = ic.associato_id
                    LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = ?
                    WHERE ic.corso_id = ?
                      AND {iscrizione_corso_year_relevance_sql('ic')}
                    ORDER BY a.cognome, a.nome
                """,
                "params": (work_year, int(target_id), *iscrizione_corso_year_relevance_params(work_year)),
                "export_name": f"partecipanti_corso_{slugify(course['nome'])}.xlsx",
                "filters": [
                    {"label": "Anno di lavoro", "value": str(work_year)},
                    {"label": "Corso", "value": course["nome"]},
                ],
            }
        )
        return base_definition

    if area == "campi-estivi":
        camp = fetch_one(
            """
            SELECT nome, anno
            FROM campi_estivi
            WHERE id = ?
              AND anno = ?
            """,
            (int(target_id), work_year),
        )
        if camp is None:
            return base_definition
        base_definition.update(
            {
                "subtitle": f"Partecipanti iscritti al Campo estivo {camp['anno']}.",
                "query": f"""
                    SELECT
                        t.codice_tesseramento,
                        {associato_display_sql('a')} AS associato,
                        COALESCE(a.telefono, '') AS telefono,
                        COALESCE(a.email, '') AS email,
                        ice.data_iscrizione,
                        ice.stato_iscrizione,
                        ice.quota_partecipazione AS quota
                    FROM iscrizioni_campi_estivi ice
                    JOIN associati a ON a.id = ice.associato_id
                    LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ?
                    WHERE ice.campo_estivo_id = ?
                    ORDER BY a.cognome, a.nome
                """,
                "params": (work_year, int(target_id)),
                "export_name": f"partecipanti_campo_estivo_{camp['anno']}.xlsx",
                "filters": [
                    {"label": "Anno di lavoro", "value": str(work_year)},
                    {"label": ESTATE_LABEL, "value": f"{ESTATE_LABEL} {camp['anno']}"},
                ],
            }
        )
        return base_definition

    if area == "oratorio":
        oratorio_row = fetch_one(
            """
            SELECT nome, anno
            FROM oratorio
            WHERE id = ?
              AND anno = ?
            """,
            (int(target_id), work_year),
        )
        if oratorio_row is None:
            return base_definition
        base_definition.update(
            {
                "subtitle": f"Partecipanti iscritti a Oratorio {oratorio_row['anno']}.",
                "query": f"""
                    SELECT
                        t.codice_tesseramento,
                        {associato_display_sql('a')} AS associato,
                        COALESCE(a.telefono, '') AS telefono,
                        COALESCE(a.email, '') AS email,
                        io.data_iscrizione,
                        io.stato_iscrizione,
                        io.quota_partecipazione AS quota
                    FROM iscrizioni_oratorio io
                    JOIN associati a ON a.id = io.associato_id
                    LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = ?
                    WHERE io.oratorio_id = ?
                    ORDER BY a.cognome, a.nome
                """,
                "params": (work_year, int(target_id)),
                "export_name": f"partecipanti_oratorio_{oratorio_row['anno']}.xlsx",
                "filters": [
                    {"label": "Anno di lavoro", "value": str(work_year)},
                    {"label": ORATORIO_LABEL, "value": f"{ORATORIO_LABEL} {oratorio_row['anno']}"},
                ],
            }
        )
        return base_definition

    if area == "eventi":
        event = fetch_one(
            """
            SELECT nome, data_evento
            FROM eventi
            WHERE id = ?
              AND substr(COALESCE(data_evento, ''), 1, 4) = ?
            """,
            (int(target_id), str(work_year)),
        )
        if event is None:
            return base_definition
        base_definition.update(
            {
                "subtitle": f"Partecipanti iscritti all'evento {event['nome']}.",
                "query": f"""
                    SELECT
                        t.codice_tesseramento,
                        {associato_display_sql('a')} AS associato,
                        COALESCE(a.telefono, '') AS telefono,
                        COALESCE(a.email, '') AS email,
                        ie.data_iscrizione,
                        ie.stato_iscrizione,
                        ie.quota_partecipazione AS quota
                    FROM iscrizioni_eventi ie
                    JOIN associati a ON a.id = ie.associato_id
                    JOIN eventi e ON e.id = ie.evento_id
                    LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
                    WHERE ie.evento_id = ?
                      AND substr(COALESCE(e.data_evento, ''), 1, 4) = ?
                    ORDER BY a.cognome, a.nome
                """,
                "params": (int(target_id), str(work_year)),
                "export_name": f"partecipanti_evento_{slugify(event['nome'])}.xlsx",
                "filters": [
                    {"label": "Anno di lavoro", "value": str(work_year)},
                    {"label": "Evento", "value": f"{event['nome']} ({event['data_evento']})"},
                ],
            }
        )
    return base_definition


def incassi_dataset_sql() -> str:
    return f"""
        SELECT
            'Tesseramento annuale' AS area,
            t.anno_sociale AS anno_riferimento,
            pt.data_pagamento,
            pt.importo,
            mp.nome AS metodo_pagamento,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            'Anno ' || t.anno_sociale AS riferimento,
            'tesseramenti' AS payment_type,
            pt.id AS payment_id,
            '' AS gruppo_ricevuta
        FROM pagamenti_tesseramenti pt
        JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
        JOIN associati a ON a.id = t.associato_id
        LEFT JOIN metodi_pagamento mp ON mp.id = pt.metodo_pagamento_id

        UNION ALL

        SELECT
            'Corso - quota mensile' AS area,
            r.anno AS anno_riferimento,
            prc.data_pagamento,
            prc.importo,
            mp.nome AS metodo_pagamento,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            c.nome || ' ' || printf('%04d-%02d', r.anno, r.mese) AS riferimento,
            'corsi-rata' AS payment_type,
            prc.id AS payment_id,
            COALESCE(prc.gruppo_ricevuta, '') AS gruppo_ricevuta
        FROM pagamenti_rate_corsi prc
        JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
        JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id

        UNION ALL

        SELECT
            'Campo estivo' AS area,
            ce.anno AS anno_riferimento,
            pce.data_pagamento,
            pce.importo,
            mp.nome AS metodo_pagamento,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            'Campo estivo ' || ce.anno AS riferimento,
            'campi-estivi' AS payment_type,
            pce.id AS payment_id,
            '' AS gruppo_ricevuta
        FROM pagamenti_campi_estivi pce
        JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
        JOIN associati a ON a.id = ice.associato_id
        JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = pce.metodo_pagamento_id

        UNION ALL

        SELECT
            'Oratorio' AS area,
            o.anno AS anno_riferimento,
            po.data_pagamento,
            po.importo,
            mp.nome AS metodo_pagamento,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            'Oratorio ' || o.anno AS riferimento,
            'oratorio' AS payment_type,
            po.id AS payment_id,
            '' AS gruppo_ricevuta
        FROM pagamenti_oratorio po
        JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
        JOIN associati a ON a.id = io.associato_id
        JOIN oratorio o ON o.id = io.oratorio_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = po.metodo_pagamento_id

        UNION ALL

        SELECT
            'Evento' AS area,
            CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER) AS anno_riferimento,
            pe.data_pagamento,
            pe.importo,
            mp.nome AS metodo_pagamento,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            e.nome AS riferimento,
            'eventi' AS payment_type,
            pe.id AS payment_id,
            '' AS gruppo_ricevuta
        FROM pagamenti_eventi pe
        JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
        JOIN associati a ON a.id = ie.associato_id
        JOIN eventi e ON e.id = ie.evento_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
        LEFT JOIN metodi_pagamento mp ON mp.id = pe.metodo_pagamento_id
    """


def lookup_label(table_name: str, record_id: str, label_query: str, default: str = "") -> str:
    if not record_id.isdigit():
        return default
    row = fetch_one(label_query, (int(record_id),))
    if row is None:
        return default
    return row[0]


def report_link(label: str, path: str, query_params: dict[str, str]) -> str:
    return f'<a href="{esc(with_query(path, query_params))}">{esc(label)}</a>'


def report_action_link(label: str, path: str, query_params: dict[str, str]) -> str:
    return action_links_html(extra_links=[(with_query(path, query_params), label)])


def format_money_plain(value: object) -> str:
    return money(value).replace(" EUR", "")


def chart_card(title: str, subtitle: str, chart_html: str, legend_html: str = "") -> str:
    return f"""
    <section class="card chart-card">
      <div class="card-head">
        <div class="card-head-copy">
          <h2>{esc(title)}</h2>
          <p>{esc(subtitle)}</p>
        </div>
      </div>
      <div class="chart-wrap">
        {chart_html}
        {legend_html}
      </div>
    </section>
    """


def pie_chart_svg(paid_total: float, due_total: float, incassi_url: str, scadenze_url: str) -> str:
    total = max(paid_total + due_total, 0.0)
    if total <= 0:
        return '<div class="empty-state">Nessun movimento disponibile per il grafico.</div>'

    circumference = 2 * 3.14159 * 72
    paid_ratio = paid_total / total
    paid_dash = circumference * paid_ratio
    due_dash = max(circumference - paid_dash, 0)
    svg = f"""
    <svg class="chart-svg pie-chart" viewBox="0 0 220 220" role="img" aria-label="Grafico a torta incassi e scadenze">
      <circle cx="110" cy="110" r="72" fill="none" stroke="#ffe6ce" stroke-width="26" class="chart-clickable" ondblclick="window.location.href='{esc(scadenze_url)}'"><title>Scadenze: {esc(money(due_total))}</title></circle>
      <circle cx="110" cy="110" r="72" fill="none" stroke="#ef7f1a" stroke-width="26" stroke-linecap="round"
              stroke-dasharray="{paid_dash:.2f} {due_dash:.2f}" transform="rotate(-90 110 110)" class="chart-clickable" ondblclick="window.location.href='{esc(incassi_url)}'"><title>Incassi: {esc(money(paid_total))}</title></circle>
      <circle cx="110" cy="110" r="48" fill="#ffffff"></circle>
      <text x="110" y="82" text-anchor="middle" class="chart-subtle">Totale complessivo<title>Totale complessivo: {esc(money(total))}</title></text>
      <text x="110" y="104" text-anchor="middle" class="chart-number">{esc(format_money_plain(total))}<title>Totale complessivo: {esc(money(total))}</title></text>
      <text x="110" y="126" text-anchor="middle" class="chart-label chart-clickable" ondblclick="window.location.href='{esc(incassi_url)}'">Incassi {esc(format_money_plain(paid_total))}<title>Incassi: {esc(money(paid_total))}</title></text>
      <text x="110" y="144" text-anchor="middle" class="chart-subtle chart-clickable" ondblclick="window.location.href='{esc(scadenze_url)}'">Scadenze {esc(format_money_plain(due_total))}<title>Scadenze: {esc(money(due_total))}</title></text>
    </svg>
    """
    legend = """
    <div class="chart-legend">
      <span><i class="legend-dot incassi"></i> Incassi</span>
      <span><i class="legend-dot scadenze"></i> Scadenze</span>
    </div>
    """
    return svg + legend


def bar_chart_svg(rows: list[tuple[str, float, float, str, str]]) -> str:
    if not rows:
        return '<div class="empty-state">Nessun movimento disponibile per il grafico.</div>'

    def chart_label_lines(label: str) -> list[str]:
        words = str(label or "").split()
        if len(words) <= 1:
            return [str(label or "")]
        lines: list[str] = []
        current = ""
        for word in words:
            candidate = word if not current else f"{current} {word}"
            if len(candidate) <= 14 or not current:
                current = candidate
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)
        if len(lines) > 2:
            lines = [lines[0], " ".join(lines[1:])]
        return lines[:2]

    max_value = max(max(incassi, scadenze) for _, incassi, scadenze, _, _ in rows) or 1
    column_step = 138
    width = max(480, len(rows) * column_step + 52)
    height = 288
    bars = []
    labels = []
    base_y = 198
    for index, (label, incassi, scadenze, incassi_url, scadenze_url) in enumerate(rows):
        x = 40 + index * column_step
        incassi_height = (incassi / max_value) * 128 if max_value else 0
        scadenze_height = (scadenze / max_value) * 128 if max_value else 0
        bars.append(
            f'<rect x="{x}" y="{base_y - incassi_height:.1f}" width="32" height="{incassi_height:.1f}" rx="8" fill="#ef7f1a" '
            f'class="chart-clickable" ondblclick="window.location.href=\'{esc(incassi_url)}\'"><title>{esc(label)} - Incassi: {esc(money(incassi))}</title></rect>'
        )
        bars.append(
            f'<rect x="{x + 40}" y="{base_y - scadenze_height:.1f}" width="32" height="{scadenze_height:.1f}" rx="8" fill="#ffd2a8" '
            f'class="chart-clickable" ondblclick="window.location.href=\'{esc(scadenze_url)}\'"><title>{esc(label)} - Scadenze: {esc(money(scadenze))}</title></rect>'
        )
        label_lines = chart_label_lines(label)
        tspans = []
        for line_index, line in enumerate(label_lines):
            dy = "0" if line_index == 0 else "1.18em"
            tspans.append(f'<tspan x="{x + 36}" dy="{dy}">{esc(line)}</tspan>')
        labels.append(
            f'<text x="{x + 36}" y="226" text-anchor="middle" class="chart-axis" '
            f'style="font-size:15px;font-weight:700;fill:#6b4a32;">{"".join(tspans)}'
            f'<title>{esc(label)} - Totale area: {esc(money(incassi + scadenze))}</title></text>'
        )

    total_incassi = sum(incassi for _, incassi, _, _, _ in rows)
    total_scadenze = sum(scadenze for _, _, scadenze, _, _ in rows)
    grand_total = total_incassi + total_scadenze

    return f"""
    <div class="bar-chart-wrap">
      <svg class="chart-svg bar-chart" viewBox="0 0 {width} {height}" role="img" aria-label="Grafico a barre incassi e scadenze per area">
        <line x1="24" y1="{base_y}" x2="{width - 16}" y2="{base_y}" stroke="#d8c1aa" stroke-width="1.5"></line>
        {''.join(bars)}
        {''.join(labels)}
      </svg>
      <div class="chart-legend">
        <span><i class="legend-dot incassi"></i> Incassi</span>
        <span><i class="legend-dot scadenze"></i> Scadenze</span>
      </div>
      <div class="chart-summary-strip">
        <article class="chart-summary-card incassi">
          <span>Totale incassi</span>
          <strong>{esc(money(total_incassi))}</strong>
        </article>
        <article class="chart-summary-card scadenze">
          <span>Totale scadenze</span>
          <strong>{esc(money(total_scadenze))}</strong>
        </article>
        <article class="chart-summary-card totale">
          <span>Totale complessivo</span>
          <strong>{esc(money(grand_total))}</strong>
        </article>
      </div>
    </div>
    """


def dashboard_charts(query_params: dict[str, str]) -> str:
    work_year = current_work_year(query_params)
    date_from, date_to = year_start_end(work_year)
    incassi_total_url = with_query(
        "/report/incassi",
        {"anno_lavoro": str(work_year), "date_from": date_from, "date_to": date_to},
    )
    scadenze_total_url = with_query("/report/scadenze", {"anno_lavoro": str(work_year)})
    incassi_total = float(
        fetch_scalar(
            f"""
            SELECT COALESCE(SUM(importo), 0)
            FROM ({incassi_dataset_sql()}) incassi
            WHERE anno_riferimento = ?
              AND data_pagamento BETWEEN ? AND ?
            """,
            (work_year, date_from, date_to),
        )
        or 0
    )
    scadenze_total = float(
        fetch_scalar(
            """
            SELECT COALESCE(SUM(saldo_residuo), 0)
            FROM v_scadenze_da_incassare
            WHERE substr(COALESCE(scadenza, ''), 1, 4) = ?
              AND area <> 'Corso - iscrizione'
            """,
            (str(work_year),),
        )
        or 0
    )
    area_rows = fetch_all(
        f"""
        WITH incassi AS (
            SELECT area, COALESCE(SUM(importo), 0) AS totale_incassi
            FROM ({incassi_dataset_sql()}) dati
            WHERE anno_riferimento = ?
              AND data_pagamento BETWEEN ? AND ?
            GROUP BY area
        ),
        scadenze AS (
            SELECT
                CASE WHEN area = 'Estate' THEN 'Campo estivo' ELSE area END AS area,
                COALESCE(SUM(saldo_residuo), 0) AS totale_scadenze
            FROM v_scadenze_da_incassare
            WHERE substr(COALESCE(scadenza, ''), 1, 4) = ?
              AND area <> 'Corso - iscrizione'
            GROUP BY 1
        ),
        aree AS (
            SELECT area FROM incassi
            UNION
            SELECT area FROM scadenze
        )
        SELECT
            area,
            COALESCE((SELECT totale_incassi FROM incassi WHERE incassi.area = aree.area), 0) AS totale_incassi,
            COALESCE((SELECT totale_scadenze FROM scadenze WHERE scadenze.area = aree.area), 0) AS totale_scadenze
        FROM aree
        ORDER BY area
        """,
        (work_year, date_from, date_to, str(work_year)),
    )
    chart_rows = [
        (
            row["area"],
            float(row["totale_incassi"] or 0),
            float(row["totale_scadenze"] or 0),
            with_query(
                "/report/incassi",
                {
                    "anno_lavoro": str(work_year),
                    "date_from": date_from,
                    "date_to": date_to,
                    "area": row["area"],
                },
            ),
            with_query(
                "/report/scadenze",
                {
                    "anno_lavoro": str(work_year),
                    "area": row["area"],
                },
            ),
        )
        for row in area_rows
    ]
    return f"""
    <div class="charts-grid">
      {chart_card("Incassi e Scadenze Totali", f"Anno di lavoro {work_year}. Doppio clic sui valori per aprire il report collegato.", pie_chart_svg(incassi_total, scadenze_total, incassi_total_url, scadenze_total_url))}
      {chart_card("Incassi e Scadenze per Area", "Doppio clic su ogni barra per aprire il report incassi o scadenze dell'area selezionata.", bar_chart_svg(chart_rows))}
    </div>
    """


def associato_detail_href(associato_id: object, query_params: dict[str, str]) -> str:
    return with_query(f"/report/associato/{associato_id}", work_year_query(query_params))


def report_share_message(definition: dict, rows: list[sqlite3.Row], *, max_length: int | None = None) -> str:
    filter_lines = [f"{filter_row['label']}: {filter_row['value']}" for filter_row in definition.get("filters", [])]
    summary_parts: list[str] = []
    summary_rows = summary_rows_for_table(rows, definition["columns"])
    if summary_rows:
        summary_row = summary_rows[0]
        for index, cell in enumerate(summary_row):
            if not cell or cell == "Totali":
                continue
            summary_parts.append(f"{definition['columns'][index][1]} {cell}")

    lines = [
        f"Report {definition['title']} - {APP_NAME}",
        f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ]
    if filter_lines:
        lines.append("Filtri: " + " | ".join(filter_lines))
    lines.append(f"Righe visualizzate: {len(rows)}")
    if summary_parts:
        lines.append("Totali: " + " | ".join(summary_parts))

    detail_lines = [report_share_row_line(row, definition["columns"], index) for index, row in enumerate(rows, start=1)]
    footer = "Cordiali saluti,\nOratorio Carlo Acutis"
    if detail_lines:
        lines.append("Dettaglio righe visibili:")
        included = 0
        for detail_line in detail_lines:
            candidate = "\n".join(lines + [detail_line, footer])
            if max_length and len(candidate) > max_length:
                break
            lines.append(detail_line)
            included += 1
        if included < len(detail_lines):
            remaining = len(detail_lines) - included
            notice = (
                f"Dettaglio parziale per limiti del messaggio: incluse {included} righe su {len(detail_lines)}."
                if included
                else f"Dettaglio troppo esteso per il canale scelto: {remaining} righe non incluse nel messaggio."
            )
            candidate = "\n".join(lines + [notice, footer])
            if not max_length or len(candidate) <= max_length:
                lines.append(notice)
    lines.append(footer)
    return "\n".join(lines)


def build_report_share_payload(
    report_key: str,
    query_params: dict[str, str],
    recipient_id: str,
    channel: str,
) -> dict[str, str] | None:
    if channel not in {"email", "whatsapp", "whatsapp-group"}:
        raise ValueError("Canale di invio non valido.")
    recipient = None
    if channel == "whatsapp":
        recipient = next((row for row in report_recipient_rows() if row["id"] == recipient_id), None)
        if recipient is None:
            raise ValueError("Destinatario non valido.")

    definition, rows = build_report_table_state(report_key, query_params)
    subject = f"Report {definition['title']} - {APP_NAME}"
    message = report_share_message(
        definition,
        rows,
        max_length=6000 if channel == "email" else 2500,
    )

    if channel == "whatsapp-group":
        return {
            "channel": "whatsapp-group",
            "url": f"https://web.whatsapp.com/send?text={quote(message)}",
            "message": message,
        }

    if channel == "email":
        recipient_email = ""
        if recipient_id:
            recipient = next((row for row in report_recipient_rows() if row["id"] == recipient_id), None)
            if recipient is None:
                raise ValueError("Destinatario non valido.")
            if not recipient.get("email"):
                raise ValueError("Il destinatario selezionato non ha un indirizzo email registrato.")
            recipient_email = str(recipient["email"])
        return {
            "channel": "email",
            "url": f"mailto:{recipient_email}?subject={quote(subject)}&body={quote(message)}",
        }

    if not recipient.get("whatsapp_phone"):
        raise ValueError("Il destinatario selezionato non ha un numero cellulare registrato.")
    return {
        "channel": "whatsapp",
        "url": f"https://wa.me/{recipient['whatsapp_phone']}?text={quote(message)}",
    }


def report_toolbar(
    report_key: str,
    query_params: dict[str, str],
    *,
    definition: dict | None = None,
    rows: list[sqlite3.Row] | None = None,
) -> str:
    export_url = with_query(f"/export/excel/{report_key}", query_params)
    pdf_url = with_query(f"/export/pdf/{report_key}", query_params)
    table_id = f"report-table-{slugify(report_key)}"
    search_value = report_search_term(query_params)
    email_button = ""
    whatsapp_group_button = ""
    toolbar_copy = "Esporta il report in Excel o PDF, oppure stampalo direttamente dal browser."
    if definition is not None and rows is not None:
        toolbar_copy = (
            "Esporta il report in Excel o PDF, stampalo oppure prepara l'invio via email o WhatsApp. "
            "Per WhatsApp il testo viene preparato e passato a WhatsApp Web, cosi puoi scegliere il gruppo e inviare subito."
        )
        whatsapp_group_button = (
            f'<button type="button" class="button action" '
            f'onclick="return shareReport(this)" '
            f'data-channel="whatsapp-group" data-report-key="{esc(report_key)}">'
            f'Invia WhatsApp</button>'
        )
        email_button = (
            f'<button type="button" class="button action" '
            f'onclick="return shareReport(this)" '
            f'data-channel="email" data-report-key="{esc(report_key)}">Prepara email</button>'
        )
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Azioni report</span>
        <p>{esc(toolbar_copy)}</p>
      </div>
      <div class="report-toolbar-actions">
        <label class="report-search">
          <span>Cerca</span>
          <input type="search" class="control" data-target-table="{esc(table_id)}" value="{esc(search_value)}" placeholder="Filtra tutte le colonne..." oninput="handleReportSearch(this)">
        </label>
        <div class="report-toolbar-action-groups">
          <div class="report-toolbar-action-row">
            <a class="button action" href="{esc(export_url)}" data-search-link="excel" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta Excel</a>
            <a class="button action" href="{esc(pdf_url)}" data-search-link="pdf" data-export-table="{esc(table_id)}" onclick="return applyTableExportStateToLink(this)">Esporta PDF</a>
            <button type="button" class="button action" onclick="window.print()">Stampa report</button>
          </div>
          <div class="report-toolbar-action-row report-toolbar-action-row-secondary">
            {email_button}
            {whatsapp_group_button}
          </div>
        </div>
      </div>
    </section>
    """


def build_scadenze_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    associato_id = normalized(query_params, "associato_id", "")
    area = normalized(query_params, "area", "")
    clauses = ["substr(COALESCE(scadenza, ''), 1, 4) = ?", "area <> 'Corso - iscrizione'"]
    params: list[object] = [str(work_year)]
    filters = [{"label": "Anno di lavoro", "value": str(work_year)}]
    lead_html = ""

    if associato_id.isdigit():
        clauses.append("associato_id = ?")
        params.append(int(associato_id))
        associato_label = lookup_label(
            "associati",
            associato_id,
            f"SELECT {associato_display_sql('associati')} AS associato FROM associati WHERE id = ?",
            "Tesserato selezionato",
        )
        filters.append({"label": "Associato", "value": associato_label})
        lead_html = f"""
            <section class="card compact screen-only">
              <div class="card-head">
                <h2>Filtro tesserato</h2>
                <p>Vista limitata al tesserato selezionato dalla dashboard.</p>
              </div>
              <div class="empty-state">{esc(associato_label)}</div>
            </section>
        """

    if area:
        clauses.append("area = ?")
        params.append(area)
        filters.append({"label": "Area", "value": area})

    return {
        "title": "Scadenze da incassare",
        "current_path": "/report/scadenze",
        "subtitle": f"Vista unica con insoluti di tesseramenti, corsi, Oratorio, Campo estivo ed eventi per l'anno {work_year}.",
        "query": f"""
            SELECT area, riferimento, codice_tesseramento, associato, scadenza, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM (
                SELECT
                    CASE WHEN area = 'Estate' THEN 'Campo estivo' ELSE area END AS area,
                    riferimento,
                    codice_tesseramento,
                    associato_id,
                    associato,
                    scadenza,
                    importo_dovuto,
                    importo_pagato,
                    saldo_residuo,
                    stato_pagamento
                FROM v_scadenze_da_incassare
            ) dati
            WHERE {' AND '.join(clauses)}
            ORDER BY scadenza, associato, area
        """,
        "params": tuple(params),
        "sheet_name": "Scadenze",
        "export_name": "scadenze_da_incassare.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("area", "Area"),
            ("riferimento", "Riferimento"),
            ("associato", "Tesserato"),
            ("scadenza", "Scadenza"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_tesseramenti_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    associato_id = normalized(query_params, "associato_id", "")
    associati = associati_options()
    clauses = ["anno_sociale = ?"]
    params: list[object] = [work_year]
    filters = [{"label": "Anno di lavoro", "value": str(work_year)}]
    if associato_id.isdigit():
        clauses.append("associato_id = ?")
        params.append(int(associato_id))
        associato_label = lookup_label(
            "associati",
            associato_id,
            f"SELECT {associato_display_sql('associati')} AS associato FROM associati WHERE id = ?",
            "",
        )
        if associato_label:
            filters.append({"label": "Associato", "value": associato_label})

    return {
        "title": "Situazione tesseramenti",
        "current_path": "/report/tesseramenti",
        "subtitle": f"Situazione completa dei tesseramenti per l'anno {work_year}.",
        "query": f"""
            SELECT codice_tesseramento, associato, anno_sociale, data_tesseramento, data_scadenza, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM v_tesseramenti_saldo
            WHERE {' AND '.join(clauses)}
            ORDER BY associato
        """,
        "params": tuple(params),
        "sheet_name": "Tesseramenti",
        "export_name": f"situazione_tesseramenti_{work_year}.xlsx",
        "filters": filters,
        "lead_html": "",
        "columns": with_codice_tesserato([
            ("codice_tesseramento", "Codice tesseramento"),
            ("associato", "Associato"),
            ("anno_sociale", "Anno"),
            ("data_tesseramento", "Data tesseramento"),
            ("data_scadenza", "Scadenza"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_registro_attivita_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    date_from, date_to = year_start_end(work_year)
    data_da = normalized(query_params, "data_da", date_from) or date_from
    data_a = normalized(query_params, "data_a", date_to) or date_to
    username = normalized(query_params, "username", "")
    associato_id = normalized(query_params, "associato_id", "")
    attivita = normalized(query_params, "attivita", "")

    clauses = ["date(substr(ra.data_ora, 1, 10)) BETWEEN date(?) AND date(?)"]
    params: list[object] = [data_da, data_a]
    filters = [
        {"label": "Data da", "value": data_da},
        {"label": "Data a", "value": data_a},
    ]

    if username:
        clauses.append("ra.username = ?")
        params.append(username)
        filters.append({"label": "Utente", "value": username})

    if associato_id.isdigit():
        clauses.append("ra.associato_id = ?")
        params.append(int(associato_id))
        associato_label = lookup_label(
            "associati",
            associato_id,
            f"SELECT trim(codice_associato || ' - ' || {associato_display_sql('')}) AS associato FROM associati WHERE id = ?",
            "",
        )
        if associato_label:
            filters.append({"label": "Associato", "value": associato_label})

    if attivita:
        clauses.append("ra.descrizione_attivita = ?")
        params.append(attivita)
        filters.append({"label": "Attivita", "value": attivita})

    lead_html = f"""
        <section class="card compact screen-only">
          <div class="card-head">
            <h2>Filtro registro attivita</h2>
            <p>Filtra il registro per intervallo date, utente e associato coinvolto.</p>
          </div>
          <form method="get" action="/report/registro-attivita" class="form-grid">
            <input type="hidden" name="anno_lavoro" value="{esc(work_year)}">
            {input_field("Data da", "data_da", input_type="date", value=data_da, required_field=True)}
            {input_field("Data a", "data_a", input_type="date", value=data_a, required_field=True)}
            {select_field("Utente", "username", render_select_options(activity_log_user_options(), username, blank_label="Tutti gli utenti"))}
            {select_field("Attivita", "attivita", render_select_options(activity_log_attivita_options(), attivita, blank_label="Tutte le attivita"))}
            {select_field("Tesserato", "associato_id", render_select_options(activity_log_associato_options(), associato_id, blank_label="Tutti i tesserati", data_keys=["search_text", "autocomplete_label"]), wide=True, searchable=True, search_placeholder="Cerca tesserato nel registro...")}
            <div class="form-actions">
              <button type="submit" class="button">Aggiorna report</button>
            </div>
          </form>
        </section>
    """

    return {
        "title": "Registro attivita",
        "current_path": "/report/registro-attivita",
        "subtitle": "Storico operativo del gestionale con utente, associato coinvolto, dispositivo e risultato dell'attivita.",
        "query": f"""
            SELECT
                data_ora,
                username,
                COALESCE(t.codice_tesseramento, '') AS codice_tesseramento,
                COALESCE(associato_nominativo, '') AS associato,
                nome_pc,
                categoria,
                descrizione_attivita,
                dettaglio,
                esito,
                percorso
            FROM registro_attivita ra
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ra.associato_id AND t.anno_sociale = ra.anno_lavoro
            WHERE {' AND '.join(clauses)}
            ORDER BY ra.data_ora DESC, ra.id DESC
        """,
        "params": tuple(params),
        "sheet_name": "Registro attivita",
        "export_name": "registro_attivita.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("data_ora", "Data ora"),
            ("username", "Utente"),
            ("associato", "Associato"),
            ("nome_pc", "Nome PC"),
            ("categoria", "Categoria"),
            ("descrizione_attivita", "Attivita"),
            ("dettaglio", "Dettaglio"),
            ("esito", "Esito"),
            ("percorso", "Percorso"),
        ]),
    }


def build_corsi_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    course_id = normalized(query_params, "corso_id", "")
    comp_from = normalized(query_params, "comp_from", f"{work_year}-01-01")
    comp_to = normalized(query_params, "comp_to", f"{work_year}-12-31")
    course_name = lookup_label("corsi", course_id, "SELECT nome FROM corsi WHERE id = ?", "")
    clauses = ["anno = ?", "date(printf('%04d-%02d-01', anno, mese)) BETWEEN date(?) AND date(?)"]
    params: list[object] = [work_year, comp_from, comp_to]
    filters = [
        {"label": "Anno di lavoro", "value": str(work_year)},
        {"label": "Competenza da", "value": comp_from},
        {"label": "Competenza a", "value": comp_to},
    ]
    if course_name:
        clauses.append("corso = ?")
        params.append(course_name)
        filters.append({"label": "Corso", "value": course_name})

    lead_html = f"""
        <section class="card compact screen-only">
          <div class="card-head">
            <h2>Filtro quote mensili corsi</h2>
            <p>Filtra per corso e per competenza da data a data.</p>
          </div>
          <form method="get" action="/report/corsi" class="form-grid">
            <input type="hidden" name="anno_lavoro" value="{esc(work_year)}">
            {select_field("Corso", "corso_id", render_select_options(corsi_report_options(work_year), course_id), wide=True)}
            {input_field("Competenza dal", "comp_from", input_type="date", value=comp_from, required_field=True)}
            {input_field("Competenza al", "comp_to", input_type="date", value=comp_to, required_field=True)}
            <div class="form-actions">
              <button type="submit" class="button">Aggiorna report</button>
            </div>
          </form>
        </section>
    """

    return {
        "title": "Situazione corsi",
        "current_path": "/report/corsi",
        "subtitle": f"Vista operativa dei pagamenti mensili per competenza nell'anno {work_year}.",
        "query": f"""
            SELECT corso, codice_tesseramento, associato, competenza, data_scadenza, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM v_rate_corsi_saldo
            WHERE {' AND '.join(clauses)}
            ORDER BY anno DESC, mese DESC, corso, associato
        """,
        "params": tuple(params),
        "sheet_name": "Situazione corsi",
        "export_name": "situazione_corsi.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("corso", "Corso"),
            ("associato", "Associato"),
            ("competenza", "Competenza"),
            ("data_scadenza", "Scadenza"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_campi_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    clauses = ["anno = ?"]
    params: list[object] = [work_year]
    filters = [{"label": "Anno di lavoro", "value": str(work_year)}]
    lead_html = ""

    return {
        "title": "Situazione campo estivo",
        "current_path": "/report/campi-estivi",
        "subtitle": f"Iscritti al Campo estivo con quota una tantum e situazione pagamenti per l'anno {work_year}.",
        "query": f"""
            SELECT anno, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM v_campi_estivi_saldo
            WHERE {' AND '.join(clauses)}
            ORDER BY associato
        """,
        "params": tuple(params),
        "sheet_name": "Situazione campo estivo",
        "export_name": "situazione_campo_estivo.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("anno", "Anno"),
            ("associato", "Associato"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_oratorio_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    return {
        "title": "Situazione oratorio",
        "current_path": "/report/oratorio",
        "subtitle": f"Iscritti a Oratorio con contributo una tantum e situazione pagamenti per l'anno {work_year}.",
        "query": """
            SELECT anno, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM v_oratorio_saldo
            WHERE anno = ?
            ORDER BY associato
        """,
        "params": (work_year,),
        "sheet_name": "Situazione oratorio",
        "export_name": "situazione_oratorio.xlsx",
        "filters": [{"label": "Anno di lavoro", "value": str(work_year)}],
        "lead_html": "",
        "columns": with_codice_tesserato([
            ("associato", "Associato"),
            ("anno", "Anno"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_eventi_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    evento_id = normalized(query_params, "evento_id", "")
    event_name = ""
    if evento_id.isdigit():
        event_row = fetch_one(
            """
            SELECT nome
            FROM eventi
            WHERE id = ?
              AND substr(COALESCE(data_evento, ''), 1, 4) = ?
            """,
            (int(evento_id), str(work_year)),
        )
        if event_row is not None:
            event_name = event_row["nome"]
    clauses = ["substr(data_evento, 1, 4) = ?"]
    params: list[object] = [str(work_year)]
    filters = [{"label": "Anno di lavoro", "value": str(work_year)}]
    if event_name:
        clauses.append("evento = ?")
        params.append(event_name)
        filters.append({"label": "Evento", "value": event_name})

    lead_html = f"""
        <section class="card compact screen-only">
          <div class="card-head">
            <h2>Filtro eventi</h2>
            <p>Seleziona l'evento da visualizzare nel report.</p>
          </div>
          <form method="get" action="/report/eventi" class="form-grid">
            <input type="hidden" name="anno_lavoro" value="{esc(work_year)}">
            {select_field("Evento", "evento_id", render_select_options(eventi_options(work_year), evento_id), wide=True)}
            <div class="form-actions">
              <button type="submit" class="button">Aggiorna report</button>
            </div>
          </form>
        </section>
    """

    return {
        "title": "Situazione eventi",
        "current_path": "/report/eventi",
        "subtitle": f"Partecipanti agli eventi con quota una tantum e stato pagamenti per l'anno {work_year}.",
        "query": f"""
            SELECT evento, tipologia, data_evento, codice_tesseramento, associato, importo_dovuto, importo_pagato, saldo_residuo, stato_pagamento
            FROM v_eventi_saldo
            WHERE {' AND '.join(clauses)}
            ORDER BY data_evento DESC, evento, associato
        """,
        "params": tuple(params),
        "sheet_name": "Situazione eventi",
        "export_name": "situazione_eventi.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("evento", "Evento"),
            ("tipologia", "Tipologia"),
            ("data_evento", "Data"),
            ("associato", "Associato"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value)),
            ("importo_pagato", "Pagato", lambda value, _: money(value)),
            ("saldo_residuo", "Residuo", lambda value, _: money(value)),
            ("stato_pagamento", "Stato"),
        ]),
    }


def build_incassi_report_definition(query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    default_from, default_to = year_start_end(work_year)
    date_from = normalized(query_params, "date_from", default_from) or default_from
    date_to = normalized(query_params, "date_to", default_to) or default_to
    area = normalized(query_params, "area", "")
    riferimento = normalized(query_params, "riferimento", "")
    associato_id = normalized(query_params, "associato_id", "")
    clauses = ["anno_riferimento = ?", "data_pagamento BETWEEN ? AND ?"]
    params: list[object] = [work_year, date_from, date_to]
    filters = [
        {"label": "Anno di lavoro", "value": str(work_year)},
        {"label": "Intervallo", "value": f"{date_from} -> {date_to}"},
    ]
    if area:
        clauses.append("area = ?")
        params.append(area)
        filters.append({"label": "Area", "value": area})
    if riferimento:
        clauses.append("LOWER(riferimento) LIKE ?")
        params.append(f"%{riferimento.lower()}%")
        filters.append({"label": "Riferimento", "value": riferimento})
    if associato_id.isdigit():
        clauses.append("associato_id = ?")
        params.append(int(associato_id))
        associato_label = lookup_label(
            "associati",
            associato_id,
            f"SELECT {associato_display_sql('associati')} AS associato FROM associati WHERE id = ?",
            "",
        )
        if associato_label:
            filters.append({"label": "Associato", "value": associato_label})

    lead_html = f"""
        <section class="card compact screen-only">
          <div class="card-head">
            <h2>Filtro report incassi</h2>
            <p>Filtra per intervallo, area e riferimento.</p>
          </div>
          <form method="get" action="/report/incassi" class="form-grid">
            <input type="hidden" name="anno_lavoro" value="{esc(work_year)}">
            <input type="hidden" name="associato_id" value="{esc(associato_id)}">
            {input_field("Dal", "date_from", input_type="date", value=date_from, required_field=True)}
            {input_field("Al", "date_to", input_type="date", value=date_to, required_field=True)}
            {select_field("Area", "area", render_static_options([
                ("", "Tutte le aree"),
                ("Tesseramento annuale", "Tesseramento annuale"),
                ("Corso - quota mensile", "Corso - quota mensile"),
                ("Oratorio", ORATORIO_LABEL),
                ("Campo estivo", ESTATE_LABEL),
                ("Evento", "Evento"),
            ], area, blank_label=None))}
            {input_field("Riferimento", "riferimento", value=riferimento, placeholder="Filtra per testo")}
            <div class="form-actions">
              <button type="submit" class="button">Aggiorna report</button>
            </div>
          </form>
        </section>
    """

    return {
        "title": "Incassi",
        "current_path": "/report/incassi",
        "subtitle": "Movimenti registrati nell'intervallo selezionato.",
        "query": f"""
            SELECT area, data_pagamento, importo, metodo_pagamento, codice_tesseramento, associato, riferimento, payment_type, payment_id, gruppo_ricevuta
            FROM ({incassi_dataset_sql()}) incassi
            WHERE {' AND '.join(clauses)}
            ORDER BY data_pagamento, area, associato
        """,
        "params": tuple(params),
        "sheet_name": "Incassi",
        "export_name": f"report_incassi_{date_from}_{date_to}.xlsx",
        "filters": filters,
        "lead_html": lead_html,
        "columns": with_codice_tesserato([
            ("area", "Area"),
            ("data_pagamento", "Data"),
            ("importo", "Importo", lambda value, _: money(value)),
            ("metodo_pagamento", "Metodo"),
            ("associato", "Associato"),
            ("riferimento", "Riferimento"),
            (
                "payment_id",
                "Azioni",
                lambda value, row: (
                    report_action_link(
                        "Ricevuta",
                        f"/ricevute/corsi-rate-gruppo/{row['gruppo_ricevuta']}"
                        if row["gruppo_ricevuta"]
                        else f"/ricevute/{row['payment_type']}/{row['payment_id']}",
                        work_year_query(query_params),
                    )
                    if row["payment_id"]
                    else ""
                ),
            ),
        ]),
    }


def storico_tesseramenti_year_range(query_params: dict[str, str]) -> tuple[int, int]:
    work_year = current_work_year(query_params)
    default_start = work_year - 4
    default_end = work_year
    raw_start = normalized(query_params, "anno_da", "")
    raw_end = normalized(query_params, "anno_a", "")
    start_year = int(raw_start) if raw_start.isdigit() else default_start
    end_year = int(raw_end) if raw_end.isdigit() else default_end
    if start_year > end_year:
        start_year, end_year = end_year, start_year
    return start_year, end_year


def storico_tesseramenti_rows(start_year: int, end_year: int) -> list[dict[str, object]]:
    associati_rows = fetch_all(
        f"""
        SELECT
            associati.id,
            COALESCE(associati.codice_associato, '') AS codice_associato,
            {associato_display_sql('associati')} AS associato
        FROM associati
        ORDER BY associati.cognome, associati.nome
        """
    )
    tesseramenti_rows = fetch_all(
        """
        SELECT associato_id, anno_sociale
        FROM tesseramenti_annuali
        WHERE anno_sociale BETWEEN ? AND ?
        """,
        (start_year, end_year),
    )
    membership = {
        (int(row["associato_id"]), int(row["anno_sociale"]))
        for row in tesseramenti_rows
    }
    years = list(range(start_year, end_year + 1))
    rows: list[dict[str, object]] = []
    for associato in associati_rows:
        item: dict[str, object] = {
            "codice_associato": str(associato["codice_associato"] or ""),
            "associato": str(associato["associato"] or ""),
        }
        associato_id = int(associato["id"])
        for year in years:
            item[str(year)] = "Si" if (associato_id, year) in membership else ""
        rows.append(item)
    return rows


def build_storico_tesseramenti_report_definition(query_params: dict[str, str]) -> dict:
    start_year, end_year = storico_tesseramenti_year_range(query_params)
    year_columns = [(str(year), str(year)) for year in range(start_year, end_year + 1)]
    lead_html = f"""
        <section class="card compact screen-only">
          <div class="card-head">
            <h2>Filtro storico tesseramenti</h2>
            <p>Seleziona l'intervallo di anni da confrontare per tutti gli associati.</p>
          </div>
          <form method="get" action="/report/storico-tesseramenti" class="form-grid">
            <input type="hidden" name="anno_lavoro" value="{esc(str(current_work_year(query_params)))}">
            {input_field("Anno da", "anno_da", input_type="number", value=str(start_year), required_field=True, minimum="2000")}
            {input_field("Anno a", "anno_a", input_type="number", value=str(end_year), required_field=True, minimum="2000")}
            <div class="form-actions">
              <button type="submit" class="button">Aggiorna report</button>
            </div>
          </form>
        </section>
    """
    return {
        "title": "Storico tesseramenti",
        "current_path": "/report/storico-tesseramenti",
        "subtitle": "Tabella storica con evidenza del tesseramento per ciascun associato negli anni selezionati.",
        "rows_builder": lambda: storico_tesseramenti_rows(start_year, end_year),
        "sheet_name": "Storico tesser.",
        "export_name": f"storico_tesseramenti_{start_year}_{end_year}.xlsx",
        "filters": [
            {"label": "Anno da", "value": str(start_year)},
            {"label": "Anno a", "value": str(end_year)},
        ],
        "lead_html": lead_html,
        "columns": [
            ("codice_associato", "Codice"),
            ("associato", "Associato"),
            *year_columns,
        ],
    }


def get_report_definition(report_key: str, query_params: dict[str, str]) -> dict:
    work_year = current_work_year(query_params)
    if report_key == "registro-attivita":
        return tesserato_definition(build_registro_attivita_report_definition(query_params))

    if report_key == "tesseramenti":
        return tesserato_definition(build_tesseramenti_report_definition(query_params))

    if report_key == "scadenze":
        return tesserato_definition(build_scadenze_report_definition(query_params))

    if report_key == "corsi":
        return tesserato_definition(build_corsi_report_definition(query_params))

    if report_key == "oratorio":
        return tesserato_definition(build_oratorio_report_definition(query_params))

    if report_key == "campi-estivi":
        return tesserato_definition(build_campi_report_definition(query_params))

    if report_key == "eventi":
        return tesserato_definition(build_eventi_report_definition(query_params))

    if report_key == "incassi":
        return tesserato_definition(build_incassi_report_definition(query_params))

    if report_key == "associati":
        return tesserato_definition({
            "title": "Posizione associati",
            "current_path": "/report/associati",
            "subtitle": "Totale dovuto, totale pagato e saldo residuo dell'anno di lavoro per associato.",
            "query": posizione_associati_query(),
            "params": posizione_associati_params(work_year),
            "sheet_name": "Associati",
            "export_name": "posizione_associati.xlsx",
            "filters": [{"label": "Anno di lavoro", "value": str(work_year)}],
            "columns": [
                ("codice_tesseramento", "Codice"),
                (
                    "associato",
                    "Associato",
                    lambda value, row: report_link(value, f"/report/associato/{row['associato_id']}", work_year_query(query_params)),
                ),
                ("totale_dovuto", "Totale dovuto", lambda value, _: money(value)),
                ("totale_pagato", "Totale pagato", lambda value, _: money(value)),
                ("saldo_residuo", "Saldo residuo", lambda value, _: money(value)),
            ],
        })

    if report_key == "partecipanti":
        return tesserato_definition(build_participants_report_definition(query_params))

    if report_key == "storico-tesseramenti":
        return build_storico_tesseramenti_report_definition(query_params)

    raise KeyError(report_key)


def build_report_table_state(report_key: str, query_params: dict[str, str]) -> tuple[dict, list[sqlite3.Row]]:
    definition = get_report_definition(report_key, query_params)
    rows_builder = definition.get("rows_builder")
    if callable(rows_builder):
        rows = rows_builder()
    else:
        rows = fetch_all(definition["query"], definition.get("params", ()))
    return apply_table_export_state(definition, rows, query_params)


def report_page(report_key: str, query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    definition, rows = build_report_table_state(report_key, query_params)
    summary_rows = summary_rows_for_table(rows, definition["columns"])
    content = (
        report_toolbar(report_key, query_params, definition=definition, rows=rows)
        + definition.get("lead_html", "")
        + table_card(
            definition["title"],
            definition["subtitle"],
            rows,
            definition["columns"],
            table_class="report-table",
            summary_rows=summary_rows,
            table_id=f"report-table-{slugify(report_key)}",
            column_filters_in_header=True,
            draggable_columns=True,
        )
    )
    return page(definition["title"], definition["current_path"], content, query_params, current_user)


def report_associati(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("associati", query_params, current_user)


def report_tesseramenti(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("tesseramenti", query_params, current_user)


def report_scadenze(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("scadenze", query_params, current_user)


def report_corsi(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("corsi", query_params, current_user)


def report_oratorio(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("oratorio", query_params, current_user)


def report_campi_estivi(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("campi-estivi", query_params, current_user)


def report_eventi(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("eventi", query_params, current_user)


def report_partecipanti(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("partecipanti", query_params, current_user)


def report_storico_tesseramenti(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("storico-tesseramenti", query_params, current_user)


def report_incassi(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("incassi", query_params, current_user)


def report_registro_attivita(query_params: dict[str, str], current_user: dict[str, object] | None = None) -> bytes:
    return report_page("registro-attivita", query_params, current_user)


def associato_iscrizioni_rows(associato_id: int, work_year: int) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []

    tesseramenti = fetch_all(
        """
        SELECT
            'Tesseramento' AS area,
            'Anno ' || anno_sociale AS riferimento,
            data_tesseramento AS data_riferimento,
            CASE
                WHEN saldo_residuo <= 0 THEN 'Saldo completo'
                WHEN importo_pagato > 0 THEN 'Pagamento parziale'
                ELSE 'Da saldare'
            END AS stato,
            importo_dovuto AS importo
        FROM v_tesseramenti_saldo
        WHERE associato_id = ? AND anno_sociale = ?
        ORDER BY data_tesseramento DESC
        """,
        (associato_id, work_year),
    )
    rows.extend(dict(row) for row in tesseramenti)

    corsi = fetch_all(
        f"""
        SELECT
            'Corso' AS area,
            c.nome AS riferimento,
            ic.data_iscrizione AS data_riferimento,
            ic.stato_iscrizione AS stato,
            ic.quota_mensile AS importo
        FROM iscrizioni_corsi ic
        JOIN corsi c ON c.id = ic.corso_id
        WHERE ic.associato_id = ?
          AND {iscrizione_corso_year_relevance_sql('ic')}
        ORDER BY ic.data_iscrizione DESC, c.nome
        """,
        (associato_id, *iscrizione_corso_year_relevance_params(work_year)),
    )
    rows.extend(dict(row) for row in corsi)

    campi = fetch_all(
        """
        SELECT
            'Campo estivo' AS area,
            ce.nome AS riferimento,
            ice.data_iscrizione AS data_riferimento,
            ice.stato_iscrizione AS stato,
            ice.quota_partecipazione AS importo
        FROM iscrizioni_campi_estivi ice
        JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
        WHERE ice.associato_id = ?
          AND ce.anno = ?
        ORDER BY ice.data_iscrizione DESC, ce.nome
        """,
        (associato_id, work_year),
    )
    rows.extend(dict(row) for row in campi)

    eventi = fetch_all(
        """
        SELECT
            'Evento' AS area,
            e.nome AS riferimento,
            ie.data_iscrizione AS data_riferimento,
            ie.stato_iscrizione AS stato,
            ie.quota_partecipazione AS importo
        FROM iscrizioni_eventi ie
        JOIN eventi e ON e.id = ie.evento_id
        WHERE ie.associato_id = ?
          AND substr(COALESCE(e.data_evento, ''), 1, 4) = ?
        ORDER BY ie.data_iscrizione DESC, e.nome
        """,
        (associato_id, str(work_year)),
    )
    rows.extend(dict(row) for row in eventi)
    return rows


def build_associato_detail_export_payload(
    associato: sqlite3.Row,
    scoped_query: dict[str, str],
    iscrizioni_rows: list[dict[str, object]],
    scadenze_rows: list[sqlite3.Row],
    incassi_rows: list[sqlite3.Row],
) -> tuple[dict[str, object], list[dict[str, object]]]:
    work_year = current_work_year(scoped_query)
    filters = [
        {"label": "Anno di lavoro", "value": str(work_year)},
        {"label": "Associato", "value": plain_text(associato["associato"])},
        {"label": "Codice", "value": plain_text(associato["codice_associato"])},
    ]
    rows: list[dict[str, object]] = []

    for row in iscrizioni_rows:
        rows.append(
            {
                "sezione": "Iscrizioni",
                "area": row.get("area", ""),
                "riferimento": row.get("riferimento", ""),
                "data_riferimento": row.get("data_riferimento", ""),
                "stato": row.get("stato", ""),
                "importo_dovuto": row.get("importo", 0),
                "importo_pagato": "",
                "saldo_residuo": "",
                "importo": "",
            }
        )

    for row in scadenze_rows:
        rows.append(
            {
                "sezione": "Scadenze",
                "area": row["area"],
                "riferimento": row["riferimento"],
                "data_riferimento": row["scadenza"],
                "stato": row["stato_pagamento"],
                "importo_dovuto": row["importo_dovuto"],
                "importo_pagato": row["importo_pagato"],
                "saldo_residuo": row["saldo_residuo"],
                "importo": "",
            }
        )

    for row in incassi_rows:
        rows.append(
            {
                "sezione": "Incassi",
                "area": row["area"],
                "riferimento": row["riferimento"],
                "data_riferimento": row["data_pagamento"],
                "stato": row["metodo_pagamento"],
                "importo_dovuto": "",
                "importo_pagato": "",
                "saldo_residuo": "",
                "importo": row["importo"],
            }
        )

    definition = {
        "title": "Dettaglio tesserato",
        "subtitle": f"Riepilogo iscrizioni, scadenze e incassi di {plain_text(associato['associato'])}.",
        "sheet_name": "Dettaglio tesserato",
        "export_name": f"dettaglio_associato_{associato['codice_associato']}.xlsx",
        "filters": filters,
        "columns": [
            ("sezione", "Sezione"),
            ("area", "Area"),
            ("riferimento", "Riferimento"),
            ("data_riferimento", "Data"),
            ("stato", "Stato"),
            ("importo_dovuto", "Dovuto", lambda value, _: money(value) if value not in ("", None) else ""),
            ("importo_pagato", "Pagato", lambda value, _: money(value) if value not in ("", None) else ""),
            ("saldo_residuo", "Residuo", lambda value, _: money(value) if value not in ("", None) else ""),
            ("importo", "Importo", lambda value, _: money(value) if value not in ("", None) else ""),
        ],
    }
    return definition, rows


def associato_detail_toolbar(associato_id: int, scoped_query: dict[str, str]) -> str:
    excel_url = with_query(f"/export/excel/associato/{associato_id}", scoped_query)
    pdf_url = with_query(f"/export/pdf/associato/{associato_id}", scoped_query)
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Azioni dettaglio</span>
        <p>Esporta il dettaglio tesserato in Excel o PDF, oppure stampalo direttamente dal browser.</p>
      </div>
      <div class="report-toolbar-actions">
        <div class="report-toolbar-action-row">
          <a class="button action" href="{esc(excel_url)}">Esporta Excel</a>
          <a class="button action" href="{esc(pdf_url)}">Esporta PDF</a>
          <button type="button" class="button action" onclick="window.print()">Stampa report</button>
        </div>
      </div>
    </section>
    """


def associato_report_page(
    associato_id: int,
    query_params: dict[str, str],
    current_user: dict[str, object] | None = None,
) -> bytes:
    work_year = current_work_year(query_params)
    associato = fetch_one(
        f"""
        SELECT
            a.id,
            a.codice_associato,
            COALESCE(t.codice_tesseramento, '') AS codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            a.email,
            a.telefono,
            a.stato_associato
        FROM associati a
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = a.id AND t.anno_sociale = ?
        WHERE a.id = ?
        """,
        (work_year, associato_id),
    )
    if associato is None:
        raise KeyError(associato_id)

    scoped_query = dict(work_year_query(query_params))
    scoped_query["associato_id"] = str(associato_id)
    iscrizioni_rows = associato_iscrizioni_rows(associato_id, current_work_year(scoped_query))
    incassi_definition = tesserato_definition(build_incassi_report_definition(scoped_query))
    scadenze_definition = tesserato_definition(build_scadenze_report_definition(scoped_query))
    incassi_rows = fetch_all(incassi_definition["query"], incassi_definition["params"])
    scadenze_rows = fetch_all(scadenze_definition["query"], scadenze_definition["params"])
    incassi_columns = list(incassi_definition["columns"])
    scadenze_columns = list(scadenze_definition["columns"])
    iscrizioni_columns = [
        ("area", "Area"),
        ("riferimento", "Riferimento"),
        ("data_riferimento", "Data iscrizione"),
        ("stato", "Stato"),
        ("importo", "Importo", lambda value, _: money(value)),
    ]
    iscrizioni_summary = summary_rows_for_table(iscrizioni_rows, iscrizioni_columns)
    incassi_summary = summary_rows_for_table(incassi_rows, incassi_columns)
    scadenze_summary = summary_rows_for_table(scadenze_rows, scadenze_columns)
    scadenze_share_actions = associato_scadenze_share_actions(
        associato,
        scadenze_rows,
        current_work_year(scoped_query),
    )

    content = f"""
    <section class="hero">
      <div>
        <span class="eyebrow">Dettaglio tesserato</span>
        <h2>{esc(associato['associato'])}</h2>
        <p>Codice {esc(associato['codice_tesseramento'] or '-')} | Stato {esc(associato['stato_associato'])} | Email {esc(associato['email'] or '-')} | Cellulare {esc(associato['telefono'] or '-')}</p>
      </div>
      <div class="hero-actions">
        <a class="button ghost" href="{esc(with_query('/report/incassi', scoped_query))}">Apri incassi filtrati</a>
        <a class="button ghost" href="{esc(with_query('/report/scadenze', scoped_query))}">Apri scadenze filtrate</a>
      </div>
    </section>
    {associato_detail_toolbar(associato_id, scoped_query)}
    {table_card(
        "Iscrizioni del tesserato",
        "Riepilogo di tesseramento e iscrizioni del tesserato nell'anno di lavoro selezionato.",
        iscrizioni_rows,
        iscrizioni_columns,
        table_class="report-table",
        summary_rows=iscrizioni_summary,
    )}
    {table_card(
        "Scadenze del tesserato",
        "Quote ancora aperte o parzialmente saldate.",
        scadenze_rows,
        scadenze_columns,
        table_class="report-table",
        summary_rows=scadenze_summary,
        head_actions_html=scadenze_share_actions,
    )}
    {table_card(
        "Incassi del tesserato",
        "Movimenti registrati per il tesserato selezionato.",
        incassi_rows,
        incassi_columns,
        table_class="report-table",
        summary_rows=incassi_summary,
    )}
    """
    return page("Dettaglio tesserato", "/report/associati", content, query_params, current_user)


def export_rows(rows: list[sqlite3.Row], columns: list[tuple]) -> list[list[object]]:
    exported_rows = [[report_display_value(row, column) for column in columns] for row in rows]
    exported_rows.extend(summary_rows_for_table(rows, columns))
    return exported_rows


def generate_report_xlsx(definition: dict, rows: list[sqlite3.Row]) -> tuple[str, bytes]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def safe_sheet_name(value: str) -> str:
        cleaned = re.sub(r'[\\/*?:\\[\\]]', " ", str(value or "")).strip()
        return (cleaned or "Report")[:31]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_sheet_name(definition.get("sheet_name", "Report"))

    title_fill = PatternFill("solid", fgColor="FFF4E8")
    header_fill = PatternFill("solid", fgColor="CB5F07")
    summary_fill = PatternFill("solid", fgColor="FFF1E2")
    thin_border = Border(
        left=Side(style="thin", color="E5D4C5"),
        right=Side(style="thin", color="E5D4C5"),
        top=Side(style="thin", color="E5D4C5"),
        bottom=Side(style="thin", color="E5D4C5"),
    )

    current_row = 1
    worksheet.cell(current_row, 1, APP_NAME)
    worksheet.cell(current_row, 1).font = Font(bold=True, size=15, color="CB5F07")
    current_row += 1
    worksheet.cell(current_row, 1, definition["title"])
    worksheet.cell(current_row, 1).font = Font(bold=True, size=13, color="8F4100")
    current_row += 1
    worksheet.cell(current_row, 1, definition["subtitle"])
    worksheet.cell(current_row, 1).font = Font(size=10, color="5F4A3C")
    current_row += 1
    worksheet.cell(current_row, 1, f"Generato il: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    worksheet.cell(current_row, 1).font = Font(size=9, color="5F4A3C")
    current_row += 1

    for filter_row in definition.get("filters", []):
        worksheet.cell(current_row, 1, f"{filter_row['label']}: {filter_row['value']}")
        worksheet.cell(current_row, 1).font = Font(size=9, color="5F4A3C")
        current_row += 1

    current_row += 1

    headers = [column[1] for column in definition["columns"]]
    for col_index, label in enumerate(headers, start=1):
        cell = worksheet.cell(current_row, col_index, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    worksheet.row_dimensions[current_row].height = 24
    current_row += 1

    data_rows = export_rows(rows, definition["columns"])
    if not data_rows:
        data_rows = [["Nessun dato disponibile."] + [""] * (len(headers) - 1)]

    for row_index, values in enumerate(data_rows, start=0):
        is_summary = row_index >= len(rows) and values and values[summary_label_index(definition["columns"])] == "Totali"
        for col_index, value in enumerate(values, start=1):
            text_value = "" if value is None else str(value)
            cell = worksheet.cell(current_row, col_index, text_value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_summary:
                cell.font = Font(bold=True, color="6E3D14")
                cell.fill = summary_fill
        current_row += 1

    total_columns = max(1, len(headers))
    worksheet.freeze_panes = "A{}".format(max(2, current_row - len(data_rows)))
    for column_index in range(1, total_columns + 1):
        letter = get_column_letter(column_index)
        max_length = 10
        for row in worksheet.iter_rows(min_col=column_index, max_col=column_index):
            cell_value = row[0].value
            if cell_value is None:
                continue
            for line in str(cell_value).splitlines() or [""]:
                max_length = max(max_length, len(line))
        worksheet.column_dimensions[letter].width = min(max_length + 2, 42)

    worksheet.sheet_view.showGridLines = False
    for row_index in range(1, min(current_row, 5)):
        worksheet.cell(row_index, 1).fill = title_fill

    buffer = BytesIO()
    workbook.save(buffer)
    return definition["export_name"], buffer.getvalue()


def export_report_excel(start_response, report_key: str, query_params: dict[str, str]):
    try:
        definition, rows = build_report_table_state(report_key, query_params)
    except KeyError:
        return not_found(start_response)

    try:
        filename, content = generate_report_xlsx(definition, rows)
        start_response(
            "200 OK",
            [
                (
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response(
            "500 Internal Server Error",
            [("Content-Type", "text/plain; charset=utf-8")],
        )
        return [f"Errore durante l'export Excel: {error}".encode("utf-8")]


def export_associato_detail_excel(start_response, associato_id: int, query_params: dict[str, str]):
    associato = fetch_one(
        f"""
        SELECT id, codice_associato, {associato_display_sql('associati')} AS associato, email, telefono, stato_associato
        FROM associati
        WHERE id = ?
        """,
        (associato_id,),
    )
    if associato is None:
        return not_found(start_response)

    scoped_query = dict(work_year_query(query_params))
    scoped_query["associato_id"] = str(associato_id)
    iscrizioni_rows = associato_iscrizioni_rows(associato_id, current_work_year(scoped_query))
    incassi_definition = build_incassi_report_definition(scoped_query)
    scadenze_definition = build_scadenze_report_definition(scoped_query)
    incassi_rows = fetch_all(incassi_definition["query"], incassi_definition["params"])
    scadenze_rows = fetch_all(scadenze_definition["query"], scadenze_definition["params"])
    definition, rows = build_associato_detail_export_payload(associato, scoped_query, iscrizioni_rows, scadenze_rows, incassi_rows)
    try:
        filename, content = generate_report_xlsx(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export Excel: {error}".encode("utf-8")]


def export_tesserati_excel(start_response, query_params: dict[str, str]):
    definition, rows = build_tesserati_export_definition(query_params)
    try:
        filename, content = generate_report_xlsx(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export Excel: {error}".encode("utf-8")]


def export_associati_storici_excel(start_response, query_params: dict[str, str]):
    definition, rows = build_associati_storici_export_definition(query_params)
    try:
        filename, content = generate_report_xlsx(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export Excel: {error}".encode("utf-8")]


def generate_report_pdf(definition: dict, rows: list[sqlite3.Row]) -> tuple[str, bytes]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=20,
        textColor=colors.HexColor("#cb5f07"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#5f4a3c"),
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#5f4a3c"),
        spaceAfter=3,
    )
    cell_style = ParagraphStyle(
        "ReportCell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=7.7,
        leading=9.2,
        textColor=colors.HexColor("#3d2d22"),
        wordWrap="CJK",
    )
    header_cell_style = ParagraphStyle(
        "ReportHeaderCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        fontSize=8.1,
        leading=9.6,
    )
    summary_cell_style = ParagraphStyle(
        "ReportSummaryCell",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#6e3d14"),
    )

    story = []
    logo_path = STATIC_DIR / "logo-ca.jpg"
    if logo_path.exists():
        header_logo = Image(str(logo_path), width=24 * mm, height=24 * mm)
    else:
        header_logo = ""

    filter_lines = [f"<b>Generato il:</b> {esc(datetime.now().strftime('%d/%m/%Y %H:%M'))}"]
    for filter_row in definition.get("filters", []):
        filter_lines.append(f"<b>{esc(filter_row['label'])}:</b> {esc(filter_row['value'])}")

    header_table = Table(
        [
            [
                header_logo,
                [
                    Paragraph(APP_NAME, title_style),
                    Paragraph(definition["title"], title_style),
                    Paragraph(definition["subtitle"], subtitle_style),
                    Paragraph(" | ".join(filter_lines), meta_style),
                ],
            ]
        ],
        colWidths=[30 * mm, 240 * mm],
    )
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 5 * mm))

    summary_rows = summary_rows_for_table(rows, definition["columns"])
    body_rows = [[report_display_value(row, column) for column in definition["columns"]] for row in rows]
    if rows:
        raw_table_rows = body_rows + summary_rows
    else:
        raw_table_rows = [[definition.get("empty_message", "Nessun dato disponibile.")] + [""] * (len(definition["columns"]) - 1)]

    def pdf_text(value: object) -> str:
        return esc(plain_text(value)).replace("\n", "<br/>")

    def distribute_column_widths(weights: list[float], total_width: float) -> list[float]:
        column_count = max(len(weights), 1)
        min_width = min(18 * mm, total_width / column_count)
        max_width = max(34 * mm, total_width * 0.32)
        widths = [min_width] * column_count
        remaining = max(total_width - sum(widths), 0)
        adjustable = set(range(column_count))
        normalized_weights = [max(float(weight), 1.0) for weight in weights]

        while adjustable and remaining > 0.1:
            total_weight = sum(normalized_weights[index] for index in adjustable) or float(len(adjustable))
            capped_any = False
            for index in list(adjustable):
                share = remaining * (normalized_weights[index] / total_weight)
                target_width = widths[index] + share
                if target_width >= max_width:
                    remaining -= max_width - widths[index]
                    widths[index] = max_width
                    adjustable.remove(index)
                    capped_any = True
            if not capped_any:
                total_weight = sum(normalized_weights[index] for index in adjustable) or float(len(adjustable))
                for index in adjustable:
                    widths[index] += remaining * (normalized_weights[index] / total_weight)
                remaining = 0

        total_current = sum(widths)
        if total_current > total_width and total_current > 0:
            scale = total_width / total_current
            widths = [width * scale for width in widths]
        return widths

    def column_weight(column_index: int, header_label: str) -> float:
        samples = [plain_text(header_label)]
        for raw_row in raw_table_rows:
            if column_index < len(raw_row):
                samples.append(plain_text(raw_row[column_index]))
        longest = max((max(len(part) for part in sample.splitlines()) if sample else 0) for sample in samples)
        return min(max(longest, 8), 42)

    column_count = max(len(definition["columns"]), 1)
    available_width = 277 * mm
    col_widths = distribute_column_widths(
        [column_weight(index, column[1]) for index, column in enumerate(definition["columns"])],
        available_width,
    )

    table_data: list[list[object]] = [[Paragraph(pdf_text(column[1]), header_cell_style) for column in definition["columns"]]]
    if rows:
        for raw_row in body_rows:
            table_data.append([Paragraph(pdf_text(value), cell_style) for value in raw_row])
        for raw_summary in summary_rows:
            table_data.append([Paragraph(pdf_text(value), summary_cell_style) for value in raw_summary])
    else:
        table_data.append(
            [Paragraph(pdf_text(value), cell_style) for value in raw_table_rows[0]]
        )

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ef7f1a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#f0d2b8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff7f0")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if rows and summary_rows:
        first_summary_row = len(table_data) - len(summary_rows)
        style_commands.extend(
            [
                ("BACKGROUND", (0, first_summary_row), (-1, -1), colors.HexColor("#fff0de")),
                ("FONTNAME", (0, first_summary_row), (-1, -1), "Helvetica-Bold"),
            ]
        )
    if not rows:
        style_commands.extend(
            [
                ("SPAN", (0, 1), (-1, 1)),
                ("ALIGN", (0, 1), (-1, 1), "CENTER"),
            ]
        )
    table.setStyle(TableStyle(style_commands))
    story.append(table)

    document.build(story)
    filename = definition.get("pdf_name") or definition["export_name"].replace(".xlsx", ".pdf")
    return filename, buffer.getvalue()


def export_report_pdf(start_response, report_key: str, query_params: dict[str, str]):
    try:
        definition, rows = build_report_table_state(report_key, query_params)
    except KeyError:
        return not_found(start_response)

    try:
        filename, content = generate_report_pdf(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/pdf"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response(
            "500 Internal Server Error",
            [("Content-Type", "text/plain; charset=utf-8")],
        )
        return [f"Errore durante l'export PDF: {error}".encode("utf-8")]


def export_associato_detail_pdf(start_response, associato_id: int, query_params: dict[str, str]):
    associato = fetch_one(
        f"""
        SELECT id, codice_associato, {associato_display_sql('associati')} AS associato, email, telefono, stato_associato
        FROM associati
        WHERE id = ?
        """,
        (associato_id,),
    )
    if associato is None:
        return not_found(start_response)

    scoped_query = dict(work_year_query(query_params))
    scoped_query["associato_id"] = str(associato_id)
    iscrizioni_rows = associato_iscrizioni_rows(associato_id, current_work_year(scoped_query))
    incassi_definition = build_incassi_report_definition(scoped_query)
    scadenze_definition = build_scadenze_report_definition(scoped_query)
    incassi_rows = fetch_all(incassi_definition["query"], incassi_definition["params"])
    scadenze_rows = fetch_all(scadenze_definition["query"], scadenze_definition["params"])
    definition, rows = build_associato_detail_export_payload(associato, scoped_query, iscrizioni_rows, scadenze_rows, incassi_rows)
    try:
        filename, content = generate_report_pdf(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/pdf"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export PDF: {error}".encode("utf-8")]


def export_tesserati_pdf(start_response, query_params: dict[str, str]):
    definition, rows = build_tesserati_export_definition(query_params)
    try:
        filename, content = generate_report_pdf(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/pdf"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export PDF: {error}".encode("utf-8")]


def export_associati_storici_pdf(start_response, query_params: dict[str, str]):
    definition, rows = build_associati_storici_export_definition(query_params)
    try:
        filename, content = generate_report_pdf(definition, rows)
        start_response(
            "200 OK",
            [
                ("Content-Type", "application/pdf"),
                ("Content-Disposition", f'attachment; filename="{filename}"'),
                ("Content-Length", str(len(content))),
            ],
        )
        return [content]
    except Exception as error:
        start_response("500 Internal Server Error", [("Content-Type", "text/plain; charset=utf-8")])
        return [f"Errore durante l'export PDF: {error}".encode("utf-8")]


def receipt_link(payment_type: str, payment_id: object, query_params: dict[str, str]) -> str:
    return with_query(f"/ricevute/{payment_type}/{payment_id}", work_year_query(query_params))


def grouped_receipt_link(group_code: str, query_params: dict[str, str]) -> str:
    return with_query(f"/ricevute/corsi-rate-gruppo/{group_code}", work_year_query(query_params))


def multi_area_receipt_link(group_code: str, query_params: dict[str, str]) -> str:
    return with_query(f"/ricevute/multi-area-gruppo/{group_code}", work_year_query(query_params))


def popup_payment_requested(form_data: dict[str, str]) -> bool:
    return normalized(form_data, "procedi_pagamento", "") == "1"


def popup_payment_payload(form_data: dict[str, str], default_date: str) -> tuple[str, Decimal, str]:
    metodo_pagamento_id = required(form_data, "pagamento_metodo_id", "Metodo pagamento")
    importo = decimal_amount(required(form_data, "pagamento_importo", "Importo pagato"), minimum="0.01")
    data_pagamento = default_date or date.today().isoformat()
    return metodo_pagamento_id, importo, data_pagamento


def build_receipt_context(payment_type: str, payment_id: int) -> dict | None:
    payment_id = int(payment_id)
    if payment_type == "tesseramenti":
        row = fetch_one(
            f"""
            SELECT
                pt.id,
                pt.data_pagamento,
                pt.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pt.riferimento, '') AS riferimento,
                COALESCE(pt.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Tesseramento annuale' AS area,
                'Quota associativa anno ' || t.anno_sociale AS causale,
                t.anno_sociale AS work_year
            FROM pagamenti_tesseramenti pt
            JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
            JOIN associati a ON a.id = t.associato_id
            LEFT JOIN metodi_pagamento mp ON mp.id = pt.metodo_pagamento_id
            WHERE pt.id = ?
            """,
            (payment_id,),
        )
        prefix = "TES"
    elif payment_type == "corsi-iscrizione":
        row = fetch_one(
            f"""
            SELECT
                pic.id,
                pic.data_pagamento,
                pic.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pic.riferimento, '') AS riferimento,
                COALESCE(pic.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Corso - iscrizione' AS area,
                'Quota iscrizione corso ' || c.nome AS causale,
                CAST(substr(COALESCE(ic.data_iscrizione, ic.data_inizio, ''), 1, 4) AS INTEGER) AS work_year
            FROM pagamenti_iscrizioni_corsi pic
            JOIN iscrizioni_corsi ic ON ic.id = pic.iscrizione_corso_id
            JOIN associati a ON a.id = ic.associato_id
            JOIN corsi c ON c.id = ic.corso_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = CAST(substr(COALESCE(ic.data_iscrizione, ic.data_inizio, ''), 1, 4) AS INTEGER)
            LEFT JOIN metodi_pagamento mp ON mp.id = pic.metodo_pagamento_id
            WHERE pic.id = ?
            """,
            (payment_id,),
        )
        prefix = "CIS"
    elif payment_type == "corsi-rata":
        row = fetch_one(
            f"""
            SELECT
                prc.id,
                prc.data_pagamento,
                prc.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(prc.riferimento, '') AS riferimento,
                COALESCE(prc.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Corso - quota mensile' AS area,
                'Quota mensile corso ' || c.nome || ' ' || printf('%04d-%02d', r.anno, r.mese) AS causale,
                r.anno AS work_year
            FROM pagamenti_rate_corsi prc
            JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
            JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
            JOIN associati a ON a.id = ic.associato_id
            JOIN corsi c ON c.id = ic.corso_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id
            WHERE prc.id = ?
            """,
            (payment_id,),
        )
        prefix = "CRM"
    elif payment_type == "campi-estivi":
        row = fetch_one(
            f"""
            SELECT
                pce.id,
                pce.data_pagamento,
                pce.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pce.riferimento, '') AS riferimento,
                COALESCE(pce.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Campo estivo' AS area,
                'Quota partecipazione Campo estivo ' || ce.anno AS causale,
                ce.anno AS work_year
            FROM pagamenti_campi_estivi pce
            JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
            JOIN associati a ON a.id = ice.associato_id
            JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = pce.metodo_pagamento_id
            WHERE pce.id = ?
            """,
            (payment_id,),
        )
        prefix = "CES"
    elif payment_type == "oratorio":
        row = fetch_one(
            f"""
            SELECT
                po.id,
                po.data_pagamento,
                po.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(po.riferimento, '') AS riferimento,
                COALESCE(po.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Oratorio' AS area,
                'Contributo Oratorio ' || o.anno AS causale,
                o.anno AS work_year
            FROM pagamenti_oratorio po
            JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
            JOIN associati a ON a.id = io.associato_id
            JOIN oratorio o ON o.id = io.oratorio_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = po.metodo_pagamento_id
            WHERE po.id = ?
            """,
            (payment_id,),
        )
        prefix = "ORA"
    elif payment_type == "eventi":
        row = fetch_one(
            f"""
            SELECT
                pe.id,
                pe.data_pagamento,
                pe.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pe.riferimento, '') AS riferimento,
                COALESCE(pe.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Evento' AS area,
                'Quota partecipazione ' || e.nome AS causale,
                CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER) AS work_year
            FROM pagamenti_eventi pe
            JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
            JOIN associati a ON a.id = ie.associato_id
            JOIN eventi e ON e.id = ie.evento_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
            LEFT JOIN metodi_pagamento mp ON mp.id = pe.metodo_pagamento_id
            WHERE pe.id = ?
            """,
            (payment_id,),
        )
        prefix = "EVT"
    else:
        return None

    if row is None:
        return None

    context = dict(row)
    contact_channels = resolve_minor_contact_channels(context)
    context["email"] = contact_channels["email"]
    context["telefono"] = contact_channels["telefono"]
    context["payment_type"] = payment_type
    context["receipt_number"] = f"RCP-{prefix}-{payment_id:06d}"
    context["whatsapp_phone"] = clean_phone_number(context.get("telefono"))
    return context


def build_grouped_rate_receipt_context(group_code: str) -> dict | None:
    rows = fetch_all(
        f"""
        SELECT
            prc.id,
            prc.data_pagamento,
            prc.importo,
            COALESCE(mp.nome, '') AS metodo_pagamento,
            COALESCE(prc.riferimento, '') AS riferimento,
            COALESCE(prc.note, '') AS note,
            a.id AS associato_id,
            t.codice_tesseramento,
            {associato_display_sql('a')} AS associato,
            COALESCE(a.email, '') AS email,
            COALESCE(a.telefono, '') AS telefono,
            COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
            COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
            COALESCE(a.data_nascita, '') AS data_nascita,
            c.nome AS corso,
            printf('%04d-%02d', r.anno, r.mese) AS competenza,
            r.anno AS work_year
        FROM pagamenti_rate_corsi prc
        JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
        JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
        JOIN associati a ON a.id = ic.associato_id
        JOIN corsi c ON c.id = ic.corso_id
        LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
        LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id
        WHERE prc.gruppo_ricevuta = ?
        ORDER BY r.anno, r.mese, c.nome
        """,
        (group_code,),
    )
    if not rows:
        return None

    first = dict(rows[0])
    items = [
        {
            "corso": row["corso"],
            "competenza": row["competenza"],
            "importo": float(row["importo"] or 0),
        }
        for row in rows
    ]
    total_amount = sum(item["importo"] for item in items)
    contact_channels = resolve_minor_contact_channels(first)
    return {
        "id": group_code,
        "payment_type": "corsi-rate-gruppo",
        "receipt_number": f"RCP-CRMSET-{group_code}",
        "data_pagamento": first["data_pagamento"],
        "importo": total_amount,
        "metodo_pagamento": first["metodo_pagamento"],
        "riferimento": first["riferimento"],
        "note": first["note"],
        "codice_tesseramento": first["codice_tesseramento"],
        "associato": first["associato"],
        "email": contact_channels["email"],
        "telefono": contact_channels["telefono"],
        "area": "Corso - quote mensili",
        "causale": "SALDO QUOTE MENSILI CORSI",
        "items": items,
        "items_mode": "mensilita",
        "associato_id": int(first["associato_id"]),
        "work_year": int(first["work_year"] or date.today().year),
        "whatsapp_phone": clean_phone_number(contact_channels["telefono"]),
    }


def build_multi_area_receipt_context(group_code: str) -> dict | None:
    rows = fetch_all(
        f"""
        SELECT *
        FROM (
            SELECT
                pt.id,
                pt.data_pagamento,
                pt.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pt.riferimento, '') AS riferimento,
                COALESCE(pt.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Tesseramento annuale' AS area,
                'Anno ' || t.anno_sociale AS scadenza_riferimento,
                COALESCE(t.data_scadenza, t.data_tesseramento) AS scadenza,
                t.anno_sociale AS work_year
            FROM pagamenti_tesseramenti pt
            JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
            JOIN associati a ON a.id = t.associato_id
            LEFT JOIN metodi_pagamento mp ON mp.id = pt.metodo_pagamento_id
            WHERE pt.gruppo_ricevuta = ?

            UNION ALL

            SELECT
                prc.id,
                prc.data_pagamento,
                prc.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(prc.riferimento, '') AS riferimento,
                COALESCE(prc.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Corso - quota mensile' AS area,
                c.nome || ' ' || printf('%04d-%02d', r.anno, r.mese) AS scadenza_riferimento,
                COALESCE(r.data_scadenza, printf('%04d-%02d-01', r.anno, r.mese)) AS scadenza,
                r.anno AS work_year
            FROM pagamenti_rate_corsi prc
            JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
            JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
            JOIN associati a ON a.id = ic.associato_id
            JOIN corsi c ON c.id = ic.corso_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ic.associato_id AND t.anno_sociale = r.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = prc.metodo_pagamento_id
            WHERE prc.gruppo_ricevuta = ?

            UNION ALL

            SELECT
                pce.id,
                pce.data_pagamento,
                pce.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pce.riferimento, '') AS riferimento,
                COALESCE(pce.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Campo estivo' AS area,
                ce.nome AS scadenza_riferimento,
                COALESCE(ce.data_inizio, ice.data_iscrizione) AS scadenza,
                ce.anno AS work_year
            FROM pagamenti_campi_estivi pce
            JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
            JOIN associati a ON a.id = ice.associato_id
            JOIN campi_estivi ce ON ce.id = ice.campo_estivo_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ice.associato_id AND t.anno_sociale = ce.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = pce.metodo_pagamento_id
            WHERE pce.gruppo_ricevuta = ?

            UNION ALL

            SELECT
                po.id,
                po.data_pagamento,
                po.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(po.riferimento, '') AS riferimento,
                COALESCE(po.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Oratorio' AS area,
                o.nome AS scadenza_riferimento,
                COALESCE(o.data_inizio, io.data_iscrizione) AS scadenza,
                o.anno AS work_year
            FROM pagamenti_oratorio po
            JOIN iscrizioni_oratorio io ON io.id = po.iscrizione_oratorio_id
            JOIN associati a ON a.id = io.associato_id
            JOIN oratorio o ON o.id = io.oratorio_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = io.associato_id AND t.anno_sociale = o.anno
            LEFT JOIN metodi_pagamento mp ON mp.id = po.metodo_pagamento_id
            WHERE po.gruppo_ricevuta = ?

            UNION ALL

            SELECT
                pe.id,
                pe.data_pagamento,
                pe.importo,
                COALESCE(mp.nome, '') AS metodo_pagamento,
                COALESCE(pe.riferimento, '') AS riferimento,
                COALESCE(pe.note, '') AS note,
                a.id AS associato_id,
                t.codice_tesseramento,
                {associato_display_sql('a')} AS associato,
                COALESCE(a.email, '') AS email,
                COALESCE(a.telefono, '') AS telefono,
                COALESCE(a.genitore_tutore_email, '') AS genitore_tutore_email,
                COALESCE(a.genitore_tutore_cellulare, '') AS genitore_tutore_cellulare,
                COALESCE(a.data_nascita, '') AS data_nascita,
                'Evento' AS area,
                e.nome AS scadenza_riferimento,
                e.data_evento AS scadenza,
                CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER) AS work_year
            FROM pagamenti_eventi pe
            JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
            JOIN associati a ON a.id = ie.associato_id
            JOIN eventi e ON e.id = ie.evento_id
            LEFT JOIN tesseramenti_annuali t ON t.associato_id = ie.associato_id AND t.anno_sociale = CAST(substr(COALESCE(e.data_evento, ''), 1, 4) AS INTEGER)
            LEFT JOIN metodi_pagamento mp ON mp.id = pe.metodo_pagamento_id
            WHERE pe.gruppo_ricevuta = ?
        ) pagamenti
        ORDER BY scadenza, area, scadenza_riferimento
        """,
        (group_code, group_code, group_code, group_code, group_code),
    )
    if not rows:
        return None

    first = dict(rows[0])
    items = [
        {
            "area": row["area"],
            "riferimento": row["scadenza_riferimento"],
            "scadenza": row["scadenza"],
            "importo": float(row["importo"] or 0),
        }
        for row in rows
    ]
    total_amount = sum(item["importo"] for item in items)
    contact_channels = resolve_minor_contact_channels(first)
    return {
        "id": group_code,
        "payment_type": "multi-area-gruppo",
        "receipt_number": f"RCP-MULTI-{group_code}",
        "data_pagamento": first["data_pagamento"],
        "importo": total_amount,
        "metodo_pagamento": first["metodo_pagamento"],
        "riferimento": first["riferimento"],
        "note": first["note"],
        "codice_tesseramento": first["codice_tesseramento"],
        "associato": first["associato"],
        "email": contact_channels["email"],
        "telefono": contact_channels["telefono"],
        "area": "Pagamento multi-area",
        "causale": "Saldo contemporaneo di scadenze provenienti da aree diverse",
        "items": items,
        "items_mode": "scadenze",
        "associato_id": int(first["associato_id"]),
        "work_year": int(first["work_year"] or date.today().year),
        "whatsapp_phone": clean_phone_number(contact_channels["telefono"]),
    }


def receipt_pending_dues_message(context: dict) -> str:
    associato_id = context.get("associato_id")
    work_year = context.get("work_year")
    if not associato_id or not work_year:
        return ""

    scoped_query = {
        "anno_lavoro": str(work_year),
        "associato_id": str(associato_id),
    }
    definition = build_scadenze_report_definition(scoped_query)
    rows = fetch_all(definition["query"], definition.get("params", ()))
    if not rows:
        return "\n\nAlla data odierna non risultano ulteriori scadenze da saldare."

    lines = ["\n\nSi riporta il riepilogo delle scadenze ancora da saldare:"]
    for row in rows[:8]:
        lines.append(
            f"- {row['area']} | {row['riferimento']} | residuo {money(row['saldo_residuo'])}"
        )
    if len(rows) > 8:
        lines.append(f"- ...e altre {len(rows) - 8} scadenze ancora aperte.")
    return "\n".join(lines)


def associato_scadenze_share_message(
    associato: sqlite3.Row,
    rows: list[sqlite3.Row],
    work_year: int,
) -> str:
    total_residuo = sum((decimal_amount(row["saldo_residuo"]) for row in rows), Decimal("0.00")).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )
    lines = [
        f"Buongiorno {plain_text(associato['associato'])},",
        f"si riporta di seguito il riepilogo delle scadenze attualmente aperte per l'anno di lavoro {work_year}.",
    ]
    if rows:
        lines.extend(
            [
                f"Totale residuo: {money(total_residuo)}",
                "Dettaglio scadenze:",
            ]
        )
        for row in rows[:12]:
            lines.append(
                f"- {row['area']} | {row['riferimento']} | scadenza {row['scadenza']} | residuo {money(row['saldo_residuo'])}"
            )
        if len(rows) > 12:
            lines.append(f"- Ulteriori scadenze aperte non riportate in questo messaggio: {len(rows) - 12}.")
    else:
        lines.extend(
            [
                "Alla data odierna non risultano scadenze aperte.",
            ]
        )
    lines.extend(
        [
            "Restiamo a disposizione per eventuali chiarimenti.",
            "Cordiali saluti,",
            APP_NAME,
        ]
    )
    return "\n".join(lines)


def associato_scadenze_share_actions(
    associato: sqlite3.Row,
    rows: list[sqlite3.Row],
    work_year: int,
) -> str:
    message = associato_scadenze_share_message(associato, rows, work_year)
    subject = quote(f"Riepilogo scadenze aperte - {APP_NAME}")
    body = quote(message)
    phone = clean_phone_number(associato["telefono"])

    email_button = (
        f'<a class="button action" href="mailto:{esc(associato["email"])}?subject={subject}&body={body}">Prepara email</a>'
        if associato["email"]
        else '<span class="button action disabled">Email non presente</span>'
    )
    whatsapp_button = (
        f'<a class="button action" href="https://wa.me/{esc(phone)}?text={body}" target="_blank" rel="noopener">Invia WhatsApp</a>'
        if phone
        else '<span class="button action disabled">Cellulare non presente</span>'
    )
    return f'<div class="mini-actions">{email_button}{whatsapp_button}</div>'


def receipt_message(context: dict) -> str:
    closing = "\n\nCordiali saluti,\nOratorio Carlo Acutis"
    base_message = (
        f"Ricevuta {context['receipt_number']} - {APP_NAME}\n"
        f"Associato: {context['associato']}\n"
        f"Causale: {context['causale']}\n"
        f"Data pagamento: {context['data_pagamento']}\n"
        f"Importo: {money(context['importo'])}\n"
        f"Metodo: {context['metodo_pagamento'] or 'Non indicato'}"
    )
    if context.get("items"):
        if context.get("items_mode") == "scadenze":
            details = "\n".join(
                f"- {item['area']} | {item['riferimento']} | {item['scadenza']} ({money(item['importo'])})"
                for item in context["items"]
            )
            return base_message + "\nScadenze saldate:\n" + details + receipt_pending_dues_message(context) + closing

        details = "\n".join(
            f"- {item['corso']} {item['competenza']} ({money(item['importo'])})"
            for item in context["items"]
        )
        return base_message + "\nMensilita saldate:\n" + details + receipt_pending_dues_message(context) + closing
    return base_message + receipt_pending_dues_message(context) + closing


def receipt_toolbar(context: dict) -> str:
    pdf_url = f"/export/pdf/ricevuta/{context['payment_type']}/{context['id']}"
    email_url = ""
    whatsapp_url = ""
    if context.get("email"):
        subject = quote(f"Ricevuta pagamento {context['receipt_number']} - {APP_NAME}")
        body = quote(receipt_message(context))
        email_url = f"mailto:{context['email']}?subject={subject}&body={body}"
    if context.get("whatsapp_phone"):
        whatsapp_url = f"https://wa.me/{context['whatsapp_phone']}?text={quote(receipt_message(context))}"
    email_button = (
        f'<button type="button" class="button action" '
        f'onclick="return shareReceipt(this)" '
        f'data-channel="email" data-mailto="{esc(email_url)}">Prepara email</button>'
        if email_url
        else '<span class="button action disabled">Email non presente</span>'
    )
    whatsapp_button = (
        f'<button type="button" class="button action" '
        f'onclick="return shareReceipt(this)" '
        f'data-channel="whatsapp" data-whatsapp="{esc(whatsapp_url)}">Invia WhatsApp</button>'
        if whatsapp_url
        else '<span class="button action disabled">Cellulare non presente</span>'
    )
    return f"""
    <section class="report-toolbar screen-only">
      <div class="report-toolbar-copy">
        <span class="eyebrow">Azioni ricevuta</span>
        <p>Stampa la ricevuta, esportala in PDF oppure prepara l'invio via email o WhatsApp senza allegato automatico.</p>
      </div>
      <div class="report-toolbar-actions">
        <div class="report-toolbar-action-row">
          <a class="button action" href="{esc(pdf_url)}">Esporta PDF</a>
          <button type="button" class="button action" onclick="window.print()">Stampa ricevuta</button>
        </div>
        <div class="report-toolbar-action-row report-toolbar-action-row-secondary">
          {email_button}
          {whatsapp_button}
        </div>
      </div>
    </section>
    """


def receipt_page(
    payment_type: str,
    payment_id: object,
    query_params: dict[str, str],
    current_user: dict[str, object] | None = None,
) -> bytes:
    if payment_type == "corsi-rate-gruppo":
        context = build_grouped_rate_receipt_context(str(payment_id))
    elif payment_type == "multi-area-gruppo":
        context = build_multi_area_receipt_context(str(payment_id))
    else:
        context = build_receipt_context(payment_type, payment_id)
    if context is None:
        raise KeyError(payment_type)

    items_block = ""
    if context.get("items"):
        if context.get("items_mode") == "scadenze":
            items_rows = "".join(
                f"<tr><td>{esc(item['area'])}</td><td>{esc(item['riferimento'])}</td><td>{esc(item['scadenza'])}</td><td>{esc(money(item['importo']))}</td></tr>"
                for item in context["items"]
            )
            items_block = f"""
            <div class="receipt-note">
              <h3>Scadenze saldate</h3>
              <div class="table-wrap">
                <table class="data-table">
                  <thead><tr><th>Area</th><th>Riferimento</th><th>Scadenza</th><th>Importo</th></tr></thead>
                  <tbody>{items_rows}</tbody>
                </table>
              </div>
            </div>
            """
        else:
            items_rows = "".join(
                f"<tr><td>{esc(item['corso'])}</td><td>{esc(item['competenza'])}</td><td>{esc(money(item['importo']))}</td></tr>"
                for item in context["items"]
            )
            items_block = f"""
            <div class="receipt-note">
              <h3>Mensilita saldate</h3>
              <div class="table-wrap">
                <table class="data-table">
                  <thead><tr><th>Corso</th><th>Competenza</th><th>Importo</th></tr></thead>
                  <tbody>{items_rows}</tbody>
                </table>
              </div>
            </div>
            """

    content = f"""
    {receipt_toolbar(context)}
    <section class="card receipt-card">
      <div class="receipt-head">
        <div>
          <span class="eyebrow">Ricevuta di pagamento</span>
          <h2>{esc(APP_NAME)}</h2>
          <p>{esc(context['area'])}</p>
        </div>
        <div class="receipt-number-box">
          <strong>{esc(context['receipt_number'])}</strong>
          <span>Data {esc(context['data_pagamento'])}</span>
        </div>
      </div>
      <div class="receipt-grid">
        <div class="receipt-block">
          <h3>Associato</h3>
          <p><strong>{esc(context['associato'])}</strong></p>
          <p>Codice: {esc(context.get('codice_tesseramento') or '-')}</p>
          <p>Email: {esc(context['email'] or '-')}</p>
          <p>Cellulare: {esc(context['telefono'] or '-')}</p>
        </div>
        <div class="receipt-block">
          <h3>Dettagli pagamento</h3>
          <p>Causale: {esc(context['causale'])}</p>
          <p>Importo: <strong>{esc(money(context['importo']))}</strong></p>
          <p>Metodo: {esc(context['metodo_pagamento'] or 'Non indicato')}</p>
          <p>Riferimento: {esc(context['riferimento'] or '-')}</p>
        </div>
      </div>
      <div class="receipt-note">
        <h3>Note</h3>
        <p>{esc(context['note'] or 'Nessuna nota inserita.')}</p>
      </div>
      {items_block}
      <div class="receipt-footer">
        <p>Documento generato dal gestionale {esc(APP_NAME)}.</p>
      </div>
    </section>
    """
    return page("Ricevuta pagamento", "/report/incassi", content, query_params, current_user)


def generate_receipt_pdf(context: dict) -> tuple[str, bytes]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReceiptTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#cb5f07"),
        spaceAfter=4,
    )
    normal_style = ParagraphStyle(
        "ReceiptBody",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#2c231d"),
    )

    story = []
    logo_path = STATIC_DIR / "logo-ca.jpg"
    header_logo = Image(str(logo_path), width=28 * mm, height=28 * mm) if logo_path.exists() else ""
    header = Table(
        [
            [
                header_logo,
                [
                    Paragraph(APP_NAME, title_style),
                    Paragraph("Ricevuta di pagamento", title_style),
                    Paragraph(f"Numero {context['receipt_number']} - Data {context['data_pagamento']}", normal_style),
                ],
            ]
        ],
        colWidths=[34 * mm, 130 * mm],
    )
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header)
    story.append(Spacer(1, 8 * mm))

    details = [
        ["Tesserato", context["associato"]],
        ["Codice", context.get("codice_tesseramento") or "-"],
        ["Area", context["area"]],
        ["Causale", context["causale"]],
        ["Importo", money(context["importo"])],
        ["Metodo pagamento", context["metodo_pagamento"] or "Non indicato"],
        ["Riferimento", context["riferimento"] or "-"],
        ["Email", context["email"] or "-"],
        ["Cellulare", context["telefono"] or "-"],
        ["Note", context["note"] or "Nessuna nota inserita."],
    ]
    detail_table = Table(details, colWidths=[45 * mm, 115 * mm])
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#fff3e7")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#cb5f07")),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#f0d2b8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(detail_table)
    if context.get("items"):
        story.append(Spacer(1, 6 * mm))
        if context.get("items_mode") == "scadenze":
            story.append(Paragraph("Scadenze saldate", title_style))
            item_rows = [["Area", "Riferimento", "Scadenza", "Importo"]] + [
                [item["area"], item["riferimento"], item["scadenza"], money(item["importo"])] for item in context["items"]
            ]
            item_table = Table(item_rows, colWidths=[40 * mm, 60 * mm, 30 * mm, 30 * mm])
        else:
            story.append(Paragraph("Mensilita saldate", title_style))
            item_rows = [["Corso", "Competenza", "Importo"]] + [
                [item["corso"], item["competenza"], money(item["importo"])] for item in context["items"]
            ]
            item_table = Table(item_rows, colWidths=[75 * mm, 35 * mm, 50 * mm])
        item_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ef7f1a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#f0d2b8")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fff7f0")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(item_table)
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph("Documento generato dal gestionale dell'Oratorio Carlo Acutis.", normal_style))

    document.build(story)
    return f"ricevuta_{context['receipt_number'].lower()}.pdf", buffer.getvalue()


def export_receipt_pdf(start_response, payment_type: str, payment_id: str):
    if payment_type == "corsi-rate-gruppo":
        context = build_grouped_rate_receipt_context(payment_id)
    elif payment_type == "multi-area-gruppo":
        context = build_multi_area_receipt_context(payment_id)
    else:
        context = build_receipt_context(payment_type, int(payment_id))
    if context is None:
        return not_found(start_response)
    filename, content = generate_receipt_pdf(context)
    start_response(
        "200 OK",
        [
            ("Content-Type", "application/pdf"),
            ("Content-Disposition", f'attachment; filename="{filename}"'),
            ("Content-Length", str(len(content))),
        ],
    )
    return [content]


def handle_post(
    path: str,
    form_data: dict[str, str],
    start_response,
    current_user: dict[str, object] | None = None,
    request_cookies: dict[str, str] | None = None,
    form_files: dict[str, UploadedFile] | None = None,
):
    context_query = work_year_query_from_form(form_data)
    if path in {"/azioni/accesso/primo-admin", "/azioni/accesso/login", "/azioni/accesso/recupera-password"}:
        next_target = login_destination(normalized(form_data, "next", "/"))
        if next_target and next_target != "/":
            context_query["next"] = next_target
    try:
        if path == "/azioni/accesso/primo-admin":
            if not bootstrap_admin_required():
                raise ValueError("L'amministratore iniziale e gia stato configurato.")
            username = validate_username(required(form_data, "username", "Username"))
            password = validate_password(
                required(form_data, "password", "Password"),
                required(form_data, "password_conferma", "Conferma password"),
            )
            with get_connection() as connection:
                user_id = create_access_user(connection, username=username, password=password, is_admin=True)
                session_token = create_user_session(connection, user_id)
                connection.commit()
            return redirect(
                start_response,
                "/",
                ok="Amministratore creato e accesso effettuato.",
                extra_headers=[session_cookie_header(session_token)],
            )

        if path == "/azioni/accesso/login":
            if bootstrap_admin_required():
                raise ValueError("Completa prima la creazione dell'amministratore.")
            username = validate_username(required(form_data, "username", "Username"))
            password = required(form_data, "password", "Password")
            next_target = login_destination(normalized(form_data, "next", "/"))
            with get_connection() as connection:
                cleanup_expired_sessions(connection)
                user_row = connection.execute(
                    """
                    SELECT *
                    FROM utenti_accesso
                    WHERE username = ?
                    """,
                    (username,),
                ).fetchone()
                if user_row is None or not verify_password(password, user_row):
                    raise ValueError("Username o password non validi.")
                if not user_row["attivo"]:
                    raise ValueError("Questo utente non e attivo.")
                session_token = create_user_session(connection, int(user_row["id"]))
                connection.commit()
            return redirect(
                start_response,
                next_target,
                ok="Accesso effettuato.",
                extra_headers=[session_cookie_header(session_token)],
            )

        if path == "/azioni/accesso/logout":
            session_token = (request_cookies or {}).get(SESSION_COOKIE_NAME, "")
            if session_token:
                with get_connection() as connection:
                    connection.execute("DELETE FROM sessioni_accesso WHERE session_token = ?", (session_token,))
                    connection.commit()
            return redirect(
                start_response,
                "/login",
                ok="Sessione chiusa.",
                extra_headers=[clear_session_cookie_header()],
            )

        if path == "/azioni/accesso/cambia-password":
            if current_user is None:
                return redirect(start_response, "/login", err="Accedi per modificare la password.")
            current_password = required(form_data, "password_attuale", "Password attuale")
            new_password = validate_password(
                required(form_data, "password", "Nuova password"),
                required(form_data, "password_conferma", "Conferma nuova password"),
            )
            with get_connection() as connection:
                user_row = connection.execute(
                    "SELECT * FROM utenti_accesso WHERE id = ?",
                    (int(current_user["id"]),),
                ).fetchone()
                if user_row is None or not verify_password(current_password, user_row):
                    raise ValueError("La password attuale non e corretta.")
                set_access_user_password(connection, int(current_user["id"]), new_password)
                connection.commit()
            return redirect(start_response, "/maschere/accesso", ok="Password aggiornata.", extra_query=context_query)

        if path == "/azioni/accesso/recupera-password":
            if bootstrap_admin_required():
                raise ValueError("Configura prima l'utente amministratore.")
            username = validate_username(required(form_data, "username", "Username"))
            recovery_email = validate_recovery_email(required(form_data, "email_recupero", "Email recupero"))
            new_password = validate_password(
                required(form_data, "password", "Nuova password"),
                required(form_data, "password_conferma", "Conferma nuova password"),
            )
            with get_connection() as connection:
                user_row = connection.execute(
                    """
                    SELECT *
                    FROM utenti_accesso
                    WHERE username = ?
                      AND is_admin = 0
                      AND attivo = 1
                    """,
                    (username,),
                ).fetchone()
                if user_row is None:
                    raise ValueError("Utente standard non trovato.")
                stored_email = (user_row["email_recupero"] or "").strip().lower()
                if stored_email != recovery_email.strip().lower():
                    raise ValueError("Email di recupero non corrispondente.")
                set_access_user_password(connection, int(user_row["id"]), new_password)
                connection.commit()
            return redirect(start_response, "/login", ok="Password aggiornata. Ora puoi accedere.", extra_query=context_query)

        if path == "/azioni/utenti/crea":
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo creare nuovi utenti.",
                    extra_query=context_query,
                )
            username = validate_username(required(form_data, "username", "Username"))
            password = validate_password(
                required(form_data, "password", "Password"),
                required(form_data, "password_conferma", "Conferma password"),
            )
            email_recupero = validate_recovery_email(required(form_data, "email_recupero", "Email recupero"))
            with get_connection() as connection:
                create_access_user(
                    connection,
                    username=username,
                    password=password,
                    is_admin=False,
                    email_recupero=email_recupero,
                )
                connection.commit()
            return redirect(start_response, "/maschere/utenti", ok="Utente creato.", extra_query=context_query)

        if path.startswith("/azioni/utenti/aggiorna/"):
            if not current_user or not current_user.get("is_admin"):
                return redirect(start_response, "/", err="Solo l'amministratore puo modificare gli utenti.", extra_query=context_query)
            user_id = int(path.removeprefix("/azioni/utenti/aggiorna/"))
            user_row = access_user_row(user_id)
            if user_row is None or user_row["is_admin"]:
                raise ValueError("Puoi gestire solo utenti standard.")
            with get_connection() as connection:
                connection.execute(
                    """
                    UPDATE utenti_accesso
                    SET username = ?, email_recupero = ?, aggiornato_il = ?
                    WHERE id = ?
                    """,
                    (
                        validate_username(required(form_data, "username", "Username")),
                        validate_recovery_email(required(form_data, "email_recupero", "Email recupero")),
                        current_timestamp(),
                        user_id,
                    ),
                )
                connection.commit()
            return redirect(start_response, f"/maschere/utenti/gestione/{user_id}", ok="Utente aggiornato.", extra_query=context_query)

        if path.startswith("/azioni/utenti/password/"):
            if not current_user or not current_user.get("is_admin"):
                return redirect(start_response, "/", err="Solo l'amministratore puo reimpostare le password.", extra_query=context_query)
            user_id = int(path.removeprefix("/azioni/utenti/password/"))
            user_row = access_user_row(user_id)
            if user_row is None or user_row["is_admin"]:
                raise ValueError("Puoi reimpostare solo password di utenti standard.")
            new_password = validate_password(
                required(form_data, "password", "Nuova password"),
                required(form_data, "password_conferma", "Conferma nuova password"),
            )
            with get_connection() as connection:
                set_access_user_password(connection, user_id, new_password)
                connection.commit()
            return redirect(start_response, f"/maschere/utenti/gestione/{user_id}", ok="Password utente aggiornata.", extra_query=context_query)

        if path.startswith("/azioni/utenti/stato/"):
            if not current_user or not current_user.get("is_admin"):
                return redirect(start_response, "/", err="Solo l'amministratore puo disattivare o riattivare gli utenti.", extra_query=context_query)
            user_id = int(path.removeprefix("/azioni/utenti/stato/"))
            user_row = access_user_row(user_id)
            if user_row is None or user_row["is_admin"]:
                raise ValueError("Puoi modificare lo stato solo degli utenti standard.")
            desired_action = normalized(form_data, "azione_stato", "")
            if desired_action not in {"disattiva", "riattiva"}:
                raise ValueError("Azione stato non valida.")
            attivo_value = 0 if desired_action == "disattiva" else 1
            with get_connection() as connection:
                connection.execute(
                    """
                    UPDATE utenti_accesso
                    SET attivo = ?, aggiornato_il = ?
                    WHERE id = ?
                    """,
                    (attivo_value, current_timestamp(), user_id),
                )
                if attivo_value == 0:
                    connection.execute("DELETE FROM sessioni_accesso WHERE utente_id = ?", (user_id,))
                connection.commit()
            message = "Utente standard disattivato." if attivo_value == 0 else "Utente standard riattivato."
            return redirect(start_response, f"/maschere/utenti/gestione/{user_id}", ok=message, extra_query=context_query)

        if path == "/azioni/backup/crea":
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo creare un backup completo.",
                    extra_query=context_query,
                )
            backup_path = create_backup_archive()
            return file_download_response(
                start_response,
                backup_path,
                content_type="application/zip",
                download_name=backup_path.name,
            )

        if path.startswith("/azioni/backup/elimina/"):
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo eliminare un backup.",
                    extra_query=context_query,
                )
            file_name = unquote(path.removeprefix("/azioni/backup/elimina/")).strip()
            if delete_backup_archive(file_name):
                return redirect(
                    start_response,
                    "/maschere/backup",
                    ok="Backup eliminato.",
                    extra_query=work_year_query(context_query),
                )
            return redirect(
                start_response,
                "/maschere/backup",
                err="Backup non trovato o non eliminabile.",
                extra_query=work_year_query(context_query),
            )

        if path == "/azioni/aggiornamenti/installa":
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo installare un pacchetto aggiornamento.",
                    extra_query=context_query,
                )
            uploaded = (form_files or {}).get("file_zip_aggiornamento")
            if not uploaded:
                raise ValueError("Seleziona un file ZIP del pacchetto aggiornamento.")
            file_name = str(uploaded.get("filename") or "")
            if not file_name.lower().endswith(".zip"):
                raise ValueError("Il file selezionato non e un pacchetto ZIP valido.")
            package_bytes = bytes(uploaded.get("content") or b"")
            if not package_bytes:
                raise ValueError("Il file ZIP selezionato e vuoto.")
            package_version = extract_update_package_version(package_bytes, file_name)
            package_path = store_uploaded_update_package(file_name, package_bytes)
            schedule_update_package_install(package_path)
            if package_version:
                message = (
                    f"Aggiornamento avviato. Versione pacchetto: {package_version}. "
                    "Se Windows chiede una conferma amministratore, autorizza l'operazione. "
                    "Attendi qualche secondo e riapri il software."
                )
            else:
                message = (
                    "Aggiornamento avviato. Se Windows chiede una conferma amministratore, "
                    "autorizza l'operazione. Attendi qualche secondo e riapri il software."
                )
            return redirect(
                start_response,
                "/login",
                ok=message,
                extra_query=context_query,
            )

        if path == "/azioni/associati/template-excel":
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo generare il modello Excel.",
                    extra_query=context_query,
                )
            filename, content = build_associati_template_xlsx()
            start_response(
                "200 OK",
                [
                    ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("Content-Disposition", f'attachment; filename="{filename}"'),
                    ("Content-Length", str(len(content))),
                ],
            )
            return [content]

        if path == "/azioni/associati/importa-excel":
            if not current_user or not current_user.get("is_admin"):
                return redirect(
                    start_response,
                    "/",
                    err="Solo l'amministratore puo importare dati da Excel.",
                    extra_query=context_query,
                )
            uploaded = (form_files or {}).get("file_excel_associati")
            if not uploaded:
                raise ValueError("Seleziona un file Excel da importare.")
            file_name = str(uploaded.get("filename") or "")
            if not file_name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                raise ValueError("Il file selezionato non e un Excel supportato.")
            result = import_associati_from_excel(bytes(uploaded.get("content") or b""), current_user=current_user)
            ok_message = (
                f"Importazione completata. Inseriti: {result['inserted']}. "
                f"Duplicati saltati: {result['duplicates']}. "
                f"Righe con errori: {len(result['errors'])}."
            )
            if result["errors"]:
                ok_message += f" Primo errore: {result['errors'][0]}"
            return redirect(start_response, "/maschere/importa-associati", ok=ok_message, extra_query=context_query)

        if path.startswith("/azioni/crud/aggiorna/"):
            entity_key, record_id = path.removeprefix("/azioni/crud/aggiorna/").split("/", 1)
            return handle_crud_update(entity_key, int(record_id), form_data, start_response, current_user)

        if path.startswith("/azioni/crud/elimina/"):
            entity_key, record_id = path.removeprefix("/azioni/crud/elimina/").split("/", 1)
            return handle_crud_delete(entity_key, int(record_id), form_data, start_response)

        if path == "/azioni/associati/crea":
            with get_connection() as connection:
                progressive_number = reserve_progressive_number(connection, "associati")
                associati_insert_placeholders = ", ".join(["?"] * 40)
                cursor = connection.execute(
                    f"""
                    INSERT INTO associati (
                        numero_progressivo, codice_associato, nome, cognome, codice_fiscale, data_nascita,
                        sesso, comune_nascita, provincia_nascita, carica, email, telefono, indirizzo, cap, citta, provincia, impiego,
                        data_prima_iscrizione, stato_associato, liberatoria_video, patologie,
                        genitore_tutore_cognome, genitore_tutore_nome, genitore_tutore_cellulare, genitore_tutore_email, genitore_tutore_impiego, genitore_tutore_tipo_documento, genitore_tutore_numero_documento,
                        prelievo_altro_genitore_nome, prelievo_altro_genitore_cognome, prelievo_altro_genitore_cellulare, prelievo_altro_genitore_impiego, prelievo_altro_genitore_tipo_documento, prelievo_altro_genitore_numero_documento,
                        prelievo_altra_persona_nome, prelievo_altra_persona_cognome, prelievo_altra_persona_cellulare, prelievo_altra_persona_tipo_documento, prelievo_altra_persona_numero_documento,
                        note
                    ) VALUES ({associati_insert_placeholders})
                    """,
                    (
                        progressive_number,
                        format_progressive_code("associati", progressive_number),
                        required(form_data, "nome", "Nome"),
                        required(form_data, "cognome", "Cognome"),
                        optional(form_data, "codice_fiscale"),
                        optional(form_data, "data_nascita"),
                        normalized(form_data, "sesso", "M") or "M",
                        optional(form_data, "comune_nascita"),
                        optional(form_data, "provincia_nascita"),
                        resolved_carica_value(form_data, current_user, existing_value="Associato"),
                        optional(form_data, "email"),
                        optional(form_data, "telefono"),
                        optional(form_data, "indirizzo"),
                        optional(form_data, "cap"),
                        optional(form_data, "citta"),
                        optional(form_data, "provincia"),
                        optional(form_data, "impiego"),
                        required(form_data, "data_prima_iscrizione", "Data prima iscrizione"),
                        normalized(form_data, "stato_associato", "Attivo") or "Attivo",
                        normalized(form_data, "liberatoria_video", "Si") or "Si",
                        optional(form_data, "patologie"),
                        optional(form_data, "genitore_tutore_cognome"),
                        optional(form_data, "genitore_tutore_nome"),
                        optional(form_data, "genitore_tutore_cellulare"),
                        optional(form_data, "genitore_tutore_email"),
                        optional(form_data, "genitore_tutore_impiego"),
                        normalized(form_data, "genitore_tutore_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
                        optional(form_data, "genitore_tutore_numero_documento"),
                        optional(form_data, "prelievo_altro_genitore_nome"),
                        optional(form_data, "prelievo_altro_genitore_cognome"),
                        optional(form_data, "prelievo_altro_genitore_cellulare"),
                        optional(form_data, "prelievo_altro_genitore_impiego"),
                        normalized(form_data, "prelievo_altro_genitore_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
                        optional(form_data, "prelievo_altro_genitore_numero_documento"),
                        optional(form_data, "prelievo_altra_persona_nome"),
                        optional(form_data, "prelievo_altra_persona_cognome"),
                        optional(form_data, "prelievo_altra_persona_cellulare"),
                        normalized(form_data, "prelievo_altra_persona_tipo_documento", DEFAULT_DOCUMENT_TYPE) or DEFAULT_DOCUMENT_TYPE,
                        optional(form_data, "prelievo_altra_persona_numero_documento"),
                        optional(form_data, "note"),
                    ),
                )
                associato_id = int(cursor.lastrowid)
                work_year = current_work_year(context_query)
                data_prima_iscrizione = required(form_data, "data_prima_iscrizione", "Data prima iscrizione")
                tesseramento_data = data_prima_iscrizione or date.today().isoformat()
                tesseramento_importo = decimal_amount(default_tesseramento_quota_importo(), minimum="0.00")
                tesseramento_progressivo, codice_tesseramento = assign_tesseramento_identifier(connection, work_year)
                tesseramento_cursor = connection.execute(
                    """
                    INSERT INTO tesseramenti_annuali (
                        associato_id, anno_sociale, numero_progressivo_anno, codice_tesseramento, data_tesseramento, importo_dovuto, data_scadenza, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        associato_id,
                        work_year,
                        tesseramento_progressivo,
                        codice_tesseramento,
                        tesseramento_data,
                        format(tesseramento_importo, ".2f"),
                        f"{work_year}-12-31",
                        "Tesseramento generato automaticamente al salvataggio del nuovo associato",
                    ),
                )
                tesseramento_id = int(tesseramento_cursor.lastrowid)

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        tesseramento_data,
                    )
                    if importo_pagato > tesseramento_importo:
                        raise ValueError("L'importo pagato non puo superare l'importo dovuto del tesseramento.")
                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_tesseramenti (
                            tesseramento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            tesseramento_id,
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            "Pagamento registrato contestualmente alla creazione del nuovo associato",
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/tesseramenti/{payment_id}",
                        ok=f"Associato, tesseramento e pagamento registrati. Codice assegnato: {codice_tesseramento}.",
                        extra_query=context_query,
                    )

                connection.commit()
            return redirect(
                start_response,
                "/maschere/associati",
                ok=f"Associato e tesseramento salvati. Codice assegnato: {codice_tesseramento}.",
                extra_query=context_query,
            )

        if path == "/azioni/quote/crea":
            area = required(form_data, "area", "Area")
            if area not in {"tesseramenti", "campi-estivi", "oratorio"}:
                raise ValueError("Area quota non valida.")
            execute(
                """
                INSERT INTO quote_predefinite (
                    area, descrizione, importo, note
                ) VALUES (?, ?, ?, ?)
                """,
                (
                    area,
                    required(form_data, "descrizione", "Descrizione"),
                    required(form_data, "importo", "Importo"),
                    optional(form_data, "note"),
                ),
            )
            if area == "tesseramenti":
                redirect_path = "/maschere/tesseramenti"
                success_message = "Quota tesseramento salvata."
            elif area == "campi-estivi":
                redirect_path = "/maschere/campi-estivi"
                success_message = "Quota Campo estivo salvata."
            else:
                redirect_path = "/maschere/oratorio"
                success_message = "Quota Oratorio salvata."
            return redirect(start_response, redirect_path, ok=success_message, extra_query=context_query)

        if path == "/azioni/tesseramenti/crea":
            with get_connection() as connection:
                anno_sociale = int(required(form_data, "anno_sociale", "Anno sociale"))
                numero_progressivo_anno, codice_tesseramento = assign_tesseramento_identifier(connection, anno_sociale)
                cursor = connection.execute(
                    """
                    INSERT INTO tesseramenti_annuali (
                        associato_id, anno_sociale, numero_progressivo_anno, codice_tesseramento, data_tesseramento, importo_dovuto, data_scadenza, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        required(form_data, "associato_id", "Associato"),
                        anno_sociale,
                        numero_progressivo_anno,
                        codice_tesseramento,
                        required(form_data, "data_tesseramento", "Data tesseramento"),
                        required(form_data, "importo_dovuto", "Importo dovuto"),
                        optional(form_data, "data_scadenza"),
                        optional(form_data, "note"),
                    ),
                )
                tesseramento_id = int(cursor.lastrowid)

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        optional(form_data, "data_tesseramento") or date.today().isoformat(),
                    )
                    extra_tokens = popup_payment_extra_tokens(form_data)
                    importo_dovuto = decimal_amount(required(form_data, "importo_dovuto", "Importo dovuto"), minimum="0")
                    if extra_tokens:
                        base_row = build_manual_scadenza_row(
                            kind="tesseramenti",
                            source_id=tesseramento_id,
                            associato_id=int(required(form_data, "associato_id", "Associato")),
                            area="Tesseramento annuale",
                            riferimento=f"Anno {required(form_data, 'anno_sociale', 'Anno sociale')}",
                            scadenza=optional(form_data, "data_scadenza") or optional(form_data, "data_tesseramento") or data_pagamento,
                            importo_dovuto=importo_dovuto,
                            is_current=True,
                        )
                        group_code = register_grouped_multi_area_payment(
                            connection,
                            [base_row, *load_multi_area_scadenze(extra_tokens)],
                            importo_totale=importo_pagato,
                            data_pagamento=data_pagamento,
                            metodo_pagamento_id=metodo_pagamento_id,
                            riferimento="",
                            note=optional(form_data, "note"),
                        )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/multi-area-gruppo/{group_code}",
                            ok="Tesseramento e pagamento registrati.",
                            extra_query=context_query,
                        )

                    if importo_pagato > importo_dovuto:
                        raise ValueError("L'importo pagato non puo superare l'importo dovuto del tesseramento.")
                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_tesseramenti (
                            tesseramento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            tesseramento_id,
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            optional(form_data, "note"),
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/tesseramenti/{payment_id}",
                        ok="Tesseramento e pagamento registrati.",
                        extra_query=context_query,
                    )

                connection.commit()
            return redirect(start_response, "/maschere/tesseramenti", ok="Tesseramento salvato.", extra_query=context_query)

        if path == "/azioni/tesseramenti/pagamento":
            with get_connection() as connection:
                cursor = connection.execute(
                    """
                    INSERT INTO pagamenti_tesseramenti (
                        tesseramento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        required(form_data, "tesseramento_id", "Tesseramento"),
                        required(form_data, "data_pagamento", "Data pagamento"),
                        required(form_data, "importo", "Importo"),
                        required(form_data, "metodo_pagamento_id", "Metodo"),
                        optional(form_data, "riferimento"),
                        optional(form_data, "note"),
                    ),
                )
                payment_id = cursor.lastrowid
                connection.commit()
            return redirect(
                start_response,
                f"/ricevute/tesseramenti/{payment_id}",
                ok="Pagamento tesseramento registrato.",
                extra_query=context_query,
            )

        if path == "/azioni/corsi/tipologia":
            raise ValueError("Le tipologie corsi sono fisse e non e possibile inserirne di nuove.")

        if path == "/azioni/corsi/crea":
            data_inizio, data_fine = validate_course_date_range(
                required(form_data, "data_inizio", "Data inizio"),
                required(form_data, "data_fine", "Data fine"),
            )
            with get_connection() as connection:
                progressive_number = reserve_progressive_number(connection, "corsi")
                connection.execute(
                    """
                    INSERT INTO corsi (
                        numero_progressivo, codice_corso, nome, tipologia_corso_id, descrizione, quota_iscrizione_standard,
                        quota_mensile_standard, data_inizio, data_fine, sede, giorno_settimana, orario
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        progressive_number,
                        format_progressive_code("corsi", progressive_number),
                        required(form_data, "nome", "Nome corso"),
                        None,
                        optional(form_data, "descrizione"),
                        "0",
                        normalized(form_data, "quota_mensile_standard", "0"),
                        data_inizio,
                        data_fine,
                        optional(form_data, "sede"),
                        optional(form_data, "giorno_settimana"),
                        optional(form_data, "orario"),
                    ),
                )
                connection.commit()
            return redirect(start_response, "/maschere/corsi", ok="Corso salvato.", extra_query=context_query)

        if path == "/azioni/corsi/iscrizione":
            associato_id_value = required(form_data, "associato_id", "Associato")
            corso_id_value = int(required(form_data, "corso_id", "Corso"))
            work_year = current_work_year(context_query)
            if not associato_has_tesseramento_for_year(associato_id_value, work_year):
                return redirect_missing_tesseramento(
                    start_response,
                    associato_id=associato_id_value,
                    work_year=work_year,
                    area_label="Corsi",
                    query_params=context_query,
                )
            data_iscrizione_value = required(form_data, "data_iscrizione", "Data iscrizione")
            try:
                data_iscrizione_date = date.fromisoformat(data_iscrizione_value)
            except ValueError:
                raise ValueError("La data di iscrizione del corso non e valida.")

            with get_connection() as connection:
                effective_data_inizio, effective_data_fine = resolve_course_enrollment_window(
                    connection,
                    corso_id_value,
                    data_iscrizione_value,
                    optional(form_data, "data_inizio"),
                    None,
                )
                effective_start_date = date.fromisoformat(effective_data_inizio)
                cursor = connection.execute(
                    """
                    INSERT INTO iscrizioni_corsi (
                        associato_id, corso_id, data_iscrizione, data_inizio, data_fine, quota_iscrizione,
                        quota_mensile, stato_iscrizione, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        associato_id_value,
                        corso_id_value,
                        data_iscrizione_value,
                        effective_data_inizio,
                        effective_data_fine,
                        "0",
                        required(form_data, "quota_mensile", "Quota mensile"),
                        normalized(form_data, "stato_iscrizione", "Attiva") or "Attiva",
                        optional(form_data, "note"),
                    ),
                )
                iscrizione_id = int(cursor.lastrowid)
                quote_note = "Quota generata automaticamente al salvataggio dell'iscrizione"
                payment_scope = normalized(form_data, "pagamento_scope", "mese-iscrizione") or "mese-iscrizione"
                rate_ids = [ensure_course_rate_for_enrollment(
                    connection,
                    iscrizione_id,
                    effective_start_date.year,
                    effective_start_date.month,
                    note=quote_note,
                )]

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        data_iscrizione_value or date.today().isoformat(),
                    )
                    extra_tokens = popup_payment_extra_tokens(form_data)
                    if payment_scope == "mensilita-future":
                        fine_competenza = required(form_data, "pagamento_competenza_fine", "Ultimo mese da pagare")
                        end_year, end_month = parse_year_month_value(fine_competenza, "Ultimo mese da pagare")
                        rate_ids = ensure_course_rates_for_enrollment_range(
                            connection,
                            iscrizione_id,
                            effective_start_date.year,
                            effective_start_date.month,
                            end_year,
                            end_month,
                            note=quote_note,
                        )

                    placeholders = ",".join("?" for _ in rate_ids)
                    rate_rows = connection.execute(
                        f"""
                        SELECT
                            r.id,
                            r.anno,
                            r.mese,
                            ic.associato_id,
                            c.nome AS corso,
                            COALESCE(r.data_scadenza, printf('%04d-%02d-01', r.anno, r.mese)) AS scadenza,
                            r.importo_dovuto,
                            COALESCE((
                                SELECT SUM(prc.importo)
                                FROM pagamenti_rate_corsi prc
                                WHERE prc.rata_corso_id = r.id
                            ), 0) AS importo_pagato
                        FROM rate_corsi_mensili r
                        JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
                        JOIN corsi c ON c.id = ic.corso_id
                        WHERE r.id IN ({placeholders})
                        ORDER BY r.anno, r.mese, r.id
                        """,
                        tuple(rate_ids),
                    ).fetchall()
                    if len(rate_rows) != len(rate_ids):
                        raise ValueError("Una o piu quote mensili da pagare non sono piu disponibili.")

                    residui: list[tuple[int, Decimal]] = []
                    totale_residuo = Decimal("0.00")
                    for row in rate_rows:
                        residuo_rata = (
                            decimal_amount(row["importo_dovuto"]) - decimal_amount(row["importo_pagato"])
                        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                        residui.append((int(row["id"]), residuo_rata))
                        totale_residuo = (totale_residuo + residuo_rata).quantize(
                            Decimal("0.01"),
                            rounding=ROUND_HALF_UP,
                        )

                    if extra_tokens:
                        base_rows = [
                            build_manual_scadenza_row(
                                kind="corsi-rate",
                                source_id=int(row["id"]),
                                associato_id=int(row["associato_id"]),
                                area="Corso - quota mensile",
                                riferimento=f"{plain_text(row['corso'])} {month_label(int(row['mese']))} {int(row['anno'])}",
                                scadenza=str(row["scadenza"] or ""),
                                importo_dovuto=row["importo_dovuto"],
                                importo_pagato=row["importo_pagato"],
                                is_current=True,
                            )
                            for row in rate_rows
                        ]
                        group_code = register_grouped_multi_area_payment(
                            connection,
                            [*base_rows, *load_multi_area_scadenze(extra_tokens)],
                            importo_totale=importo_pagato,
                            data_pagamento=data_pagamento,
                            metodo_pagamento_id=metodo_pagamento_id,
                            riferimento="",
                            note=optional(form_data, "note"),
                        )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/multi-area-gruppo/{group_code}",
                            ok="Iscrizione corso e pagamento registrati.",
                            extra_query=context_query,
                        )

                    if importo_pagato > totale_residuo:
                        if payment_scope == "mensilita-future":
                            raise ValueError("L'importo pagato non puo superare il totale delle mensilita selezionate.")
                        raise ValueError("L'importo pagato non puo superare la quota mensile del mese di iscrizione.")

                    if payment_scope == "mensilita-future":
                        group_code = generate_receipt_group_code()
                        remaining = importo_pagato
                        for rate_id, residuo_rata in residui:
                            if remaining <= Decimal("0.00"):
                                break
                            importo_rata = min(remaining, residuo_rata).quantize(
                                Decimal("0.01"),
                                rounding=ROUND_HALF_UP,
                            )
                            if importo_rata <= Decimal("0.00"):
                                continue
                            connection.execute(
                                """
                                INSERT INTO pagamenti_rate_corsi (
                                    rata_corso_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    rate_id,
                                    data_pagamento,
                                    format(importo_rata, ".2f"),
                                    metodo_pagamento_id,
                                    "",
                                    optional(form_data, "note"),
                                    group_code,
                                ),
                            )
                            remaining = (remaining - importo_rata).quantize(
                                Decimal("0.01"),
                                rounding=ROUND_HALF_UP,
                            )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/corsi-rate-gruppo/{group_code}",
                            ok="Iscrizione corso, quote mensili e pagamento registrati.",
                            extra_query=context_query,
                        )

                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_rate_corsi (
                            rata_corso_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            rate_ids[0],
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            optional(form_data, "note"),
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/corsi-rata/{payment_id}",
                        ok="Iscrizione corso, prima quota utile e pagamento registrati.",
                        extra_query=context_query,
                    )

                connection.commit()
                first_quota_label = f"{month_label(effective_start_date.month)} {effective_start_date.year}"
            return redirect(
                start_response,
                "/maschere/corsi",
                ok=f"Iscrizione corso salvata. La quota mensile di {first_quota_label} e stata generata automaticamente.",
                extra_query=context_query,
            )

        if path == "/azioni/corsi/pagamento-iscrizione":
            raise ValueError("Il pagamento di iscrizione al corso e stato disattivato.")

        if path == "/azioni/corsi/rata":
            raise ValueError("L'inserimento della singola quota mensile e disattivato. Le quote corsi vengono generate automaticamente dal gestionale.")

        if path in ("/azioni/corsi/pagamento-rata", "/azioni/corsi/pagamento-rate"):
            rate_ids = multi_values(form_data, "rata_corso_id")
            if not rate_ids:
                raise ValueError("Seleziona almeno una mensilita da saldare.")

            associato_id = required(form_data, "associato_id", "Associato")
            importo_totale = decimal_amount(required(form_data, "importo", "Importo"), minimum="0.01")
            placeholders = ",".join("?" for _ in rate_ids)
            rows = fetch_all(
                f"""
                SELECT id, associato_id, corso, competenza, saldo_residuo
                FROM v_rate_corsi_saldo
                WHERE id IN ({placeholders})
                ORDER BY anno, mese, corso
                """,
                tuple(int(rate_id) for rate_id in rate_ids),
            )
            if len(rows) != len(rate_ids):
                raise ValueError("Una o piu mensilita selezionate non sono piu disponibili.")

            selected_associati = {str(row["associato_id"]) for row in rows}
            if selected_associati != {associato_id}:
                raise ValueError("Le mensilita selezionate devono appartenere tutte allo stesso associato.")

            if any(float(row["saldo_residuo"] or 0) <= 0 for row in rows):
                raise ValueError("Una o piu mensilita selezionate risultano gia saldate.")

            totale_residuo = sum(decimal_amount(row["saldo_residuo"]) for row in rows)
            if importo_totale > totale_residuo:
                raise ValueError("L'importo inserito supera il residuo totale delle mensilita selezionate.")

            group_code = generate_receipt_group_code()
            remaining = importo_totale
            with get_connection() as connection:
                for row in rows:
                    residuo_rata = decimal_amount(row["saldo_residuo"])
                    if remaining <= Decimal("0.00"):
                        break
                    importo_rata = min(remaining, residuo_rata).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    if importo_rata <= Decimal("0.00"):
                        continue
                    connection.execute(
                        """
                        INSERT INTO pagamenti_rate_corsi (
                            rata_corso_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note, gruppo_ricevuta
                        ) VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row["id"],
                            required(form_data, "data_pagamento", "Data pagamento"),
                            format(importo_rata, ".2f"),
                            required(form_data, "metodo_pagamento_id", "Metodo"),
                            optional(form_data, "riferimento"),
                            optional(form_data, "note"),
                            group_code,
                        ),
                    )
                    remaining = (remaining - importo_rata).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                connection.commit()

            return redirect(
                start_response,
                f"/ricevute/corsi-rate-gruppo/{group_code}",
                ok="Pagamento quote mensili registrato.",
                extra_query=context_query,
            )

        if path == "/azioni/corsi/genera-rate":
            anno = int(required(form_data, "anno", "Anno"))
            mese = int(required(form_data, "mese", "Mese"))
            data_scadenza = optional(form_data, "data_scadenza")
            note = optional(form_data, "note")
            with get_connection() as connection:
                inserted, skipped = generate_course_rates_for_month(
                    connection,
                    anno,
                    mese,
                    data_scadenza=data_scadenza,
                    note=note,
                )
                connection.commit()

            message = (
                f"Sono state generate {inserted} quote mensili per {month_label(mese)} {anno}."
                if inserted
                else f"Nessuna nuova quota da generare per {month_label(mese)} {anno}."
            )
            if skipped:
                message += f" {skipped} gia presenti."
            return redirect(start_response, "/maschere/corsi", ok=message, extra_query=context_query)

        if path == "/azioni/corsi/genera-rate-mancanti":
            work_year = int(required(form_data, "anno", "Anno"))
            reminder = pending_course_monthly_generations(work_year)
            if reminder is None:
                return redirect(
                    start_response,
                    "/",
                    ok="Non risultano quote mensili mancanti da generare.",
                    extra_query=context_query or {"anno_lavoro": str(work_year)},
                )

            inserted_total = 0
            skipped_total = 0
            generated_labels: list[str] = []
            with get_connection() as connection:
                for anno, mese in reminder["missing_months"]:
                    inserted, skipped = generate_course_rates_for_month(
                        connection,
                        int(anno),
                        int(mese),
                        data_scadenza=default_mass_rate_due_date(int(anno), int(mese)),
                        note="Generazione automatica mesi mancanti",
                    )
                    inserted_total += inserted
                    skipped_total += skipped
                    generated_labels.append(f"{month_label(int(mese))} {int(anno)}")
                connection.commit()

            context = context_query or {"anno_lavoro": str(work_year)}
            message = (
                f"Generazione automatica completata per: {', '.join(generated_labels)}. "
                f"Nuove quote create: {inserted_total}."
            )
            if skipped_total:
                message += f" Quote gia presenti: {skipped_total}."
            return redirect(start_response, "/", ok=message, extra_query=context)

        if path == "/azioni/campi-estivi/crea":
            year = int(required(form_data, "anno", "Anno"))
            existing_camp = fetch_scalar("SELECT id FROM campi_estivi WHERE anno = ?", (year,))
            if existing_camp:
                raise ValueError("Per ogni anno di lavoro e previsto un solo Campo estivo.")
            with get_connection() as connection:
                progressive_number = reserve_progressive_number(connection, "campi_estivi")
                connection.execute(
                    """
                    INSERT INTO campi_estivi (
                        numero_progressivo, codice_campo, nome, anno, data_inizio, data_fine, sede,
                        quota_partecipazione_standard, posti_massimi, descrizione
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        progressive_number,
                        format_progressive_code("campi_estivi", progressive_number),
                        required(form_data, "nome", "Nome"),
                        year,
                        required(form_data, "data_inizio", "Data inizio"),
                        required(form_data, "data_fine", "Data fine"),
                        optional(form_data, "sede"),
                        normalized(form_data, "quota_partecipazione_standard", "0"),
                        optional(form_data, "posti_massimi"),
                        optional(form_data, "descrizione"),
                    ),
                )
                connection.commit()
            return redirect(start_response, "/maschere/campi-estivi", ok="Campo estivo salvato.", extra_query=context_query)

        if path == "/azioni/campi-estivi/quota-standard":
            work_year = current_work_year(context_query)
            standard_fee = required(form_data, "quota_partecipazione_standard", "Quota di partecipazione standard")
            with get_connection() as connection:
                ensure_estate_record(connection, work_year, standard_fee)
                connection.commit()
            return redirect(
                start_response,
                "/maschere/campi-estivi",
                ok="Quota standard Campo estivo aggiornata.",
                extra_query=context_query,
            )

        if path == "/azioni/campi-estivi/iscrizione":
            work_year = current_work_year(context_query)
            associato_id_value = required(form_data, "associato_id", "Associato")
            if not associato_has_tesseramento_for_year(associato_id_value, work_year):
                return redirect_missing_tesseramento(
                    start_response,
                    associato_id=associato_id_value,
                    work_year=work_year,
                    area_label="Campo estivo",
                    query_params=context_query,
                )
            with get_connection() as connection:
                estate_id = connection.execute(
                    "SELECT id FROM campi_estivi WHERE anno = ? ORDER BY id LIMIT 1",
                    (work_year,),
                ).fetchone()
                campo_estivo_id = (
                    int(estate_id["id"])
                    if estate_id is not None
                    else ensure_estate_record(connection, work_year)
                )
                cursor = connection.execute(
                    """
                    INSERT INTO iscrizioni_campi_estivi (
                        associato_id, campo_estivo_id, data_iscrizione, quota_partecipazione, stato_iscrizione, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        associato_id_value,
                        campo_estivo_id,
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                        required(form_data, "quota_partecipazione", "Quota partecipazione"),
                        normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
                        optional(form_data, "note"),
                    ),
                )
                iscrizione_id = int(cursor.lastrowid)

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                    )
                    extra_tokens = popup_payment_extra_tokens(form_data)
                    quota_dovuta = decimal_amount(required(form_data, "quota_partecipazione", "Importo"), minimum="0")
                    if extra_tokens:
                        campo_row = connection.execute(
                            "SELECT nome, COALESCE(data_inizio, ?) AS scadenza FROM campi_estivi WHERE id = ?",
                            (required(form_data, "data_iscrizione", "Data iscrizione"), campo_estivo_id),
                        ).fetchone()
                        base_row = build_manual_scadenza_row(
                            kind="campi-estivi",
                            source_id=iscrizione_id,
                            associato_id=int(associato_id_value),
                            area="Campo estivo",
                            riferimento=str(campo_row["nome"] if campo_row is not None else f"Campo estivo {work_year}"),
                            scadenza=str(campo_row["scadenza"] if campo_row is not None else required(form_data, "data_iscrizione", "Data iscrizione")),
                            importo_dovuto=quota_dovuta,
                            is_current=True,
                        )
                        group_code = register_grouped_multi_area_payment(
                            connection,
                            [base_row, *load_multi_area_scadenze(extra_tokens)],
                            importo_totale=importo_pagato,
                            data_pagamento=data_pagamento,
                            metodo_pagamento_id=metodo_pagamento_id,
                            riferimento="",
                            note=optional(form_data, "note"),
                        )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/multi-area-gruppo/{group_code}",
                            ok="Iscrizione e pagamento Campo estivo registrati.",
                            extra_query=context_query,
                        )

                    if importo_pagato > quota_dovuta:
                        raise ValueError("L'importo pagato non puo superare l'importo dovuto del Campo estivo.")
                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_campi_estivi (
                            iscrizione_campo_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            iscrizione_id,
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            optional(form_data, "note"),
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/campi-estivi/{payment_id}",
                        ok="Iscrizione e pagamento Campo estivo registrati.",
                        extra_query=context_query,
                    )
                connection.commit()
            return redirect(start_response, "/maschere/campi-estivi", ok="Iscrizione Campo estivo salvata.", extra_query=context_query)

        if path == "/azioni/campi-estivi/pagamento":
            with get_connection() as connection:
                cursor = connection.execute(
                    """
                    INSERT INTO pagamenti_campi_estivi (
                        iscrizione_campo_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        required(form_data, "iscrizione_campo_id", "Iscrizione Campo estivo"),
                        required(form_data, "data_pagamento", "Data pagamento"),
                        required(form_data, "importo", "Importo"),
                        required(form_data, "metodo_pagamento_id", "Metodo"),
                        optional(form_data, "riferimento"),
                        optional(form_data, "note"),
                    ),
                )
                payment_id = cursor.lastrowid
                connection.commit()
            return redirect(
                start_response,
                f"/ricevute/campi-estivi/{payment_id}",
                ok="Pagamento Campo estivo registrato.",
                extra_query=context_query,
            )

        if path == "/azioni/oratorio/iscrizione":
            work_year = current_work_year(context_query)
            associato_id_value = required(form_data, "associato_id", "Associato")
            if not associato_has_tesseramento_for_year(associato_id_value, work_year):
                return redirect_missing_tesseramento(
                    start_response,
                    associato_id=associato_id_value,
                    work_year=work_year,
                    area_label=ORATORIO_LABEL,
                    query_params=context_query,
                )
            with get_connection() as connection:
                oratorio_id_row = connection.execute(
                    "SELECT id FROM oratorio WHERE anno = ? ORDER BY id LIMIT 1",
                    (work_year,),
                ).fetchone()
                oratorio_id = (
                    int(oratorio_id_row["id"])
                    if oratorio_id_row is not None
                    else ensure_oratorio_record(connection, work_year)
                )
                cursor = connection.execute(
                    """
                    INSERT INTO iscrizioni_oratorio (
                        associato_id, oratorio_id, data_iscrizione, quota_partecipazione, stato_iscrizione, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        associato_id_value,
                        oratorio_id,
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                        required(form_data, "quota_partecipazione", "Quota partecipazione"),
                        normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
                        optional(form_data, "note"),
                    ),
                )
                iscrizione_id = int(cursor.lastrowid)

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                    )
                    extra_tokens = popup_payment_extra_tokens(form_data)
                    quota_dovuta = decimal_amount(required(form_data, "quota_partecipazione", "Importo"), minimum="0")
                    if extra_tokens:
                        oratorio_row = connection.execute(
                            "SELECT nome, COALESCE(data_inizio, ?) AS scadenza FROM oratorio WHERE id = ?",
                            (required(form_data, "data_iscrizione", "Data iscrizione"), oratorio_id),
                        ).fetchone()
                        base_row = build_manual_scadenza_row(
                            kind="oratorio",
                            source_id=iscrizione_id,
                            associato_id=int(associato_id_value),
                            area=ORATORIO_LABEL,
                            riferimento=str(oratorio_row["nome"] if oratorio_row is not None else f"{ORATORIO_LABEL} {work_year}"),
                            scadenza=str(oratorio_row["scadenza"] if oratorio_row is not None else required(form_data, "data_iscrizione", "Data iscrizione")),
                            importo_dovuto=quota_dovuta,
                            is_current=True,
                        )
                        group_code = register_grouped_multi_area_payment(
                            connection,
                            [base_row, *load_multi_area_scadenze(extra_tokens)],
                            importo_totale=importo_pagato,
                            data_pagamento=data_pagamento,
                            metodo_pagamento_id=metodo_pagamento_id,
                            riferimento="",
                            note=optional(form_data, "note"),
                        )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/multi-area-gruppo/{group_code}",
                            ok="Iscrizione e pagamento Oratorio registrati.",
                            extra_query=context_query,
                        )

                    if importo_pagato > quota_dovuta:
                        raise ValueError("L'importo pagato non puo superare l'importo dovuto di Oratorio.")
                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_oratorio (
                            iscrizione_oratorio_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            iscrizione_id,
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            optional(form_data, "note"),
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/oratorio/{payment_id}",
                        ok="Iscrizione e pagamento Oratorio registrati.",
                        extra_query=context_query,
                    )
                connection.commit()
            return redirect(start_response, "/maschere/oratorio", ok="Iscrizione Oratorio salvata.", extra_query=context_query)

        if path == "/azioni/oratorio/pagamento":
            with get_connection() as connection:
                cursor = connection.execute(
                    """
                    INSERT INTO pagamenti_oratorio (
                        iscrizione_oratorio_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        required(form_data, "iscrizione_oratorio_id", "Iscrizione Oratorio"),
                        required(form_data, "data_pagamento", "Data pagamento"),
                        required(form_data, "importo", "Importo"),
                        required(form_data, "metodo_pagamento_id", "Metodo"),
                        optional(form_data, "riferimento"),
                        optional(form_data, "note"),
                    ),
                )
                payment_id = cursor.lastrowid
                connection.commit()
            return redirect(
                start_response,
                f"/ricevute/oratorio/{payment_id}",
                ok="Pagamento Oratorio registrato.",
                extra_query=context_query,
            )

        if path == "/azioni/eventi/crea":
            with get_connection() as connection:
                progressive_number = reserve_progressive_number(connection, "eventi")
                connection.execute(
                    """
                    INSERT INTO eventi (
                        numero_progressivo, codice_evento, nome, tipologia, data_evento, luogo,
                        quota_partecipazione_standard, posti_massimi, descrizione
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        progressive_number,
                        format_progressive_code("eventi", progressive_number),
                        required(form_data, "nome", "Nome evento"),
                        optional(form_data, "tipologia"),
                        required(form_data, "data_evento", "Data evento"),
                        optional(form_data, "luogo"),
                        normalized(form_data, "quota_partecipazione_standard", "0"),
                        optional(form_data, "posti_massimi"),
                        optional(form_data, "descrizione"),
                    ),
                )
                connection.commit()
            return redirect(start_response, "/maschere/eventi", ok="Evento salvato.", extra_query=context_query)

        if path == "/azioni/eventi/iscrizione":
            associato_id_value = required(form_data, "associato_id", "Associato")
            work_year = current_work_year(context_query)
            if not associato_has_tesseramento_for_year(associato_id_value, work_year):
                return redirect_missing_tesseramento(
                    start_response,
                    associato_id=associato_id_value,
                    work_year=work_year,
                    area_label="Eventi",
                    query_params=context_query,
                )
            with get_connection() as connection:
                cursor = connection.execute(
                    """
                    INSERT INTO iscrizioni_eventi (
                        associato_id, evento_id, data_iscrizione, quota_partecipazione, stato_iscrizione, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        associato_id_value,
                        required(form_data, "evento_id", "Evento"),
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                        required(form_data, "quota_partecipazione", "Quota partecipazione"),
                        normalized(form_data, "stato_iscrizione", "Iscritto") or "Iscritto",
                        optional(form_data, "note"),
                    ),
                )
                iscrizione_id = int(cursor.lastrowid)

                if popup_payment_requested(form_data):
                    metodo_pagamento_id, importo_pagato, data_pagamento = popup_payment_payload(
                        form_data,
                        required(form_data, "data_iscrizione", "Data iscrizione"),
                    )
                    extra_tokens = popup_payment_extra_tokens(form_data)
                    quota_dovuta = decimal_amount(required(form_data, "quota_partecipazione", "Quota dovuta"), minimum="0")
                    if extra_tokens:
                        evento_row = connection.execute(
                            "SELECT nome, COALESCE(data_evento, ?) AS scadenza FROM eventi WHERE id = ?",
                            (required(form_data, "data_iscrizione", "Data iscrizione"), required(form_data, "evento_id", "Evento")),
                        ).fetchone()
                        base_row = build_manual_scadenza_row(
                            kind="eventi",
                            source_id=iscrizione_id,
                            associato_id=int(associato_id_value),
                            area="Evento",
                            riferimento=str(evento_row["nome"] if evento_row is not None else "Evento"),
                            scadenza=str(evento_row["scadenza"] if evento_row is not None else required(form_data, "data_iscrizione", "Data iscrizione")),
                            importo_dovuto=quota_dovuta,
                            is_current=True,
                        )
                        group_code = register_grouped_multi_area_payment(
                            connection,
                            [base_row, *load_multi_area_scadenze(extra_tokens)],
                            importo_totale=importo_pagato,
                            data_pagamento=data_pagamento,
                            metodo_pagamento_id=metodo_pagamento_id,
                            riferimento="",
                            note=optional(form_data, "note"),
                        )
                        connection.commit()
                        return redirect(
                            start_response,
                            f"/ricevute/multi-area-gruppo/{group_code}",
                            ok="Iscrizione e pagamento evento registrati.",
                            extra_query=context_query,
                        )

                    if importo_pagato > quota_dovuta:
                        raise ValueError("L'importo pagato non puo superare l'importo dovuto dell'evento.")
                    payment_cursor = connection.execute(
                        """
                        INSERT INTO pagamenti_eventi (
                            iscrizione_evento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                        ) VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (
                            iscrizione_id,
                            data_pagamento,
                            format(importo_pagato, ".2f"),
                            metodo_pagamento_id,
                            "",
                            optional(form_data, "note"),
                        ),
                    )
                    payment_id = int(payment_cursor.lastrowid)
                    connection.commit()
                    return redirect(
                        start_response,
                        f"/ricevute/eventi/{payment_id}",
                        ok="Iscrizione e pagamento evento registrati.",
                        extra_query=context_query,
                    )
                connection.commit()
            return redirect(start_response, "/maschere/eventi", ok="Iscrizione evento salvata.", extra_query=context_query)

        if path == "/azioni/eventi/pagamento":
            with get_connection() as connection:
                cursor = connection.execute(
                    """
                    INSERT INTO pagamenti_eventi (
                        iscrizione_evento_id, data_pagamento, importo, metodo_pagamento_id, riferimento, note
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        required(form_data, "iscrizione_evento_id", "Iscrizione evento"),
                        required(form_data, "data_pagamento", "Data pagamento"),
                        required(form_data, "importo", "Importo"),
                        required(form_data, "metodo_pagamento_id", "Metodo"),
                        optional(form_data, "riferimento"),
                        optional(form_data, "note"),
                    ),
                )
                payment_id = cursor.lastrowid
                connection.commit()
            return redirect(
                start_response,
                f"/ricevute/eventi/{payment_id}",
                ok="Pagamento evento registrato.",
                extra_query=context_query,
            )

        if path == "/azioni/pagamenti-multi-area/crea":
            scadenza_tokens = multi_values(form_data, "scadenza_id")
            if not scadenza_tokens:
                raise ValueError("Seleziona almeno una scadenza da saldare.")

            associato_id = required(form_data, "associato_id", "Associato")
            importo_totale = decimal_amount(required(form_data, "importo", "Importo"), minimum="0.01")
            rows = load_multi_area_scadenze(scadenza_tokens)
            with get_connection() as connection:
                group_code = register_grouped_multi_area_payment(
                    connection,
                    rows,
                    importo_totale=importo_totale,
                    data_pagamento=required(form_data, "data_pagamento", "Data pagamento"),
                    metodo_pagamento_id=required(form_data, "metodo_pagamento_id", "Metodo"),
                    riferimento=optional(form_data, "riferimento"),
                    note=optional(form_data, "note"),
                )
                connection.commit()

            return redirect(
                start_response,
                f"/ricevute/multi-area-gruppo/{group_code}",
                ok="Pagamento multi-area registrato.",
                extra_query=context_query,
            )

        if path.startswith("/azioni/pagamenti-multi-area/elimina/"):
            group_code = path.removeprefix("/azioni/pagamenti-multi-area/elimina/").strip()
            if not group_code:
                raise ValueError("Pagamento multi-area non valido.")
            delete_multi_area_group(group_code)
            return redirect(
                start_response,
                "/maschere/pagamenti-multi-area",
                ok="Pagamento multi-area eliminato e scadenze ripristinate.",
                extra_query={**context_query, "vista": "dati"},
            )

        return redirect(start_response, "/", err="Azione non riconosciuta.", extra_query=context_query)

    except ValueError as error:
        return redirect(start_response, fallback_path(path), err=str(error), extra_query=context_query)
    except sqlite3.IntegrityError as error:
        return redirect(start_response, fallback_path(path), err=friendly_db_error(error), extra_query=context_query)
    except sqlite3.DatabaseError as error:
        return redirect(start_response, fallback_path(path), err=friendly_db_error(error), extra_query=context_query)


def fallback_path(path: str) -> str:
    if path.startswith("/azioni/crud/aggiorna/") or path.startswith("/azioni/crud/elimina/"):
        record_path = path.split("/")
        entity_key = record_path[4] if len(record_path) > 4 else ""
        config = CRUD_CONFIG.get(entity_key)
        if config:
            return config["return_path"]

    mapping = {
        "/azioni/associati/crea": "/maschere/associati",
        "/azioni/accesso/primo-admin": "/login",
        "/azioni/accesso/login": "/login",
        "/azioni/accesso/logout": "/login",
        "/azioni/accesso/cambia-password": "/maschere/accesso",
        "/azioni/accesso/recupera-password": "/recupera-password",
        "/azioni/utenti/crea": "/maschere/utenti",
        "/azioni/tesseramenti/crea": "/maschere/tesseramenti",
        "/azioni/tesseramenti/pagamento": "/maschere/tesseramenti",
        "/azioni/corsi/tipologia": "/maschere/corsi",
        "/azioni/corsi/crea": "/maschere/corsi",
        "/azioni/corsi/iscrizione": "/maschere/corsi",
        "/azioni/corsi/pagamento-iscrizione": "/maschere/corsi",
        "/azioni/corsi/rata": "/maschere/corsi",
        "/azioni/corsi/pagamento-rata": "/maschere/corsi",
        "/azioni/corsi/pagamento-rate": "/maschere/corsi",
        "/azioni/corsi/genera-rate": "/maschere/corsi",
        "/azioni/corsi/genera-rate-mancanti": "/",
        "/azioni/campi-estivi/crea": "/maschere/campi-estivi",
        "/azioni/campi-estivi/iscrizione": "/maschere/campi-estivi",
        "/azioni/campi-estivi/pagamento": "/maschere/campi-estivi",
        "/azioni/oratorio/iscrizione": "/maschere/oratorio",
        "/azioni/oratorio/pagamento": "/maschere/oratorio",
        "/azioni/eventi/crea": "/maschere/eventi",
        "/azioni/eventi/iscrizione": "/maschere/eventi",
        "/azioni/eventi/pagamento": "/maschere/eventi",
        "/azioni/pagamenti-multi-area/crea": "/maschere/pagamenti-multi-area",
    }
    if path.startswith("/azioni/pagamenti-multi-area/elimina/"):
        return "/maschere/pagamenti-multi-area"
    if path.startswith("/azioni/utenti/aggiorna/") or path.startswith("/azioni/utenti/password/") or path.startswith("/azioni/utenti/stato/"):
        return f"/maschere/utenti/gestione/{path.rsplit('/', 1)[-1]}"
    return mapping.get(path, "/")


def response_message_from_headers(headers: list[tuple[str, str]]) -> tuple[str, str]:
    location = next((value for name, value in headers if name.lower() == "location"), "")
    if not location:
        return "", ""
    parsed = urlsplit(location)
    query = parse_qs(parsed.query, keep_blank_values=True)
    ok_message = " ".join(query.get("ok", [])).strip()
    err_message = " ".join(query.get("err", [])).strip()
    return ok_message, err_message


def request_client_ip(environ: dict) -> str:
    forwarded = (environ.get("HTTP_X_FORWARDED_FOR") or "").strip()
    if forwarded:
        return forwarded.split(",", 1)[0].strip()
    return (environ.get("REMOTE_ADDR") or "").strip()


def request_user_agent(environ: dict) -> str:
    return (environ.get("HTTP_USER_AGENT") or "").strip()


def request_work_year_for_log(query_params: dict[str, str], form_data: dict[str, str]) -> int | None:
    raw_value = (form_data.get("anno_lavoro") or query_params.get("anno_lavoro") or "").strip()
    if raw_value.isdigit():
        return int(raw_value)
    return None


def activity_username_for_log(
    current_user: dict[str, object] | None,
    path: str,
    form_data: dict[str, str],
) -> str:
    if current_user and current_user.get("username"):
        return str(current_user["username"])
    if path in {"/azioni/accesso/login", "/azioni/accesso/primo-admin", "/azioni/accesso/recupera-password"}:
        return normalized(form_data, "username", "") or "anonimo"
    return "anonimo"


def associato_log_context_by_id(associato_id: int) -> tuple[int | None, str, str]:
    row = fetch_one(
        f"""
        SELECT id, codice_associato, {associato_display_sql('associati')} AS associato
        FROM associati
        WHERE id = ?
        """,
        (associato_id,),
    )
    if row is None:
        return None, "", ""
    return int(row["id"]), str(row["codice_associato"] or ""), str(row["associato"] or "")


def resolve_activity_associato_context(
    path: str,
    query_params: dict[str, str],
    form_data: dict[str, str],
) -> tuple[int | None, str, str]:
    direct_value = normalized(form_data, "associato_id", "") or normalized(query_params, "associato_id", "")
    if direct_value.isdigit():
        return associato_log_context_by_id(int(direct_value))

    if path.startswith("/report/associato/"):
        associato_id = path.removeprefix("/report/associato/")
        if associato_id.isdigit():
            return associato_log_context_by_id(int(associato_id))

    if path.startswith("/export/excel/associato/"):
        associato_id = path.removeprefix("/export/excel/associato/")
        if associato_id.isdigit():
            return associato_log_context_by_id(int(associato_id))

    if path.startswith("/export/pdf/associato/"):
        associato_id = path.removeprefix("/export/pdf/associato/")
        if associato_id.isdigit():
            return associato_log_context_by_id(int(associato_id))

    if path.startswith("/modifica/associati/"):
        associato_id = path.removeprefix("/modifica/associati/")
        if associato_id.isdigit():
            return associato_log_context_by_id(int(associato_id))

    if path.startswith("/azioni/crud/aggiorna/associati/") or path.startswith("/azioni/crud/elimina/associati/"):
        associato_id = path.rsplit("/", 1)[-1]
        if associato_id.isdigit():
            return associato_log_context_by_id(int(associato_id))

    tesseramento_id = normalized(form_data, "tesseramento_id", "")
    if tesseramento_id.isdigit():
        row = fetch_one("SELECT associato_id FROM tesseramenti_annuali WHERE id = ?", (int(tesseramento_id),))
        if row is not None:
            return associato_log_context_by_id(int(row["associato_id"]))

    iscrizione_campo_id = normalized(form_data, "iscrizione_campo_id", "")
    if iscrizione_campo_id.isdigit():
        row = fetch_one("SELECT associato_id FROM iscrizioni_campi_estivi WHERE id = ?", (int(iscrizione_campo_id),))
        if row is not None:
            return associato_log_context_by_id(int(row["associato_id"]))

    iscrizione_evento_id = normalized(form_data, "iscrizione_evento_id", "")
    if iscrizione_evento_id.isdigit():
        row = fetch_one("SELECT associato_id FROM iscrizioni_eventi WHERE id = ?", (int(iscrizione_evento_id),))
        if row is not None:
            return associato_log_context_by_id(int(row["associato_id"]))

    rata_ids = multi_values(form_data, "rata_corso_id")
    if rata_ids:
        placeholders = ",".join("?" for _ in rata_ids)
        rows = fetch_all(
            f"""
            SELECT DISTINCT ic.associato_id
            FROM rate_corsi_mensili r
            JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
            WHERE r.id IN ({placeholders})
            """,
            tuple(int(rata_id) for rata_id in rata_ids if rata_id.isdigit()),
        )
        associato_ids = {int(row["associato_id"]) for row in rows if row["associato_id"] is not None}
        if len(associato_ids) == 1:
            return associato_log_context_by_id(next(iter(associato_ids)))

    if path.startswith("/azioni/pagamenti-multi-area/elimina/"):
        group_code = path.removeprefix("/azioni/pagamenti-multi-area/elimina/").strip()
        if group_code:
            row = fetch_one(
                """
                SELECT associato_id
                FROM (
                    SELECT t.associato_id AS associato_id
                    FROM pagamenti_tesseramenti pt
                    JOIN tesseramenti_annuali t ON t.id = pt.tesseramento_id
                    WHERE pt.gruppo_ricevuta = ?
                    UNION ALL
                    SELECT ic.associato_id AS associato_id
                    FROM pagamenti_rate_corsi prc
                    JOIN rate_corsi_mensili r ON r.id = prc.rata_corso_id
                    JOIN iscrizioni_corsi ic ON ic.id = r.iscrizione_corso_id
                    WHERE prc.gruppo_ricevuta = ?
                    UNION ALL
                    SELECT ice.associato_id AS associato_id
                    FROM pagamenti_campi_estivi pce
                    JOIN iscrizioni_campi_estivi ice ON ice.id = pce.iscrizione_campo_id
                    WHERE pce.gruppo_ricevuta = ?
                    UNION ALL
                    SELECT ie.associato_id AS associato_id
                    FROM pagamenti_eventi pe
                    JOIN iscrizioni_eventi ie ON ie.id = pe.iscrizione_evento_id
                    WHERE pe.gruppo_ricevuta = ?
                ) groups
                LIMIT 1
                """,
                (group_code, group_code, group_code, group_code),
            )
            if row is not None and row["associato_id"] is not None:
                return associato_log_context_by_id(int(row["associato_id"]))

    return None, "", ""


def report_title_for_key(report_key: str, query_params: dict[str, str]) -> str:
    try:
        return get_report_definition(report_key, query_params)["title"]
    except Exception:
        return report_key or "Report"


def describe_logged_activity(
    method: str,
    path: str,
    query_params: dict[str, str],
    form_data: dict[str, str],
) -> tuple[str, str, str]:
    work_year = request_work_year_for_log(query_params, form_data)
    work_year_detail = f"Anno di lavoro: {work_year}" if work_year else ""

    if method == "POST":
        if path == "/azioni/accesso/primo-admin":
            return "Accesso", "Configurazione utente amministratore iniziale", ""
        if path == "/azioni/accesso/login":
            return "Accesso", "Accesso al software", ""
        if path == "/azioni/accesso/logout":
            return "Accesso", "Logout dal software", ""
        if path == "/azioni/accesso/cambia-password":
            return "Accesso", "Cambio password del profilo corrente", ""
        if path == "/azioni/accesso/recupera-password":
            return "Accesso", "Recupero password utente standard", ""
        if path == "/azioni/utenti/crea":
            return "Utenti", "Creazione nuovo utente standard", ""
        if path.startswith("/azioni/utenti/aggiorna/"):
            return "Utenti", "Aggiornamento dati utente standard", ""
        if path.startswith("/azioni/utenti/password/"):
            return "Utenti", "Reimpostazione password utente standard", ""
        if path.startswith("/azioni/utenti/stato/"):
            action = normalized(form_data, "azione_stato", "")
            detail = "Azione: disattivazione" if action == "disattiva" else "Azione: riattivazione"
            return "Utenti", "Aggiornamento stato utente standard", detail
        if path == "/azioni/backup/crea":
            return "Amministrazione", "Creazione backup completo del gestionale", ""
        if path.startswith("/azioni/backup/elimina/"):
            return "Amministrazione", "Eliminazione backup completo del gestionale", ""
        if path == "/azioni/aggiornamenti/installa":
            return "Amministrazione", "Installazione guidata aggiornamento", ""
        if path == "/azioni/associati/template-excel":
            return "Amministrazione", "Generazione modello Excel import associati", ""
        if path == "/azioni/associati/importa-excel":
            return "Amministrazione", "Importazione guidata associati da Excel", ""
        if path == "/azioni/associati/crea":
            return "Associati", "Creazione nuovo associato", plain_text(
                f"{normalized(form_data, 'cognome', '')} {normalized(form_data, 'nome', '')}"
            ).strip()
        if path.startswith("/azioni/crud/aggiorna/associati/"):
            return "Associati", "Aggiornamento anagrafica associato", ""
        if path.startswith("/azioni/crud/elimina/associati/"):
            return "Associati", "Eliminazione associato", ""
        if path == "/azioni/tesseramenti/crea":
            detail = work_year_detail or f"Anno sociale: {normalized(form_data, 'anno_sociale', '')}"
            return "Tesseramenti", "Registrazione nuovo tesseramento", detail
        if path == "/azioni/tesseramenti/pagamento":
            return "Tesseramenti", "Registrazione pagamento tesseramento", ""
        if path == "/azioni/corsi/crea":
            return "Corsi", "Creazione nuovo corso", normalized(form_data, "nome", "")
        if path == "/azioni/corsi/iscrizione":
            return "Corsi", "Registrazione iscrizione corso", work_year_detail
        if path in {"/azioni/corsi/pagamento-rata", "/azioni/corsi/pagamento-rate"}:
            return "Corsi", "Registrazione pagamento quote mensili corso", work_year_detail
        if path == "/azioni/corsi/genera-rate":
            month_value = normalized(form_data, "mese", "")
            month_text = month_label(int(month_value)) if month_value.isdigit() else month_value
            return "Corsi", "Generazione massiva quote mensili corsi", (
                f"Competenza: {month_text} {normalized(form_data, 'anno', '')}".strip()
            )
        if path == "/azioni/corsi/genera-rate-mancanti":
            return "Corsi", "Generazione automatica quote mensili mancanti", work_year_detail
        if path == "/azioni/campi-estivi/crea":
            return ESTATE_LABEL, "Creazione scheda annuale Campo estivo", work_year_detail
        if path == "/azioni/campi-estivi/iscrizione":
            return ESTATE_LABEL, "Registrazione iscrizione Campo estivo", work_year_detail
        if path == "/azioni/campi-estivi/pagamento":
            return ESTATE_LABEL, "Registrazione pagamento Campo estivo", work_year_detail
        if path == "/azioni/oratorio/iscrizione":
            return ORATORIO_LABEL, "Registrazione iscrizione Oratorio", work_year_detail
        if path == "/azioni/oratorio/pagamento":
            return ORATORIO_LABEL, "Registrazione pagamento Oratorio", work_year_detail
        if path == "/azioni/eventi/crea":
            return "Eventi", "Creazione nuovo evento", normalized(form_data, "nome", "")
        if path == "/azioni/eventi/iscrizione":
            return "Eventi", "Registrazione iscrizione evento", work_year_detail
        if path == "/azioni/eventi/pagamento":
            return "Eventi", "Registrazione pagamento evento", work_year_detail
        if path == "/azioni/pagamenti-multi-area/crea":
            return "Pagamenti", "Registrazione pagamento multi-area", work_year_detail
        if path.startswith("/azioni/pagamenti-multi-area/elimina/"):
            return "Pagamenti", "Annullamento pagamento multi-area", work_year_detail
        if path.startswith("/azioni/crud/aggiorna/"):
            entity_key = path.removeprefix("/azioni/crud/aggiorna/").split("/", 1)[0]
            return "Dati", f"Aggiornamento record {entity_key}", work_year_detail
        if path.startswith("/azioni/crud/elimina/"):
            entity_key = path.removeprefix("/azioni/crud/elimina/").split("/", 1)[0]
            return "Dati", f"Eliminazione record {entity_key}", work_year_detail
        return "Operazioni", f"Azione {path}", work_year_detail

    if path == "/api/report-share":
        channel = normalized(query_params, "channel", "")
        report_title = report_title_for_key(normalized(query_params, "report_key", ""), query_params)
        detail = f"Canale: {channel}" if channel else ""
        if channel == "whatsapp-group":
            detail = "Canale: gruppo WhatsApp"
        return "Report", f"Preparazione invio report {report_title}", detail

    if path.startswith("/export/excel/"):
        if path.startswith("/export/excel/associato/"):
            return "Report", "Export Excel dettaglio associato", work_year_detail
        report_key = path.removeprefix("/export/excel/")
        return "Report", f"Export Excel report {report_title_for_key(report_key, query_params)}", work_year_detail

    if path.startswith("/export/pdf/ricevuta/"):
        return "Ricevute", "Export PDF ricevuta di pagamento", ""

    if path.startswith("/export/pdf/"):
        if path.startswith("/export/pdf/associato/"):
            return "Report", "Export PDF dettaglio associato", work_year_detail
        report_key = path.removeprefix("/export/pdf/")
        return "Report", f"Export PDF report {report_title_for_key(report_key, query_params)}", work_year_detail

    if path.startswith("/download/backup/"):
        return "Amministrazione", "Download backup completo del gestionale", ""

    if path.startswith("/download/aggiornamento/"):
        return "Amministrazione", "Download pacchetto aggiornamento guidato", ""

    if path == "/":
        return "Navigazione", "Apertura dashboard", work_year_detail

    page_titles = {
        "/login": "Apertura pagina di accesso",
        "/recupera-password": "Apertura pagina recupero password",
        "/maschere/accesso": "Apertura profilo accesso",
        "/maschere/utenti": "Apertura maschera utenti",
        "/maschere/backup": "Apertura maschera backup",
        "/maschere/aggiornamenti": "Apertura maschera aggiornamenti",
        "/maschere/importa-associati": "Apertura maschera importa associati",
        "/maschere/guida": "Apertura tutorial",
        "/maschere/consiglio-direttivo": "Apertura maschera Consiglio Direttivo",
        "/maschere/tesserati": "Apertura maschera tesserati",
        "/maschere/associati": "Apertura maschera anagrafica associati",
        "/maschere/tesseramenti": "Apertura maschera tesseramenti",
        "/maschere/oratorio": "Apertura maschera Oratorio",
        "/maschere/corsi": "Apertura maschera corsi",
        "/maschere/campi-estivi": "Apertura maschera Campo estivo",
        "/maschere/eventi": "Apertura maschera eventi",
        "/maschere/pagamenti-multi-area": "Apertura maschera pagamenti",
        "/report/associati": "Apertura report Posizione associati",
        "/report/tesseramenti": "Apertura report Situazione tesseramenti",
        "/report/scadenze": "Apertura report Scadenze da incassare",
        "/report/corsi": "Apertura report Situazione corsi",
        "/report/oratorio": "Apertura report Situazione oratorio",
        "/report/campi-estivi": "Apertura report Situazione campo estivo",
        "/report/eventi": "Apertura report Situazione eventi",
        "/report/partecipanti": "Apertura report Partecipanti attivitÃ ",
        "/report/incassi": "Apertura report Incassi totali",
        "/report/registro-attivita": "Apertura report Registro attivita",
    }
    if path in page_titles:
        category = "Report" if path.startswith("/report/") else "Navigazione"
        return category, page_titles[path], work_year_detail
    if path.startswith("/report/associato/"):
        return "Report", "Apertura dettaglio associato", work_year_detail
    if path.startswith("/ricevute/"):
        return "Ricevute", "Apertura ricevuta di pagamento", work_year_detail
    if path.startswith("/modifica/"):
        return "Navigazione", "Apertura maschera modifica record", work_year_detail
    return "Navigazione", f"Accesso a {path}", work_year_detail


def should_log_activity(method: str, path: str) -> bool:
    if path.startswith("/static/"):
        return False
    if path in {"/api/codice-fiscale", "/api/codice-fiscale/calcola", "/api/comuni", "/api/cap"}:
        return False
    if path == "/favicon.ico":
        return False
    return (
        path.startswith("/azioni/")
        or path.startswith("/report/")
        or path.startswith("/export/")
        or path.startswith("/download/backup/")
        or path.startswith("/download/aggiornamento/")
        or path.startswith("/ricevute/")
        or path.startswith("/modifica/")
        or path.startswith("/maschere/")
        or path in {"/", "/login", "/recupera-password", "/api/report-share"}
    )


def log_request_activity(
    environ: dict,
    method: str,
    path: str,
    query_params: dict[str, str],
    form_data: dict[str, str],
    current_user: dict[str, object] | None,
    response_status: str,
    response_headers: list[tuple[str, str]],
) -> None:
    if not should_log_activity(method, path):
        return
    category, description, detail = describe_logged_activity(method, path, query_params, form_data)
    associato_id, associato_codice, associato_nominativo = resolve_activity_associato_context(path, query_params, form_data)
    ok_message, err_message = response_message_from_headers(response_headers)
    if err_message:
        outcome = f"Errore: {err_message}"
    elif ok_message:
        outcome = f"OK: {ok_message}"
    elif response_status.startswith("2"):
        outcome = "Completata"
    else:
        outcome = response_status
    record_activity(
        username=activity_username_for_log(current_user, path, form_data),
        path=path,
        method=method,
        description=description,
        category=category,
        outcome=outcome,
        work_year=request_work_year_for_log(query_params, form_data),
        detail=detail,
        associato_id=associato_id,
        associato_codice=associato_codice,
        associato_nominativo=associato_nominativo,
        ip_client=request_client_ip(environ),
        user_agent=request_user_agent(environ),
    )


def serve_static(start_response, path: str):
    relative_path = path.removeprefix("/static/")
    target = (STATIC_DIR / relative_path).resolve()

    if not target.is_file() or STATIC_DIR.resolve() not in target.parents:
        return not_found(start_response)

    mime_type, _ = mimetypes.guess_type(str(target))
    if mime_type == "text/css":
        content_type = "text/css; charset=utf-8"
    else:
        content_type = mime_type or "application/octet-stream"

    start_response("200 OK", [("Content-Type", content_type)])
    return [target.read_bytes()]


def file_download_response(
    start_response,
    file_path: Path,
    *,
    content_type: str = "application/octet-stream",
    download_name: str | None = None,
):
    if not file_path.is_file():
        return not_found(start_response)
    file_size = file_path.stat().st_size
    target_name = download_name or file_path.name
    start_response(
        "200 OK",
        [
            ("Content-Type", content_type),
            ("Content-Length", str(file_size)),
            ("Content-Disposition", f'attachment; filename="{target_name}"'),
        ],
    )

    def iterator():
        with file_path.open("rb") as handle:
            while True:
                chunk = handle.read(64 * 1024)
                if not chunk:
                    break
                yield chunk

    return iterator()


def not_found(start_response) -> list[bytes]:
    start_response("404 Not Found", [("Content-Type", "text/plain; charset=utf-8")])
    return [b"Pagina non trovata."]


def dispatch_request(
    environ,
    start_response,
    *,
    parsed_request: ParsedRequest | None = None,
    current_user: dict[str, object] | None = None,
):
    if parsed_request is None:
        path, method, query_params, form_data, form_files, request_cookies = parse_request(environ)
    else:
        path, method, query_params, form_data, form_files, request_cookies = parsed_request
    if current_user is None:
        current_user = current_user_from_cookies(request_cookies)

    if path.startswith("/static/"):
        return serve_static(start_response, path)

    if method == "GET" and path == "/login":
        if current_user is not None:
            return redirect(start_response, "/", extra_query=work_year_query(query_params))
        start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
        return [login_page(query_params)]

    if method == "GET" and path == "/recupera-password":
        if current_user is not None:
            return redirect(start_response, "/", extra_query=work_year_query(query_params))
        start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
        return [recover_password_page(query_params)]

    if method == "POST" and path in {
        "/azioni/accesso/primo-admin",
        "/azioni/accesso/login",
        "/azioni/accesso/logout",
        "/azioni/accesso/recupera-password",
    }:
        return handle_post(path, form_data, start_response, current_user, request_cookies, form_files)

    if current_user is None:
        if path.startswith("/api/"):
            return json_response(
                start_response,
                {"ok": False, "error": "Sessione non valida. Accedi di nuovo."},
                status="401 Unauthorized",
            )
        request_query = {key: value for key, value in query_params.items() if key not in {"ok", "err", "next"}}
        requested_target = login_destination(urlunsplit(("", "", path, urlencode(request_query), "")))
        login_query = {"next": requested_target}
        message = "Primo accesso: crea l'utente amministratore." if bootstrap_admin_required() else None
        return redirect(start_response, "/login", ok=message, extra_query=login_query)

    if method == "GET" and path == "/api/codice-fiscale":
        payload = decode_codice_fiscale_birth_data(query_params.get("value", ""))
        return json_response(start_response, {"found": bool(payload), **(payload or {})})

    if method == "GET" and path == "/api/codice-fiscale/calcola":
        payload = calculate_codice_fiscale(
            nome=query_params.get("nome", ""),
            cognome=query_params.get("cognome", ""),
            data_nascita=query_params.get("data_nascita", ""),
            sesso=query_params.get("sesso", ""),
            comune_nascita=query_params.get("comune_nascita", ""),
        )
        return json_response(start_response, {"found": bool(payload), **(payload or {})})

    if method == "GET" and path == "/api/comuni":
        payload = lookup_comune_details(query_params.get("nome", ""))
        return json_response(start_response, {"found": bool(payload), **(payload or {})})

    if method == "GET" and path == "/api/cap":
        payload = lookup_cap_details(query_params.get("value", ""))
        return json_response(start_response, {"found": bool(payload), **(payload or {})})

    if method == "GET" and path == "/api/report-share":
        try:
            report_key = query_params.get("report_key", "")
            if report_requires_admin(report_key) and not current_user.get("is_admin"):
                return json_response(
                    start_response,
                    {"ok": False, "error": "Solo l'amministratore puo usare questo report."},
                    status="403 Forbidden",
                )
            recipient_id = query_params.get("recipient_id", "")
            channel = query_params.get("channel", "")
            share_query = {
                key: value
                for key, value in query_params.items()
                if key not in {"report_key", "recipient_id", "channel"}
            }
            payload = build_report_share_payload(report_key, share_query, recipient_id, channel)
            if payload is None:
                return json_response(start_response, {"ok": False, "error": "Report non disponibile."}, status="400 Bad Request")
            return json_response(start_response, {"ok": True, **payload})
        except KeyError:
            return json_response(start_response, {"ok": False, "error": "Report non disponibile."}, status="404 Not Found")
        except ValueError as error:
            return json_response(start_response, {"ok": False, "error": str(error)}, status="400 Bad Request")

    if method == "GET" and path == "/api/pagamenti-multi-area/scadenze-aperte":
        try:
            work_year = current_work_year(query_params)
            associato_id = int(required(query_params, "associato_id", "Associato"))
            return json_response(
                start_response,
                {
                    "ok": True,
                    "options": scadenze_multi_area_payload_rows(work_year, associato_id),
                },
            )
        except ValueError as error:
            return json_response(start_response, {"ok": False, "error": str(error)}, status="400 Bad Request")

    if method == "GET" and path == "/api/tesseramenti/codice-anteprima":
        try:
            anno_sociale = int(required(query_params, "anno_sociale", "Anno sociale"))
            return json_response(
                start_response,
                {
                    "ok": True,
                    "codice_tesseramento": peek_next_tesseramento_code(anno_sociale),
                },
            )
        except ValueError as error:
            return json_response(start_response, {"ok": False, "error": str(error)}, status="400 Bad Request")

    if method == "GET" and path == "/api/pagamenti-multi-area/quote-corso-future":
        try:
            work_year = current_work_year(query_params)
            associato_id = int(required(query_params, "associato_id", "Associato"))
            iscrizione_corso_id = int(required(query_params, "iscrizione_corso_id", "Corso"))
            fine_competenza = required(query_params, "fine_competenza", "Competenza finale")
            payload = generate_multi_area_future_course_payload(
                work_year,
                associato_id,
                iscrizione_corso_id,
                fine_competenza,
            )
            return json_response(start_response, payload)
        except ValueError as error:
            return json_response(start_response, {"ok": False, "error": str(error)}, status="400 Bad Request")

    if method == "GET" and path.startswith("/export/pdf/ricevuta/"):
        receipt_path = path.removeprefix("/export/pdf/ricevuta/")
        if "/" not in receipt_path:
            return not_found(start_response)
        payment_type, payment_id = receipt_path.split("/", 1)
        return export_receipt_pdf(start_response, payment_type, payment_id)

    if method == "GET" and path.startswith("/export/excel/"):
        if path.startswith("/export/excel/associato/"):
            associato_id = path.removeprefix("/export/excel/associato/")
            if not associato_id.isdigit():
                return not_found(start_response)
            return export_associato_detail_excel(start_response, int(associato_id), query_params)
        if path == "/export/excel/tesserati":
            return export_tesserati_excel(start_response, query_params)
        if path == "/export/excel/anagrafica-associati":
            return export_associati_storici_excel(start_response, query_params)
        report_key = path.removeprefix("/export/excel/")
        if report_requires_admin(report_key) and not current_user.get("is_admin"):
            return redirect(
                start_response,
                "/",
                err="Solo l'amministratore puo esportare questo report.",
                extra_query=work_year_query(query_params),
            )
        return export_report_excel(start_response, report_key, query_params)

    if method == "GET" and path.startswith("/export/pdf/"):
        if path.startswith("/export/pdf/associato/"):
            associato_id = path.removeprefix("/export/pdf/associato/")
            if not associato_id.isdigit():
                return not_found(start_response)
            return export_associato_detail_pdf(start_response, int(associato_id), query_params)
        if path == "/export/pdf/tesserati":
            return export_tesserati_pdf(start_response, query_params)
        if path == "/export/pdf/anagrafica-associati":
            return export_associati_storici_pdf(start_response, query_params)
        report_key = path.removeprefix("/export/pdf/")
        if report_requires_admin(report_key) and not current_user.get("is_admin"):
            return redirect(
                start_response,
                "/",
                err="Solo l'amministratore puo esportare questo report.",
                extra_query=work_year_query(query_params),
            )
        return export_report_pdf(start_response, report_key, query_params)

    if method == "GET" and path.startswith("/download/backup/"):
        if not current_user.get("is_admin"):
            return redirect(
                start_response,
                "/",
                err="Solo l'amministratore puo scaricare i backup completi.",
                extra_query=work_year_query(query_params),
            )
        file_name = unquote(path.removeprefix("/download/backup/")).strip()
        if not re.fullmatch(r"[A-Za-z0-9._-]+\.zip", file_name):
            return not_found(start_response)
        return file_download_response(
            start_response,
            BACKUP_DIR / file_name,
            content_type="application/zip",
            download_name=file_name,
        )

    if method == "GET" and path.startswith("/download/tutorial/"):
        file_name = unquote(path.removeprefix("/download/tutorial/")).strip()
        course_key = file_name.removesuffix(".mp4")
        if not re.fullmatch(r"[A-Za-z0-9._-]+", course_key):
            return not_found(start_response)
        try:
            video_path, download_name = ensure_tutorial_video_file(course_key)
        except KeyError:
            return not_found(start_response)
        except ValueError as error:
            message = str(error).encode("utf-8", "replace")
            start_response(
                "500 Internal Server Error",
                [
                    ("Content-Type", "text/plain; charset=utf-8"),
                    ("Content-Length", str(len(message))),
                ],
            )
            return [message]
        return file_download_response(
            start_response,
            video_path,
            content_type="video/mp4",
            download_name=download_name,
        )

    if method == "GET" and path.startswith("/download/aggiornamento/"):
        if not current_user.get("is_admin"):
            return redirect(
                start_response,
                "/",
                err="Solo l'amministratore puo scaricare i pacchetti aggiornamento.",
                extra_query=work_year_query(query_params),
            )
        file_name = unquote(path.removeprefix("/download/aggiornamento/")).strip()
        if not re.fullmatch(r"[A-Za-z0-9._-]+\.zip", file_name):
            return not_found(start_response)
        return file_download_response(
            start_response,
            UPDATE_DIR / file_name,
            content_type="application/zip",
            download_name=file_name,
        )

    if method == "POST":
        return handle_post(path, form_data, start_response, current_user, request_cookies, form_files)

    if method == "GET" and path.startswith("/modifica/"):
        edit_path_value = path.removeprefix("/modifica/")
        if "/" not in edit_path_value:
            return not_found(start_response)
        entity_key, record_id = edit_path_value.split("/", 1)
        try:
            start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
            return [render_crud_edit_page(entity_key, int(record_id), query_params, current_user)]
        except (KeyError, ValueError):
            return not_found(start_response)

    if method == "GET" and path.startswith("/ricevute/"):
        receipt_path = path.removeprefix("/ricevute/")
        if "/" not in receipt_path:
            return not_found(start_response)
        payment_type, payment_id = receipt_path.split("/", 1)
        try:
            start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
            if payment_type in {"corsi-rate-gruppo", "multi-area-gruppo"}:
                return [receipt_page(payment_type, payment_id, query_params, current_user)]
            return [receipt_page(payment_type, int(payment_id), query_params, current_user)]
        except (KeyError, ValueError):
            return not_found(start_response)

    if method == "GET" and path.startswith("/report/associato/"):
        associato_id = path.removeprefix("/report/associato/")
        try:
            start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
            return [associato_report_page(int(associato_id), query_params, current_user)]
        except (KeyError, ValueError):
            return not_found(start_response)

    if method == "GET" and path.startswith("/maschere/utenti/gestione/"):
        if not current_user.get("is_admin"):
            return redirect(
                start_response,
                "/",
                err="Solo l'amministratore puo accedere alla gestione utenti.",
                extra_query=work_year_query(query_params),
            )
        user_id = path.removeprefix("/maschere/utenti/gestione/")
        try:
            start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
            return [gestione_utente_page(int(user_id), query_params, current_user)]
        except (KeyError, ValueError):
            return not_found(start_response)

    if path == "/maschere/utenti" and not current_user.get("is_admin"):
        return redirect(
            start_response,
            "/",
            err="Solo l'amministratore puo accedere alla gestione utenti.",
            extra_query=work_year_query(query_params),
        )

    if path in {"/maschere/backup", "/maschere/aggiornamenti", "/maschere/importa-associati"} and not current_user.get("is_admin"):
        return redirect(
            start_response,
            "/",
            err="Solo l'amministratore puo accedere a questa area amministrativa.",
            extra_query=work_year_query(query_params),
        )

    if path == "/report/registro-attivita" and not current_user.get("is_admin"):
        return redirect(
            start_response,
            "/",
            err="Solo l'amministratore puo accedere al registro attivita.",
            extra_query=work_year_query(query_params),
        )

    routes = {
        "/": dashboard_page,
        "/maschere/accesso": access_profile_page,
        "/maschere/utenti": utenti_page,
        "/maschere/backup": backup_page,
        "/maschere/aggiornamenti": aggiornamenti_page,
        "/maschere/importa-associati": importa_associati_page,
        "/maschere/guida": video_tutorial_page,
        "/maschere/consiglio-direttivo": consiglio_direttivo_page,
        "/maschere/tesserati": tesserati_page,
        "/maschere/associati": associati_page,
        "/maschere/tesseramenti": tesseramenti_page,
        "/maschere/oratorio": oratorio_page,
        "/maschere/corsi": corsi_page,
        "/maschere/campi-estivi": campi_estivi_page,
        "/maschere/eventi": eventi_page,
        "/maschere/pagamenti-multi-area": pagamenti_multi_area_page,
        "/report/associati": report_associati,
        "/report/tesseramenti": report_tesseramenti,
        "/report/scadenze": report_scadenze,
        "/report/corsi": report_corsi,
        "/report/oratorio": report_oratorio,
        "/report/campi-estivi": report_campi_estivi,
        "/report/eventi": report_eventi,
        "/report/partecipanti": report_partecipanti,
        "/report/storico-tesseramenti": report_storico_tesseramenti,
        "/report/incassi": report_incassi,
        "/report/registro-attivita": report_registro_attivita,
    }

    handler = routes.get(path)
    if handler is None:
        return not_found(start_response)

    start_response("200 OK", [("Content-Type", "text/html; charset=utf-8")])
    return [handler(query_params, current_user=current_user)]


def app(environ, start_response):
    parsed_request = parse_request(environ)
    path, method, query_params, form_data, _form_files, request_cookies = parsed_request
    current_user = current_user_from_cookies(request_cookies)
    captured_response: dict[str, object] = {
        "status": "",
        "headers": [],
    }

    def logging_start_response(status, headers, exc_info=None):
        captured_response["status"] = status
        captured_response["headers"] = list(headers)
        if exc_info is None:
            return start_response(status, headers)
        return start_response(status, headers, exc_info)

    response = dispatch_request(
        environ,
        logging_start_response,
        parsed_request=parsed_request,
        current_user=current_user,
    )
    try:
        log_request_activity(
            environ,
            method,
            path,
            query_params,
            form_data,
            current_user,
            str(captured_response.get("status", "")),
            list(captured_response.get("headers", [])),
        )
    except Exception:
        pass
    return response


ensure_schema()


class ThreadingWSGIServer(ThreadingMixIn, WSGIServer):
    daemon_threads = True


def main() -> None:
    try:
        auto_generate_course_rates_on_startup()
    except Exception:
        pass
    host = os.environ.get("ASSOCIAZIONE_HOST", "127.0.0.1")
    port = int(os.environ.get("ASSOCIAZIONE_PORT", "8000"))

    print(f"Database: {DB_PATH}")
    print(f"Web app disponibile su http://{host}:{port}")
    with make_server(host, port, app, server_class=ThreadingWSGIServer) as server:
        server.serve_forever()


if __name__ == "__main__":
    main()

